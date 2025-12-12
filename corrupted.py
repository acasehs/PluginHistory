canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_remediation_effectiveness_viz(self):
        """Create remediation effectiveness analysis"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Remediation")
        self.viz_frames['remediation'] = frame
        
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        if not self.lifecycle_df.empty:
            # Mean Time to Resolution by severity
            resolved_findings = self.lifecycle_df[self.lifecycle_df['status'] == 'Resolved']
            if not resolved_findings.empty:
                mttr_by_severity = resolved_findings.groupby('severity_text')['days_open'].mean().sort_values(ascending=False)
                colors = self.get_severity_colors()
                severity_colors = [colors.get(sev, '#6c757d') for sev in mttr_by_severity.index]
                
                bars = ax1.barh(range(len(mttr_by_severity)), mttr_by_severity.values, color=severity_colors)
                ax1.set_yticks(range(len(mttr_by_severity)))
                ax1.set_yticklabels(mttr_by_severity.index, color='white')
                ax1.set_title('Mean Time to Resolution (MTTR) by Severity', fontsize=12, fontweight='bold', color='white')
                ax1.set_xlabel('Average Days to Resolution', color='white')
                ax1.set_facecolor('#2b2b2b')
                ax1.tick_params(colors='white')
                
                for bar in bars:
                    width = bar.get_width()
                    ax1.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                            f'{width:.1f}', ha='left', va='center', color='white', fontweight='bold')
            
            # Remediation velocity over time
            resolved_by_month = resolved_findings.groupby(
                resolved_findings['last_seen'].dt.to_period('M')
            ).size().reset_index(name='resolved_count')
            resolved_by_month['last_seen'] = resolved_by_month['last_seen'].dt.to_timestamp()
            
            if not resolved_by_month.empty:
                ax2.plot(resolved_by_month['last_seen'], resolved_by_month['resolved_count'], 
                        marker='o', linewidth=2, color='#28a745')
                ax2.set_title('Remediation Velocity (Resolved/Month)', fontsize=12, fontweight='bold', color='white')
                ax2.set_xlabel('Month', color='white')
                ax2.set_ylabel('Findings Resolved', color='white')
                ax2.set_facecolor('#2b2b2b')
                ax2.tick_params(colors='white')
                ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
                plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # Recurring vs one-time findings
            recurring = len(self.lifecycle_df[self.lifecycle_df['reappearances'] > 0])
            one_time = len(self.lifecycle_df[self.lifecycle_df['reappearances'] == 0])
            
            bars = ax3.bar(['Recurring', 'One-time'], [recurring, one_time], color=['#dc3545', '#28a745'])
            ax3.set_title('Recurring vs One-time Findings', fontsize=12, fontweight='bold', color='white')
            ax3.set_ylabel('Number of Findings', color='white')
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            
            for bar in bars:
                height = bar.get_height()
                ax3.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                        f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
            
            # OPDIR due date compliance (if OPDIR data available)
            if not self.opdir_df.empty and 'opdir_days_to_due' in self.lifecycle_df.columns:
                opdir_findings = self.lifecycle_df[self.lifecycle_df['opdir_days_to_due'].notna()]
                if not opdir_findings.empty:
                    overdue = len(opdir_findings[opdir_findings['opdir_days_to_due'] < 0])
                    due_soon = len(opdir_findings[(opdir_findings['opdir_days_to_due'] >= 0) & 
                                                 (opdir_findings['opdir_days_to_due'] <= 30)])
                    on_track = len(opdir_findings[opdir_findings['opdir_days_to_due'] > 30])
                    
                    bars = ax4.bar(['Overdue', 'Due Soon (<30d)', 'On Track (>30d)'], 
                                  [overdue, due_soon, on_track], 
                                  color=['#dc3545', '#ffc107', '#28a745'])
                    ax4.set_title('OPDIR Due Date Compliance', fontsize=12, fontweight='bold', color='white')
                    ax4.set_ylabel('Number of Findings', color='white')
                    ax4.set_facecolor('#2b2b2b')
                    ax4.tick_params(colors='white')
                    
                    for bar in bars:
                        height = bar.get_height()
                        if height > 0:
                            ax4.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                                    f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
                else:
                    ax4.text(0.5, 0.5, 'No OPDIR due date data', ha='center', va='center', 
                            transform=ax4.transAxes, color='white', fontsize=12)
                    ax4.set_facecolor('#2b2b2b')
            else:
                ax4.text(0.5, 0.5, 'OPDIR data not available', ha='center', va='center', 
                        transform=ax4.transAxes, color='white', fontsize=12)
                ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_risk_heatmap_viz(self):
        """Create risk assessment heatmaps"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Risk Heatmaps")
        self.viz_frames['risk_heatmaps'] = frame
        
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        filtered_df = self.get_filtered_data(self.historical_df)
        
        if not filtered_df.empty:
            # Host risk matrix (findings count vs highest severity)
            host_risk = filtered_df.groupby('hostname').agg({
                'severity_value': 'max',
                'plugin_id': 'count'
            }).reset_index()
            host_risk.columns = ['hostname', 'max_severity', 'finding_count']
            
            # Create risk buckets
            severity_bins = [0, 1, 2, 3, 4, 5]
            severity_labels = ['Info', 'Low', 'Medium', 'High', 'Critical']
            count_bins = [0, 5, 15, 50, 100, float('inf')]
            count_labels = ['1-5', '6-15', '16-50', '51-100', '100+']
            
            host_risk['severity_bucket'] = pd.cut(host_risk['max_severity'], bins=severity_bins, 
                                                 labels=severity_labels, right=False)
            host_risk['count_bucket'] = pd.cut(host_risk['finding_count'], bins=count_bins, 
                                              labels=count_labels, right=False)
            
            # Create heatmap data
            heatmap_data = host_risk.groupby(['severity_bucket', 'count_bucket']).size().unstack(fill_value=0)
            
            if not heatmap_data.empty:
                import numpy as np
                im = ax1.imshow(heatmap_data.values, cmap='Reds', aspect='auto')
                ax1.set_xticks(range(len(heatmap_data.columns)))
                ax1.set_xticklabels(heatmap_data.columns, color='white')
                ax1.set_yticks(range(len(heatmap_data.index)))
                ax1.set_yticklabels(heatmap_data.index, color='white')
                ax1.set_title('Host Risk Matrix (Severity vs Finding Count)', fontsize=12, fontweight='bold', color='white')
                ax1.set_xlabel('Finding Count', color='white')
                ax1.set_ylabel('Max Severity', color='white')
                
                # Add text annotations
                for i in range(len(heatmap_data.index)):
                    for j in range(len(heatmap_data.columns)):
                        text = ax1.text(j, i, int(heatmap_data.iloc[i, j]),
                                       ha="center", va="center", color="white", fontweight='bold')
                
                ax1.set_facecolor('#2b2b2b')
                ax1.tick_params(colors='white')
            
            # Plugin family prevalence
            if 'family' in filtered_df.columns:
                family_counts = filtered_df['family'].value_counts().head(10)
                bars = ax2.barh(range(len(family_counts)), family_counts.values, color='#007bff')
                ax2.set_yticks(range(len(family_counts)))
                ax2.set_yticklabels([f[:20] + '...' if len(f) > 20 else f for f in family_counts.index], color='white')
                ax2.set_title('Top 10 Plugin Families', fontsize=12, fontweight='bold', color='white')
                ax2.set_xlabel('Number of Findings', color='white')
                ax2.set_facecolor('#2b2b2b')
                ax2.tick_params(colors='white')
                
                for bar in bars:
                    width = bar.get_width()
                    ax2.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                            f'{int(width)}', ha='left', va='center', color='white', fontweight='bold')
            
            # Time-based risk evolution
            risk_timeline = filtered_df.groupby(['scan_date', 'severity_text']).size().unstack(fill_value=0)
            if not risk_timeline.empty and 'Critical' in risk_timeline.columns:
                ax3.plot(risk_timeline.index, risk_timeline['Critical'], marker='o', 
                        linewidth=2, color='#dc3545', label='Critical')
                if 'High' in risk_timeline.columns:
                    ax3.plot(risk_timeline.index, risk_timeline['High'], marker='s', 
                            linewidth=2, color='#fd7e14', label='High')
                
                ax3.set_title('Critical/High Risk Trends', fontsize=12, fontweight='bold', color='white')
                ax3.set_xlabel('Scan Date', color='white')
                ax3.set_ylabel('Number of Findings', color='white')
                ax3.legend()
                ax3.set_facecolor('#2b2b2b')
                ax3.tick_params(colors='white')
                ax3.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
                plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # CVSS vs Exploitability (if exploit data available)
            if 'exploit_available' in filtered_df.columns and 'cvss3_base_score' in filtered_df.columns:
                exploit_data = filtered_df[filtered_df['cvss3_base_score'].notna()].copy()
                exploit_data['cvss_numeric'] = pd.to_numeric(exploit_data['cvss3_base_score'], errors='coerce')
                exploit_data = exploit_data[exploit_data['cvss_numeric'].notna()]
                
                if not exploit_data.empty:
                    exploitable = exploit_data[exploit_data['exploit_available'] == 'Yes']
                    not_exploitable = exploit_data[exploit_data['exploit_available'] == 'No']
                    
                    if not exploitable.empty:
                        ax4.scatter(exploitable['cvss_numeric'], [1]*len(exploitable), 
                                   alpha=0.6, color='#dc3545', label='Exploitable', s=50)
                    if not not_exploitable.empty:
                        ax4.scatter(not_exploitable['cvss_numeric'], [0]*len(not_exploitable), 
                                   alpha=0.6, color='#28a745', label='Not Exploitable', s=50)
                    
                    ax4.set_title('CVSS vs Exploitability', fontsize=12, fontweight='bold', color='white')
                    ax4.set_xlabel('CVSS v3 Score', color='white')
                    ax4.set_ylabel('Exploitable', color='white')
                    ax4.set_yticks([0, 1])
                    ax4.set_yticklabels(['No', 'Yes'], color='white')
                    ax4.legend()
                    ax4.set_facecolor('#2b2b2b')
                    ax4.tick_params(colors='white')
                else:
                    ax4.text(0.5, 0.5, 'No CVSS/exploit data', ha='center', va='center', 
                            transform=ax4.transAxes, color='white', fontsize=12)
                    ax4.set_facecolor('#2b2b2b')
            else:
                ax4.text(0.5, 0.5, 'Exploit data not available', ha='center', va='center', 
                        transform=ax4.transAxes, color='white', fontsize=12)
                ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_compliance_dashboard_viz(self):
        """Create compliance and executive dashboard"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Compliance")
        self.viz_frames['compliance'] = frame
        
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        filtered_df = self.get_filtered_data(self.historical_df)
        
        if not filtered_df.empty:
            # STIG severity distribution
            if 'stig_severity' in filtered_df.columns:
                stig_counts = filtered_df[filtered_df['stig_severity'].notna()]['stig_severity'].value_counts()
                if not stig_counts.empty:
                    colors_stig = ['#dc3545', '#fd7e14', '#ffc107']
                    ax1.pie(stig_counts.values, labels=stig_counts.index, autopct='%1.1f%%',
                           colors=colors_stig[:len(stig_counts)], startangle=90)
                    ax1.set_title('STIG Severity Distribution', fontsize=12, fontweight='bold', color='white')
                    ax1.set_facecolor('#2b2b2b')
                else:
                    ax1.text(0.5, 0.5, 'No STIG data available', ha='center', va='center', 
                            transform=ax1.transAxes, color='white', fontsize=12)
                    ax1.set_facecolor('#2b2b2b')
            
            # Patch age analysis
            if 'patch_publication_date' in filtered_df.columns:
                patch_data = filtered_df[filtered_df['patch_publication_date'].notna()].copy()
                if not patch_data.empty:
                    patch_data['patch_date'] = pd.to_datetime(patch_data['patch_publication_date'], errors='coerce')
                    patch_data = patch_data[patch_data['patch_date'].notna()]
                    
                    if not patch_data.empty:
                        current_date = pd.Timestamp.now()
                        patch_data['patch_age_days'] = (current_date - patch_data['patch_date']).dt.days
                        
                        age_bins = [0, 30, 90, 180, 365, float('inf')]
                        age_labels = ['0-30 days', '31-90 days', '91-180 days', '181-365 days', '365+ days']
                        patch_data['age_category'] = pd.cut(patch_data['patch_age_days'], bins=age_bins, 
                                                           labels=age_labels, right=False)
                        
                        age_counts = patch_data['age_category'].value_counts()
                        bars = ax2.bar(range(len(age_counts)), age_counts.values, color='#ffc107')
                        ax2.set_xticks(range(len(age_counts)))
                        ax2.set_xticklabels(age_counts.index, rotation=45, ha='right', color='white')
                        ax2.set_title('Patch Age Distribution', fontsize=12, fontweight='bold', color='white')
                        ax2.set_ylabel('Number of Findings', color='white')
                        ax2.set_facecolor('#2b2b2b')
                        ax2.tick_params(colors='white')
                        
                        for bar in bars:
                            height = bar.get_height()
                            if height > 0:
                                ax2.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                                        f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
                    else:
                        ax2.text(0.5, 0.5, 'No valid patch dates', ha='center', va='center', 
                                transform=ax2.transAxes, color='white', fontsize=12)
                        ax2.set_facecolor('#2b2b2b')
                else:
                    ax2.text(0.5, 0.5, 'No patch date data', ha='center', va='center', 
                            transform=ax2.transAxes, color='white', fontsize=12)
                    ax2.set_facecolor('#2b2b2b')
            
            # Executive scorecard (key metrics)
            total_findings = len(filtered_df)
            critical_high = len(filtered_df[filtered_df['severity_text'].isin(['Critical', 'High'])])
            with_cve = len(filtered_df[filtered_df['cves'].notna() & (filtered_df['cves'] != '')])
            
            metrics = ['Total Findings', 'Critical/High', 'With CVE']
            values = [total_findings, critical_high, with_cve]
            colors_metrics = ['#007bff', '#dc3545', '#fd7e14']
            
            bars = ax3.bar(metrics, values, color=colors_metrics)
            ax3.set_title('Executive Scorecard', fontsize=12, fontweight='bold', color='white')
            ax3.set_ylabel('Count', color='white')
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            
            for bar in bars:
                height = bar.get_height()
                ax3.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                        f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
            
            # SLA compliance tracking
            if not self.lifecycle_df.empty:
                # Define SLA targets by severity (example: Critical=7days, High=30days, Medium=90days)
                sla_targets = {'Critical': 7, 'High': 30, 'Medium': 90, 'Low': 180}
                
                compliance_data = []
                for severity, target_days in sla_targets.items():
                    severity_findings = self.lifecycle_df[self.lifecycle_df['severity_text'] == severity]
                    if not severity_findings.empty:
                        within_sla = len(severity_findings[severity_findings['days_open'] <= target_days])
                        total_severity = len(severity_findings)
                        compliance_rate = (within_sla / total_severity) * 100 if total_severity > 0 else 0
                        compliance_data.append({'severity': severity, 'compliance': compliance_rate})
                
                if compliance_data:
                    compliance_df = pd.DataFrame(compliance_data)
                    colors = self.get_severity_colors()
                    severity_colors = [colors.get(sev, '#6c757d') for sev in compliance_df['severity']]
                    
                    bars = ax4.bar(compliance_df['severity'], compliance_df['compliance'], color=severity_colors)
                    ax4.set_title('SLA Compliance by Severity', fontsize=12, fontweight='bold', color='white')
                    ax4.set_ylabel('Compliance Rate (%)', color='white')
                    ax4.set_ylim(0, 100)
                    ax4.axhline(y=80, color='#ffc107', linestyle='--', alpha=0.7, label='80% Target')
                    ax4.legend()
                    ax4.set_facecolor('#2b2b2b')
                    ax4.tick_params(colors='white')
                    
                    for bar in bars:
                        height = bar.get_height()
                        ax4.text(bar.get_x() + bar.get_width()/2., height + 2,
                                f'{height:.1f}%', ha='center', va='bottom', color='white', fontweight='bold')
                else:
                    ax4.text(0.5, 0.5, 'No SLA data available', ha='center', va='center', 
                            transform=ax4.transAxes, color='white', fontsize=12)
                    ax4.set_facecolor('#2b2b2b')
            else:
                ax4.text(0.5, 0.5, 'No lifecycle data available', ha='center', va='center', 
                        transform=ax4.transAxes, color='white', fontsize=12)
                ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True).set_xticks(range(len(age_counts)))
                        ax2.set_xticklabels(age_counts.index, rotation=45, ha='right', color='white')
                        ax2"""
Enhanced Nessus Historical Analysis and Visualization System
Tracks vulnerability findings and host presence across multiple scans over time.
"""

import os
import json
import re
import sqlite3
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any
from collections import defaultdict
import pandas as pd  # pip install pandas
import tkinter as tk  # pip install tk
from tkinter import filedialog, messagebox, ttk
import matplotlib.pyplot as plt  # pip install matplotlib
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates

try:
    from archive_extraction import extract_nested_archives, find_files_by_extension, cleanup_temp_directory
    from nessus_parser import parse_multiple_nessus_files
    from plugin_database import load_plugins_database
    from data_processing import enrich_findings_with_severity
except ImportError:
    print("Warning: Could not import all modules. Ensure they are in the same directory.")


def load_and_process_opdir(opdir_file_path: str) -> pd.DataFrame:
    """
    Load and process OPDIR spreadsheet data.
    
    Args:
        opdir_file_path: Path to OPDIR Excel/CSV file
        
    Returns:
        DataFrame with processed OPDIR data
    """
    try:
        # Try to load as Excel first, then CSV
        if opdir_file_path.lower().endswith('.xlsx') or opdir_file_path.lower().endswith('.xls'):
            opdir_df = pd.read_excel(opdir_file_path)
        else:
            opdir_df = pd.read_csv(opdir_file_path)
        
        print(f"Loaded OPDIR file with {len(opdir_df)} records")
        
        # Standardize column names (case insensitive matching)
        column_mapping = {}
        for col in opdir_df.columns:
            col_lower = col.lower().strip()
            if 'opdir' in col_lower and 'number' in col_lower:
                column_mapping[col] = 'opdir_number'
            elif 'iava' in col_lower or 'iav' in col_lower:
                column_mapping[col] = 'iavx_reference'
            elif 'subject' in col_lower:
                column_mapping[col] = 'subject'
            elif 'release' in col_lower and 'date' in col_lower:
                column_mapping[col] = 'release_date'
            elif 'acknowledge' in col_lower and 'date' in col_lower:
                column_mapping[col] = 'acknowledge_date'
            elif 'poa' in col_lower and 'due' in col_lower:
                column_mapping[col] = 'poam_due_date'
            elif 'final' in col_lower and 'due' in col_lower:
                column_mapping[col] = 'final_due_date'
        
        opdir_df = opdir_df.rename(columns=column_mapping)
        
        # Parse OPDIR number to extract year
        if 'opdir_number' in opdir_df.columns:
            opdir_df['opdir_year'] = opdir_df['opdir_number'].astype(str).str.extract(r'-(\d{2})
    """
    Extract scan date from filename using common patterns.
    
    Args:
        filename: Name of the file
        
    Returns:
        datetime object or None
    """
    # Common date patterns in filenames
    patterns = [
        r'(\d{4})[-_](\d{2})[-_](\d{2})',  # YYYY-MM-DD or YYYY_MM_DD
        r'(\d{2})[-_](\d{2})[-_](\d{4})',  # MM-DD-YYYY or MM_DD_YYYY
        r'(\d{8})',  # YYYYMMDD
        r'(\d{6})',  # YYMMDD or MMDDYY
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            try:
                if len(match.groups()) == 3:
                    # YYYY-MM-DD or MM-DD-YYYY
                    parts = match.groups()
                    if len(parts[0]) == 4:  # YYYY-MM-DD
                        return datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                    else:  # MM-DD-YYYY
                        return datetime(int(parts[2]), int(parts[0]), int(parts[1]))
                elif len(match.group(1)) == 8:  # YYYYMMDD
                    date_str = match.group(1)
                    return datetime.strptime(date_str, '%Y%m%d')
                elif len(match.group(1)) == 6:  # Try both YYMMDD and MMDDYY
                    date_str = match.group(1)
                    try:
                        return datetime.strptime(date_str, '%y%m%d')
                    except ValueError:
                        return datetime.strptime(date_str, '%m%d%y')
            except (ValueError, IndexError):
                continue
    
    return None


def extract_scan_date_from_nessus(nessus_file: str) -> Optional[datetime]:
    """
    Extract scan date from .nessus file content.
    
    Args:
        nessus_file: Path to .nessus file
        
    Returns:
        datetime object or None
    """
    try:
        import xml.etree.ElementTree as ET
        tree = ET.parse(nessus_file)
        root = tree.getroot()
        
        # Try to find scan end date
        policy_elem = root.find(".//Policy/policyName")
        if policy_elem is not None and policy_elem.text:
            date = extract_scan_date_from_filename(policy_elem.text)
            if date:
                return date
        
        # Try Report name attribute
        report_elem = root.find(".//Report")
        if report_elem is not None and 'name' in report_elem.attrib:
            date = extract_scan_date_from_filename(report_elem.attrib['name'])
            if date:
                return date
        
        # Try HOST_END timestamps
        host_end = root.find(".//tag[@name='HOST_END']")
        if host_end is not None and host_end.text:
            try:
                timestamp = int(host_end.text)
                return datetime.fromtimestamp(timestamp)
            except (ValueError, TypeError):
                pass
        
    except Exception as e:
        print(f"Error extracting date from {nessus_file}: {e}")
    
    return None


def create_host_presence_analysis(historical_df: pd.DataFrame) -> pd.DataFrame:
    """
    Analyze host presence across scans to identify missing hosts.
    
    Args:
        historical_df: DataFrame with historical findings
        
    Returns:
        DataFrame with host presence analysis
    """
    if historical_df.empty:
        return pd.DataFrame()
    
    # Get all unique combinations of scan dates and hosts
    scan_dates = sorted(historical_df['scan_date'].unique())
    all_hosts = historical_df.groupby(['hostname', 'ip_address']).size().reset_index(name='count')
    
    presence_records = []
    
    for _, host_row in all_hosts.iterrows():
        hostname = host_row['hostname']
        ip_address = host_row['ip_address']
        
        # Get scan dates where this host appeared
        host_scans = historical_df[
            (historical_df['hostname'] == hostname) & 
            (historical_df['ip_address'] == ip_address)
        ]['scan_date'].unique()
        
        first_seen = min(host_scans)
        last_seen = max(host_scans)
        total_scans = len(scan_dates)
        present_scans = len(host_scans)
        missing_scans = total_scans - present_scans
        
        # Identify missing scan periods
        missing_periods = []
        for scan_date in scan_dates:
            if scan_date not in host_scans:
                missing_periods.append(scan_date.strftime('%Y-%m-%d'))
        
        # Determine status
        if last_seen == max(scan_dates):
            status = 'Active'
        else:
            status = 'Missing'
        
        # Calculate presence percentage
        presence_percentage = (present_scans / total_scans) * 100 if total_scans > 0 else 0
        
        presence_records.append({
            'hostname': hostname,
            'ip_address': ip_address,
            'first_seen': first_seen,
            'last_seen': last_seen,
            'total_scans_available': total_scans,
            'scans_present': present_scans,
            'scans_missing': missing_scans,
            'presence_percentage': round(presence_percentage, 1),
            'status': status,
            'missing_scan_dates': ', '.join(missing_periods) if missing_periods else 'None'
        })
    
    presence_df = pd.DataFrame(presence_records)
    presence_df = presence_df.sort_values(['status', 'presence_percentage'], ascending=[True, False])
    
    return presence_df


def analyze_scan_changes(historical_df: pd.DataFrame) -> pd.DataFrame:
    """
    Analyze changes between consecutive scans (hosts added/removed).
    
    Args:
        historical_df: DataFrame with historical findings
        
    Returns:
        DataFrame with scan-to-scan changes
    """
    if historical_df.empty:
        return pd.DataFrame()
    
    scan_dates = sorted(historical_df['scan_date'].unique())
    change_records = []
    
    for i in range(1, len(scan_dates)):
        prev_scan = scan_dates[i-1]
        curr_scan = scan_dates[i]
        
        # Get hosts in each scan
        prev_hosts = set(historical_df[historical_df['scan_date'] == prev_scan]['hostname'].unique())
        curr_hosts = set(historical_df[historical_df['scan_date'] == curr_scan]['hostname'].unique())
        
        # Find changes
        added_hosts = curr_hosts - prev_hosts
        removed_hosts = prev_hosts - curr_hosts
        unchanged_hosts = curr_hosts & prev_hosts
        
        change_records.append({
            'scan_date': curr_scan,
            'previous_scan': prev_scan,
            'hosts_added': len(added_hosts),
            'hosts_removed': len(removed_hosts),
            'hosts_unchanged': len(unchanged_hosts),
            'total_hosts_current': len(curr_hosts),
            'total_hosts_previous': len(prev_hosts),
            'net_change': len(curr_hosts) - len(prev_hosts),
            'added_host_list': ', '.join(sorted(added_hosts)) if added_hosts else 'None',
            'removed_host_list': ', '.join(sorted(removed_hosts)) if removed_hosts else 'None'
        })
    
    changes_df = pd.DataFrame(change_records)
    return changes_df


def load_existing_database(db_path: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Load existing historical database.
    
    Args:
        db_path: Path to SQLite database file
        
    Returns:
        Tuple of (historical_df, lifecycle_df, host_presence_df, scan_changes_df)
    """
    try:
        if not os.path.exists(db_path):
            print(f"Database file {db_path} does not exist")
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        
        conn = sqlite3.connect(db_path)
        
        # Load historical findings
        try:
            historical_df = pd.read_sql_query("SELECT * FROM historical_findings", conn)
            historical_df['scan_date'] = pd.to_datetime(historical_df['scan_date'])
            print(f"Loaded {len(historical_df)} historical findings from database")
        except Exception as e:
            print(f"Error loading historical_findings: {e}")
            historical_df = pd.DataFrame()
        
        # Load lifecycle analysis
        try:
            lifecycle_df = pd.read_sql_query("SELECT * FROM finding_lifecycle", conn)
            lifecycle_df['first_seen'] = pd.to_datetime(lifecycle_df['first_seen'])
            lifecycle_df['last_seen'] = pd.to_datetime(lifecycle_df['last_seen'])
            print(f"Loaded {len(lifecycle_df)} lifecycle records from database")
        except Exception as e:
            print(f"Error loading finding_lifecycle: {e}")
            lifecycle_df = pd.DataFrame()
        
        # Load host presence analysis
        try:
            host_presence_df = pd.read_sql_query("SELECT * FROM host_presence", conn)
            host_presence_df['first_seen'] = pd.to_datetime(host_presence_df['first_seen'])
            host_presence_df['last_seen'] = pd.to_datetime(host_presence_df['last_seen'])
            print(f"Loaded {len(host_presence_df)} host presence records from database")
        except Exception as e:
            print(f"Error loading host_presence: {e}")
            host_presence_df = pd.DataFrame()
        
        # Load scan changes
        try:
            scan_changes_df = pd.read_sql_query("SELECT * FROM scan_changes", conn)
            scan_changes_df['scan_date'] = pd.to_datetime(scan_changes_df['scan_date'])
            scan_changes_df['previous_scan'] = pd.to_datetime(scan_changes_df['previous_scan'])
            print(f"Loaded {len(scan_changes_df)} scan change records from database")
        except Exception as e:
            print(f"Error loading scan_changes: {e}")
            scan_changes_df = pd.DataFrame()
        
        conn.close()
        return historical_df, lifecycle_df, host_presence_df, scan_changes_df
        
    except Exception as e:
        print(f"Error loading existing database: {e}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()


def check_for_duplicates(new_df: pd.DataFrame, existing_df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    """
    Check for and remove duplicate scans based on scan_date and scan_file.
    
    Args:
        new_df: New findings DataFrame
        existing_df: Existing findings DataFrame
        
    Returns:
        Tuple of (filtered_new_df, duplicate_count)
    """
    if existing_df.empty:
        return new_df, 0
    
    # Create composite key for duplicate detection
    new_df['composite_key'] = new_df['scan_date'].astype(str) + '|' + new_df['scan_file'].astype(str)
    existing_df['composite_key'] = existing_df['scan_date'].astype(str) + '|' + existing_df['scan_file'].astype(str)
    
    # Find duplicates
    duplicate_keys = set(existing_df['composite_key'].unique())
    new_unique_mask = ~new_df['composite_key'].isin(duplicate_keys)
    
    filtered_new_df = new_df[new_unique_mask].copy()
    duplicate_count = len(new_df) - len(filtered_new_df)
    
    # Remove the temporary composite key
    filtered_new_df = filtered_new_df.drop('composite_key', axis=1)
    
    return filtered_new_df, duplicate_count


def process_historical_scans(archive_paths: List[str], plugins_db_path: Optional[str] = None, 
                           existing_db_path: Optional[str] = None, include_info: bool = False) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Process multiple Nessus archives and track findings history with enhanced features.
    
    Args:
        archive_paths: List of paths to Nessus archives or .nessus files
        plugins_db_path: Optional path to plugins database
        existing_db_path: Optional path to existing database to update
        include_info: Whether to include Info-level findings in analysis (default: False)
        
    Returns:
        Tuple of (historical_df, lifecycle_df, host_presence_df, scan_changes_df)
    """
    import tempfile
    
    # Load existing database if provided
    existing_historical_df = pd.DataFrame()
    if existing_db_path and os.path.exists(existing_db_path):
        print(f"Loading existing database from {existing_db_path}")
        existing_historical_df, _, _, _ = load_existing_database(existing_db_path)
    
    # Load plugins database if provided
    plugins_dict = None
    if plugins_db_path:
        plugins_dict = load_plugins_database(plugins_db_path)
    
    all_historical_findings = []
    
    for archive_path in archive_paths:
        print(f"\n{'='*60}")
        print(f"Processing: {os.path.basename(archive_path)}")
        print(f"{'='*60}")
        
        temp_dir = None
        nessus_files = []
        
        try:
            # Determine scan date from filename first
            scan_date = extract_scan_date_from_filename(os.path.basename(archive_path))
            
            # Handle different input types
            if archive_path.lower().endswith('.zip'):
                temp_dir = tempfile.mkdtemp()
                extract_nested_archives(archive_path, temp_dir)
                nessus_files = find_files_by_extension(temp_dir, '.nessus')
            elif archive_path.lower().endswith('.nessus'):
                nessus_files = [archive_path]
            else:
                print(f"Unsupported file format: {archive_path}")
                continue
            
            if not nessus_files:
                print(f"No .nessus files found in {archive_path}")
                continue
            
            # If no date from filename, try extracting from first .nessus file
            if not scan_date and nessus_files:
                scan_date = extract_scan_date_from_nessus(nessus_files[0])
            
            # Fallback to file modification time
            if not scan_date:
                scan_date = datetime.fromtimestamp(os.path.getmtime(archive_path))
                print(f"Warning: Using file modification time for scan date: {scan_date.strftime('%Y-%m-%d')}")
            else:
                print(f"Scan date: {scan_date.strftime('%Y-%m-%d')}")
            
            # Parse nessus files with improved host display
            findings_df, host_summary_df = parse_multiple_nessus_files(nessus_files, plugins_dict)
            
            if not findings_df.empty:
                # Add scan metadata
                findings_df['scan_date'] = scan_date
                findings_df['scan_file'] = os.path.basename(archive_path)
                
                # Enrich with severity
                findings_df = enrich_findings_with_severity(findings_df)
                
                all_historical_findings.append(findings_df)
                
                # Display with hostname (IP) format
                for _, host in host_summary_df.iterrows():
                    hostname = host.get('hostname', 'Unknown')
                    ip = host.get('host_name', 'Unknown')
                    print(f"  Processed: {hostname} ({ip})")
                print(f"Extracted {len(findings_df)} findings from {len(host_summary_df)} hosts")
            
        finally:
            if temp_dir:
                cleanup_temp_directory(temp_dir)
    
    if not all_historical_findings:
        print("No findings extracted from any archive")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    # Combine all new findings
    new_historical_df = pd.concat(all_historical_findings, ignore_index=True)
    
    # Check for duplicates if we have existing data
    duplicate_count = 0
    if not existing_historical_df.empty:
        print(f"\nChecking for duplicates against existing database...")
        new_historical_df, duplicate_count = check_for_duplicates(new_historical_df, existing_historical_df)
        if duplicate_count > 0:
            print(f"Filtered out {duplicate_count} duplicate findings")
    
    # Combine with existing data
    if not existing_historical_df.empty and not new_historical_df.empty:
        combined_df = pd.concat([existing_historical_df, new_historical_df], ignore_index=True)
    elif not existing_historical_df.empty:
        combined_df = existing_historical_df
    else:
        combined_df = new_historical_df
    
    combined_df = combined_df.sort_values('scan_date')
    
    print(f"\n{'='*60}")
    print(f"Total findings across all scans: {len(combined_df)}")
    if not new_historical_df.empty:
        print(f"New findings added: {len(new_historical_df)}")
    if duplicate_count > 0:
        print(f"Duplicate findings filtered: {duplicate_count}")
    print(f"Date range: {combined_df['scan_date'].min()} to {combined_df['scan_date'].max()}")
    print(f"{'='*60}\n")
    
    # Perform analysis (filter Info findings for analysis only, not database storage)
    analysis_df = combined_df.copy()
    if not include_info:
        analysis_df = analysis_df[analysis_df['severity_text'] != 'Info']
        print(f"Filtered out {len(combined_df) - len(analysis_df)} Info-level findings from analysis")
    
    lifecycle_df = analyze_finding_lifecycle(analysis_df)
    host_presence_df = create_host_presence_analysis(analysis_df)
    scan_changes_df = analyze_scan_changes(analysis_df)
    
    return combined_df, lifecycle_df, host_presence_df, scan_changes_df


def analyze_finding_lifecycle(historical_df: pd.DataFrame) -> pd.DataFrame:
    """
    Analyze the lifecycle of findings (first seen, last seen, resolved, reappeared).
    
    Args:
        historical_df: DataFrame with historical findings
        
    Returns:
        DataFrame with lifecycle analysis
    """
    if historical_df.empty:
        return pd.DataFrame()
    
    # Create unique finding identifier
    historical_df['finding_key'] = (
        historical_df['hostname'].astype(str) + '|' + 
        historical_df['plugin_id'].astype(str)
    )
    
    lifecycle_records = []
    
    # Group by finding_key
    for finding_key, group in historical_df.groupby('finding_key'):
        hostname, plugin_id = finding_key.split('|')
        
        # Sort by scan date
        group = group.sort_values('scan_date')
        
        scan_dates = group['scan_date'].tolist()
        scan_files = group['scan_file'].tolist()
        
        # Get finding details from most recent observation
        latest = group.iloc[-1]
        
        first_seen = scan_dates[0]
        last_seen = scan_dates[-1]
        total_observations = len(scan_dates)
        
        # Check for gaps (resolved then reappeared)
        gaps = []
        reappearances = 0

        if len(scan_dates) > 1:
            for i in range(1, len(scan_dates)):
                days_gap = (scan_dates[i] - scan_dates[i-1]).days
                if days_gap > 45:  # More than 45 days gap suggests resolution
                    gaps.append({
                        'resolved_after': scan_dates[i-1].strftime('%Y-%m-%d'),
                        'reappeared_on': scan_dates[i].strftime('%Y-%m-%d'),
                        'days_resolved': days_gap
                    })
                    reappearances += 1
        
        # Determine current status
        if latest['scan_date'] == historical_df['scan_date'].max():
            status = 'Active'
        else:
            status = 'Resolved'
        
        lifecycle_records.append({
            'hostname': hostname,
            'ip_address': latest.get('ip_address', ''),
            'plugin_id': plugin_id,
            'plugin_name': latest['name'],
            'severity_text': latest.get('severity_text', 'Unknown'),
            'severity_value': latest.get('severity_value', 0),
            'first_seen': first_seen,
            'last_seen': last_seen,
            'days_open': (last_seen - first_seen).days,
            'total_observations': total_observations,
            'reappearances': reappearances,
            'status': status,
            'gap_details': json.dumps(gaps) if gaps else '',
            'cvss3_base_score': latest.get('cvss3_base_score'),
            'cves': latest.get('cves', ''),
            'iavx': latest.get('iavx', '')
        })
    
    lifecycle_df = pd.DataFrame(lifecycle_records)
    lifecycle_df = lifecycle_df.sort_values(['severity_value', 'days_open'], ascending=[False, False])
    
    return lifecycle_df


def export_to_sqlite(historical_df: pd.DataFrame, lifecycle_df: pd.DataFrame, 
                     host_presence_df: pd.DataFrame, scan_changes_df: pd.DataFrame,
                     db_path: str) -> None:
    """
    Export data to SQLite database with enhanced tables.
    
    Args:
        historical_df: Historical findings DataFrame
        lifecycle_df: Lifecycle analysis DataFrame
        host_presence_df: Host presence analysis DataFrame
        scan_changes_df: Scan changes DataFrame
        db_path: Path to SQLite database file
    """
    try:
        conn = sqlite3.connect(db_path)
        
        # Export all DataFrames
        historical_df.to_sql('historical_findings', conn, if_exists='replace', index=False)
        lifecycle_df.to_sql('finding_lifecycle', conn, if_exists='replace', index=False)
        host_presence_df.to_sql('host_presence', conn, if_exists='replace', index=False)
        scan_changes_df.to_sql('scan_changes', conn, if_exists='replace', index=False)
        
        # Create indexes for better query performance
        cursor = conn.cursor()
        
        # Historical findings indexes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_historical_hostname ON historical_findings(hostname)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_historical_plugin ON historical_findings(plugin_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_historical_date ON historical_findings(scan_date)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_historical_ip ON historical_findings(ip_address)')
        
        # Lifecycle indexes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_lifecycle_hostname ON finding_lifecycle(hostname)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_lifecycle_plugin ON finding_lifecycle(plugin_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_lifecycle_status ON finding_lifecycle(status)')
        
        # Host presence indexes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_host_presence_hostname ON host_presence(hostname)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_host_presence_status ON host_presence(status)')
        
        # Scan changes indexes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_scan_changes_date ON scan_changes(scan_date)')
        
        conn.commit()
        conn.close()
        
        print(f"Successfully exported to SQLite database: {db_path}")
        
    except Exception as e:
        print(f"Error exporting to SQLite: {e}")
        import traceback
        traceback.print_exc()


def export_to_json(historical_df: pd.DataFrame, lifecycle_df: pd.DataFrame, 
                   host_presence_df: pd.DataFrame, scan_changes_df: pd.DataFrame,
                   json_path: str) -> None:
    """
    Export data to JSON file with enhanced structure.
    
    Args:
        historical_df: Historical findings DataFrame
        lifecycle_df: Lifecycle analysis DataFrame
        host_presence_df: Host presence analysis DataFrame
        scan_changes_df: Scan changes DataFrame
        json_path: Path to JSON output file
    """
    try:
        # Convert datetime columns to ISO format strings before export
        historical_export = historical_df.copy()
        lifecycle_export = lifecycle_df.copy()
        host_presence_export = host_presence_df.copy()
        scan_changes_export = scan_changes_df.copy()
        
        # Convert datetime columns in historical_df
        if 'scan_date' in historical_export.columns:
            historical_export['scan_date'] = historical_export['scan_date'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Convert datetime columns in lifecycle_df
        if 'first_seen' in lifecycle_export.columns:
            lifecycle_export['first_seen'] = lifecycle_export['first_seen'].dt.strftime('%Y-%m-%d %H:%M:%S')
        if 'last_seen' in lifecycle_export.columns:
            lifecycle_export['last_seen'] = lifecycle_export['last_seen'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Convert datetime columns in host_presence_df
        if 'first_seen' in host_presence_export.columns:
            host_presence_export['first_seen'] = host_presence_export['first_seen'].dt.strftime('%Y-%m-%d %H:%M:%S')
        if 'last_seen' in host_presence_export.columns:
            host_presence_export['last_seen'] = host_presence_export['last_seen'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Convert datetime columns in scan_changes_df
        if 'scan_date' in scan_changes_export.columns:
            scan_changes_export['scan_date'] = scan_changes_export['scan_date'].dt.strftime('%Y-%m-%d %H:%M:%S')
        if 'previous_scan' in scan_changes_export.columns:
            scan_changes_export['previous_scan'] = scan_changes_export['previous_scan'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Replace NaN with None for JSON compatibility
        historical_export = historical_export.where(pd.notnull(historical_export), None)
        lifecycle_export = lifecycle_export.where(pd.notnull(lifecycle_export), None)
        host_presence_export = host_presence_export.where(pd.notnull(host_presence_export), None)
        scan_changes_export = scan_changes_export.where(pd.notnull(scan_changes_export), None)
        
        export_data = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'total_findings': len(historical_df),
                'unique_findings': len(lifecycle_df),
                'total_hosts': len(host_presence_df),
                'total_scans': len(scan_changes_df) + 1 if len(scan_changes_df) > 0 else len(historical_df['scan_date'].unique()) if not historical_df.empty else 0,
                'date_range': {
                    'start': historical_df['scan_date'].min().isoformat() if not historical_df.empty and 'scan_date' in historical_df.columns else None,
                    'end': historical_df['scan_date'].max().isoformat() if not historical_df.empty and 'scan_date' in historical_df.columns else None
                }
            },
            'historical_findings': historical_export.to_dict('records'),
            'lifecycle_analysis': lifecycle_export.to_dict('records'),
            'host_presence': host_presence_export.to_dict('records'),
            'scan_changes': scan_changes_export.to_dict('records')
        }
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        print(f"Successfully exported to JSON: {json_path}")
        
    except Exception as e:
        print(f"Error exporting to JSON: {e}")
        import traceback
        traceback.print_exc()


def export_json_schema(schema_path: str) -> None:
    """
    Export JSON schema definition for the export format.
    
    Args:
        schema_path: Path to save the JSON schema file
    """
    schema = {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "title": "Nessus Historical Analysis Export",
        "type": "object",
        "properties": {
            "metadata": {
                "type": "object",
                "properties": {
                    "export_date": {"type": "string", "format": "date-time"},
                    "total_findings": {"type": "integer"},
                    "unique_findings": {"type": "integer"},
                    "total_hosts": {"type": "integer"},
                    "total_scans": {"type": "integer"},
                    "date_range": {
                        "type": "object",
                        "properties": {
                            "start": {"type": ["string", "null"], "format": "date-time"},
                            "end": {"type": ["string", "null"], "format": "date-time"}
                        }
                    }
                }
            },
            "historical_findings": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "plugin_id": {"type": "string"},
                        "hostname": {"type": "string"},
                        "ip_address": {"type": "string"},
                        "scan_date": {"type": "string", "format": "date-time"},
                        "scan_file": {"type": "string"},
                        "name": {"type": "string"},
                        "severity_text": {"type": "string"},
                        "severity_value": {"type": "integer"},
                        "cvss3_base_score": {"type": ["string", "null"]},
                        "cvss2_base_score": {"type": ["string", "null"]},
                        "cves": {"type": ["string", "null"]},
                        "iavx": {"type": ["string", "null"]}
                    }
                }
            },
            "lifecycle_analysis": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "hostname": {"type": "string"},
                        "ip_address": {"type": "string"},
                        "plugin_id": {"type": "string"},
                        "plugin_name": {"type": "string"},
                        "severity_text": {"type": "string"},
                        "severity_value": {"type": "integer"},
                        "first_seen": {"type": "string", "format": "date-time"},
                        "last_seen": {"type": "string", "format": "date-time"},
                        "days_open": {"type": "integer"},
                        "total_observations": {"type": "integer"},
                        "reappearances": {"type": "integer"},
                        "status": {"type": "string", "enum": ["Active", "Resolved"]},
                        "gap_details": {"type": "string"}
                    }
                }
            },
            "host_presence": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "hostname": {"type": "string"},
                        "ip_address": {"type": "string"},
                        "first_seen": {"type": "string", "format": "date-time"},
                        "last_seen": {"type": "string", "format": "date-time"},
                        "total_scans_available": {"type": "integer"},
                        "scans_present": {"type": "integer"},
                        "scans_missing": {"type": "integer"},
                        "presence_percentage": {"type": "number"},
                        "status": {"type": "string", "enum": ["Active", "Missing"]},
                        "missing_scan_dates": {"type": "string"}
                    }
                }
            },
            "scan_changes": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "scan_date": {"type": "string", "format": "date-time"},
                        "previous_scan": {"type": "string", "format": "date-time"},
                        "hosts_added": {"type": "integer"},
                        "hosts_removed": {"type": "integer"},
                        "hosts_unchanged": {"type": "integer"},
                        "total_hosts_current": {"type": "integer"},
                        "total_hosts_previous": {"type": "integer"},
                        "net_change": {"type": "integer"},
                        "added_host_list": {"type": "string"},
                        "removed_host_list": {"type": "string"}
                    }
                }
            }
        }
    }
    
    try:
        with open(schema_path, 'w', encoding='utf-8') as f:
            json.dump(schema, f, indent=2, ensure_ascii=False)
        
        print(f"Successfully exported JSON schema: {schema_path}")
        
    except Exception as e:
        print(f"Error exporting JSON schema: {e}")


def export_to_excel(historical_df: pd.DataFrame, lifecycle_df: pd.DataFrame, 
                    host_presence_df: pd.DataFrame, scan_changes_df: pd.DataFrame,
                    excel_path: str) -> None:
    """
    Export data to Excel with multiple sheets and formatting.
    
    Args:
        historical_df: Historical findings DataFrame
        lifecycle_df: Lifecycle analysis DataFrame
        host_presence_df: Host presence analysis DataFrame
        scan_changes_df: Scan changes DataFrame
        excel_path: Path to Excel output file
    """
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Export sheets
            lifecycle_df.to_excel(writer, sheet_name='Finding_Lifecycle', index=False)
            host_presence_df.to_excel(writer, sheet_name='Host_Presence', index=False)
            scan_changes_df.to_excel(writer, sheet_name='Scan_Changes', index=False)
            historical_df.to_excel(writer, sheet_name='Historical_Data', index=False)
            
            # Create summary statistics
            include_info_val = getattr(self, 'include_info', tk.BooleanVar(value=False)).get()
            info_status = "Included" if include_info_val else "Excluded from Analysis"
            
            summary_data = {
                'Metric': [
                    'Analysis Mode',
                    'Total Scans',
                    'Total Findings (Database)',
                    'Info Findings (Database)',
                    'Findings in Analysis',
                    'Unique Findings',
                    'Active Findings',
                    'Resolved Findings',
                    'Reappeared Findings',
                    'Total Hosts Ever Seen',
                    'Currently Active Hosts',
                    'Missing Hosts',
                    'Critical Findings',
                    'High Findings',
                    'Medium Findings',
                    'Low Findings'
                ],
                'Count': [
                    f"Info findings {info_status}",
                    historical_df['scan_date'].nunique() if not historical_df.empty else 0,
                    len(historical_df),
                    len(historical_df[historical_df['severity_text'] == 'Info']) if not historical_df.empty else 0,
                    len(lifecycle_df),
                    len(lifecycle_df),
                    len(lifecycle_df[lifecycle_df['status'] == 'Active']) if not lifecycle_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['status'] == 'Resolved']) if not lifecycle_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['reappearances'] > 0]) if not lifecycle_df.empty else 0,
                    len(host_presence_df),
                    len(host_presence_df[host_presence_df['status'] == 'Active']) if not host_presence_df.empty else 0,
                    len(host_presence_df[host_presence_df['status'] == 'Missing']) if not host_presence_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['severity_text'] == 'Critical']) if not lifecycle_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['severity_text'] == 'High']) if not lifecycle_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['severity_text'] == 'Medium']) if not lifecycle_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['severity_text'] == 'Low']) if not lifecycle_df.empty else 0
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Auto-fit columns and add filters
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Auto-fit columns
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add autofilter and banding
                if worksheet.max_row > 0:
                    worksheet.auto_filter.ref = worksheet.dimensions
                    
                    # Add table formatting with banding
                    if worksheet.max_row > 1:
                        from openpyxl.worksheet.table import Table, TableStyleInfo
                        tab = Table(displayName=f"Table_{sheet_name}", ref=worksheet.dimensions)
                        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                             showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                        tab.tableStyleInfo = style
                        worksheet.add_table(tab)
        
        print(f"Successfully exported to Excel: {excel_path}")
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()


class EnhancedHistoricalAnalysisGUI:
    """Enhanced GUI for Nessus Historical Analysis System with host tracking"""
    
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Enhanced Nessus Historical Analysis System")
        self.window.geometry("1400x900")
        self.window.configure(bg='#2b2b2b')  # Dark theme background
        
        # Configure style for dark theme
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure('.', background='#2b2b2b', foreground='white')
        self.style.configure('TLabel', background='#2b2b2b', foreground='white')
        self.style.configure('TFrame', background='#2b2b2b')
        self.style.configure('TLabelFrame', background='#2b2b2b', foreground='white')
        self.style.configure('TButton', background='#404040', foreground='white')
        self.style.map('TButton', background=[('active', '#505050')])
        
        self.archive_paths = []
        self.plugins_db_path = None
        self.existing_db_path = None
        self.opdir_file_path = None
        self.include_info = tk.BooleanVar(value=False)  # Default to exclude Info findings
        
        # Date filter variables
        self.filter_start_date = tk.StringVar()
        self.filter_end_date = tk.StringVar()
        self.use_date_filter = tk.BooleanVar(value=False)
        
        self.historical_df = pd.DataFrame()
        self.lifecycle_df = pd.DataFrame()
        self.host_presence_df = pd.DataFrame()
        self.scan_changes_df = pd.DataFrame()
        self.opdir_df = pd.DataFrame()
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup the user interface with dark theme"""
        # Main container with padding
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        
        # File selection frame
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        file_frame.columnconfigure(1, weight=1)
        
        # Archive selection
        ttk.Label(file_frame, text="Nessus Archives:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.archives_label = ttk.Label(file_frame, text="No files selected (optional if database loaded)", foreground="gray")
        self.archives_label.grid(row=0, column=1, sticky=tk.W, padx=5)
        ttk.Button(file_frame, text="Select Archives", command=self.select_archives).grid(row=0, column=2, padx=5)
        
        # Plugins DB selection
        ttk.Label(file_frame, text="Plugins DB (optional):").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.plugins_label = ttk.Label(file_frame, text="None selected", foreground="gray")
        self.plugins_label.grid(row=1, column=1, sticky=tk.W, padx=5)
        ttk.Button(file_frame, text="Select Plugins DB", command=self.select_plugins_db).grid(row=1, column=2, padx=5)
        
        # Existing DB selection
        ttk.Label(file_frame, text="Existing DB (optional):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.existing_db_label = ttk.Label(file_frame, text="None selected", foreground="gray")
        self.existing_db_label.grid(row=2, column=1, sticky=tk.W, padx=5)
        ttk.Button(file_frame, text="Load Existing DB", command=self.select_existing_db).grid(row=2, column=2, padx=5)
        
        # OPDIR file selection
        ttk.Label(file_frame, text="OPDIR File (optional):").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.opdir_label = ttk.Label(file_frame, text="None selected", foreground="gray")
        self.opdir_label.grid(row=3, column=1, sticky=tk.W, padx=5)
        ttk.Button(file_frame, text="Select OPDIR File", command=self.select_opdir_file).grid(row=3, column=2, padx=5)
        
        # Analysis options frame
        options_frame = ttk.LabelFrame(main_frame, text="Analysis Options", padding="10")
        options_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=5)
        options_frame.columnconfigure(1, weight=1)
        
        # Include Info findings checkbox
        info_checkbox = ttk.Checkbutton(
            options_frame, 
            text="Include Info-level findings in analysis and visualizations", 
            variable=self.include_info
        )
        info_checkbox.grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=5)
        
        # Info label
        info_label = ttk.Label(
            options_frame, 
            text="Note: Info findings are always stored in database, this only affects analysis/visuals",
            foreground="gray"
        )
        info_label.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=2)
        
        # Date filtering controls
        ttk.Separator(options_frame, orient='horizontal').grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        date_filter_checkbox = ttk.Checkbutton(
            options_frame,
            text="Filter visualizations by date range",
            variable=self.use_date_filter,
            command=self.toggle_date_filter
        )
        date_filter_checkbox.grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=5)
        
        # Start date
        ttk.Label(options_frame, text="Start Date:").grid(row=4, column=0, sticky=tk.W, padx=(20, 5), pady=2)
        self.start_date_entry = ttk.Entry(options_frame, textvariable=self.filter_start_date, width=12, state='disabled')
        self.start_date_entry.grid(row=4, column=1, sticky=tk.W, padx=5, pady=2)
        ttk.Label(options_frame, text="(YYYY-MM-DD)", foreground="gray").grid(row=4, column=2, sticky=tk.W, padx=5, pady=2)
        
        # End date
        ttk.Label(options_frame, text="End Date:").grid(row=5, column=0, sticky=tk.W, padx=(20, 5), pady=2)
        self.end_date_entry = ttk.Entry(options_frame, textvariable=self.filter_end_date, width=12, state='disabled')
        self.end_date_entry.grid(row=5, column=1, sticky=tk.W, padx=5, pady=2)
        ttk.Label(options_frame, text="(YYYY-MM-DD)", foreground="gray").grid(row=5, column=2, sticky=tk.W, padx=5, pady=2)
        
        # Action buttons frame
        action_frame = ttk.Frame(main_frame, padding="10")
        action_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Button(action_frame, text="Process/Analyze", command=self.process_archives).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Refresh Analysis", command=self.refresh_analysis).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Export to Excel", command=self.export_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Export to SQLite", command=self.export_sqlite).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Export to JSON", command=self.export_json).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Export JSON Schema", command=self.export_schema).pack(side=tk.LEFT, padx=5)
        
        # Notebook for different views
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # Status tab
        self.status_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.status_frame, text="Status")
        
        self.status_text = tk.Text(self.status_frame, wrap=tk.WORD, height=20, 
                                   bg='#1e1e1e', fg='white', insertbackground='white')
        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        status_scroll = ttk.Scrollbar(self.status_frame, command=self.status_text.yview)
        status_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.config(yscrollcommand=status_scroll.set)
        
        # Visualization tabs (created after processing)
        self.viz_frames = {}
    
    def select_archives(self):
        """Select Nessus archive files with dark theme dialog"""
        filetypes = (
            ('Archive files', '*.zip'),
            ('Nessus files', '*.nessus'),
            ('All files', '*.*')
        )
        
        paths = filedialog.askopenfilenames(
            title='Select Nessus Archives',
            filetypes=filetypes
        )
        
        if paths:
            self.archive_paths = list(paths)
            self.archives_label.config(
                text=f"{len(self.archive_paths)} file(s) selected",
                foreground="white"
            )
            self.log(f"Selected {len(self.archive_paths)} archive(s)")
        else:
            # Update archives label to show current mode when no files selected
            self.archives_label.config(
                text="No files selected (archives optional if database loaded)",
                foreground="gray"
            )
    
    def select_plugins_db(self):
        """Select plugins database file with dark theme dialog"""
        filetypes = (
            ('XML files', '*.xml'),
            ('JSON files', '*.json'),
            ('All files', '*.*')
        )
        
        path = filedialog.askopenfilename(
            title='Select Plugins Database',
            filetypes=filetypes
        )
        
        if path:
            self.plugins_db_path = path
            self.plugins_label.config(
                text=os.path.basename(path),
                foreground="white"
            )
            self.log(f"Selected plugins DB: {os.path.basename(path)}")
    
    def toggle_date_filter(self):
        """Enable/disable date filter controls"""
        state = 'normal' if self.use_date_filter.get() else 'disabled'
        self.start_date_entry.config(state=state)
        self.end_date_entry.config(state=state)
        
        # Set default dates when enabling
        if self.use_date_filter.get() and not self.historical_df.empty:
            if not self.filter_start_date.get():
                start_date = self.historical_df['scan_date'].min().strftime('%Y-%m-%d')
                self.filter_start_date.set(start_date)
            if not self.filter_end_date.get():
                end_date = self.historical_df['scan_date'].max().strftime('%Y-%m-%d')
                self.filter_end_date.set(end_date)
    
    def get_filtered_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply date and info filtering to DataFrame"""
        filtered_df = df.copy()
        
        # Apply Info filtering
        if not self.include_info.get():
            filtered_df = filtered_df[filtered_df['severity_text'] != 'Info']
        
        # Apply date filtering
        if self.use_date_filter.get() and not filtered_df.empty:
            try:
                start_date = pd.to_datetime(self.filter_start_date.get())
                end_date = pd.to_datetime(self.filter_end_date.get())
                
                # Ensure we include the full end date
                end_date = end_date + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                
                filtered_df = filtered_df[
                    (filtered_df['scan_date'] >= start_date) & 
                    (filtered_df['scan_date'] <= end_date)
                ]
            except (ValueError, TypeError) as e:
                self.log(f"Date filter error: {e}")
        
        return filtered_df
    
    def get_severity_colors(self):
        """Get the standard severity color scheme"""
        return {
            'Critical': '#dc3545',  # Red
            'High': '#fd7e14',      # Orange  
            'Medium': '#ffc107',    # Yellow
            'Low': '#007bff',       # Blue
            'Info': '#6c757d'       # Gray
        }
    
    def get_filter_status_text(self):
        """Get descriptive text for current filter settings"""
        status_parts = []
        
        # Info filter status
        if not self.include_info.get():
            if not self.historical_df.empty:
                info_count = len(self.historical_df[self.historical_df['severity_text'] == 'Info'])
                total_count = len(self.historical_df)
                status_parts.append(f"Excluding Info ({info_count} of {total_count} findings)")
            else:
                status_parts.append("Excluding Info findings")
        else:
            status_parts.append("Including Info findings")
        
        # Date filter status
        if self.use_date_filter.get():
            start_date = self.filter_start_date.get()
            end_date = self.filter_end_date.get()
            if start_date and end_date:
                status_parts.append(f"Date Range: {start_date} to {end_date}")
            else:
                status_parts.append("Date filter enabled (no dates set)")
        
        return " | ".join(status_parts) if status_parts else "No filters applied"
    
    def refresh_analysis(self):
        """Refresh analysis with current Info filtering setting without reprocessing archives"""
        if self.historical_df.empty:
            messagebox.showwarning("No Data", "Please process archives first")
            return
        
        try:
            include_info_val = self.include_info.get()
            self.log(f"Refreshing analysis with current filter settings")
            self.log(f"Filter status: {self.get_filter_status_text()}")
            
            # Filter data for analysis if needed
            analysis_df = self.historical_df.copy()
            if not include_info_val:
                analysis_df = analysis_df[analysis_df['severity_text'] != 'Info']
                info_filtered = len(self.historical_df) - len(analysis_df)
                self.log(f"Filtered out {info_filtered} Info-level findings from analysis")
            
            # Regenerate analysis DataFrames
            self.lifecycle_df = analyze_finding_lifecycle(analysis_df)
            self.host_presence_df = create_host_presence_analysis(analysis_df)
            self.scan_changes_df = analyze_scan_changes(analysis_df)
            
            # Update date filter defaults if enabled
            if self.use_date_filter.get():
                self.toggle_date_filter()
            
            # Update visualizations
            self.create_visualizations()
            
            self.log(f"Analysis refreshed!")
            self.log(f"Total findings in database: {len(self.historical_df)}")
            if not include_info_val:
                info_count = len(self.historical_df[self.historical_df['severity_text'] == 'Info'])
                self.log(f"Info findings in database (excluded from analysis): {info_count}")
            self.log(f"Findings in analysis: {len(self.lifecycle_df)}")
            
            messagebox.showinfo("Success", "Analysis refreshed with current settings!")
            
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            messagebox.showerror("Error", f"Refresh failed: {str(e)}")
    
    
        """Refresh analysis with current Info filtering setting without reprocessing archives"""
        if self.historical_df.empty:
            messagebox.showwarning("No Data", "Please process archives first")
            return
        
        try:
            include_info_val = self.include_info.get()
            self.log(f"Refreshing analysis with Info findings {'included' if include_info_val else 'excluded'}")
            
            # Filter data for analysis if needed
            analysis_df = self.historical_df.copy()
            if not include_info_val:
                analysis_df = analysis_df[analysis_df['severity_text'] != 'Info']
                info_filtered = len(self.historical_df) - len(analysis_df)
                self.log(f"Filtered out {info_filtered} Info-level findings from analysis")
            
            # Regenerate analysis DataFrames
            self.lifecycle_df = analyze_finding_lifecycle(analysis_df)
            self.host_presence_df = create_host_presence_analysis(analysis_df)
            self.scan_changes_df = analyze_scan_changes(analysis_df)
            
            # Update visualizations
            self.create_visualizations()
            
            self.log(f"Analysis refreshed!")
            self.log(f"Total findings in database: {len(self.historical_df)}")
            if not include_info_val:
                info_count = len(self.historical_df[self.historical_df['severity_text'] == 'Info'])
                self.log(f"Info findings in database (excluded from analysis): {info_count}")
            self.log(f"Findings in analysis: {len(self.lifecycle_df)}")
            
            messagebox.showinfo("Success", "Analysis refreshed with current settings!")
            
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            messagebox.showerror("Error", f"Refresh failed: {str(e)}")
    
    def select_existing_db(self):
    
        """Select existing database file for updating"""
        filetypes = (
            ('SQLite database', '*.db'),
            ('All files', '*.*')
        )
        
        path = filedialog.askopenfilename(
            title='Select Existing Database',
            filetypes=filetypes
        )
        
        if path:
            self.existing_db_path = path
            self.existing_db_label.config(
                text=os.path.basename(path),
                foreground="white"
            )
            self.log(f"Selected existing DB: {os.path.basename(path)}")
    
    def select_opdir_file(self):
        """Select OPDIR spreadsheet file"""
        filetypes = (
            ('Excel files', '*.xlsx'),
            ('Excel files', '*.xls'),
            ('CSV files', '*.csv'),
            ('All files', '*.*')
        )
        
        path = filedialog.askopenfilename(
            title='Select OPDIR Spreadsheet',
            filetypes=filetypes
        )
        
        if path:
            self.opdir_file_path = path
            self.opdir_label.config(
                text=os.path.basename(path),
                foreground="white"
            )
            self.log(f"Selected OPDIR file: {os.path.basename(path)}")
            
            # Load and process OPDIR data
            try:
                self.opdir_df = load_and_process_opdir(path)
                if not self.opdir_df.empty:
                    self.log(f"Loaded {len(self.opdir_df)} OPDIR records")
                else:
                    self.log("Warning: OPDIR file loaded but no valid records found")
            except Exception as e:
                self.log(f"Error processing OPDIR file: {e}")
                self.opdir_df = pd.DataFrame()
    
    
    def log(self, message: str):
        """Add message to status log"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.window.update()
    
    def process_archives(self):
        """Process selected archives with enhanced features or analyze existing database"""
        # Check if we have either archives or an existing database
        has_archives = bool(self.archive_paths)
        has_existing_db = bool(self.existing_db_path and os.path.exists(self.existing_db_path))
        
        if not has_archives and not has_existing_db:
            messagebox.showwarning("No Data Source", "Please select archive files or load an existing database")
            return
        
        if not has_archives and has_existing_db:
            # Load and analyze existing database only
            self.log("="*60)
            self.log("Loading and analyzing existing database...")
            self.log("="*60)
            
            try:
                # Load existing database
                existing_historical_df, existing_lifecycle_df, existing_host_presence_df, existing_scan_changes_df = load_existing_database(self.existing_db_path)
                
                if existing_historical_df.empty:
                    messagebox.showerror("Error", "No data found in the existing database")
                    return
                
                # Set the loaded data
                self.historical_df = existing_historical_df
                
                # Re-run analysis with current filter settings
                include_info_val = self.include_info.get()
                self.log(f"Re-analyzing with current filter settings")
                self.log(f"Filter status: {self.get_filter_status_text()}")
                
                # Filter data for analysis if needed
                analysis_df = self.historical_df.copy()
                if not include_info_val:
                    analysis_df = analysis_df[analysis_df['severity_text'] != 'Info']
                    info_filtered = len(self.historical_df) - len(analysis_df)
                    self.log(f"Filtered out {info_filtered} Info-level findings from analysis")
                
                # Regenerate analysis DataFrames with current settings
                self.lifecycle_df = analyze_finding_lifecycle(analysis_df)
                self.host_presence_df = create_host_presence_analysis(analysis_df)
                self.scan_changes_df = analyze_scan_changes(analysis_df)
                
                # Set up date filter defaults
                if not self.filter_start_date.get() and not self.historical_df.empty:
                    start_date = self.historical_df['scan_date'].min().strftime('%Y-%m-%d')
                    self.filter_start_date.set(start_date)
                if not self.filter_end_date.get() and not self.historical_df.empty:
                    end_date = self.historical_df['scan_date'].max().strftime('%Y-%m-%d')
                    self.filter_end_date.set(end_date)
                
                self.log(f"Database analysis complete!")
                self.log(f"Total findings in database: {len(self.historical_df)}")
                self.log(f"Date range: {self.historical_df['scan_date'].min()} to {self.historical_df['scan_date'].max()}")
                if not include_info_val:
                    info_count = len(self.historical_df[self.historical_df['severity_text'] == 'Info'])
                    self.log(f"Info findings in database (excluded from analysis): {info_count}")
                self.log(f"Findings in analysis: {len(self.lifecycle_df)}")
                self.log(f"Hosts tracked: {len(self.host_presence_df)}")
                self.log(f"Scan changes: {len(self.scan_changes_df)}")
                
                # Create visualizations
                self.create_visualizations()
                
                messagebox.showinfo("Success", "Existing database loaded and analyzed successfully!")
                
            except Exception as e:
                self.log(f"ERROR: {str(e)}")
                messagebox.showerror("Error", f"Database analysis failed: {str(e)}")
                import traceback
                traceback.print_exc()
            
            return
        
        # Original archive processing logic (when archives are provided)
        self.log("="*60)
        self.log("Starting enhanced archive processing...")
        self.log("="*60)
        
        try:
            # Process archives with enhanced features
            include_info_val = self.include_info.get()
            self.log(f"Processing with Info findings {'included' if include_info_val else 'excluded'} from analysis")
            
            self.historical_df, self.lifecycle_df, self.host_presence_df, self.scan_changes_df = process_historical_scans(
                self.archive_paths,
                self.plugins_db_path,
                self.existing_db_path,
                include_info_val
            )
            
            if self.historical_df.empty:
                messagebox.showerror("Error", "No findings extracted from archives")
                return
            
            # Enrich with OPDIR data if available
            if not self.opdir_df.empty:
                self.log("Enriching findings with OPDIR data...")
                self.historical_df = enrich_findings_with_opdir(self.historical_df, self.opdir_df)
                self.lifecycle_df = enrich_findings_with_opdir(self.lifecycle_df, self.opdir_df)
            
            # Set up date filter defaults
            if not self.filter_start_date.get():
                start_date = self.historical_df['scan_date'].min().strftime('%Y-%m-%d')
                self.filter_start_date.set(start_date)
            if not self.filter_end_date.get():
                end_date = self.historical_df['scan_date'].max().strftime('%Y-%m-%d')
                self.filter_end_date.set(end_date)
            
            self.log(f"Analysis complete!")
            self.log(f"Total findings in database: {len(self.historical_df)}")
            if not include_info_val:
                info_count = len(self.historical_df[self.historical_df['severity_text'] == 'Info'])
                self.log(f"Info findings in database (excluded from analysis): {info_count}")
            self.log(f"Findings in analysis: {len(self.lifecycle_df)}")
            self.log(f"Hosts tracked: {len(self.host_presence_df)}")
            self.log(f"Scan changes: {len(self.scan_changes_df)}")
            
            # Create visualizations
            self.create_visualizations()
            
            messagebox.showinfo("Success", "Archives processed successfully with enhanced features!")
            
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            messagebox.showerror("Error", f"Processing failed: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def create_visualizations(self):
        """Create enhanced visualization tabs"""
        # Remove old visualization tabs
        for tab_name in list(self.viz_frames.keys()):
            self.notebook.forget(self.viz_frames[tab_name])
        self.viz_frames = {}
        
        # Create new visualization tabs
        self.create_timeline_viz()
        self.create_host_presence_viz()
        self.create_scan_changes_viz()
        self.create_severity_trend_viz()
        self.create_source_file_viz()
        self.create_cvss_analysis_viz()
        self.create_network_analysis_viz()
        self.create_threat_intelligence_viz()
        self.create_remediation_effectiveness_viz()
        self.create_risk_heatmap_viz()
        self.create_compliance_dashboard_viz()
        self.create_host_viz()
        self.create_plugin_viz()
        self.create_lifecycle_viz()
    
    def create_host_presence_viz(self):
        """Create host presence visualization with date filtering"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Host Tracking")
        self.viz_frames['host_presence'] = frame
        
        # Add filter status
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        # Get filtered host presence data based on date range
        filtered_host_presence = self.host_presence_df.copy()
        if self.use_date_filter.get() and not filtered_host_presence.empty:
            try:
                start_date = pd.to_datetime(self.filter_start_date.get())
                end_date = pd.to_datetime(self.filter_end_date.get()) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                
                # Filter based on last_seen date for host presence
                filtered_host_presence = filtered_host_presence[
                    (filtered_host_presence['last_seen'] >= start_date) | 
                    (filtered_host_presence['first_seen'] <= end_date)
                ]
            except (ValueError, TypeError):
                pass
        
        # Host status distribution
        if not filtered_host_presence.empty:
            status_counts = filtered_host_presence['status'].value_counts()
            ax1.pie(status_counts.values, labels=status_counts.index, autopct='%1.1f%%',
                   colors=['#28a745', '#dc3545'], startangle=90)
            ax1.set_title('Host Status Distribution', fontsize=12, fontweight='bold', color='white')
            ax1.set_facecolor('#2b2b2b')
        else:
            ax1.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax1.transAxes, color='white', fontsize=12)
            ax1.set_facecolor('#2b2b2b')
        
        # Presence percentage distribution
        if not filtered_host_presence.empty:
            bins = [0, 25, 50, 75, 90, 100]
            labels = ['0-25%', '26-50%', '51-75%', '76-90%', '91-100%']
            filtered_host_presence['presence_category'] = pd.cut(
                filtered_host_presence['presence_percentage'], 
                bins=bins, labels=labels, right=True
            )
            presence_dist = filtered_host_presence['presence_category'].value_counts().sort_index()
            
            bars = ax2.bar(range(len(presence_dist)), presence_dist.values, color='#007bff')
            ax2.set_xticks(range(len(presence_dist)))
            ax2.set_xticklabels(presence_dist.index, rotation=45, ha='right', color='white')
            ax2.set_title('Host Presence Percentage Distribution', fontsize=12, fontweight='bold', color='white')
            ax2.set_ylabel('Number of Hosts', color='white')
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                if height > 0:  # Only show label if there's a value
                    ax2.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                            f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        else:
            ax2.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax2.transAxes, color='white', fontsize=12)
            ax2.set_facecolor('#2b2b2b')
        
        # Most unreliable hosts (lowest presence percentage)
        if not filtered_host_presence.empty:
            unreliable_hosts = filtered_host_presence.nsmallest(10, 'presence_percentage')
            display_names = [f"{row['hostname']} ({row['ip_address']})" for _, row in unreliable_hosts.iterrows()]
            
            bars = ax3.barh(range(len(unreliable_hosts)), unreliable_hosts['presence_percentage'].values, color='#dc3545')
            ax3.set_yticks(range(len(unreliable_hosts)))
            ax3.set_yticklabels(display_names, color='white')
            ax3.set_title('10 Most Unreliable Hosts', fontsize=12, fontweight='bold', color='white')
            ax3.set_xlabel('Presence Percentage', color='white')
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            
            # Add data labels on horizontal bars
            for i, bar in enumerate(bars):
                width = bar.get_width()
                ax3.text(width + 1, bar.get_y() + bar.get_height()/2.,
                        f'{width:.1f}%', ha='left', va='center', color='white', fontweight='bold')
        else:
            ax3.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax3.transAxes, color='white', fontsize=12)
            ax3.set_facecolor('#2b2b2b')
        
        # Recently missing hosts
        if not filtered_host_presence.empty:
            missing_hosts = filtered_host_presence[filtered_host_presence['status'] == 'Missing'].head(10)
            if not missing_hosts.empty:
                display_names = [f"{row['hostname']} ({row['ip_address']})" for _, row in missing_hosts.iterrows()]
                days_since_last = [(datetime.now() - row['last_seen']).days for _, row in missing_hosts.iterrows()]
                
                bars = ax4.barh(range(len(missing_hosts)), days_since_last, color='#fd7e14')
                ax4.set_yticks(range(len(missing_hosts)))
                ax4.set_yticklabels(display_names, color='white')
                ax4.set_title('Recently Missing Hosts', fontsize=12, fontweight='bold', color='white')
                ax4.set_xlabel('Days Since Last Seen', color='white')
                ax4.set_facecolor('#2b2b2b')
                ax4.tick_params(colors='white')
                
                # Add data labels on horizontal bars
                for bar in bars:
                    width = bar.get_width()
                    ax4.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                            f'{int(width)}', ha='left', va='center', color='white', fontweight='bold')
            else:
                ax4.text(0.5, 0.5, 'No missing hosts', ha='center', va='center', 
                        transform=ax4.transAxes, color='white', fontsize=12)
                ax4.set_facecolor('#2b2b2b')
        else:
            ax4.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax4.transAxes, color='white', fontsize=12)
            ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_network_analysis_viz(self):
        """Create network/port analysis visualization"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Network Analysis")
        self.viz_frames['network_analysis'] = frame
        
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        filtered_df = self.get_filtered_data(self.historical_df)
        
        if not filtered_df.empty:
            # Most vulnerable ports
            port_counts = filtered_df['port'].value_counts().head(15)
            bars = ax1.barh(range(len(port_counts)), port_counts.values, color='#dc3545')
            ax1.set_yticks(range(len(port_counts)))
            ax1.set_yticklabels(port_counts.index, color='white')
            ax1.set_title('Most Vulnerable Ports (Top 15)', fontsize=12, fontweight='bold', color='white')
            ax1.set_xlabel('Number of Findings', color='white')
            ax1.set_facecolor('#2b2b2b')
            ax1.tick_params(colors='white')
            
            for bar in bars:
                width = bar.get_width()
                ax1.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                        f'{int(width)}', ha='left', va='center', color='white', fontweight='bold')
            
            # Protocol distribution
            protocol_counts = filtered_df['protocol'].value_counts()
            colors_protocol = ['#007bff', '#28a745', '#ffc107', '#fd7e14', '#dc3545']
            ax2.pie(protocol_counts.values, labels=protocol_counts.index, autopct='%1.1f%%',
                   colors=colors_protocol[:len(protocol_counts)], startangle=90)
            ax2.set_title('Protocol Distribution', fontsize=12, fontweight='bold', color='white')
            ax2.set_facecolor('#2b2b2b')
            
            # Service name analysis
            if 'svc_name' in filtered_df.columns:
                service_counts = filtered_df[filtered_df['svc_name'].notna()]['svc_name'].value_counts().head(10)
                if not service_counts.empty:
                    bars = ax3.bar(range(len(service_counts)), service_counts.values, color='#28a745')
                    ax3.set_xticks(range(len(service_counts)))
                    ax3.set_xticklabels(service_counts.index, rotation=45, ha='right', color='white')
                    ax3.set_title('Most Vulnerable Services (Top 10)', fontsize=12, fontweight='bold', color='white')
                    ax3.set_ylabel('Number of Findings', color='white')
                    ax3.set_facecolor('#2b2b2b')
                    ax3.tick_params(colors='white')
                    
                    for bar in bars:
                        height = bar.get_height()
                        ax3.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                                f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
                else:
                    ax3.text(0.5, 0.5, 'No service name data', ha='center', va='center', 
                            transform=ax3.transAxes, color='white', fontsize=12)
                    ax3.set_facecolor('#2b2b2b')
            
            # Port vs Severity heatmap
            port_severity = filtered_df.groupby(['port', 'severity_text']).size().unstack(fill_value=0)
            if not port_severity.empty:
                # Get top 10 ports by total findings
                top_ports = port_severity.sum(axis=1).nlargest(10)
                heatmap_data = port_severity.loc[top_ports.index]
                
                # Create heatmap manually
                import numpy as np
                severity_order = ['Critical', 'High', 'Medium', 'Low', 'Info']
                available_severities = [s for s in severity_order if s in heatmap_data.columns]
                
                if available_severities:
                    heatmap_array = heatmap_data[available_severities].values
                    
                    im = ax4.imshow(heatmap_array, cmap='Reds', aspect='auto')
                    ax4.set_xticks(range(len(available_severities)))
                    ax4.set_xticklabels(available_severities, color='white')
                    ax4.set_yticks(range(len(heatmap_data)))
                    ax4.set_yticklabels(heatmap_data.index, color='white')
                    ax4.set_title('Port vs Severity Heatmap', fontsize=12, fontweight='bold', color='white')
                    
                    # Add text annotations
                    for i in range(len(heatmap_data)):
                        for j in range(len(available_severities)):
                            text = ax4.text(j, i, int(heatmap_array[i, j]),
                                           ha="center", va="center", color="white", fontweight='bold')
                    
                    ax4.set_facecolor('#2b2b2b')
                    ax4.tick_params(colors='white')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_threat_intelligence_viz(self):
        """Create threat intelligence and CVE/IAVx analysis"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Threat Intel")
        self.viz_frames['threat_intel'] = frame
        
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        filtered_df = self.get_filtered_data(self.historical_df)
        
        if not filtered_df.empty:
            # Most frequent CVEs
            cve_findings = filtered_df[filtered_df['cves'].notna() & (filtered_df['cves'] != '')]
            if not cve_findings.empty:
                all_cves = []
                for cves in cve_findings['cves']:
                    all_cves.extend([cve.strip() for cve in str(cves).split('\n') if cve.strip()])
                
                if all_cves:
                    cve_counts = pd.Series(all_cves).value_counts().head(15)
                    bars = ax1.barh(range(len(cve_counts)), cve_counts.values, color='#dc3545')
                    ax1.set_yticks(range(len(cve_counts)))
                    ax1.set_yticklabels(cve_counts.index, color='white')
                    ax1.set_title('Most Frequent CVEs (Top 15)', fontsize=12, fontweight='bold', color='white')
                    ax1.set_xlabel('Number of Affected Hosts', color='white')
                    ax1.set_facecolor('#2b2b2b')
                    ax1.tick_params(colors='white')
                    
                    for bar in bars:
                        width = bar.get_width()
                        ax1.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                                f'{int(width)}', ha='left', va='center', color='white', fontweight='bold')
            
            # IAVx distribution
            iavx_findings = filtered_df[filtered_df['iavx'].notna() & (filtered_df['iavx'] != '')]
            if not iavx_findings.empty:
                all_iavx = []
                for iavx in iavx_findings['iavx']:
                    all_iavx.extend([ref.strip() for ref in str(iavx).split('\n') if ref.strip()])
                
                if all_iavx:
                    # Extract IAVx types (A, B, T)
                    iavx_types = [ref.split(':')[0] if ':' in ref else ref[:4] for ref in all_iavx]
                    iavx_type_counts = pd.Series(iavx_types).value_counts()
                    
                    colors_iavx = ['#fd7e14', '#007bff', '#28a745']
                    ax2.pie(iavx_type_counts.values, labels=iavx_type_counts.index, autopct='%1.1f%%',
                           colors=colors_iavx[:len(iavx_type_counts)], startangle=90)
                    ax2.set_title('IAVx Type Distribution', fontsize=12, fontweight='bold', color='white')
                    ax2.set_facecolor('#2b2b2b')
            
            # CVE vs No CVE findings
            has_cve = len(filtered_df[filtered_df['cves'].notna() & (filtered_df['cves'] != '')])
            no_cve = len(filtered_df) - has_cve
            
            bars = ax3.bar(['Has CVE', 'No CVE'], [has_cve, no_cve], color=['#dc3545', '#6c757d'])
            ax3.set_title('CVE Coverage', fontsize=12, fontweight='bold', color='white')
            ax3.set_ylabel('Number of Findings', color='white')
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            
            for bar in bars:
                height = bar.get_height()
                ax3.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                        f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
            
            # OPDIR compliance (if OPDIR data available)
            if not self.opdir_df.empty and 'opdir_number' in filtered_df.columns:
                has_opdir = len(filtered_df[filtered_df['opdir_number'].notna()])
                no_opdir = len(filtered_df) - has_opdir
                
                bars = ax4.bar(['Has OPDIR', 'No OPDIR'], [has_opdir, no_opdir], color=['#28a745', '#6c757d'])
                ax4.set_title('OPDIR Coverage', fontsize=12, fontweight='bold', color='white')
                ax4.set_ylabel('Number of Findings', color='white')
                ax4.set_facecolor('#2b2b2b')
                ax4.tick_params(colors='white')
                
                for bar in bars:
                    height = bar.get_height()
                    ax4.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                            f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
            else:
                ax4.text(0.5, 0.5, 'OPDIR data not available', ha='center', va='center', 
                        transform=ax4.transAxes, color='white', fontsize=12)
                ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_source_file_viz(self):
        """Create source file analysis visualization"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Source Files")
        self.viz_frames['source_files'] = frame
        
        # Add filter status
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        filtered_df = self.get_filtered_data(self.historical_df)
        
        if not filtered_df.empty:
            # Findings per source file
            source_counts = filtered_df['scan_file'].value_counts().head(15)
            bars = ax1.barh(range(len(source_counts)), source_counts.values, color='#007bff')
            ax1.set_yticks(range(len(source_counts)))
            ax1.set_yticklabels([f[:30] + '...' if len(f) > 30 else f for f in source_counts.index], color='white')
            ax1.set_title('Findings per Source File (Top 15)', fontsize=12, fontweight='bold', color='white')
            ax1.set_xlabel('Number of Findings', color='white')
            ax1.set_facecolor('#2b2b2b')
            ax1.tick_params(colors='white')
            
            for bar in bars:
                width = bar.get_width()
                ax1.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                        f'{int(width)}', ha='left', va='center', color='white', fontweight='bold')
            
            # Scan frequency timeline
            scan_timeline = filtered_df.groupby(['scan_date', 'scan_file']).size().reset_index(name='findings')
            unique_scans = scan_timeline.groupby('scan_date')['scan_file'].nunique().reset_index(name='file_count')
            
            ax2.plot(unique_scans['scan_date'], unique_scans['file_count'], marker='o', linewidth=2, color='#28a745')
            ax2.set_title('Source Files per Scan Date', fontsize=12, fontweight='bold', color='white')
            ax2.set_xlabel('Scan Date', color='white')
            ax2.set_ylabel('Number of Source Files', color='white')
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # Scan coverage by severity
            severity_by_source = filtered_df.groupby(['scan_file', 'severity_text']).size().unstack(fill_value=0)
            colors = self.get_severity_colors()
            
            if not severity_by_source.empty:
                top_sources = severity_by_source.sum(axis=1).nlargest(10)
                plot_data = severity_by_source.loc[top_sources.index]
                
                bottom = None
                for severity in ['Critical', 'High', 'Medium', 'Low', 'Info']:
                    if severity in plot_data.columns:
                        bars = ax3.bar(range(len(plot_data)), plot_data[severity], 
                                      bottom=bottom, label=severity, color=colors.get(severity))
                        
                        # Add data labels for non-zero values
                        for i, bar in enumerate(bars):
                            height = bar.get_height()
                            if height > 0:
                                ax3.text(bar.get_x() + bar.get_width()/2., 
                                        bar.get_y() + height/2.,
                                        f'{int(height)}', ha='center', va='center', 
                                        color='white', fontweight='bold', fontsize=8)
                        
                        bottom = plot_data[severity] if bottom is None else bottom + plot_data[severity]
                
                ax3.set_xticks(range(len(plot_data)))
                ax3.set_xticklabels([f[:15] + '...' if len(f) > 15 else f for f in plot_data.index], 
                                   rotation=45, ha='right', color='white')
                ax3.set_title('Severity Distribution by Source (Top 10)', fontsize=12, fontweight='bold', color='white')
                ax3.set_ylabel('Number of Findings', color='white')
                ax3.legend(loc='upper right')
                ax3.set_facecolor('#2b2b2b')
                ax3.tick_params(colors='white')
            
            # Scan gaps analysis
            scan_dates = sorted(filtered_df['scan_date'].unique())
            if len(scan_dates) > 1:
                gaps = []
                gap_labels = []
                for i in range(1, len(scan_dates)):
                    gap_days = (scan_dates[i] - scan_dates[i-1]).days
                    gaps.append(gap_days)
                    gap_labels.append(f"{scan_dates[i-1].strftime('%m/%d')} to {scan_dates[i].strftime('%m/%d')}")
                
                bars = ax4.bar(range(len(gaps)), gaps, color='#fd7e14')
                ax4.set_xticks(range(len(gaps)))
                ax4.set_xticklabels(gap_labels, rotation=45, ha='right', color='white')
                ax4.set_title('Days Between Scans', fontsize=12, fontweight='bold', color='white')
                ax4.set_ylabel('Days', color='white')
                ax4.set_facecolor('#2b2b2b')
                ax4.tick_params(colors='white')
                
                for bar in bars:
                    height = bar.get_height()
                    ax4.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                            f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_cvss_analysis_viz(self):
        """Create CVSS score analysis visualization"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="CVSS Analysis")
        self.viz_frames['cvss_analysis'] = frame
        
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        filtered_df = self.get_filtered_data(self.historical_df)
        
        if not filtered_df.empty:
            # CVSS v3 vs v2 distribution
            cvss3_data = pd.to_numeric(filtered_df['cvss3_base_score'], errors='coerce').dropna()
            cvss2_data = pd.to_numeric(filtered_df['cvss2_base_score'], errors='coerce').dropna()
            
            if not cvss3_data.empty:
                ax1.hist(cvss3_data, bins=20, alpha=0.7, label='CVSS v3', color='#007bff', edgecolor='white')
            if not cvss2_data.empty:
                ax1.hist(cvss2_data, bins=20, alpha=0.7, label='CVSS v2', color='#fd7e14', edgecolor='white')
            
            ax1.set_title('CVSS Score Distribution', fontsize=12, fontweight='bold', color='white')
            ax1.set_xlabel('CVSS Score', color='white')
            ax1.set_ylabel('Number of Findings', color='white')
            ax1.legend()
            ax1.set_facecolor('#2b2b2b')
            ax1.tick_params(colors='white')
            
            # CVSS scores over time
            if not cvss3_data.empty:
                cvss_timeline = filtered_df[filtered_df['cvss3_base_score'].notna()].copy()
                cvss_timeline['cvss3_numeric'] = pd.to_numeric(cvss_timeline['cvss3_base_score'], errors='coerce')
                cvss_avg_by_date = cvss_timeline.groupby('scan_date')['cvss3_numeric'].mean().reset_index()
                
                ax2.plot(cvss_avg_by_date['scan_date'], cvss_avg_by_date['cvss3_numeric'], 
                        marker='o', linewidth=2, color='#dc3545')
                ax2.set_title('Average CVSS v3 Score Over Time', fontsize=12, fontweight='bold', color='white')
                ax2.set_xlabel('Scan Date', color='white')
                ax2.set_ylabel('Average CVSS Score', color='white')
                ax2.set_facecolor('#2b2b2b')
                ax2.tick_params(colors='white')
                ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
                plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # CVSS ranges by severity
            if 'severity_text' in filtered_df.columns and not cvss3_data.empty:
                cvss_by_severity = filtered_df[filtered_df['cvss3_base_score'].notna()].copy()
                cvss_by_severity['cvss3_numeric'] = pd.to_numeric(cvss_by_severity['cvss3_base_score'], errors='coerce')
                
                severities = ['Critical', 'High', 'Medium', 'Low']
                colors = self.get_severity_colors()
                
                for i, severity in enumerate(severities):
                    severity_data = cvss_by_severity[cvss_by_severity['severity_text'] == severity]['cvss3_numeric']
                    if not severity_data.empty:
                        ax3.boxplot(severity_data, positions=[i], widths=0.6, 
                                   patch_artist=True, boxprops=dict(facecolor=colors.get(severity)))
                
                ax3.set_xticks(range(len(severities)))
                ax3.set_xticklabels(severities, color='white')
                ax3.set_title('CVSS Score Ranges by Severity', fontsize=12, fontweight='bold', color='white')
                ax3.set_ylabel('CVSS v3 Score', color='white')
                ax3.set_facecolor('#2b2b2b')
                ax3.tick_params(colors='white')
            
            # VPR vs CVSS correlation (if VPR data exists)
            if 'vpr_score' in filtered_df.columns:
                vpr_data = pd.to_numeric(filtered_df['vpr_score'], errors='coerce')
                cvss_data = pd.to_numeric(filtered_df['cvss3_base_score'], errors='coerce')
                
                valid_data = pd.DataFrame({'vpr': vpr_data, 'cvss': cvss_data}).dropna()
                
                if len(valid_data) > 10:
                    ax4.scatter(valid_data['cvss'], valid_data['vpr'], alpha=0.6, color='#28a745')
                    ax4.set_title('VPR vs CVSS v3 Correlation', fontsize=12, fontweight='bold', color='white')
                    ax4.set_xlabel('CVSS v3 Score', color='white')
                    ax4.set_ylabel('VPR Score', color='white')
                    ax4.set_facecolor('#2b2b2b')
                    ax4.tick_params(colors='white')
                else:
                    ax4.text(0.5, 0.5, 'Insufficient VPR data', ha='center', va='center', 
                            transform=ax4.transAxes, color='white', fontsize=12)
                    ax4.set_facecolor('#2b2b2b')
            else:
                ax4.text(0.5, 0.5, 'VPR data not available', ha='center', va='center', 
                        transform=ax4.transAxes, color='white', fontsize=12)
                ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_scan_changes_viz(self):
        """Create scan changes visualization"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Scan Changes")
        self.viz_frames['scan_changes'] = frame
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        if not self.scan_changes_df.empty:
            # Host additions over time
            ax1.plot(self.scan_changes_df['scan_date'], self.scan_changes_df['hosts_added'], 
                    marker='o', linewidth=2, color='#28a745', label='Added')
            ax1.plot(self.scan_changes_df['scan_date'], self.scan_changes_df['hosts_removed'], 
                    marker='s', linewidth=2, color='#dc3545', label='Removed')
            ax1.set_title('Host Additions/Removals Over Time', fontsize=12, fontweight='bold', color='white')
            ax1.set_xlabel('Scan Date', color='white')
            ax1.set_ylabel('Number of Hosts', color='white')
            ax1.legend()
            ax1.grid(True, alpha=0.3)
            ax1.set_facecolor('#2b2b2b')
            ax1.tick_params(colors='white')
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # Net change over time
            bars = ax2.bar(range(len(self.scan_changes_df)), self.scan_changes_df['net_change'].values,
                   color=['#28a745' if x >= 0 else '#dc3545' for x in self.scan_changes_df['net_change']])
            ax2.set_title('Net Host Change by Scan', fontsize=12, fontweight='bold', color='white')
            ax2.set_xlabel('Scan Number', color='white')
            ax2.set_ylabel('Net Change', color='white')
            ax2.axhline(y=0, color='white', linestyle='-', alpha=0.5)
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                label_y = height + (abs(height) * 0.02) if height >= 0 else height - (abs(height) * 0.02)
                ax2.text(bar.get_x() + bar.get_width()/2., label_y,
                        f'{int(height)}', ha='center', va='bottom' if height >= 0 else 'top', 
                        color='white', fontweight='bold')
            
            # Total hosts over time
            ax3.plot(self.scan_changes_df['scan_date'], self.scan_changes_df['total_hosts_current'], 
                    marker='o', linewidth=2, color='#007bff')
            ax3.set_title('Total Hosts Over Time', fontsize=12, fontweight='bold', color='white')
            ax3.set_xlabel('Scan Date', color='white')
            ax3.set_ylabel('Total Hosts', color='white')
            ax3.grid(True, alpha=0.3)
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            ax3.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # Stability metrics
            stability_data = {
                'High Turnover': len(self.scan_changes_df[abs(self.scan_changes_df['net_change']) > 5]),
                'Medium Turnover': len(self.scan_changes_df[(abs(self.scan_changes_df['net_change']) > 1) & 
                                                           (abs(self.scan_changes_df['net_change']) <= 5)]),
                'Stable': len(self.scan_changes_df[abs(self.scan_changes_df['net_change']) <= 1])
            }
            
            ax4.pie(stability_data.values(), labels=stability_data.keys(), autopct='%1.1f%%',
                   colors=['#dc3545', '#ffc107', '#28a745'], startangle=90)
            ax4.set_title('Scan Stability Distribution', fontsize=12, fontweight='bold', color='white')
            ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    # [Continue with other visualization methods with dark theme...]
    # [For brevity, I'll include the remaining methods in a follow-up if needed]
    
    def export_excel(self):
        """Export to Excel with enhanced data"""
        if self.historical_df.empty:
            messagebox.showwarning("No Data", "Please process archives first")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Excel File"
        )
        
        if filepath:
            try:
                self.log(f"Exporting to Excel: {filepath}")
                export_to_excel(self.historical_df, self.lifecycle_df, 
                               self.host_presence_df, self.scan_changes_df, filepath)
                self.log("Excel export complete!")
                messagebox.showinfo("Success", f"Enhanced data exported to:\n{filepath}")
            except Exception as e:
                self.log(f"ERROR: {str(e)}")
                messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def export_sqlite(self):
        """Export to SQLite with enhanced data"""
        if self.historical_df.empty:
            messagebox.showwarning("No Data", "Please process archives first")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".db",
            filetypes=[("SQLite database", "*.db"), ("All files", "*.*")],
            title="Save SQLite Database"
        )
        
        if filepath:
            try:
                self.log(f"Exporting to SQLite: {filepath}")
                export_to_sqlite(self.historical_df, self.lifecycle_df, 
                                self.host_presence_df, self.scan_changes_df, filepath)
                self.log("SQLite export complete!")
                messagebox.showinfo("Success", f"Enhanced database exported to:\n{filepath}")
            except Exception as e:
                self.log(f"ERROR: {str(e)}")
                messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def export_json(self):
        """Export to JSON with enhanced data"""
        if self.historical_df.empty:
            messagebox.showwarning("No Data", "Please process archives first")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save JSON File"
        )
        
        if filepath:
            try:
                self.log(f"Exporting to JSON: {filepath}")
                export_to_json(self.historical_df, self.lifecycle_df, 
                              self.host_presence_df, self.scan_changes_df, filepath)
                self.log("JSON export complete!")
                messagebox.showinfo("Success", f"Enhanced data exported to:\n{filepath}")
            except Exception as e:
                self.log(f"ERROR: {str(e)}")
                messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def export_schema(self):
        """Export JSON schema"""
        filepath = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save JSON Schema File"
        )
        
        if filepath:
            try:
                self.log(f"Exporting JSON schema: {filepath}")
                export_json_schema(filepath)
                self.log("JSON schema export complete!")
                messagebox.showinfo("Success", f"JSON schema exported to:\n{filepath}")
            except Exception as e:
                self.log(f"ERROR: {str(e)}")
                messagebox.showerror("Error", f"Schema export failed: {str(e)}")
    
    def create_timeline_viz(self):
        """Create findings timeline visualization with date filtering and corrected colors"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Timeline")
        self.viz_frames['timeline'] = frame
        
        # Add filter status and controls
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        # Filter status
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        # Refresh button
        ttk.Button(control_frame, text="Apply Filters", command=self.create_visualizations).pack(side=tk.RIGHT, padx=5)
        
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8))
        fig.patch.set_facecolor('#2b2b2b')
        
        # Apply all filters to get visualization data
        filtered_df = self.get_filtered_data(self.historical_df)
        
        if filtered_df.empty:
            ax1.text(0.5, 0.5, 'No data matches current filters', ha='center', va='center', 
                    transform=ax1.transAxes, color='white', fontsize=14)
            ax2.text(0.5, 0.5, 'No data matches current filters', ha='center', va='center', 
                    transform=ax2.transAxes, color='white', fontsize=14)
        else:
            # Total findings over time
            timeline_data = filtered_df.groupby('scan_date').size().reset_index(name='count')
            ax1.plot(timeline_data['scan_date'], timeline_data['count'], marker='o', linewidth=2, color='#007bff')
            
            filter_suffix = ""
            if self.use_date_filter.get():
                filter_suffix += f" ({self.filter_start_date.get()} to {self.filter_end_date.get()})"
            if not self.include_info.get():
                filter_suffix += " (Excluding Info)"
            
            ax1.set_title(f'Total Findings Over Time{filter_suffix}', fontsize=14, fontweight='bold', color='white')
            ax1.set_xlabel('Scan Date', color='white')
            ax1.set_ylabel('Number of Findings', color='white')
            ax1.grid(True, alpha=0.3)
            ax1.set_facecolor('#2b2b2b')
            ax1.tick_params(colors='white')
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # Findings by severity over time
            severity_timeline = filtered_df.groupby(['scan_date', 'severity_text']).size().unstack(fill_value=0)
            
            colors = self.get_severity_colors()
            for severity in ['Critical', 'High', 'Medium', 'Low', 'Info']:
                if severity in severity_timeline.columns:
                    ax2.plot(severity_timeline.index, severity_timeline[severity], 
                            marker='o', label=severity, color=colors.get(severity, 'gray'), linewidth=2)
            
            ax2.set_title(f'Findings by Severity Over Time{filter_suffix}', fontsize=14, fontweight='bold', color='white')
            ax2.set_xlabel('Scan Date', color='white')
            ax2.set_ylabel('Number of Findings', color='white')
            ax2.legend(loc='upper left')
            ax2.grid(True, alpha=0.3)
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_severity_trend_viz(self):
        """Create severity trend visualization with date filtering and corrected colors"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Severity Trends")
        self.viz_frames['severity'] = frame
        
        # Add filter status
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        # Get filtered lifecycle data
        filtered_lifecycle = self.lifecycle_df.copy()
        if self.use_date_filter.get() and not filtered_lifecycle.empty:
            try:
                start_date = pd.to_datetime(self.filter_start_date.get())
                end_date = pd.to_datetime(self.filter_end_date.get()) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                
                # Filter by first_seen date for lifecycle data
                filtered_lifecycle = filtered_lifecycle[
                    (filtered_lifecycle['first_seen'] >= start_date) & 
                    (filtered_lifecycle['first_seen'] <= end_date)
                ]
            except (ValueError, TypeError):
                pass
        
        colors = self.get_severity_colors()
        
        # Current severity distribution
        if not filtered_lifecycle.empty:
            severity_counts = filtered_lifecycle['severity_text'].value_counts()
            severity_colors = [colors.get(sev, '#6c757d') for sev in severity_counts.index]
            ax1.pie(severity_counts.values, labels=severity_counts.index, autopct='%1.1f%%', 
                   colors=severity_colors, startangle=90)
            ax1.set_title('Severity Distribution', fontsize=12, fontweight='bold', color='white')
            ax1.set_facecolor('#2b2b2b')
        else:
            ax1.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax1.transAxes, color='white', fontsize=12)
            ax1.set_facecolor('#2b2b2b')
        
        # Active vs Resolved
        if not filtered_lifecycle.empty:
            status_counts = filtered_lifecycle['status'].value_counts()
            bars = ax2.bar(status_counts.index, status_counts.values, color=['#dc3545', '#28a745'])
            ax2.set_title('Active vs Resolved Findings', fontsize=12, fontweight='bold', color='white')
            ax2.set_ylabel('Count', color='white')
            ax2.grid(True, alpha=0.3, axis='y')
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                        f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        else:
            ax2.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax2.transAxes, color='white', fontsize=12)
            ax2.set_facecolor('#2b2b2b')
        
        # Average days open by severity
        if not filtered_lifecycle.empty:
            avg_days = filtered_lifecycle.groupby('severity_text')['days_open'].mean().sort_values(ascending=False)
            severity_colors = [colors.get(sev, '#6c757d') for sev in avg_days.index]
            bars = ax3.barh(avg_days.index, avg_days.values, color=severity_colors)
            ax3.set_title('Average Days Open by Severity', fontsize=12, fontweight='bold', color='white')
            ax3.set_xlabel('Average Days', color='white')
            ax3.grid(True, alpha=0.3, axis='x')
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            
            # Add data labels on horizontal bars
            for bar in bars:
                width = bar.get_width()
                ax3.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                        f'{width:.0f}', ha='left', va='center', color='white', fontweight='bold')
        else:
            ax3.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax3.transAxes, color='white', fontsize=12)
            ax3.set_facecolor('#2b2b2b')
        
        # Reappearance rate
        if not filtered_lifecycle.empty:
            reappeared = len(filtered_lifecycle[filtered_lifecycle['reappearances'] > 0])
            never_reappeared = len(filtered_lifecycle[filtered_lifecycle['reappearances'] == 0])
            ax4.pie([reappeared, never_reappeared], labels=['Reappeared', 'Never Reappeared'],
                   autopct='%1.1f%%', colors=['#dc3545', '#28a745'], startangle=90)
            ax4.set_title('Finding Reappearance Rate', fontsize=12, fontweight='bold', color='white')
            ax4.set_facecolor('#2b2b2b')
        else:
            ax4.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax4.transAxes, color='white', fontsize=12)
            ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_host_viz(self):
        """Create host-specific visualization with enhanced display"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Host Analysis")
        self.viz_frames['host'] = frame
        
        # Create controls frame
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        
        ttk.Label(control_frame, text="Select Host:").pack(side=tk.LEFT, padx=5)
        
        # Get unique hostnames with IP addresses for better display
        host_data = self.historical_df[['hostname', 'ip_address']].drop_duplicates()
        host_options = [f"{row['hostname']} ({row['ip_address']})" for _, row in host_data.iterrows()]
        host_options = sorted(host_options)
        
        self.host_var = tk.StringVar(value=host_options[0] if host_options else "")
        host_combo = ttk.Combobox(control_frame, textvariable=self.host_var, values=host_options, width=50)
        host_combo.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(control_frame, text="Update", command=self.update_host_viz).pack(side=tk.LEFT, padx=5)
        
        # Create figure
        self.host_fig, ((self.host_ax1, self.host_ax2), (self.host_ax3, self.host_ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        self.host_fig.patch.set_facecolor('#2b2b2b')
        
        self.host_canvas = FigureCanvasTkAgg(self.host_fig, frame)
        self.host_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Initial update
        if host_options:
            self.update_host_viz()
    
    def update_host_viz(self):
        """Update host visualization based on selected host with enhanced display"""
        host_str = self.host_var.get()
        if not host_str:
            return
        
        # Extract hostname from "hostname (ip)" format
        hostname = host_str.split(' (')[0]
        
        # Clear axes
        for ax in [self.host_ax1, self.host_ax2, self.host_ax3, self.host_ax4]:
            ax.clear()
            ax.set_facecolor('#2b2b2b')
        
        # Filter data for selected host
        host_data = self.historical_df[self.historical_df['hostname'] == hostname]
        host_lifecycle = self.lifecycle_df[self.lifecycle_df['hostname'] == hostname]
        
        # Findings over time for this host
        timeline = host_data.groupby('scan_date').size().reset_index(name='count')
        self.host_ax1.plot(timeline['scan_date'], timeline['count'], marker='o', linewidth=2, color='#007bff')
        self.host_ax1.set_title(f'Findings Timeline - {host_str}', fontsize=12, fontweight='bold', color='white')
        self.host_ax1.set_xlabel('Scan Date', color='white')
        self.host_ax1.set_ylabel('Number of Findings', color='white')
        self.host_ax1.grid(True, alpha=0.3)
        self.host_ax1.tick_params(colors='white')
        self.host_ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.setp(self.host_ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        # Severity distribution
        if not host_lifecycle.empty:
            severity_counts = host_lifecycle['severity_text'].value_counts()
            colors = ['#dc3545', '#fd7e14', '#ffc107', '#28a745', '#6c757d']
            self.host_ax2.pie(severity_counts.values, labels=severity_counts.index, autopct='%1.1f%%',
                             colors=colors[:len(severity_counts)], startangle=90)
            self.host_ax2.set_title(f'Severity Distribution - {hostname}', fontsize=12, fontweight='bold', color='white')
        
        # Top 10 plugins for this host
        if not host_lifecycle.empty:
            top_plugins = host_lifecycle.nlargest(10, 'days_open')[['plugin_name', 'days_open']]
            self.host_ax3.barh(range(len(top_plugins)), top_plugins['days_open'].values, color='#dc3545')
            self.host_ax3.set_yticks(range(len(top_plugins)))
            self.host_ax3.set_yticklabels([name[:40] + '...' if len(name) > 40 else name 
                                          for name in top_plugins['plugin_name'].values], color='white')
            self.host_ax3.set_title(f'Top 10 Longest-Running Findings - {hostname}', fontsize=12, fontweight='bold', color='white')
            self.host_ax3.set_xlabel('Days Open', color='white')
            self.host_ax3.grid(True, alpha=0.3, axis='x')
            self.host_ax3.tick_params(colors='white')
        
        # Status breakdown
        if not host_lifecycle.empty:
            status_counts = host_lifecycle['status'].value_counts()
            colors_status = {'Active': '#dc3545', 'Resolved': '#28a745'}
            bars = self.host_ax4.bar(status_counts.index, status_counts.values, 
                             color=[colors_status.get(s, 'gray') for s in status_counts.index])
            self.host_ax4.set_title(f'Finding Status - {hostname}', fontsize=12, fontweight='bold', color='white')
            self.host_ax4.set_ylabel('Count', color='white')
            self.host_ax4.grid(True, alpha=0.3, axis='y')
            self.host_ax4.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                self.host_ax4.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                        f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        self.host_canvas.draw()
    
    def create_plugin_viz(self):
        """Create plugin-specific visualization"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Plugin Analysis")
        self.viz_frames['plugin'] = frame
        
        # Create controls frame
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        
        ttk.Label(control_frame, text="Select Plugin ID:").pack(side=tk.LEFT, padx=5)
        
        # Get unique plugin IDs with names
        plugin_list = self.lifecycle_df[['plugin_id', 'plugin_name']].drop_duplicates()
        plugin_list = plugin_list.sort_values('plugin_id')
        plugin_options = [f"{row['plugin_id']} - {row['plugin_name'][:50]}" 
                         for _, row in plugin_list.iterrows()]
        
        self.plugin_var = tk.StringVar(value=plugin_options[0] if plugin_options else "")
        plugin_combo = ttk.Combobox(control_frame, textvariable=self.plugin_var, 
                                   values=plugin_options, width=60)
        plugin_combo.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(control_frame, text="Update", command=self.update_plugin_viz).pack(side=tk.LEFT, padx=5)
        
        # Create figure
        self.plugin_fig, ((self.plugin_ax1, self.plugin_ax2), (self.plugin_ax3, self.plugin_ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        self.plugin_fig.patch.set_facecolor('#2b2b2b')
        
        self.plugin_canvas = FigureCanvasTkAgg(self.plugin_fig, frame)
        self.plugin_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Initial update
        if plugin_options:
            self.update_plugin_viz()
    
    def update_plugin_viz(self):
        """Update plugin visualization based on selected plugin"""
        plugin_str = self.plugin_var.get()
        if not plugin_str:
            return
        
        plugin_id = plugin_str.split(' - ')[0]
        
        # Clear axes
        for ax in [self.plugin_ax1, self.plugin_ax2, self.plugin_ax3, self.plugin_ax4]:
            ax.clear()
            ax.set_facecolor('#2b2b2b')
        
        # Filter data for selected plugin
        plugin_data = self.historical_df[self.historical_df['plugin_id'] == plugin_id]
        plugin_lifecycle = self.lifecycle_df[self.lifecycle_df['plugin_id'] == plugin_id]
        
        # Affected hosts over time
        timeline = plugin_data.groupby('scan_date')['hostname'].nunique().reset_index(name='host_count')
        self.plugin_ax1.plot(timeline['scan_date'], timeline['host_count'], marker='o', linewidth=2, color='#dc3545')
        self.plugin_ax1.set_title(f'Affected Hosts Over Time - Plugin {plugin_id}', fontsize=12, fontweight='bold', color='white')
        self.plugin_ax1.set_xlabel('Scan Date', color='white')
        self.plugin_ax1.set_ylabel('Number of Hosts', color='white')
        self.plugin_ax1.grid(True, alpha=0.3)
        self.plugin_ax1.tick_params(colors='white')
        self.plugin_ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.setp(self.plugin_ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        # Host status distribution
        if not plugin_lifecycle.empty:
            status_counts = plugin_lifecycle['status'].value_counts()
            colors_status = {'Active': '#dc3545', 'Resolved': '#28a745'}
            self.plugin_ax2.pie(status_counts.values, labels=status_counts.index, autopct='%1.1f%%',
                               colors=[colors_status.get(s, 'gray') for s in status_counts.index], 
                               startangle=90)
            self.plugin_ax2.set_title(f'Status Distribution - Plugin {plugin_id}', fontsize=12, fontweight='bold', color='white')
        
        # Top affected hosts with hostname (IP) format
        if not plugin_lifecycle.empty:
            top_hosts = plugin_lifecycle.nlargest(10, 'days_open')[['hostname', 'ip_address', 'days_open']]
            display_names = [f"{row['hostname']} ({row['ip_address']})" for _, row in top_hosts.iterrows()]
            
            self.plugin_ax3.barh(range(len(top_hosts)), top_hosts['days_open'].values, color='#fd7e14')
            self.plugin_ax3.set_yticks(range(len(top_hosts)))
            self.plugin_ax3.set_yticklabels(display_names, color='white')
            self.plugin_ax3.set_title(f'Top 10 Hosts by Days Open - Plugin {plugin_id}', fontsize=12, fontweight='bold', color='white')
            self.plugin_ax3.set_xlabel('Days Open', color='white')
            self.plugin_ax3.grid(True, alpha=0.3, axis='x')
            self.plugin_ax3.tick_params(colors='white')
        
        # Reappearance statistics
        if not plugin_lifecycle.empty:
            reappeared = len(plugin_lifecycle[plugin_lifecycle['reappearances'] > 0])
            never_reappeared = len(plugin_lifecycle[plugin_lifecycle['reappearances'] == 0])
            bars = self.plugin_ax4.bar(['Reappeared', 'Never Reappeared'], [reappeared, never_reappeared],
                               color=['#dc3545', '#28a745'])
            self.plugin_ax4.set_title(f'Reappearance Statistics - Plugin {plugin_id}', fontsize=12, fontweight='bold', color='white')
            self.plugin_ax4.set_ylabel('Number of Hosts', color='white')
            self.plugin_ax4.grid(True, alpha=0.3, axis='y')
            self.plugin_ax4.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    self.plugin_ax4.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                            f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        self.plugin_canvas.draw()
        
    def create_lifecycle_viz(self):
        """Create lifecycle analysis visualization with dark theme"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Lifecycle Analysis")
        self.viz_frames['lifecycle'] = frame
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        # Distribution of days open
        if not self.lifecycle_df.empty:
            max_days = self.lifecycle_df['days_open'].max()
            
            # Create bins that are guaranteed to be monotonically increasing
            base_bins = [0, 30, 60, 90, 120, 180, 365]
            bins = [b for b in base_bins if b <= max_days]
            bins.append(max_days + 1)  # Always add max as the final bin
            
            # Create labels for the bins
            labels = []
            for i in range(len(bins) - 1):
                if i == len(bins) - 2:  # Last label
                    if bins[i] >= 365:
                        labels.append('365+')
                    else:
                        labels.append(f'{bins[i]}-{int(max_days)}')
                else:
                    labels.append(f'{bins[i]}-{bins[i+1]-1}')
            
            self.lifecycle_df['age_category'] = pd.cut(self.lifecycle_df['days_open'], 
                                                        bins=bins, labels=labels, right=False)
            age_dist = self.lifecycle_df['age_category'].value_counts().sort_index()
            
            bars = ax1.bar(range(len(age_dist)), age_dist.values, color='#007bff')
            ax1.set_xticks(range(len(age_dist)))
            ax1.set_xticklabels(age_dist.index, rotation=45, ha='right', color='white')
            ax1.set_title('Distribution of Finding Age (Days Open)', fontsize=12, fontweight='bold', color='white')
            ax1.set_ylabel('Number of Findings', color='white')
            ax1.grid(True, alpha=0.3, axis='y')
            ax1.set_facecolor('#2b2b2b')
            ax1.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax1.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                            f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        
        # Top 15 most prevalent plugins
        if not self.lifecycle_df.empty:
            top_plugins = self.lifecycle_df['plugin_id'].value_counts().head(15)
            plugin_names = [self.lifecycle_df[self.lifecycle_df['plugin_id'] == pid]['plugin_name'].iloc[0][:30] 
                        for pid in top_plugins.index]
            
            bars = ax2.barh(range(len(top_plugins)), top_plugins.values, color='#dc3545')
            ax2.set_yticks(range(len(top_plugins)))
            ax2.set_yticklabels(plugin_names, color='white')
            ax2.set_title('Top 15 Most Prevalent Findings', fontsize=12, fontweight='bold', color='white')
            ax2.set_xlabel('Number of Affected Hosts', color='white')
            ax2.grid(True, alpha=0.3, axis='x')
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            
            # Add data labels on horizontal bars
            for bar in bars:
                width = bar.get_width()
                ax2.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                        f'{int(width)}', ha='left', va='center', color='white', fontweight='bold')
        
        # Findings by first seen date
        if not self.lifecycle_df.empty:
            first_seen_timeline = self.lifecycle_df.groupby(
                self.lifecycle_df['first_seen'].dt.to_period('M')
            ).size().reset_index(name='count')
            first_seen_timeline['first_seen'] = first_seen_timeline['first_seen'].dt.to_timestamp()
            
            ax3.plot(first_seen_timeline['first_seen'], first_seen_timeline['count'], 
                    marker='o', linewidth=2, color='#28a745')
            ax3.set_title('New Findings by Month (First Seen)', fontsize=12, fontweight='bold', color='white')
            ax3.set_xlabel('Month', color='white')
            ax3.set_ylabel('New Findings', color='white')
            ax3.grid(True, alpha=0.3)
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            ax3.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
            plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        # Remediation effectiveness (resolved findings by severity)
        if not self.lifecycle_df.empty:
            resolved = self.lifecycle_df[self.lifecycle_df['status'] == 'Resolved']
            if not resolved.empty:
                remediation = resolved.groupby('severity_text').size().sort_values(ascending=True)
                colors_sev = {'Critical': '#dc3545', 'High': '#fd7e14', 'Medium': '#ffc107', 
                            'Low': '#007bff', 'Info': '#6c757d'}
                
                bars = ax4.barh(range(len(remediation)), remediation.values,
                        color=[colors_sev.get(s, 'gray') for s in remediation.index])
                ax4.set_yticks(range(len(remediation)))
                ax4.set_yticklabels(remediation.index, color='white')
                ax4.set_title('Resolved Findings by Severity', fontsize=12, fontweight='bold', color='white')
                ax4.set_xlabel('Number of Findings Resolved', color='white')
                ax4.grid(True, alpha=0.3, axis='x')
                ax4.set_facecolor('#2b2b2b')
                ax4.tick_params(colors='white')
                
                # Add data labels on horizontal bars
                for bar in bars:
                    width = bar.get_width()
                    ax4.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                            f'{int(width)}', ha='left', va='center', color='white', fontweight='bold')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def run(self):
        """Run the enhanced GUI"""
        self.window.mainloop()


def main():
    """Main entry point for the enhanced application"""
    app = EnhancedHistoricalAnalysisGUI()
    app.run()


if __name__ == "__main__":
    main())[0]
            opdir_df['opdir_sequence'] = opdir_df['opdir_number'].astype(str).str.extract(r'^(\d+)-')[0]
        
        # Create normalized IAVx reference for matching
        if 'iavx_reference' in opdir_df.columns:
            # Extract the base reference (e.g., "B-0146" from "IAVB-2025-B-0146")
            opdir_df['iavx_normalized'] = opdir_df['iavx_reference'].astype(str).str.extract(r'([A-Z]-\d+)')[0]
            
            # Create full IAVx format for matching (e.g., "IAVB:2025-B-0146")
            opdir_df['iavx_full'] = opdir_df.apply(lambda row: 
                f"IAV{row['iavx_reference'][0]}:20{row['opdir_year']}-{row['iavx_normalized']}" 
                if pd.notna(row.get('iavx_normalized')) and pd.notna(row.get('opdir_year')) 
                else None, axis=1)
        
        # Convert date columns
        date_columns = ['release_date', 'acknowledge_date', 'poam_due_date', 'final_due_date']
        for col in date_columns:
            if col in opdir_df.columns:
                opdir_df[col] = pd.to_datetime(opdir_df[col], errors='coerce')
        
        print(f"Processed OPDIR data with {len(opdir_df)} valid records")
        return opdir_df
        
    except Exception as e:
        print(f"Error loading OPDIR file: {e}")
        return pd.DataFrame()


def enrich_findings_with_opdir(findings_df: pd.DataFrame, opdir_df: pd.DataFrame) -> pd.DataFrame:
    """
    Enrich findings with OPDIR data based on IAVx references.
    
    Args:
        findings_df: Historical findings DataFrame
        opdir_df: OPDIR DataFrame
        
    Returns:
        Enhanced findings DataFrame with OPDIR information
    """
    if opdir_df.empty or findings_df.empty:
        return findings_df
    
    enriched_df = findings_df.copy()
    
    # Initialize OPDIR columns
    enriched_df['opdir_number'] = None
    enriched_df['opdir_subject'] = None
    enriched_df['opdir_release_date'] = None
    enriched_df['opdir_poam_due_date'] = None
    enriched_df['opdir_final_due_date'] = None
    enriched_df['opdir_days_to_due'] = None
    
    # Match IAVx references
    for idx, finding in enriched_df.iterrows():
        if pd.notna(finding.get('iavx')) and finding['iavx']:
            iavx_list = finding['iavx'].split('\n')
            
            for iavx in iavx_list:
                iavx = iavx.strip()
                
                # Try to match with OPDIR data
                matching_opdir = opdir_df[
                    (opdir_df['iavx_full'].notna()) & 
                    (opdir_df['iavx_full'].str.contains(iavx.replace(':', ':'), regex=False, na=False))
                ]
                
                if not matching_opdir.empty:
                    opdir_record = matching_opdir.iloc[0]
                    enriched_df.at[idx, 'opdir_number'] = opdir_record.get('opdir_number')
                    enriched_df.at[idx, 'opdir_subject'] = opdir_record.get('subject')
                    enriched_df.at[idx, 'opdir_release_date'] = opdir_record.get('release_date')
                    enriched_df.at[idx, 'opdir_poam_due_date'] = opdir_record.get('poam_due_date')
                    enriched_df.at[idx, 'opdir_final_due_date'] = opdir_record.get('final_due_date')
                    
                    # Calculate days to due date
                    if pd.notna(opdir_record.get('final_due_date')):
                        days_to_due = (opdir_record['final_due_date'] - pd.Timestamp.now()).days
                        enriched_df.at[idx, 'opdir_days_to_due'] = days_to_due
                    
                    break  # Use first match
    
    print(f"Enriched {len(enriched_df[enriched_df['opdir_number'].notna()])} findings with OPDIR data")
    return enriched_df
    """
    Extract scan date from filename using common patterns.
    
    Args:
        filename: Name of the file
        
    Returns:
        datetime object or None
    """
    # Common date patterns in filenames
    patterns = [
        r'(\d{4})[-_](\d{2})[-_](\d{2})',  # YYYY-MM-DD or YYYY_MM_DD
        r'(\d{2})[-_](\d{2})[-_](\d{4})',  # MM-DD-YYYY or MM_DD_YYYY
        r'(\d{8})',  # YYYYMMDD
        r'(\d{6})',  # YYMMDD or MMDDYY
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            try:
                if len(match.groups()) == 3:
                    # YYYY-MM-DD or MM-DD-YYYY
                    parts = match.groups()
                    if len(parts[0]) == 4:  # YYYY-MM-DD
                        return datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                    else:  # MM-DD-YYYY
                        return datetime(int(parts[2]), int(parts[0]), int(parts[1]))
                elif len(match.group(1)) == 8:  # YYYYMMDD
                    date_str = match.group(1)
                    return datetime.strptime(date_str, '%Y%m%d')
                elif len(match.group(1)) == 6:  # Try both YYMMDD and MMDDYY
                    date_str = match.group(1)
                    try:
                        return datetime.strptime(date_str, '%y%m%d')
                    except ValueError:
                        return datetime.strptime(date_str, '%m%d%y')
            except (ValueError, IndexError):
                continue
    
    return None


def extract_scan_date_from_nessus(nessus_file: str) -> Optional[datetime]:
    """
    Extract scan date from .nessus file content.
    
    Args:
        nessus_file: Path to .nessus file
        
    Returns:
        datetime object or None
    """
    try:
        import xml.etree.ElementTree as ET
        tree = ET.parse(nessus_file)
        root = tree.getroot()
        
        # Try to find scan end date
        policy_elem = root.find(".//Policy/policyName")
        if policy_elem is not None and policy_elem.text:
            date = extract_scan_date_from_filename(policy_elem.text)
            if date:
                return date
        
        # Try Report name attribute
        report_elem = root.find(".//Report")
        if report_elem is not None and 'name' in report_elem.attrib:
            date = extract_scan_date_from_filename(report_elem.attrib['name'])
            if date:
                return date
        
        # Try HOST_END timestamps
        host_end = root.find(".//tag[@name='HOST_END']")
        if host_end is not None and host_end.text:
            try:
                timestamp = int(host_end.text)
                return datetime.fromtimestamp(timestamp)
            except (ValueError, TypeError):
                pass
        
    except Exception as e:
        print(f"Error extracting date from {nessus_file}: {e}")
    
    return None


def create_host_presence_analysis(historical_df: pd.DataFrame) -> pd.DataFrame:
    """
    Analyze host presence across scans to identify missing hosts.
    
    Args:
        historical_df: DataFrame with historical findings
        
    Returns:
        DataFrame with host presence analysis
    """
    if historical_df.empty:
        return pd.DataFrame()
    
    # Get all unique combinations of scan dates and hosts
    scan_dates = sorted(historical_df['scan_date'].unique())
    all_hosts = historical_df.groupby(['hostname', 'ip_address']).size().reset_index(name='count')
    
    presence_records = []
    
    for _, host_row in all_hosts.iterrows():
        hostname = host_row['hostname']
        ip_address = host_row['ip_address']
        
        # Get scan dates where this host appeared
        host_scans = historical_df[
            (historical_df['hostname'] == hostname) & 
            (historical_df['ip_address'] == ip_address)
        ]['scan_date'].unique()
        
        first_seen = min(host_scans)
        last_seen = max(host_scans)
        total_scans = len(scan_dates)
        present_scans = len(host_scans)
        missing_scans = total_scans - present_scans
        
        # Identify missing scan periods
        missing_periods = []
        for scan_date in scan_dates:
            if scan_date not in host_scans:
                missing_periods.append(scan_date.strftime('%Y-%m-%d'))
        
        # Determine status
        if last_seen == max(scan_dates):
            status = 'Active'
        else:
            status = 'Missing'
        
        # Calculate presence percentage
        presence_percentage = (present_scans / total_scans) * 100 if total_scans > 0 else 0
        
        presence_records.append({
            'hostname': hostname,
            'ip_address': ip_address,
            'first_seen': first_seen,
            'last_seen': last_seen,
            'total_scans_available': total_scans,
            'scans_present': present_scans,
            'scans_missing': missing_scans,
            'presence_percentage': round(presence_percentage, 1),
            'status': status,
            'missing_scan_dates': ', '.join(missing_periods) if missing_periods else 'None'
        })
    
    presence_df = pd.DataFrame(presence_records)
    presence_df = presence_df.sort_values(['status', 'presence_percentage'], ascending=[True, False])
    
    return presence_df


def analyze_scan_changes(historical_df: pd.DataFrame) -> pd.DataFrame:
    """
    Analyze changes between consecutive scans (hosts added/removed).
    
    Args:
        historical_df: DataFrame with historical findings
        
    Returns:
        DataFrame with scan-to-scan changes
    """
    if historical_df.empty:
        return pd.DataFrame()
    
    scan_dates = sorted(historical_df['scan_date'].unique())
    change_records = []
    
    for i in range(1, len(scan_dates)):
        prev_scan = scan_dates[i-1]
        curr_scan = scan_dates[i]
        
        # Get hosts in each scan
        prev_hosts = set(historical_df[historical_df['scan_date'] == prev_scan]['hostname'].unique())
        curr_hosts = set(historical_df[historical_df['scan_date'] == curr_scan]['hostname'].unique())
        
        # Find changes
        added_hosts = curr_hosts - prev_hosts
        removed_hosts = prev_hosts - curr_hosts
        unchanged_hosts = curr_hosts & prev_hosts
        
        change_records.append({
            'scan_date': curr_scan,
            'previous_scan': prev_scan,
            'hosts_added': len(added_hosts),
            'hosts_removed': len(removed_hosts),
            'hosts_unchanged': len(unchanged_hosts),
            'total_hosts_current': len(curr_hosts),
            'total_hosts_previous': len(prev_hosts),
            'net_change': len(curr_hosts) - len(prev_hosts),
            'added_host_list': ', '.join(sorted(added_hosts)) if added_hosts else 'None',
            'removed_host_list': ', '.join(sorted(removed_hosts)) if removed_hosts else 'None'
        })
    
    changes_df = pd.DataFrame(change_records)
    return changes_df


def load_existing_database(db_path: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Load existing historical database.
    
    Args:
        db_path: Path to SQLite database file
        
    Returns:
        Tuple of (historical_df, lifecycle_df, host_presence_df, scan_changes_df)
    """
    try:
        if not os.path.exists(db_path):
            print(f"Database file {db_path} does not exist")
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        
        conn = sqlite3.connect(db_path)
        
        # Load historical findings
        try:
            historical_df = pd.read_sql_query("SELECT * FROM historical_findings", conn)
            historical_df['scan_date'] = pd.to_datetime(historical_df['scan_date'])
            print(f"Loaded {len(historical_df)} historical findings from database")
        except Exception as e:
            print(f"Error loading historical_findings: {e}")
            historical_df = pd.DataFrame()
        
        # Load lifecycle analysis
        try:
            lifecycle_df = pd.read_sql_query("SELECT * FROM finding_lifecycle", conn)
            lifecycle_df['first_seen'] = pd.to_datetime(lifecycle_df['first_seen'])
            lifecycle_df['last_seen'] = pd.to_datetime(lifecycle_df['last_seen'])
            print(f"Loaded {len(lifecycle_df)} lifecycle records from database")
        except Exception as e:
            print(f"Error loading finding_lifecycle: {e}")
            lifecycle_df = pd.DataFrame()
        
        # Load host presence analysis
        try:
            host_presence_df = pd.read_sql_query("SELECT * FROM host_presence", conn)
            host_presence_df['first_seen'] = pd.to_datetime(host_presence_df['first_seen'])
            host_presence_df['last_seen'] = pd.to_datetime(host_presence_df['last_seen'])
            print(f"Loaded {len(host_presence_df)} host presence records from database")
        except Exception as e:
            print(f"Error loading host_presence: {e}")
            host_presence_df = pd.DataFrame()
        
        # Load scan changes
        try:
            scan_changes_df = pd.read_sql_query("SELECT * FROM scan_changes", conn)
            scan_changes_df['scan_date'] = pd.to_datetime(scan_changes_df['scan_date'])
            scan_changes_df['previous_scan'] = pd.to_datetime(scan_changes_df['previous_scan'])
            print(f"Loaded {len(scan_changes_df)} scan change records from database")
        except Exception as e:
            print(f"Error loading scan_changes: {e}")
            scan_changes_df = pd.DataFrame()
        
        conn.close()
        return historical_df, lifecycle_df, host_presence_df, scan_changes_df
        
    except Exception as e:
        print(f"Error loading existing database: {e}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()


def check_for_duplicates(new_df: pd.DataFrame, existing_df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    """
    Check for and remove duplicate scans based on scan_date and scan_file.
    
    Args:
        new_df: New findings DataFrame
        existing_df: Existing findings DataFrame
        
    Returns:
        Tuple of (filtered_new_df, duplicate_count)
    """
    if existing_df.empty:
        return new_df, 0
    
    # Create composite key for duplicate detection
    new_df['composite_key'] = new_df['scan_date'].astype(str) + '|' + new_df['scan_file'].astype(str)
    existing_df['composite_key'] = existing_df['scan_date'].astype(str) + '|' + existing_df['scan_file'].astype(str)
    
    # Find duplicates
    duplicate_keys = set(existing_df['composite_key'].unique())
    new_unique_mask = ~new_df['composite_key'].isin(duplicate_keys)
    
    filtered_new_df = new_df[new_unique_mask].copy()
    duplicate_count = len(new_df) - len(filtered_new_df)
    
    # Remove the temporary composite key
    filtered_new_df = filtered_new_df.drop('composite_key', axis=1)
    
    return filtered_new_df, duplicate_count


def process_historical_scans(archive_paths: List[str], plugins_db_path: Optional[str] = None, 
                           existing_db_path: Optional[str] = None, include_info: bool = False) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Process multiple Nessus archives and track findings history with enhanced features.
    
    Args:
        archive_paths: List of paths to Nessus archives or .nessus files
        plugins_db_path: Optional path to plugins database
        existing_db_path: Optional path to existing database to update
        include_info: Whether to include Info-level findings in analysis (default: False)
        
    Returns:
        Tuple of (historical_df, lifecycle_df, host_presence_df, scan_changes_df)
    """
    import tempfile
    
    # Load existing database if provided
    existing_historical_df = pd.DataFrame()
    if existing_db_path and os.path.exists(existing_db_path):
        print(f"Loading existing database from {existing_db_path}")
        existing_historical_df, _, _, _ = load_existing_database(existing_db_path)
    
    # Load plugins database if provided
    plugins_dict = None
    if plugins_db_path:
        plugins_dict = load_plugins_database(plugins_db_path)
    
    all_historical_findings = []
    
    for archive_path in archive_paths:
        print(f"\n{'='*60}")
        print(f"Processing: {os.path.basename(archive_path)}")
        print(f"{'='*60}")
        
        temp_dir = None
        nessus_files = []
        
        try:
            # Determine scan date from filename first
            scan_date = extract_scan_date_from_filename(os.path.basename(archive_path))
            
            # Handle different input types
            if archive_path.lower().endswith('.zip'):
                temp_dir = tempfile.mkdtemp()
                extract_nested_archives(archive_path, temp_dir)
                nessus_files = find_files_by_extension(temp_dir, '.nessus')
            elif archive_path.lower().endswith('.nessus'):
                nessus_files = [archive_path]
            else:
                print(f"Unsupported file format: {archive_path}")
                continue
            
            if not nessus_files:
                print(f"No .nessus files found in {archive_path}")
                continue
            
            # If no date from filename, try extracting from first .nessus file
            if not scan_date and nessus_files:
                scan_date = extract_scan_date_from_nessus(nessus_files[0])
            
            # Fallback to file modification time
            if not scan_date:
                scan_date = datetime.fromtimestamp(os.path.getmtime(archive_path))
                print(f"Warning: Using file modification time for scan date: {scan_date.strftime('%Y-%m-%d')}")
            else:
                print(f"Scan date: {scan_date.strftime('%Y-%m-%d')}")
            
            # Parse nessus files with improved host display
            findings_df, host_summary_df = parse_multiple_nessus_files(nessus_files, plugins_dict)
            
            if not findings_df.empty:
                # Add scan metadata
                findings_df['scan_date'] = scan_date
                findings_df['scan_file'] = os.path.basename(archive_path)
                
                # Enrich with severity
                findings_df = enrich_findings_with_severity(findings_df)
                
                all_historical_findings.append(findings_df)
                
                # Display with hostname (IP) format
                for _, host in host_summary_df.iterrows():
                    hostname = host.get('hostname', 'Unknown')
                    ip = host.get('host_name', 'Unknown')
                    print(f"  Processed: {hostname} ({ip})")
                print(f"Extracted {len(findings_df)} findings from {len(host_summary_df)} hosts")
            
        finally:
            if temp_dir:
                cleanup_temp_directory(temp_dir)
    
    if not all_historical_findings:
        print("No findings extracted from any archive")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    # Combine all new findings
    new_historical_df = pd.concat(all_historical_findings, ignore_index=True)
    
    # Check for duplicates if we have existing data
    duplicate_count = 0
    if not existing_historical_df.empty:
        print(f"\nChecking for duplicates against existing database...")
        new_historical_df, duplicate_count = check_for_duplicates(new_historical_df, existing_historical_df)
        if duplicate_count > 0:
            print(f"Filtered out {duplicate_count} duplicate findings")
    
    # Combine with existing data
    if not existing_historical_df.empty and not new_historical_df.empty:
        combined_df = pd.concat([existing_historical_df, new_historical_df], ignore_index=True)
    elif not existing_historical_df.empty:
        combined_df = existing_historical_df
    else:
        combined_df = new_historical_df
    
    combined_df = combined_df.sort_values('scan_date')
    
    print(f"\n{'='*60}")
    print(f"Total findings across all scans: {len(combined_df)}")
    if not new_historical_df.empty:
        print(f"New findings added: {len(new_historical_df)}")
    if duplicate_count > 0:
        print(f"Duplicate findings filtered: {duplicate_count}")
    print(f"Date range: {combined_df['scan_date'].min()} to {combined_df['scan_date'].max()}")
    print(f"{'='*60}\n")
    
    # Perform analysis (filter Info findings for analysis only, not database storage)
    analysis_df = combined_df.copy()
    if not include_info:
        analysis_df = analysis_df[analysis_df['severity_text'] != 'Info']
        print(f"Filtered out {len(combined_df) - len(analysis_df)} Info-level findings from analysis")
    
    lifecycle_df = analyze_finding_lifecycle(analysis_df)
    host_presence_df = create_host_presence_analysis(analysis_df)
    scan_changes_df = analyze_scan_changes(analysis_df)
    
    return combined_df, lifecycle_df, host_presence_df, scan_changes_df


def analyze_finding_lifecycle(historical_df: pd.DataFrame) -> pd.DataFrame:
    """
    Analyze the lifecycle of findings (first seen, last seen, resolved, reappeared).
    
    Args:
        historical_df: DataFrame with historical findings
        
    Returns:
        DataFrame with lifecycle analysis
    """
    if historical_df.empty:
        return pd.DataFrame()
    
    # Create unique finding identifier
    historical_df['finding_key'] = (
        historical_df['hostname'].astype(str) + '|' + 
        historical_df['plugin_id'].astype(str)
    )
    
    lifecycle_records = []
    
    # Group by finding_key
    for finding_key, group in historical_df.groupby('finding_key'):
        hostname, plugin_id = finding_key.split('|')
        
        # Sort by scan date
        group = group.sort_values('scan_date')
        
        scan_dates = group['scan_date'].tolist()
        scan_files = group['scan_file'].tolist()
        
        # Get finding details from most recent observation
        latest = group.iloc[-1]
        
        first_seen = scan_dates[0]
        last_seen = scan_dates[-1]
        total_observations = len(scan_dates)
        
        # Check for gaps (resolved then reappeared)
        gaps = []
        reappearances = 0

        if len(scan_dates) > 1:
            for i in range(1, len(scan_dates)):
                days_gap = (scan_dates[i] - scan_dates[i-1]).days
                if days_gap > 45:  # More than 45 days gap suggests resolution
                    gaps.append({
                        'resolved_after': scan_dates[i-1].strftime('%Y-%m-%d'),
                        'reappeared_on': scan_dates[i].strftime('%Y-%m-%d'),
                        'days_resolved': days_gap
                    })
                    reappearances += 1
        
        # Determine current status
        if latest['scan_date'] == historical_df['scan_date'].max():
            status = 'Active'
        else:
            status = 'Resolved'
        
        lifecycle_records.append({
            'hostname': hostname,
            'ip_address': latest.get('ip_address', ''),
            'plugin_id': plugin_id,
            'plugin_name': latest['name'],
            'severity_text': latest.get('severity_text', 'Unknown'),
            'severity_value': latest.get('severity_value', 0),
            'first_seen': first_seen,
            'last_seen': last_seen,
            'days_open': (last_seen - first_seen).days,
            'total_observations': total_observations,
            'reappearances': reappearances,
            'status': status,
            'gap_details': json.dumps(gaps) if gaps else '',
            'cvss3_base_score': latest.get('cvss3_base_score'),
            'cves': latest.get('cves', ''),
            'iavx': latest.get('iavx', '')
        })
    
    lifecycle_df = pd.DataFrame(lifecycle_records)
    lifecycle_df = lifecycle_df.sort_values(['severity_value', 'days_open'], ascending=[False, False])
    
    return lifecycle_df


def export_to_sqlite(historical_df: pd.DataFrame, lifecycle_df: pd.DataFrame, 
                     host_presence_df: pd.DataFrame, scan_changes_df: pd.DataFrame,
                     db_path: str) -> None:
    """
    Export data to SQLite database with enhanced tables.
    
    Args:
        historical_df: Historical findings DataFrame
        lifecycle_df: Lifecycle analysis DataFrame
        host_presence_df: Host presence analysis DataFrame
        scan_changes_df: Scan changes DataFrame
        db_path: Path to SQLite database file
    """
    try:
        conn = sqlite3.connect(db_path)
        
        # Export all DataFrames
        historical_df.to_sql('historical_findings', conn, if_exists='replace', index=False)
        lifecycle_df.to_sql('finding_lifecycle', conn, if_exists='replace', index=False)
        host_presence_df.to_sql('host_presence', conn, if_exists='replace', index=False)
        scan_changes_df.to_sql('scan_changes', conn, if_exists='replace', index=False)
        
        # Create indexes for better query performance
        cursor = conn.cursor()
        
        # Historical findings indexes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_historical_hostname ON historical_findings(hostname)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_historical_plugin ON historical_findings(plugin_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_historical_date ON historical_findings(scan_date)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_historical_ip ON historical_findings(ip_address)')
        
        # Lifecycle indexes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_lifecycle_hostname ON finding_lifecycle(hostname)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_lifecycle_plugin ON finding_lifecycle(plugin_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_lifecycle_status ON finding_lifecycle(status)')
        
        # Host presence indexes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_host_presence_hostname ON host_presence(hostname)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_host_presence_status ON host_presence(status)')
        
        # Scan changes indexes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_scan_changes_date ON scan_changes(scan_date)')
        
        conn.commit()
        conn.close()
        
        print(f"Successfully exported to SQLite database: {db_path}")
        
    except Exception as e:
        print(f"Error exporting to SQLite: {e}")
        import traceback
        traceback.print_exc()


def export_to_json(historical_df: pd.DataFrame, lifecycle_df: pd.DataFrame, 
                   host_presence_df: pd.DataFrame, scan_changes_df: pd.DataFrame,
                   json_path: str) -> None:
    """
    Export data to JSON file with enhanced structure.
    
    Args:
        historical_df: Historical findings DataFrame
        lifecycle_df: Lifecycle analysis DataFrame
        host_presence_df: Host presence analysis DataFrame
        scan_changes_df: Scan changes DataFrame
        json_path: Path to JSON output file
    """
    try:
        # Convert datetime columns to ISO format strings before export
        historical_export = historical_df.copy()
        lifecycle_export = lifecycle_df.copy()
        host_presence_export = host_presence_df.copy()
        scan_changes_export = scan_changes_df.copy()
        
        # Convert datetime columns in historical_df
        if 'scan_date' in historical_export.columns:
            historical_export['scan_date'] = historical_export['scan_date'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Convert datetime columns in lifecycle_df
        if 'first_seen' in lifecycle_export.columns:
            lifecycle_export['first_seen'] = lifecycle_export['first_seen'].dt.strftime('%Y-%m-%d %H:%M:%S')
        if 'last_seen' in lifecycle_export.columns:
            lifecycle_export['last_seen'] = lifecycle_export['last_seen'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Convert datetime columns in host_presence_df
        if 'first_seen' in host_presence_export.columns:
            host_presence_export['first_seen'] = host_presence_export['first_seen'].dt.strftime('%Y-%m-%d %H:%M:%S')
        if 'last_seen' in host_presence_export.columns:
            host_presence_export['last_seen'] = host_presence_export['last_seen'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Convert datetime columns in scan_changes_df
        if 'scan_date' in scan_changes_export.columns:
            scan_changes_export['scan_date'] = scan_changes_export['scan_date'].dt.strftime('%Y-%m-%d %H:%M:%S')
        if 'previous_scan' in scan_changes_export.columns:
            scan_changes_export['previous_scan'] = scan_changes_export['previous_scan'].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Replace NaN with None for JSON compatibility
        historical_export = historical_export.where(pd.notnull(historical_export), None)
        lifecycle_export = lifecycle_export.where(pd.notnull(lifecycle_export), None)
        host_presence_export = host_presence_export.where(pd.notnull(host_presence_export), None)
        scan_changes_export = scan_changes_export.where(pd.notnull(scan_changes_export), None)
        
        export_data = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'total_findings': len(historical_df),
                'unique_findings': len(lifecycle_df),
                'total_hosts': len(host_presence_df),
                'total_scans': len(scan_changes_df) + 1 if len(scan_changes_df) > 0 else len(historical_df['scan_date'].unique()) if not historical_df.empty else 0,
                'date_range': {
                    'start': historical_df['scan_date'].min().isoformat() if not historical_df.empty and 'scan_date' in historical_df.columns else None,
                    'end': historical_df['scan_date'].max().isoformat() if not historical_df.empty and 'scan_date' in historical_df.columns else None
                }
            },
            'historical_findings': historical_export.to_dict('records'),
            'lifecycle_analysis': lifecycle_export.to_dict('records'),
            'host_presence': host_presence_export.to_dict('records'),
            'scan_changes': scan_changes_export.to_dict('records')
        }
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        print(f"Successfully exported to JSON: {json_path}")
        
    except Exception as e:
        print(f"Error exporting to JSON: {e}")
        import traceback
        traceback.print_exc()


def export_json_schema(schema_path: str) -> None:
    """
    Export JSON schema definition for the export format.
    
    Args:
        schema_path: Path to save the JSON schema file
    """
    schema = {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "title": "Nessus Historical Analysis Export",
        "type": "object",
        "properties": {
            "metadata": {
                "type": "object",
                "properties": {
                    "export_date": {"type": "string", "format": "date-time"},
                    "total_findings": {"type": "integer"},
                    "unique_findings": {"type": "integer"},
                    "total_hosts": {"type": "integer"},
                    "total_scans": {"type": "integer"},
                    "date_range": {
                        "type": "object",
                        "properties": {
                            "start": {"type": ["string", "null"], "format": "date-time"},
                            "end": {"type": ["string", "null"], "format": "date-time"}
                        }
                    }
                }
            },
            "historical_findings": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "plugin_id": {"type": "string"},
                        "hostname": {"type": "string"},
                        "ip_address": {"type": "string"},
                        "scan_date": {"type": "string", "format": "date-time"},
                        "scan_file": {"type": "string"},
                        "name": {"type": "string"},
                        "severity_text": {"type": "string"},
                        "severity_value": {"type": "integer"},
                        "cvss3_base_score": {"type": ["string", "null"]},
                        "cvss2_base_score": {"type": ["string", "null"]},
                        "cves": {"type": ["string", "null"]},
                        "iavx": {"type": ["string", "null"]}
                    }
                }
            },
            "lifecycle_analysis": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "hostname": {"type": "string"},
                        "ip_address": {"type": "string"},
                        "plugin_id": {"type": "string"},
                        "plugin_name": {"type": "string"},
                        "severity_text": {"type": "string"},
                        "severity_value": {"type": "integer"},
                        "first_seen": {"type": "string", "format": "date-time"},
                        "last_seen": {"type": "string", "format": "date-time"},
                        "days_open": {"type": "integer"},
                        "total_observations": {"type": "integer"},
                        "reappearances": {"type": "integer"},
                        "status": {"type": "string", "enum": ["Active", "Resolved"]},
                        "gap_details": {"type": "string"}
                    }
                }
            },
            "host_presence": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "hostname": {"type": "string"},
                        "ip_address": {"type": "string"},
                        "first_seen": {"type": "string", "format": "date-time"},
                        "last_seen": {"type": "string", "format": "date-time"},
                        "total_scans_available": {"type": "integer"},
                        "scans_present": {"type": "integer"},
                        "scans_missing": {"type": "integer"},
                        "presence_percentage": {"type": "number"},
                        "status": {"type": "string", "enum": ["Active", "Missing"]},
                        "missing_scan_dates": {"type": "string"}
                    }
                }
            },
            "scan_changes": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "scan_date": {"type": "string", "format": "date-time"},
                        "previous_scan": {"type": "string", "format": "date-time"},
                        "hosts_added": {"type": "integer"},
                        "hosts_removed": {"type": "integer"},
                        "hosts_unchanged": {"type": "integer"},
                        "total_hosts_current": {"type": "integer"},
                        "total_hosts_previous": {"type": "integer"},
                        "net_change": {"type": "integer"},
                        "added_host_list": {"type": "string"},
                        "removed_host_list": {"type": "string"}
                    }
                }
            }
        }
    }
    
    try:
        with open(schema_path, 'w', encoding='utf-8') as f:
            json.dump(schema, f, indent=2, ensure_ascii=False)
        
        print(f"Successfully exported JSON schema: {schema_path}")
        
    except Exception as e:
        print(f"Error exporting JSON schema: {e}")


def export_to_excel(historical_df: pd.DataFrame, lifecycle_df: pd.DataFrame, 
                    host_presence_df: pd.DataFrame, scan_changes_df: pd.DataFrame,
                    excel_path: str) -> None:
    """
    Export data to Excel with multiple sheets and formatting.
    
    Args:
        historical_df: Historical findings DataFrame
        lifecycle_df: Lifecycle analysis DataFrame
        host_presence_df: Host presence analysis DataFrame
        scan_changes_df: Scan changes DataFrame
        excel_path: Path to Excel output file
    """
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Export sheets
            lifecycle_df.to_excel(writer, sheet_name='Finding_Lifecycle', index=False)
            host_presence_df.to_excel(writer, sheet_name='Host_Presence', index=False)
            scan_changes_df.to_excel(writer, sheet_name='Scan_Changes', index=False)
            historical_df.to_excel(writer, sheet_name='Historical_Data', index=False)
            
            # Create summary statistics
            include_info_val = getattr(self, 'include_info', tk.BooleanVar(value=False)).get()
            info_status = "Included" if include_info_val else "Excluded from Analysis"
            
            summary_data = {
                'Metric': [
                    'Analysis Mode',
                    'Total Scans',
                    'Total Findings (Database)',
                    'Info Findings (Database)',
                    'Findings in Analysis',
                    'Unique Findings',
                    'Active Findings',
                    'Resolved Findings',
                    'Reappeared Findings',
                    'Total Hosts Ever Seen',
                    'Currently Active Hosts',
                    'Missing Hosts',
                    'Critical Findings',
                    'High Findings',
                    'Medium Findings',
                    'Low Findings'
                ],
                'Count': [
                    f"Info findings {info_status}",
                    historical_df['scan_date'].nunique() if not historical_df.empty else 0,
                    len(historical_df),
                    len(historical_df[historical_df['severity_text'] == 'Info']) if not historical_df.empty else 0,
                    len(lifecycle_df),
                    len(lifecycle_df),
                    len(lifecycle_df[lifecycle_df['status'] == 'Active']) if not lifecycle_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['status'] == 'Resolved']) if not lifecycle_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['reappearances'] > 0]) if not lifecycle_df.empty else 0,
                    len(host_presence_df),
                    len(host_presence_df[host_presence_df['status'] == 'Active']) if not host_presence_df.empty else 0,
                    len(host_presence_df[host_presence_df['status'] == 'Missing']) if not host_presence_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['severity_text'] == 'Critical']) if not lifecycle_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['severity_text'] == 'High']) if not lifecycle_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['severity_text'] == 'Medium']) if not lifecycle_df.empty else 0,
                    len(lifecycle_df[lifecycle_df['severity_text'] == 'Low']) if not lifecycle_df.empty else 0
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Auto-fit columns and add filters
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Auto-fit columns
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add autofilter and banding
                if worksheet.max_row > 0:
                    worksheet.auto_filter.ref = worksheet.dimensions
                    
                    # Add table formatting with banding
                    if worksheet.max_row > 1:
                        from openpyxl.worksheet.table import Table, TableStyleInfo
                        tab = Table(displayName=f"Table_{sheet_name}", ref=worksheet.dimensions)
                        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                             showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                        tab.tableStyleInfo = style
                        worksheet.add_table(tab)
        
        print(f"Successfully exported to Excel: {excel_path}")
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()


class EnhancedHistoricalAnalysisGUI:
    """Enhanced GUI for Nessus Historical Analysis System with host tracking"""
    
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Enhanced Nessus Historical Analysis System")
        self.window.geometry("1400x900")
        self.window.configure(bg='#2b2b2b')  # Dark theme background
        
        # Configure style for dark theme
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure('.', background='#2b2b2b', foreground='white')
        self.style.configure('TLabel', background='#2b2b2b', foreground='white')
        self.style.configure('TFrame', background='#2b2b2b')
        self.style.configure('TLabelFrame', background='#2b2b2b', foreground='white')
        self.style.configure('TButton', background='#404040', foreground='white')
        self.style.map('TButton', background=[('active', '#505050')])
        
        self.archive_paths = []
        self.plugins_db_path = None
        self.existing_db_path = None
        self.opdir_file_path = None
        self.include_info = tk.BooleanVar(value=False)  # Default to exclude Info findings
        
        # Date filter variables
        self.filter_start_date = tk.StringVar()
        self.filter_end_date = tk.StringVar()
        self.use_date_filter = tk.BooleanVar(value=False)
        
        self.historical_df = pd.DataFrame()
        self.lifecycle_df = pd.DataFrame()
        self.host_presence_df = pd.DataFrame()
        self.scan_changes_df = pd.DataFrame()
        self.opdir_df = pd.DataFrame()
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup the user interface with dark theme"""
        # Main container with padding
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        
        # File selection frame
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        file_frame.columnconfigure(1, weight=1)
        
        # Archive selection
        ttk.Label(file_frame, text="Nessus Archives:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.archives_label = ttk.Label(file_frame, text="No files selected (optional if database loaded)", foreground="gray")
        self.archives_label.grid(row=0, column=1, sticky=tk.W, padx=5)
        ttk.Button(file_frame, text="Select Archives", command=self.select_archives).grid(row=0, column=2, padx=5)
        
        # Plugins DB selection
        ttk.Label(file_frame, text="Plugins DB (optional):").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.plugins_label = ttk.Label(file_frame, text="None selected", foreground="gray")
        self.plugins_label.grid(row=1, column=1, sticky=tk.W, padx=5)
        ttk.Button(file_frame, text="Select Plugins DB", command=self.select_plugins_db).grid(row=1, column=2, padx=5)
        
        # Existing DB selection
        ttk.Label(file_frame, text="Existing DB (optional):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.existing_db_label = ttk.Label(file_frame, text="None selected", foreground="gray")
        self.existing_db_label.grid(row=2, column=1, sticky=tk.W, padx=5)
        ttk.Button(file_frame, text="Load Existing DB", command=self.select_existing_db).grid(row=2, column=2, padx=5)
        
        # OPDIR file selection
        ttk.Label(file_frame, text="OPDIR File (optional):").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.opdir_label = ttk.Label(file_frame, text="None selected", foreground="gray")
        self.opdir_label.grid(row=3, column=1, sticky=tk.W, padx=5)
        ttk.Button(file_frame, text="Select OPDIR File", command=self.select_opdir_file).grid(row=3, column=2, padx=5)
        
        # Analysis options frame
        options_frame = ttk.LabelFrame(main_frame, text="Analysis Options", padding="10")
        options_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=5)
        options_frame.columnconfigure(1, weight=1)
        
        # Include Info findings checkbox
        info_checkbox = ttk.Checkbutton(
            options_frame, 
            text="Include Info-level findings in analysis and visualizations", 
            variable=self.include_info
        )
        info_checkbox.grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=5)
        
        # Info label
        info_label = ttk.Label(
            options_frame, 
            text="Note: Info findings are always stored in database, this only affects analysis/visuals",
            foreground="gray"
        )
        info_label.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=2)
        
        # Date filtering controls
        ttk.Separator(options_frame, orient='horizontal').grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        date_filter_checkbox = ttk.Checkbutton(
            options_frame,
            text="Filter visualizations by date range",
            variable=self.use_date_filter,
            command=self.toggle_date_filter
        )
        date_filter_checkbox.grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=5)
        
        # Start date
        ttk.Label(options_frame, text="Start Date:").grid(row=4, column=0, sticky=tk.W, padx=(20, 5), pady=2)
        self.start_date_entry = ttk.Entry(options_frame, textvariable=self.filter_start_date, width=12, state='disabled')
        self.start_date_entry.grid(row=4, column=1, sticky=tk.W, padx=5, pady=2)
        ttk.Label(options_frame, text="(YYYY-MM-DD)", foreground="gray").grid(row=4, column=2, sticky=tk.W, padx=5, pady=2)
        
        # End date
        ttk.Label(options_frame, text="End Date:").grid(row=5, column=0, sticky=tk.W, padx=(20, 5), pady=2)
        self.end_date_entry = ttk.Entry(options_frame, textvariable=self.filter_end_date, width=12, state='disabled')
        self.end_date_entry.grid(row=5, column=1, sticky=tk.W, padx=5, pady=2)
        ttk.Label(options_frame, text="(YYYY-MM-DD)", foreground="gray").grid(row=5, column=2, sticky=tk.W, padx=5, pady=2)
        
        # Action buttons frame
        action_frame = ttk.Frame(main_frame, padding="10")
        action_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Button(action_frame, text="Process/Analyze", command=self.process_archives).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Refresh Analysis", command=self.refresh_analysis).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Export to Excel", command=self.export_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Export to SQLite", command=self.export_sqlite).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Export to JSON", command=self.export_json).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Export JSON Schema", command=self.export_schema).pack(side=tk.LEFT, padx=5)
        
        # Notebook for different views
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # Status tab
        self.status_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.status_frame, text="Status")
        
        self.status_text = tk.Text(self.status_frame, wrap=tk.WORD, height=20, 
                                   bg='#1e1e1e', fg='white', insertbackground='white')
        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        status_scroll = ttk.Scrollbar(self.status_frame, command=self.status_text.yview)
        status_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.config(yscrollcommand=status_scroll.set)
        
        # Visualization tabs (created after processing)
        self.viz_frames = {}
    
    def select_archives(self):
        """Select Nessus archive files with dark theme dialog"""
        filetypes = (
            ('Archive files', '*.zip'),
            ('Nessus files', '*.nessus'),
            ('All files', '*.*')
        )
        
        paths = filedialog.askopenfilenames(
            title='Select Nessus Archives',
            filetypes=filetypes
        )
        
        if paths:
            self.archive_paths = list(paths)
            self.archives_label.config(
                text=f"{len(self.archive_paths)} file(s) selected",
                foreground="white"
            )
            self.log(f"Selected {len(self.archive_paths)} archive(s)")
        else:
            # Update archives label to show current mode when no files selected
            self.archives_label.config(
                text="No files selected (archives optional if database loaded)",
                foreground="gray"
            )
    
    def select_plugins_db(self):
        """Select plugins database file with dark theme dialog"""
        filetypes = (
            ('XML files', '*.xml'),
            ('JSON files', '*.json'),
            ('All files', '*.*')
        )
        
        path = filedialog.askopenfilename(
            title='Select Plugins Database',
            filetypes=filetypes
        )
        
        if path:
            self.plugins_db_path = path
            self.plugins_label.config(
                text=os.path.basename(path),
                foreground="white"
            )
            self.log(f"Selected plugins DB: {os.path.basename(path)}")
    
    def toggle_date_filter(self):
        """Enable/disable date filter controls"""
        state = 'normal' if self.use_date_filter.get() else 'disabled'
        self.start_date_entry.config(state=state)
        self.end_date_entry.config(state=state)
        
        # Set default dates when enabling
        if self.use_date_filter.get() and not self.historical_df.empty:
            if not self.filter_start_date.get():
                start_date = self.historical_df['scan_date'].min().strftime('%Y-%m-%d')
                self.filter_start_date.set(start_date)
            if not self.filter_end_date.get():
                end_date = self.historical_df['scan_date'].max().strftime('%Y-%m-%d')
                self.filter_end_date.set(end_date)
    
    def get_filtered_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply date and info filtering to DataFrame"""
        filtered_df = df.copy()
        
        # Apply Info filtering
        if not self.include_info.get():
            filtered_df = filtered_df[filtered_df['severity_text'] != 'Info']
        
        # Apply date filtering
        if self.use_date_filter.get() and not filtered_df.empty:
            try:
                start_date = pd.to_datetime(self.filter_start_date.get())
                end_date = pd.to_datetime(self.filter_end_date.get())
                
                # Ensure we include the full end date
                end_date = end_date + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                
                filtered_df = filtered_df[
                    (filtered_df['scan_date'] >= start_date) & 
                    (filtered_df['scan_date'] <= end_date)
                ]
            except (ValueError, TypeError) as e:
                self.log(f"Date filter error: {e}")
        
        return filtered_df
    
    def get_severity_colors(self):
        """Get the standard severity color scheme"""
        return {
            'Critical': '#dc3545',  # Red
            'High': '#fd7e14',      # Orange  
            'Medium': '#ffc107',    # Yellow
            'Low': '#007bff',       # Blue
            'Info': '#6c757d'       # Gray
        }
    
    def get_filter_status_text(self):
        """Get descriptive text for current filter settings"""
        status_parts = []
        
        # Info filter status
        if not self.include_info.get():
            if not self.historical_df.empty:
                info_count = len(self.historical_df[self.historical_df['severity_text'] == 'Info'])
                total_count = len(self.historical_df)
                status_parts.append(f"Excluding Info ({info_count} of {total_count} findings)")
            else:
                status_parts.append("Excluding Info findings")
        else:
            status_parts.append("Including Info findings")
        
        # Date filter status
        if self.use_date_filter.get():
            start_date = self.filter_start_date.get()
            end_date = self.filter_end_date.get()
            if start_date and end_date:
                status_parts.append(f"Date Range: {start_date} to {end_date}")
            else:
                status_parts.append("Date filter enabled (no dates set)")
        
        return " | ".join(status_parts) if status_parts else "No filters applied"
    
    def refresh_analysis(self):
        """Refresh analysis with current Info filtering setting without reprocessing archives"""
        if self.historical_df.empty:
            messagebox.showwarning("No Data", "Please process archives first")
            return
        
        try:
            include_info_val = self.include_info.get()
            self.log(f"Refreshing analysis with current filter settings")
            self.log(f"Filter status: {self.get_filter_status_text()}")
            
            # Filter data for analysis if needed
            analysis_df = self.historical_df.copy()
            if not include_info_val:
                analysis_df = analysis_df[analysis_df['severity_text'] != 'Info']
                info_filtered = len(self.historical_df) - len(analysis_df)
                self.log(f"Filtered out {info_filtered} Info-level findings from analysis")
            
            # Regenerate analysis DataFrames
            self.lifecycle_df = analyze_finding_lifecycle(analysis_df)
            self.host_presence_df = create_host_presence_analysis(analysis_df)
            self.scan_changes_df = analyze_scan_changes(analysis_df)
            
            # Update date filter defaults if enabled
            if self.use_date_filter.get():
                self.toggle_date_filter()
            
            # Update visualizations
            self.create_visualizations()
            
            self.log(f"Analysis refreshed!")
            self.log(f"Total findings in database: {len(self.historical_df)}")
            if not include_info_val:
                info_count = len(self.historical_df[self.historical_df['severity_text'] == 'Info'])
                self.log(f"Info findings in database (excluded from analysis): {info_count}")
            self.log(f"Findings in analysis: {len(self.lifecycle_df)}")
            
            messagebox.showinfo("Success", "Analysis refreshed with current settings!")
            
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            messagebox.showerror("Error", f"Refresh failed: {str(e)}")
    
    
        """Refresh analysis with current Info filtering setting without reprocessing archives"""
        if self.historical_df.empty:
            messagebox.showwarning("No Data", "Please process archives first")
            return
        
        try:
            include_info_val = self.include_info.get()
            self.log(f"Refreshing analysis with Info findings {'included' if include_info_val else 'excluded'}")
            
            # Filter data for analysis if needed
            analysis_df = self.historical_df.copy()
            if not include_info_val:
                analysis_df = analysis_df[analysis_df['severity_text'] != 'Info']
                info_filtered = len(self.historical_df) - len(analysis_df)
                self.log(f"Filtered out {info_filtered} Info-level findings from analysis")
            
            # Regenerate analysis DataFrames
            self.lifecycle_df = analyze_finding_lifecycle(analysis_df)
            self.host_presence_df = create_host_presence_analysis(analysis_df)
            self.scan_changes_df = analyze_scan_changes(analysis_df)
            
            # Update visualizations
            self.create_visualizations()
            
            self.log(f"Analysis refreshed!")
            self.log(f"Total findings in database: {len(self.historical_df)}")
            if not include_info_val:
                info_count = len(self.historical_df[self.historical_df['severity_text'] == 'Info'])
                self.log(f"Info findings in database (excluded from analysis): {info_count}")
            self.log(f"Findings in analysis: {len(self.lifecycle_df)}")
            
            messagebox.showinfo("Success", "Analysis refreshed with current settings!")
            
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            messagebox.showerror("Error", f"Refresh failed: {str(e)}")
    
    def select_existing_db(self):
    
        """Select existing database file for updating"""
        filetypes = (
            ('SQLite database', '*.db'),
            ('All files', '*.*')
        )
        
        path = filedialog.askopenfilename(
            title='Select Existing Database',
            filetypes=filetypes
        )
        
        if path:
            self.existing_db_path = path
            self.existing_db_label.config(
                text=os.path.basename(path),
                foreground="white"
            )
            self.log(f"Selected existing DB: {os.path.basename(path)}")
    
    def log(self, message: str):
        """Add message to status log"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.window.update()
    
    def process_archives(self):
        """Process selected archives with enhanced features or analyze existing database"""
        # Check if we have either archives or an existing database
        has_archives = bool(self.archive_paths)
        has_existing_db = bool(self.existing_db_path and os.path.exists(self.existing_db_path))
        
        if not has_archives and not has_existing_db:
            messagebox.showwarning("No Data Source", "Please select archive files or load an existing database")
            return
        
        if not has_archives and has_existing_db:
            # Load and analyze existing database only
            self.log("="*60)
            self.log("Loading and analyzing existing database...")
            self.log("="*60)
            
            try:
                # Load existing database
                existing_historical_df, existing_lifecycle_df, existing_host_presence_df, existing_scan_changes_df = load_existing_database(self.existing_db_path)
                
                if existing_historical_df.empty:
                    messagebox.showerror("Error", "No data found in the existing database")
                    return
                
                # Set the loaded data
                self.historical_df = existing_historical_df
                
                # Re-run analysis with current filter settings
                include_info_val = self.include_info.get()
                self.log(f"Re-analyzing with current filter settings")
                self.log(f"Filter status: {self.get_filter_status_text()}")
                
                # Filter data for analysis if needed
                analysis_df = self.historical_df.copy()
                if not include_info_val:
                    analysis_df = analysis_df[analysis_df['severity_text'] != 'Info']
                    info_filtered = len(self.historical_df) - len(analysis_df)
                    self.log(f"Filtered out {info_filtered} Info-level findings from analysis")
                
                # Regenerate analysis DataFrames with current settings
                self.lifecycle_df = analyze_finding_lifecycle(analysis_df)
                self.host_presence_df = create_host_presence_analysis(analysis_df)
                self.scan_changes_df = analyze_scan_changes(analysis_df)
                
                # Set up date filter defaults
                if not self.filter_start_date.get() and not self.historical_df.empty:
                    start_date = self.historical_df['scan_date'].min().strftime('%Y-%m-%d')
                    self.filter_start_date.set(start_date)
                if not self.filter_end_date.get() and not self.historical_df.empty:
                    end_date = self.historical_df['scan_date'].max().strftime('%Y-%m-%d')
                    self.filter_end_date.set(end_date)
                
                self.log(f"Database analysis complete!")
                self.log(f"Total findings in database: {len(self.historical_df)}")
                self.log(f"Date range: {self.historical_df['scan_date'].min()} to {self.historical_df['scan_date'].max()}")
                if not include_info_val:
                    info_count = len(self.historical_df[self.historical_df['severity_text'] == 'Info'])
                    self.log(f"Info findings in database (excluded from analysis): {info_count}")
                self.log(f"Findings in analysis: {len(self.lifecycle_df)}")
                self.log(f"Hosts tracked: {len(self.host_presence_df)}")
                self.log(f"Scan changes: {len(self.scan_changes_df)}")
                
                # Create visualizations
                self.create_visualizations()
                
                messagebox.showinfo("Success", "Existing database loaded and analyzed successfully!")
                
            except Exception as e:
                self.log(f"ERROR: {str(e)}")
                messagebox.showerror("Error", f"Database analysis failed: {str(e)}")
                import traceback
                traceback.print_exc()
            
            return
        
        # Original archive processing logic (when archives are provided)
        self.log("="*60)
        self.log("Starting enhanced archive processing...")
        self.log("="*60)
        
        try:
            # Process archives with enhanced features
            include_info_val = self.include_info.get()
            self.log(f"Processing with Info findings {'included' if include_info_val else 'excluded'} from analysis")
            
            self.historical_df, self.lifecycle_df, self.host_presence_df, self.scan_changes_df = process_historical_scans(
                self.archive_paths,
                self.plugins_db_path,
                self.existing_db_path,
                include_info_val
            )
            
            if self.historical_df.empty:
                messagebox.showerror("Error", "No findings extracted from archives")
                return
            
            # Set up date filter defaults
            if not self.filter_start_date.get():
                start_date = self.historical_df['scan_date'].min().strftime('%Y-%m-%d')
                self.filter_start_date.set(start_date)
            if not self.filter_end_date.get():
                end_date = self.historical_df['scan_date'].max().strftime('%Y-%m-%d')
                self.filter_end_date.set(end_date)
            
            self.log(f"Analysis complete!")
            self.log(f"Total findings in database: {len(self.historical_df)}")
            if not include_info_val:
                info_count = len(self.historical_df[self.historical_df['severity_text'] == 'Info'])
                self.log(f"Info findings in database (excluded from analysis): {info_count}")
            self.log(f"Findings in analysis: {len(self.lifecycle_df)}")
            self.log(f"Hosts tracked: {len(self.host_presence_df)}")
            self.log(f"Scan changes: {len(self.scan_changes_df)}")
            
            # Create visualizations
            self.create_visualizations()
            
            messagebox.showinfo("Success", "Archives processed successfully with enhanced features!")
            
        except Exception as e:
            self.log(f"ERROR: {str(e)}")
            messagebox.showerror("Error", f"Processing failed: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def create_visualizations(self):
        """Create enhanced visualization tabs"""
        # Remove old visualization tabs
        for tab_name in list(self.viz_frames.keys()):
            self.notebook.forget(self.viz_frames[tab_name])
        self.viz_frames = {}
        
        # Create new visualization tabs
        self.create_timeline_viz()
        self.create_host_presence_viz()
        self.create_scan_changes_viz()
        self.create_severity_trend_viz()
        self.create_host_viz()
        self.create_plugin_viz()
        self.create_lifecycle_viz()
    
    def create_host_presence_viz(self):
        """Create host presence visualization with date filtering"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Host Tracking")
        self.viz_frames['host_presence'] = frame
        
        # Add filter status
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        # Get filtered host presence data based on date range
        filtered_host_presence = self.host_presence_df.copy()
        if self.use_date_filter.get() and not filtered_host_presence.empty:
            try:
                start_date = pd.to_datetime(self.filter_start_date.get())
                end_date = pd.to_datetime(self.filter_end_date.get()) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                
                # Filter based on last_seen date for host presence
                filtered_host_presence = filtered_host_presence[
                    (filtered_host_presence['last_seen'] >= start_date) | 
                    (filtered_host_presence['first_seen'] <= end_date)
                ]
            except (ValueError, TypeError):
                pass
        
        # Host status distribution
        if not filtered_host_presence.empty:
            status_counts = filtered_host_presence['status'].value_counts()
            ax1.pie(status_counts.values, labels=status_counts.index, autopct='%1.1f%%',
                   colors=['#28a745', '#dc3545'], startangle=90)
            ax1.set_title('Host Status Distribution', fontsize=12, fontweight='bold', color='white')
            ax1.set_facecolor('#2b2b2b')
        else:
            ax1.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax1.transAxes, color='white', fontsize=12)
            ax1.set_facecolor('#2b2b2b')
        
        # Presence percentage distribution
        if not filtered_host_presence.empty:
            bins = [0, 25, 50, 75, 90, 100]
            labels = ['0-25%', '26-50%', '51-75%', '76-90%', '91-100%']
            filtered_host_presence['presence_category'] = pd.cut(
                filtered_host_presence['presence_percentage'], 
                bins=bins, labels=labels, right=True
            )
            presence_dist = filtered_host_presence['presence_category'].value_counts().sort_index()
            
            bars = ax2.bar(range(len(presence_dist)), presence_dist.values, color='#007bff')
            ax2.set_xticks(range(len(presence_dist)))
            ax2.set_xticklabels(presence_dist.index, rotation=45, ha='right', color='white')
            ax2.set_title('Host Presence Percentage Distribution', fontsize=12, fontweight='bold', color='white')
            ax2.set_ylabel('Number of Hosts', color='white')
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                if height > 0:  # Only show label if there's a value
                    ax2.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                            f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        else:
            ax2.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax2.transAxes, color='white', fontsize=12)
            ax2.set_facecolor('#2b2b2b')
        
        # Most unreliable hosts (lowest presence percentage)
        if not filtered_host_presence.empty:
            unreliable_hosts = filtered_host_presence.nsmallest(10, 'presence_percentage')
            display_names = [f"{row['hostname']} ({row['ip_address']})" for _, row in unreliable_hosts.iterrows()]
            
            bars = ax3.barh(range(len(unreliable_hosts)), unreliable_hosts['presence_percentage'].values, color='#dc3545')
            ax3.set_yticks(range(len(unreliable_hosts)))
            ax3.set_yticklabels(display_names, color='white')
            ax3.set_title('10 Most Unreliable Hosts', fontsize=12, fontweight='bold', color='white')
            ax3.set_xlabel('Presence Percentage', color='white')
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            
            # Add data labels on horizontal bars
            for i, bar in enumerate(bars):
                width = bar.get_width()
                ax3.text(width + 1, bar.get_y() + bar.get_height()/2.,
                        f'{width:.1f}%', ha='left', va='center', color='white', fontweight='bold')
        else:
            ax3.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax3.transAxes, color='white', fontsize=12)
            ax3.set_facecolor('#2b2b2b')
        
        # Recently missing hosts
        if not filtered_host_presence.empty:
            missing_hosts = filtered_host_presence[filtered_host_presence['status'] == 'Missing'].head(10)
            if not missing_hosts.empty:
                display_names = [f"{row['hostname']} ({row['ip_address']})" for _, row in missing_hosts.iterrows()]
                days_since_last = [(datetime.now() - row['last_seen']).days for _, row in missing_hosts.iterrows()]
                
                bars = ax4.barh(range(len(missing_hosts)), days_since_last, color='#fd7e14')
                ax4.set_yticks(range(len(missing_hosts)))
                ax4.set_yticklabels(display_names, color='white')
                ax4.set_title('Recently Missing Hosts', fontsize=12, fontweight='bold', color='white')
                ax4.set_xlabel('Days Since Last Seen', color='white')
                ax4.set_facecolor('#2b2b2b')
                ax4.tick_params(colors='white')
                
                # Add data labels on horizontal bars
                for bar in bars:
                    width = bar.get_width()
                    ax4.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                            f'{int(width)}', ha='left', va='center', color='white', fontweight='bold')
            else:
                ax4.text(0.5, 0.5, 'No missing hosts', ha='center', va='center', 
                        transform=ax4.transAxes, color='white', fontsize=12)
                ax4.set_facecolor('#2b2b2b')
        else:
            ax4.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax4.transAxes, color='white', fontsize=12)
            ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_scan_changes_viz(self):
        """Create scan changes visualization"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Scan Changes")
        self.viz_frames['scan_changes'] = frame
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        if not self.scan_changes_df.empty:
            # Host additions over time
            ax1.plot(self.scan_changes_df['scan_date'], self.scan_changes_df['hosts_added'], 
                    marker='o', linewidth=2, color='#28a745', label='Added')
            ax1.plot(self.scan_changes_df['scan_date'], self.scan_changes_df['hosts_removed'], 
                    marker='s', linewidth=2, color='#dc3545', label='Removed')
            ax1.set_title('Host Additions/Removals Over Time', fontsize=12, fontweight='bold', color='white')
            ax1.set_xlabel('Scan Date', color='white')
            ax1.set_ylabel('Number of Hosts', color='white')
            ax1.legend()
            ax1.grid(True, alpha=0.3)
            ax1.set_facecolor('#2b2b2b')
            ax1.tick_params(colors='white')
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # Net change over time
            bars = ax2.bar(range(len(self.scan_changes_df)), self.scan_changes_df['net_change'].values,
                   color=['#28a745' if x >= 0 else '#dc3545' for x in self.scan_changes_df['net_change']])
            ax2.set_title('Net Host Change by Scan', fontsize=12, fontweight='bold', color='white')
            ax2.set_xlabel('Scan Number', color='white')
            ax2.set_ylabel('Net Change', color='white')
            ax2.axhline(y=0, color='white', linestyle='-', alpha=0.5)
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                label_y = height + (abs(height) * 0.02) if height >= 0 else height - (abs(height) * 0.02)
                ax2.text(bar.get_x() + bar.get_width()/2., label_y,
                        f'{int(height)}', ha='center', va='bottom' if height >= 0 else 'top', 
                        color='white', fontweight='bold')
            
            # Total hosts over time
            ax3.plot(self.scan_changes_df['scan_date'], self.scan_changes_df['total_hosts_current'], 
                    marker='o', linewidth=2, color='#007bff')
            ax3.set_title('Total Hosts Over Time', fontsize=12, fontweight='bold', color='white')
            ax3.set_xlabel('Scan Date', color='white')
            ax3.set_ylabel('Total Hosts', color='white')
            ax3.grid(True, alpha=0.3)
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            ax3.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # Stability metrics
            stability_data = {
                'High Turnover': len(self.scan_changes_df[abs(self.scan_changes_df['net_change']) > 5]),
                'Medium Turnover': len(self.scan_changes_df[(abs(self.scan_changes_df['net_change']) > 1) & 
                                                           (abs(self.scan_changes_df['net_change']) <= 5)]),
                'Stable': len(self.scan_changes_df[abs(self.scan_changes_df['net_change']) <= 1])
            }
            
            ax4.pie(stability_data.values(), labels=stability_data.keys(), autopct='%1.1f%%',
                   colors=['#dc3545', '#ffc107', '#28a745'], startangle=90)
            ax4.set_title('Scan Stability Distribution', fontsize=12, fontweight='bold', color='white')
            ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    # [Continue with other visualization methods with dark theme...]
    # [For brevity, I'll include the remaining methods in a follow-up if needed]
    
    def export_excel(self):
        """Export to Excel with enhanced data"""
        if self.historical_df.empty:
            messagebox.showwarning("No Data", "Please process archives first")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Excel File"
        )
        
        if filepath:
            try:
                self.log(f"Exporting to Excel: {filepath}")
                export_to_excel(self.historical_df, self.lifecycle_df, 
                               self.host_presence_df, self.scan_changes_df, filepath)
                self.log("Excel export complete!")
                messagebox.showinfo("Success", f"Enhanced data exported to:\n{filepath}")
            except Exception as e:
                self.log(f"ERROR: {str(e)}")
                messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def export_sqlite(self):
        """Export to SQLite with enhanced data"""
        if self.historical_df.empty:
            messagebox.showwarning("No Data", "Please process archives first")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".db",
            filetypes=[("SQLite database", "*.db"), ("All files", "*.*")],
            title="Save SQLite Database"
        )
        
        if filepath:
            try:
                self.log(f"Exporting to SQLite: {filepath}")
                export_to_sqlite(self.historical_df, self.lifecycle_df, 
                                self.host_presence_df, self.scan_changes_df, filepath)
                self.log("SQLite export complete!")
                messagebox.showinfo("Success", f"Enhanced database exported to:\n{filepath}")
            except Exception as e:
                self.log(f"ERROR: {str(e)}")
                messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def export_json(self):
        """Export to JSON with enhanced data"""
        if self.historical_df.empty:
            messagebox.showwarning("No Data", "Please process archives first")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save JSON File"
        )
        
        if filepath:
            try:
                self.log(f"Exporting to JSON: {filepath}")
                export_to_json(self.historical_df, self.lifecycle_df, 
                              self.host_presence_df, self.scan_changes_df, filepath)
                self.log("JSON export complete!")
                messagebox.showinfo("Success", f"Enhanced data exported to:\n{filepath}")
            except Exception as e:
                self.log(f"ERROR: {str(e)}")
                messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def export_schema(self):
        """Export JSON schema"""
        filepath = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save JSON Schema File"
        )
        
        if filepath:
            try:
                self.log(f"Exporting JSON schema: {filepath}")
                export_json_schema(filepath)
                self.log("JSON schema export complete!")
                messagebox.showinfo("Success", f"JSON schema exported to:\n{filepath}")
            except Exception as e:
                self.log(f"ERROR: {str(e)}")
                messagebox.showerror("Error", f"Schema export failed: {str(e)}")
    
    def create_timeline_viz(self):
        """Create findings timeline visualization with date filtering and corrected colors"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Timeline")
        self.viz_frames['timeline'] = frame
        
        # Add filter status and controls
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        # Filter status
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        # Refresh button
        ttk.Button(control_frame, text="Apply Filters", command=self.create_visualizations).pack(side=tk.RIGHT, padx=5)
        
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8))
        fig.patch.set_facecolor('#2b2b2b')
        
        # Apply all filters to get visualization data
        filtered_df = self.get_filtered_data(self.historical_df)
        
        if filtered_df.empty:
            ax1.text(0.5, 0.5, 'No data matches current filters', ha='center', va='center', 
                    transform=ax1.transAxes, color='white', fontsize=14)
            ax2.text(0.5, 0.5, 'No data matches current filters', ha='center', va='center', 
                    transform=ax2.transAxes, color='white', fontsize=14)
        else:
            # Total findings over time
            timeline_data = filtered_df.groupby('scan_date').size().reset_index(name='count')
            ax1.plot(timeline_data['scan_date'], timeline_data['count'], marker='o', linewidth=2, color='#007bff')
            
            filter_suffix = ""
            if self.use_date_filter.get():
                filter_suffix += f" ({self.filter_start_date.get()} to {self.filter_end_date.get()})"
            if not self.include_info.get():
                filter_suffix += " (Excluding Info)"
            
            ax1.set_title(f'Total Findings Over Time{filter_suffix}', fontsize=14, fontweight='bold', color='white')
            ax1.set_xlabel('Scan Date', color='white')
            ax1.set_ylabel('Number of Findings', color='white')
            ax1.grid(True, alpha=0.3)
            ax1.set_facecolor('#2b2b2b')
            ax1.tick_params(colors='white')
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # Findings by severity over time
            severity_timeline = filtered_df.groupby(['scan_date', 'severity_text']).size().unstack(fill_value=0)
            
            colors = self.get_severity_colors()
            for severity in ['Critical', 'High', 'Medium', 'Low', 'Info']:
                if severity in severity_timeline.columns:
                    ax2.plot(severity_timeline.index, severity_timeline[severity], 
                            marker='o', label=severity, color=colors.get(severity, 'gray'), linewidth=2)
            
            ax2.set_title(f'Findings by Severity Over Time{filter_suffix}', fontsize=14, fontweight='bold', color='white')
            ax2.set_xlabel('Scan Date', color='white')
            ax2.set_ylabel('Number of Findings', color='white')
            ax2.legend(loc='upper left')
            ax2.grid(True, alpha=0.3)
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_severity_trend_viz(self):
        """Create severity trend visualization with date filtering and corrected colors"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Severity Trends")
        self.viz_frames['severity'] = frame
        
        # Add filter status
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        status_label = ttk.Label(control_frame, text=f"Filters: {self.get_filter_status_text()}", foreground="orange")
        status_label.pack(side=tk.LEFT)
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        # Get filtered lifecycle data
        filtered_lifecycle = self.lifecycle_df.copy()
        if self.use_date_filter.get() and not filtered_lifecycle.empty:
            try:
                start_date = pd.to_datetime(self.filter_start_date.get())
                end_date = pd.to_datetime(self.filter_end_date.get()) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                
                # Filter by first_seen date for lifecycle data
                filtered_lifecycle = filtered_lifecycle[
                    (filtered_lifecycle['first_seen'] >= start_date) & 
                    (filtered_lifecycle['first_seen'] <= end_date)
                ]
            except (ValueError, TypeError):
                pass
        
        colors = self.get_severity_colors()
        
        # Current severity distribution
        if not filtered_lifecycle.empty:
            severity_counts = filtered_lifecycle['severity_text'].value_counts()
            severity_colors = [colors.get(sev, '#6c757d') for sev in severity_counts.index]
            ax1.pie(severity_counts.values, labels=severity_counts.index, autopct='%1.1f%%', 
                   colors=severity_colors, startangle=90)
            ax1.set_title('Severity Distribution', fontsize=12, fontweight='bold', color='white')
            ax1.set_facecolor('#2b2b2b')
        else:
            ax1.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax1.transAxes, color='white', fontsize=12)
            ax1.set_facecolor('#2b2b2b')
        
        # Active vs Resolved
        if not filtered_lifecycle.empty:
            status_counts = filtered_lifecycle['status'].value_counts()
            bars = ax2.bar(status_counts.index, status_counts.values, color=['#dc3545', '#28a745'])
            ax2.set_title('Active vs Resolved Findings', fontsize=12, fontweight='bold', color='white')
            ax2.set_ylabel('Count', color='white')
            ax2.grid(True, alpha=0.3, axis='y')
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                        f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        else:
            ax2.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax2.transAxes, color='white', fontsize=12)
            ax2.set_facecolor('#2b2b2b')
        
        # Average days open by severity
        if not filtered_lifecycle.empty:
            avg_days = filtered_lifecycle.groupby('severity_text')['days_open'].mean().sort_values(ascending=False)
            severity_colors = [colors.get(sev, '#6c757d') for sev in avg_days.index]
            bars = ax3.barh(avg_days.index, avg_days.values, color=severity_colors)
            ax3.set_title('Average Days Open by Severity', fontsize=12, fontweight='bold', color='white')
            ax3.set_xlabel('Average Days', color='white')
            ax3.grid(True, alpha=0.3, axis='x')
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            
            # Add data labels on horizontal bars
            for bar in bars:
                width = bar.get_width()
                ax3.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                        f'{width:.0f}', ha='left', va='center', color='white', fontweight='bold')
        else:
            ax3.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax3.transAxes, color='white', fontsize=12)
            ax3.set_facecolor('#2b2b2b')
        
        # Reappearance rate
        if not filtered_lifecycle.empty:
            reappeared = len(filtered_lifecycle[filtered_lifecycle['reappearances'] > 0])
            never_reappeared = len(filtered_lifecycle[filtered_lifecycle['reappearances'] == 0])
            ax4.pie([reappeared, never_reappeared], labels=['Reappeared', 'Never Reappeared'],
                   autopct='%1.1f%%', colors=['#dc3545', '#28a745'], startangle=90)
            ax4.set_title('Finding Reappearance Rate', fontsize=12, fontweight='bold', color='white')
            ax4.set_facecolor('#2b2b2b')
        else:
            ax4.text(0.5, 0.5, 'No data matches filters', ha='center', va='center', 
                    transform=ax4.transAxes, color='white', fontsize=12)
            ax4.set_facecolor('#2b2b2b')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def create_host_viz(self):
        """Create host-specific visualization with enhanced display"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Host Analysis")
        self.viz_frames['host'] = frame
        
        # Create controls frame
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        
        ttk.Label(control_frame, text="Select Host:").pack(side=tk.LEFT, padx=5)
        
        # Get unique hostnames with IP addresses for better display
        host_data = self.historical_df[['hostname', 'ip_address']].drop_duplicates()
        host_options = [f"{row['hostname']} ({row['ip_address']})" for _, row in host_data.iterrows()]
        host_options = sorted(host_options)
        
        self.host_var = tk.StringVar(value=host_options[0] if host_options else "")
        host_combo = ttk.Combobox(control_frame, textvariable=self.host_var, values=host_options, width=50)
        host_combo.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(control_frame, text="Update", command=self.update_host_viz).pack(side=tk.LEFT, padx=5)
        
        # Create figure
        self.host_fig, ((self.host_ax1, self.host_ax2), (self.host_ax3, self.host_ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        self.host_fig.patch.set_facecolor('#2b2b2b')
        
        self.host_canvas = FigureCanvasTkAgg(self.host_fig, frame)
        self.host_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Initial update
        if host_options:
            self.update_host_viz()
    
    def update_host_viz(self):
        """Update host visualization based on selected host with enhanced display"""
        host_str = self.host_var.get()
        if not host_str:
            return
        
        # Extract hostname from "hostname (ip)" format
        hostname = host_str.split(' (')[0]
        
        # Clear axes
        for ax in [self.host_ax1, self.host_ax2, self.host_ax3, self.host_ax4]:
            ax.clear()
            ax.set_facecolor('#2b2b2b')
        
        # Filter data for selected host
        host_data = self.historical_df[self.historical_df['hostname'] == hostname]
        host_lifecycle = self.lifecycle_df[self.lifecycle_df['hostname'] == hostname]
        
        # Findings over time for this host
        timeline = host_data.groupby('scan_date').size().reset_index(name='count')
        self.host_ax1.plot(timeline['scan_date'], timeline['count'], marker='o', linewidth=2, color='#007bff')
        self.host_ax1.set_title(f'Findings Timeline - {host_str}', fontsize=12, fontweight='bold', color='white')
        self.host_ax1.set_xlabel('Scan Date', color='white')
        self.host_ax1.set_ylabel('Number of Findings', color='white')
        self.host_ax1.grid(True, alpha=0.3)
        self.host_ax1.tick_params(colors='white')
        self.host_ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.setp(self.host_ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        # Severity distribution
        if not host_lifecycle.empty:
            severity_counts = host_lifecycle['severity_text'].value_counts()
            colors = ['#dc3545', '#fd7e14', '#ffc107', '#28a745', '#6c757d']
            self.host_ax2.pie(severity_counts.values, labels=severity_counts.index, autopct='%1.1f%%',
                             colors=colors[:len(severity_counts)], startangle=90)
            self.host_ax2.set_title(f'Severity Distribution - {hostname}', fontsize=12, fontweight='bold', color='white')
        
        # Top 10 plugins for this host
        if not host_lifecycle.empty:
            top_plugins = host_lifecycle.nlargest(10, 'days_open')[['plugin_name', 'days_open']]
            self.host_ax3.barh(range(len(top_plugins)), top_plugins['days_open'].values, color='#dc3545')
            self.host_ax3.set_yticks(range(len(top_plugins)))
            self.host_ax3.set_yticklabels([name[:40] + '...' if len(name) > 40 else name 
                                          for name in top_plugins['plugin_name'].values], color='white')
            self.host_ax3.set_title(f'Top 10 Longest-Running Findings - {hostname}', fontsize=12, fontweight='bold', color='white')
            self.host_ax3.set_xlabel('Days Open', color='white')
            self.host_ax3.grid(True, alpha=0.3, axis='x')
            self.host_ax3.tick_params(colors='white')
        
        # Status breakdown
        if not host_lifecycle.empty:
            status_counts = host_lifecycle['status'].value_counts()
            colors_status = {'Active': '#dc3545', 'Resolved': '#28a745'}
            bars = self.host_ax4.bar(status_counts.index, status_counts.values, 
                             color=[colors_status.get(s, 'gray') for s in status_counts.index])
            self.host_ax4.set_title(f'Finding Status - {hostname}', fontsize=12, fontweight='bold', color='white')
            self.host_ax4.set_ylabel('Count', color='white')
            self.host_ax4.grid(True, alpha=0.3, axis='y')
            self.host_ax4.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                self.host_ax4.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                        f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        self.host_canvas.draw()
    
    def create_plugin_viz(self):
        """Create plugin-specific visualization"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Plugin Analysis")
        self.viz_frames['plugin'] = frame
        
        # Create controls frame
        control_frame = ttk.Frame(frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        
        ttk.Label(control_frame, text="Select Plugin ID:").pack(side=tk.LEFT, padx=5)
        
        # Get unique plugin IDs with names
        plugin_list = self.lifecycle_df[['plugin_id', 'plugin_name']].drop_duplicates()
        plugin_list = plugin_list.sort_values('plugin_id')
        plugin_options = [f"{row['plugin_id']} - {row['plugin_name'][:50]}" 
                         for _, row in plugin_list.iterrows()]
        
        self.plugin_var = tk.StringVar(value=plugin_options[0] if plugin_options else "")
        plugin_combo = ttk.Combobox(control_frame, textvariable=self.plugin_var, 
                                   values=plugin_options, width=60)
        plugin_combo.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(control_frame, text="Update", command=self.update_plugin_viz).pack(side=tk.LEFT, padx=5)
        
        # Create figure
        self.plugin_fig, ((self.plugin_ax1, self.plugin_ax2), (self.plugin_ax3, self.plugin_ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        self.plugin_fig.patch.set_facecolor('#2b2b2b')
        
        self.plugin_canvas = FigureCanvasTkAgg(self.plugin_fig, frame)
        self.plugin_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Initial update
        if plugin_options:
            self.update_plugin_viz()
    
    def update_plugin_viz(self):
        """Update plugin visualization based on selected plugin"""
        plugin_str = self.plugin_var.get()
        if not plugin_str:
            return
        
        plugin_id = plugin_str.split(' - ')[0]
        
        # Clear axes
        for ax in [self.plugin_ax1, self.plugin_ax2, self.plugin_ax3, self.plugin_ax4]:
            ax.clear()
            ax.set_facecolor('#2b2b2b')
        
        # Filter data for selected plugin
        plugin_data = self.historical_df[self.historical_df['plugin_id'] == plugin_id]
        plugin_lifecycle = self.lifecycle_df[self.lifecycle_df['plugin_id'] == plugin_id]
        
        # Affected hosts over time
        timeline = plugin_data.groupby('scan_date')['hostname'].nunique().reset_index(name='host_count')
        self.plugin_ax1.plot(timeline['scan_date'], timeline['host_count'], marker='o', linewidth=2, color='#dc3545')
        self.plugin_ax1.set_title(f'Affected Hosts Over Time - Plugin {plugin_id}', fontsize=12, fontweight='bold', color='white')
        self.plugin_ax1.set_xlabel('Scan Date', color='white')
        self.plugin_ax1.set_ylabel('Number of Hosts', color='white')
        self.plugin_ax1.grid(True, alpha=0.3)
        self.plugin_ax1.tick_params(colors='white')
        self.plugin_ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.setp(self.plugin_ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        # Host status distribution
        if not plugin_lifecycle.empty:
            status_counts = plugin_lifecycle['status'].value_counts()
            colors_status = {'Active': '#dc3545', 'Resolved': '#28a745'}
            self.plugin_ax2.pie(status_counts.values, labels=status_counts.index, autopct='%1.1f%%',
                               colors=[colors_status.get(s, 'gray') for s in status_counts.index], 
                               startangle=90)
            self.plugin_ax2.set_title(f'Status Distribution - Plugin {plugin_id}', fontsize=12, fontweight='bold', color='white')
        
        # Top affected hosts with hostname (IP) format
        if not plugin_lifecycle.empty:
            top_hosts = plugin_lifecycle.nlargest(10, 'days_open')[['hostname', 'ip_address', 'days_open']]
            display_names = [f"{row['hostname']} ({row['ip_address']})" for _, row in top_hosts.iterrows()]
            
            self.plugin_ax3.barh(range(len(top_hosts)), top_hosts['days_open'].values, color='#fd7e14')
            self.plugin_ax3.set_yticks(range(len(top_hosts)))
            self.plugin_ax3.set_yticklabels(display_names, color='white')
            self.plugin_ax3.set_title(f'Top 10 Hosts by Days Open - Plugin {plugin_id}', fontsize=12, fontweight='bold', color='white')
            self.plugin_ax3.set_xlabel('Days Open', color='white')
            self.plugin_ax3.grid(True, alpha=0.3, axis='x')
            self.plugin_ax3.tick_params(colors='white')
        
        # Reappearance statistics
        if not plugin_lifecycle.empty:
            reappeared = len(plugin_lifecycle[plugin_lifecycle['reappearances'] > 0])
            never_reappeared = len(plugin_lifecycle[plugin_lifecycle['reappearances'] == 0])
            bars = self.plugin_ax4.bar(['Reappeared', 'Never Reappeared'], [reappeared, never_reappeared],
                               color=['#dc3545', '#28a745'])
            self.plugin_ax4.set_title(f'Reappearance Statistics - Plugin {plugin_id}', fontsize=12, fontweight='bold', color='white')
            self.plugin_ax4.set_ylabel('Number of Hosts', color='white')
            self.plugin_ax4.grid(True, alpha=0.3, axis='y')
            self.plugin_ax4.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    self.plugin_ax4.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                            f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        self.plugin_canvas.draw()
        
    def create_lifecycle_viz(self):
        """Create lifecycle analysis visualization with dark theme"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Lifecycle Analysis")
        self.viz_frames['lifecycle'] = frame
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.patch.set_facecolor('#2b2b2b')
        
        # Distribution of days open
        if not self.lifecycle_df.empty:
            max_days = self.lifecycle_df['days_open'].max()
            
            # Create bins that are guaranteed to be monotonically increasing
            base_bins = [0, 30, 60, 90, 120, 180, 365]
            bins = [b for b in base_bins if b <= max_days]
            bins.append(max_days + 1)  # Always add max as the final bin
            
            # Create labels for the bins
            labels = []
            for i in range(len(bins) - 1):
                if i == len(bins) - 2:  # Last label
                    if bins[i] >= 365:
                        labels.append('365+')
                    else:
                        labels.append(f'{bins[i]}-{int(max_days)}')
                else:
                    labels.append(f'{bins[i]}-{bins[i+1]-1}')
            
            self.lifecycle_df['age_category'] = pd.cut(self.lifecycle_df['days_open'], 
                                                        bins=bins, labels=labels, right=False)
            age_dist = self.lifecycle_df['age_category'].value_counts().sort_index()
            
            bars = ax1.bar(range(len(age_dist)), age_dist.values, color='#007bff')
            ax1.set_xticks(range(len(age_dist)))
            ax1.set_xticklabels(age_dist.index, rotation=45, ha='right', color='white')
            ax1.set_title('Distribution of Finding Age (Days Open)', fontsize=12, fontweight='bold', color='white')
            ax1.set_ylabel('Number of Findings', color='white')
            ax1.grid(True, alpha=0.3, axis='y')
            ax1.set_facecolor('#2b2b2b')
            ax1.tick_params(colors='white')
            
            # Add data labels on bars
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax1.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                            f'{int(height)}', ha='center', va='bottom', color='white', fontweight='bold')
        
        # Top 15 most prevalent plugins
        if not self.lifecycle_df.empty:
            top_plugins = self.lifecycle_df['plugin_id'].value_counts().head(15)
            plugin_names = [self.lifecycle_df[self.lifecycle_df['plugin_id'] == pid]['plugin_name'].iloc[0][:30] 
                        for pid in top_plugins.index]
            
            bars = ax2.barh(range(len(top_plugins)), top_plugins.values, color='#dc3545')
            ax2.set_yticks(range(len(top_plugins)))
            ax2.set_yticklabels(plugin_names, color='white')
            ax2.set_title('Top 15 Most Prevalent Findings', fontsize=12, fontweight='bold', color='white')
            ax2.set_xlabel('Number of Affected Hosts', color='white')
            ax2.grid(True, alpha=0.3, axis='x')
            ax2.set_facecolor('#2b2b2b')
            ax2.tick_params(colors='white')
            
            # Add data labels on horizontal bars
            for bar in bars:
                width = bar.get_width()
                ax2.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                        f'{int(width)}', ha='left', va='center', color='white', fontweight='bold')
        
        # Findings by first seen date
        if not self.lifecycle_df.empty:
            first_seen_timeline = self.lifecycle_df.groupby(
                self.lifecycle_df['first_seen'].dt.to_period('M')
            ).size().reset_index(name='count')
            first_seen_timeline['first_seen'] = first_seen_timeline['first_seen'].dt.to_timestamp()
            
            ax3.plot(first_seen_timeline['first_seen'], first_seen_timeline['count'], 
                    marker='o', linewidth=2, color='#28a745')
            ax3.set_title('New Findings by Month (First Seen)', fontsize=12, fontweight='bold', color='white')
            ax3.set_xlabel('Month', color='white')
            ax3.set_ylabel('New Findings', color='white')
            ax3.grid(True, alpha=0.3)
            ax3.set_facecolor('#2b2b2b')
            ax3.tick_params(colors='white')
            ax3.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
            plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        # Remediation effectiveness (resolved findings by severity)
        if not self.lifecycle_df.empty:
            resolved = self.lifecycle_df[self.lifecycle_df['status'] == 'Resolved']
            if not resolved.empty:
                remediation = resolved.groupby('severity_text').size().sort_values(ascending=True)
                colors_sev = {'Critical': '#dc3545', 'High': '#fd7e14', 'Medium': '#ffc107', 
                            'Low': '#007bff', 'Info': '#6c757d'}
                
                bars = ax4.barh(range(len(remediation)), remediation.values,
                        color=[colors_sev.get(s, 'gray') for s in remediation.index])
                ax4.set_yticks(range(len(remediation)))
                ax4.set_yticklabels(remediation.index, color='white')
                ax4.set_title('Resolved Findings by Severity', fontsize=12, fontweight='bold', color='white')
                ax4.set_xlabel('Number of Findings Resolved', color='white')
                ax4.grid(True, alpha=0.3, axis='x')
                ax4.set_facecolor('#2b2b2b')
                ax4.tick_params(colors='white')
                
                # Add data labels on horizontal bars
                for bar in bars:
                    width = bar.get_width()
                    ax4.text(width + width*0.01, bar.get_y() + bar.get_height()/2.,
                            f'{int(width)}', ha='left', va='center', color='white', fontweight='bold')
        
        plt.tight_layout()
        plt.style.use('dark_background')
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    def run(self):
        """Run the enhanced GUI"""
        self.window.mainloop()


def main():
    """Main entry point for the enhanced application"""
    app = EnhancedHistoricalAnalysisGUI()
    app.run()


if __name__ == "__main__":
    main()