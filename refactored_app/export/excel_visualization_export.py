"""
Excel Visualization Export Module
Creates Excel workbooks with embedded charts tied to data.
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple, Callable
from openpyxl import Workbook
from openpyxl.chart import (
    BarChart, LineChart, PieChart, DoughnutChart, AreaChart,
    ScatterChart, RadarChart, Reference, Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter

# Chart color palette
CHART_COLORS = [
    '4472C4', 'ED7D31', 'A5A5A5', 'FFC000', '5B9BD5',
    '70AD47', '264478', '9E480E', '636363', 'C55A11'
]

SEVERITY_COLORS = {
    'Critical': 'FF0000',
    'High': 'FFA500',
    'Medium': 'FFFF00',
    'Low': '00FF00',
    'Info': '0000FF'
}


class ExcelVisualizationExporter:
    """
    Exports visualization data to Excel with embedded charts.
    """

    # Chart types not directly supported in Excel via openpyxl
    UNSUPPORTED_CHARTS = {
        'sankey': 'Sankey/Flow diagrams not supported in Excel',
        'treemap': 'Treemaps require Excel 2016+ and cannot be created via openpyxl',
        'gauge': 'Gauge charts approximated using doughnut chart',
        'heatmap': 'Heatmaps implemented using conditional formatting instead of charts',
    }

    def __init__(self, lifecycle_df: pd.DataFrame,
                 historical_df: pd.DataFrame = None,
                 opdir_df: pd.DataFrame = None,
                 host_presence_df: pd.DataFrame = None):
        """
        Initialize the exporter with data sources.

        Args:
            lifecycle_df: Finding lifecycle DataFrame
            historical_df: Historical findings DataFrame
            opdir_df: OPDIR mapping DataFrame
            host_presence_df: Host presence analysis DataFrame
        """
        self.lifecycle_df = lifecycle_df if lifecycle_df is not None else pd.DataFrame()
        self.historical_df = historical_df if historical_df is not None else pd.DataFrame()
        self.opdir_df = opdir_df if opdir_df is not None else pd.DataFrame()
        self.host_presence_df = host_presence_df if host_presence_df is not None else pd.DataFrame()

        self.workbook = None
        self.sheet_count = 0

    def export_all_visualizations(self, filepath: str,
                                   filter_settings: Dict = None) -> Tuple[bool, List[str]]:
        """
        Export all available visualizations to Excel.

        Args:
            filepath: Output Excel file path
            filter_settings: Optional filter settings dict

        Returns:
            Tuple of (success, list of unsupported chart notes)
        """
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)

        unsupported_notes = []

        # Define all visualization categories and their charts
        visualization_groups = [
            ('Overview', self._create_overview_charts),
            ('Severity Analysis', self._create_severity_charts),
            ('Metrics', self._create_metrics_charts),
            ('Timeline', self._create_timeline_charts),
            ('OPDIR Compliance', self._create_opdir_charts),
            ('SLA Analysis', self._create_sla_charts),
            ('Network Analysis', self._create_network_charts),
            ('Plugin Analysis', self._create_plugin_charts),
            ('Host Tracking', self._create_host_tracking_charts),
            ('Priority Analysis', self._create_priority_charts),
            ('Efficiency', self._create_efficiency_charts),
        ]

        for group_name, create_func in visualization_groups:
            try:
                notes = create_func(filter_settings)
                if notes:
                    unsupported_notes.extend(notes)
            except Exception as e:
                print(f"Error creating {group_name} charts: {e}")

        # Add summary sheet at the beginning
        self._create_summary_sheet()

        # Save workbook
        try:
            self.workbook.save(filepath)
            return True, unsupported_notes
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False, unsupported_notes

    def _add_data_and_chart(self, sheet_name: str, data: pd.DataFrame,
                            chart_type: str, chart_title: str,
                            x_title: str = None, y_title: str = None,
                            chart_position: str = "E2",
                            data_start_row: int = 1,
                            show_legend: bool = True,
                            chart_width: int = 15,
                            chart_height: int = 10) -> None:
        """
        Add data to a worksheet and create a chart tied to it.

        Args:
            sheet_name: Name of the worksheet
            data: DataFrame with the data
            chart_type: Type of chart ('bar', 'line', 'pie', 'doughnut', 'area', 'scatter', 'radar')
            chart_title: Title for the chart
            x_title: X-axis title
            y_title: Y-axis title
            chart_position: Cell position for chart
            data_start_row: Row where data starts
            show_legend: Whether to show legend
            chart_width: Chart width in Excel units
            chart_height: Chart height in Excel units
        """
        if data.empty:
            return

        # Create or get worksheet
        safe_name = self._sanitize_sheet_name(sheet_name)
        if safe_name in self.workbook.sheetnames:
            ws = self.workbook[safe_name]
        else:
            ws = self.workbook.create_sheet(safe_name)

        # Write data to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), data_start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == data_start_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')

        # Auto-fit columns
        for col_idx, col_name in enumerate(data.columns, 1):
            max_len = max(len(str(col_name)), data[col_name].astype(str).str.len().max())
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        # Create chart
        num_rows = len(data) + 1  # +1 for header
        num_cols = len(data.columns)

        chart = self._create_chart(chart_type, chart_title, x_title, y_title, show_legend)
        if chart is None:
            return

        chart.width = chart_width
        chart.height = chart_height

        # Set up data references based on chart type
        if chart_type in ['pie', 'doughnut']:
            # Pie/Doughnut: first column is labels, second is values
            labels = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + num_rows - 1)
            values = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_start_row + num_rows - 1)
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(labels)
        elif chart_type == 'radar':
            # Radar chart
            labels = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + num_rows - 1)
            for col in range(2, num_cols + 1):
                values = Reference(ws, min_col=col, min_row=data_start_row, max_row=data_start_row + num_rows - 1)
                chart.add_data(values, titles_from_data=True)
            chart.set_categories(labels)
        elif chart_type == 'scatter':
            # Scatter chart: x values in col 1, y values in col 2
            x_values = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + num_rows - 1)
            y_values = Reference(ws, min_col=2, min_row=data_start_row + 1, max_row=data_start_row + num_rows - 1)
            series = Series(y_values, x_values, title=data.columns[1] if len(data.columns) > 1 else "Values")
            chart.series.append(series)
        else:
            # Bar, Line, Area: first column is categories, rest are series
            categories = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + num_rows - 1)
            for col in range(2, num_cols + 1):
                values = Reference(ws, min_col=col, min_row=data_start_row, max_row=data_start_row + num_rows - 1)
                chart.add_data(values, titles_from_data=True)
            chart.set_categories(categories)

        # Add chart to worksheet
        ws.add_chart(chart, chart_position)

    def _create_chart(self, chart_type: str, title: str,
                      x_title: str = None, y_title: str = None,
                      show_legend: bool = True):
        """Create a chart object of the specified type."""
        chart_classes = {
            'bar': BarChart,
            'bar_horizontal': BarChart,
            'line': LineChart,
            'pie': PieChart,
            'doughnut': DoughnutChart,
            'area': AreaChart,
            'scatter': ScatterChart,
            'radar': RadarChart,
        }

        chart_class = chart_classes.get(chart_type)
        if chart_class is None:
            return None

        chart = chart_class()
        chart.title = title
        chart.style = 10

        if chart_type == 'bar_horizontal':
            chart.type = 'bar'

        if x_title and hasattr(chart, 'x_axis'):
            chart.x_axis.title = x_title
        if y_title and hasattr(chart, 'y_axis'):
            chart.y_axis.title = y_title

        chart.legend = None if not show_legend else chart.legend

        return chart

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name for Excel."""
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        for char in invalid_chars:
            name = name.replace(char, '_')
        if len(name) > 31:
            name = name[:28] + '...'
        return name

    def _create_summary_sheet(self):
        """Create a summary sheet at the beginning of the workbook."""
        ws = self.workbook.create_sheet("Summary", 0)

        ws['A1'] = "Vulnerability Analysis - Excel Visualization Export"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Data summary
        ws['A5'] = "Data Summary"
        ws['A5'].font = Font(bold=True, size=12)

        row = 6
        if not self.lifecycle_df.empty:
            ws[f'A{row}'] = "Total Findings:"
            ws[f'B{row}'] = len(self.lifecycle_df)
            row += 1

            if 'status' in self.lifecycle_df.columns:
                active = len(self.lifecycle_df[self.lifecycle_df['status'] == 'Active'])
                remediated = len(self.lifecycle_df[self.lifecycle_df['status'] == 'Remediated'])
                ws[f'A{row}'] = "Active Findings:"
                ws[f'B{row}'] = active
                row += 1
                ws[f'A{row}'] = "Remediated Findings:"
                ws[f'B{row}'] = remediated
                row += 1

            if 'host_name' in self.lifecycle_df.columns:
                ws[f'A{row}'] = "Unique Hosts:"
                ws[f'B{row}'] = self.lifecycle_df['host_name'].nunique()
                row += 1

        # Sheet index
        row += 2
        ws[f'A{row}'] = "Sheet Index"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for sheet_name in self.workbook.sheetnames:
            if sheet_name != "Summary":
                ws[f'A{row}'] = sheet_name
                row += 1

        # Unsupported charts note
        row += 2
        ws[f'A{row}'] = "Notes on Chart Limitations"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        for chart_type, note in self.UNSUPPORTED_CHARTS.items():
            ws[f'A{row}'] = f"- {chart_type.title()}: {note}"
            row += 1

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _create_overview_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create overview charts."""
        notes = []
        df = self._apply_filters(self.lifecycle_df, filter_settings)

        if df.empty:
            return notes

        # Severity Distribution (Pie Chart)
        if 'severity' in df.columns:
            severity_data = df['severity'].value_counts().reset_index()
            severity_data.columns = ['Severity', 'Count']
            # Order by severity
            order = ['Critical', 'High', 'Medium', 'Low', 'Info']
            severity_data['order'] = severity_data['Severity'].apply(lambda x: order.index(x) if x in order else 99)
            severity_data = severity_data.sort_values('order').drop('order', axis=1)
            self._add_data_and_chart('Severity Distribution', severity_data, 'pie',
                                     'Findings by Severity', chart_position='D2')

        # Status Distribution (Doughnut Chart)
        if 'status' in df.columns:
            status_data = df['status'].value_counts().reset_index()
            status_data.columns = ['Status', 'Count']
            self._add_data_and_chart('Status Distribution', status_data, 'doughnut',
                                     'Findings by Status', chart_position='D2')

        return notes

    def _create_severity_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create severity analysis charts."""
        notes = []
        df = self._apply_filters(self.lifecycle_df, filter_settings)

        if df.empty:
            return notes

        # CVSS Score Distribution (Bar Chart)
        if 'cvss' in df.columns:
            df['cvss_range'] = pd.cut(df['cvss'].fillna(0),
                                       bins=[0, 2, 4, 6, 8, 10],
                                       labels=['0-2', '2-4', '4-6', '6-8', '8-10'])
            cvss_data = df['cvss_range'].value_counts().sort_index().reset_index()
            cvss_data.columns = ['CVSS Range', 'Count']
            self._add_data_and_chart('CVSS Distribution', cvss_data, 'bar',
                                     'CVSS Score Distribution', 'CVSS Range', 'Count')

        # Severity by Status (Stacked Bar)
        if 'severity' in df.columns and 'status' in df.columns:
            pivot = pd.crosstab(df['severity'], df['status']).reset_index()
            self._add_data_and_chart('Severity by Status', pivot, 'bar',
                                     'Remediation Status by Severity', 'Severity', 'Count')

        return notes

    def _create_metrics_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create metrics charts."""
        notes = []

        # For metrics, use all statuses for accurate calculation
        df = self.lifecycle_df.copy()
        if 'status' in df.columns:
            df = df[df['status'].isin(['Active', 'Remediated'])]

        if df.empty:
            return notes

        # Remediation Rate by Severity (Bar Chart)
        if 'severity' in df.columns and 'status' in df.columns:
            rate_data = []
            for sev in ['Critical', 'High', 'Medium', 'Low', 'Info']:
                sev_df = df[df['severity'] == sev]
                if len(sev_df) > 0:
                    rate = (sev_df['status'] == 'Remediated').sum() / len(sev_df) * 100
                    rate_data.append({'Severity': sev, 'Remediation Rate (%)': round(rate, 1)})

            if rate_data:
                rate_df = pd.DataFrame(rate_data)
                self._add_data_and_chart('Remediation Rate', rate_df, 'bar',
                                         'Remediation Rate by Severity', 'Severity', 'Rate (%)')

        # MTTR by Severity (Bar Chart) - only remediated findings
        if 'severity' in df.columns and 'days_open' in df.columns:
            remediated = df[df['status'] == 'Remediated']
            if not remediated.empty:
                mttr_data = remediated.groupby('severity')['days_open'].mean().reset_index()
                mttr_data.columns = ['Severity', 'Avg Days to Remediate']
                mttr_data['Avg Days to Remediate'] = mttr_data['Avg Days to Remediate'].round(1)
                order = ['Critical', 'High', 'Medium', 'Low', 'Info']
                mttr_data['order'] = mttr_data['Severity'].apply(lambda x: order.index(x) if x in order else 99)
                mttr_data = mttr_data.sort_values('order').drop('order', axis=1)
                self._add_data_and_chart('MTTR by Severity', mttr_data, 'bar',
                                         'Mean Time to Remediation', 'Severity', 'Days')

        # Findings by Age (Bar Chart)
        if 'days_open' in df.columns:
            active = df[df['status'] == 'Active'] if 'status' in df.columns else df
            if not active.empty:
                active['age_bucket'] = pd.cut(active['days_open'],
                                              bins=[0, 30, 60, 90, 180, 365, float('inf')],
                                              labels=['0-30', '31-60', '61-90', '91-180', '181-365', '365+'])
                age_data = active['age_bucket'].value_counts().sort_index().reset_index()
                age_data.columns = ['Age (Days)', 'Count']
                self._add_data_and_chart('Findings by Age', age_data, 'bar',
                                         'Active Findings by Age', 'Age Range', 'Count')

        # Top Risky Hosts (Horizontal Bar)
        if 'host_name' in df.columns:
            active = df[df['status'] == 'Active'] if 'status' in df.columns else df
            if not active.empty:
                host_counts = active['host_name'].value_counts().head(15).reset_index()
                host_counts.columns = ['Host', 'Vulnerability Count']
                self._add_data_and_chart('Top Risky Hosts', host_counts, 'bar_horizontal',
                                         'Top 15 Hosts by Vulnerability Count', 'Host', 'Count')

        return notes

    def _create_timeline_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create timeline/trend charts."""
        notes = []
        df = self._apply_filters(self.lifecycle_df, filter_settings)

        if df.empty or 'first_seen' not in df.columns:
            return notes

        df = df.copy()
        df['first_seen'] = pd.to_datetime(df['first_seen'])
        df['month'] = df['first_seen'].dt.to_period('M').astype(str)

        # Total Findings Over Time (Line Chart)
        monthly = df.groupby('month').size().reset_index()
        monthly.columns = ['Month', 'New Findings']
        if not monthly.empty:
            self._add_data_and_chart('Findings Over Time', monthly, 'line',
                                     'New Findings by Month', 'Month', 'Count')

        # Severity Timeline (Stacked Area)
        if 'severity' in df.columns:
            sev_monthly = pd.crosstab(df['month'], df['severity']).reset_index()
            self._add_data_and_chart('Severity Timeline', sev_monthly, 'area',
                                     'Findings by Severity Over Time', 'Month', 'Count')

        # New vs Resolved (Line Chart)
        if 'status' in df.columns and 'last_seen' in df.columns:
            df['last_seen'] = pd.to_datetime(df['last_seen'])
            df['resolved_month'] = df['last_seen'].dt.to_period('M').astype(str)

            new_by_month = df.groupby('month').size()
            resolved = df[df['status'] == 'Remediated']
            resolved_by_month = resolved.groupby('resolved_month').size()

            # Combine into single DataFrame
            all_months = sorted(set(new_by_month.index) | set(resolved_by_month.index))
            trend_data = pd.DataFrame({
                'Month': all_months,
                'New': [new_by_month.get(m, 0) for m in all_months],
                'Resolved': [resolved_by_month.get(m, 0) for m in all_months]
            })
            self._add_data_and_chart('New vs Resolved', trend_data, 'line',
                                     'New vs Resolved Findings', 'Month', 'Count')

        return notes

    def _create_opdir_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create OPDIR compliance charts."""
        notes = []
        df = self._apply_filters(self.lifecycle_df, filter_settings)

        if df.empty or 'opdir_number' not in df.columns:
            return notes

        # OPDIR Mapping Coverage (Pie Chart)
        mapped = df['opdir_number'].notna() & (df['opdir_number'] != '')
        coverage_data = pd.DataFrame({
            'Category': ['OPDIR Mapped', 'Not Mapped'],
            'Count': [mapped.sum(), (~mapped).sum()]
        })
        self._add_data_and_chart('OPDIR Coverage', coverage_data, 'pie',
                                 'OPDIR Mapping Coverage', chart_position='D2')

        # OPDIR Status Distribution (Doughnut)
        if 'opdir_status' in df.columns:
            opdir_df = df[mapped]
            if not opdir_df.empty:
                status_data = opdir_df['opdir_status'].value_counts().reset_index()
                status_data.columns = ['Status', 'Count']
                self._add_data_and_chart('OPDIR Status', status_data, 'doughnut',
                                         'OPDIR Status Distribution', chart_position='D2')

        # Compliance by OPDIR Year (Bar)
        if 'opdir_year' in df.columns:
            opdir_df = df[mapped]
            if not opdir_df.empty:
                year_data = opdir_df.groupby('opdir_year').size().reset_index()
                year_data.columns = ['Year', 'Count']
                year_data = year_data.sort_values('Year')
                self._add_data_and_chart('OPDIR by Year', year_data, 'bar',
                                         'Findings by OPDIR Year', 'Year', 'Count')

        return notes

    def _create_sla_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create SLA analysis charts."""
        notes = []
        df = self._apply_filters(self.lifecycle_df, filter_settings)

        if df.empty:
            return notes

        # SLA Compliance Status (Pie)
        if 'sla_status' in df.columns:
            sla_data = df['sla_status'].value_counts().reset_index()
            sla_data.columns = ['SLA Status', 'Count']
            self._add_data_and_chart('SLA Compliance', sla_data, 'pie',
                                     'SLA Compliance Status', chart_position='D2')

        # Overdue by Severity (Bar)
        if 'sla_status' in df.columns and 'severity' in df.columns:
            overdue = df[df['sla_status'] == 'Overdue']
            if not overdue.empty:
                overdue_sev = overdue['severity'].value_counts().reset_index()
                overdue_sev.columns = ['Severity', 'Overdue Count']
                self._add_data_and_chart('Overdue by Severity', overdue_sev, 'bar',
                                         'Overdue Findings by Severity', 'Severity', 'Count')

        # Days Until/Past SLA (Bar)
        if 'days_until_sla' in df.columns:
            active = df[df['status'] == 'Active'] if 'status' in df.columns else df
            if not active.empty:
                active = active.copy()
                active['sla_bucket'] = pd.cut(active['days_until_sla'],
                                              bins=[-float('inf'), -30, -7, 0, 7, 30, float('inf')],
                                              labels=['30+ Overdue', '7-30 Overdue', '0-7 Overdue',
                                                      '0-7 Left', '7-30 Left', '30+ Left'])
                sla_days = active['sla_bucket'].value_counts().reindex(
                    ['30+ Overdue', '7-30 Overdue', '0-7 Overdue', '0-7 Left', '7-30 Left', '30+ Left']
                ).fillna(0).reset_index()
                sla_days.columns = ['Days Until SLA', 'Count']
                self._add_data_and_chart('SLA Days Status', sla_days, 'bar',
                                         'Days Until/Past SLA', 'Category', 'Count')

        return notes

    def _create_network_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create network analysis charts."""
        notes = []
        df = self._apply_filters(self.lifecycle_df, filter_settings)

        if df.empty:
            return notes

        # Top Subnets (Bar)
        if 'subnet' in df.columns:
            subnet_data = df['subnet'].value_counts().head(15).reset_index()
            subnet_data.columns = ['Subnet', 'Vulnerability Count']
            self._add_data_and_chart('Top Subnets', subnet_data, 'bar_horizontal',
                                     'Top Subnets by Vulnerability Count', 'Subnet', 'Count')

        # Environment Distribution (Pie)
        if 'environment' in df.columns:
            env_data = df['environment'].value_counts().reset_index()
            env_data.columns = ['Environment', 'Count']
            self._add_data_and_chart('Environment Dist', env_data, 'pie',
                                     'Findings by Environment', chart_position='D2')

        # Host Criticality Distribution (Pie)
        if 'criticality' in df.columns:
            crit_data = df['criticality'].value_counts().reset_index()
            crit_data.columns = ['Criticality', 'Count']
            self._add_data_and_chart('Host Criticality', crit_data, 'pie',
                                     'Host Criticality Distribution', chart_position='D2')

        # Environment by Severity (Stacked Bar)
        if 'environment' in df.columns and 'severity' in df.columns:
            env_sev = pd.crosstab(df['environment'], df['severity']).reset_index()
            self._add_data_and_chart('Env by Severity', env_sev, 'bar',
                                     'Severity Distribution by Environment', 'Environment', 'Count')

        return notes

    def _create_plugin_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create plugin analysis charts."""
        notes = []
        df = self._apply_filters(self.lifecycle_df, filter_settings)

        if df.empty:
            return notes

        # Top Plugins (Bar)
        if 'plugin_id' in df.columns and 'plugin_name' in df.columns:
            plugin_counts = df.groupby(['plugin_id', 'plugin_name']).size().reset_index(name='Count')
            plugin_counts = plugin_counts.nlargest(15, 'Count')
            plugin_counts['Plugin'] = plugin_counts['plugin_name'].str[:40]
            top_plugins = plugin_counts[['Plugin', 'Count']]
            self._add_data_and_chart('Top Plugins', top_plugins, 'bar_horizontal',
                                     'Top 15 Most Common Plugins', 'Plugin', 'Count')

        # Plugin Severity Distribution (Pie)
        if 'plugin_family' in df.columns:
            family_data = df['plugin_family'].value_counts().head(10).reset_index()
            family_data.columns = ['Plugin Family', 'Count']
            self._add_data_and_chart('Plugin Families', family_data, 'pie',
                                     'Top Plugin Families', chart_position='D2')

        # Plugins by Host Count (Bar)
        if 'plugin_id' in df.columns and 'host_name' in df.columns:
            plugins_hosts = df.groupby('plugin_id')['host_name'].nunique().reset_index()
            plugins_hosts.columns = ['Plugin ID', 'Host Count']
            plugins_hosts = plugins_hosts.nlargest(15, 'Host Count')
            self._add_data_and_chart('Plugins by Hosts', plugins_hosts, 'bar',
                                     'Plugins Affecting Most Hosts', 'Plugin ID', 'Host Count')

        return notes

    def _create_host_tracking_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create host tracking charts."""
        notes = []

        if self.host_presence_df.empty:
            return notes

        df = self.host_presence_df.copy()

        # Host Status Distribution (Pie)
        if 'status' in df.columns:
            status_data = df['status'].value_counts().reset_index()
            status_data.columns = ['Status', 'Count']
            self._add_data_and_chart('Host Status', status_data, 'pie',
                                     'Host Status Distribution', chart_position='D2')

        # Host Presence Distribution (Bar)
        if 'presence_percentage' in df.columns:
            df['presence_bucket'] = pd.cut(df['presence_percentage'],
                                           bins=[0, 25, 50, 75, 100],
                                           labels=['0-25%', '25-50%', '50-75%', '75-100%'])
            presence_data = df['presence_bucket'].value_counts().sort_index().reset_index()
            presence_data.columns = ['Presence %', 'Host Count']
            self._add_data_and_chart('Host Presence', presence_data, 'bar',
                                     'Host Scan Presence Distribution', 'Presence Range', 'Count')

        # Missing Hosts (Bar)
        if 'status' in df.columns and 'hostname' in df.columns:
            missing = df[df['status'] == 'Missing'].head(15)
            if not missing.empty:
                missing_data = missing[['hostname']].copy()
                missing_data['Days Missing'] = missing.get('days_missing', 0)
                self._add_data_and_chart('Missing Hosts', missing_data, 'bar_horizontal',
                                         'Hosts Missing from Recent Scans', 'Host', 'Days')

        return notes

    def _create_priority_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create priority analysis charts."""
        notes = []
        df = self._apply_filters(self.lifecycle_df, filter_settings)

        if df.empty:
            return notes

        # Priority Distribution (Bar)
        if 'priority_score' in df.columns:
            df = df.copy()
            df['priority_level'] = pd.cut(df['priority_score'],
                                          bins=[0, 25, 50, 75, 100],
                                          labels=['Low', 'Medium', 'High', 'Critical'])
            priority_data = df['priority_level'].value_counts().reindex(
                ['Critical', 'High', 'Medium', 'Low']).fillna(0).reset_index()
            priority_data.columns = ['Priority Level', 'Count']
            self._add_data_and_chart('Priority Dist', priority_data, 'bar',
                                     'Priority Level Distribution', 'Priority', 'Count')

        # Priority by Severity (Grouped Bar)
        if 'priority_score' in df.columns and 'severity' in df.columns:
            priority_sev = df.groupby('severity')['priority_score'].mean().reset_index()
            priority_sev.columns = ['Severity', 'Avg Priority Score']
            priority_sev['Avg Priority Score'] = priority_sev['Avg Priority Score'].round(1)
            order = ['Critical', 'High', 'Medium', 'Low', 'Info']
            priority_sev['order'] = priority_sev['Severity'].apply(lambda x: order.index(x) if x in order else 99)
            priority_sev = priority_sev.sort_values('order').drop('order', axis=1)
            self._add_data_and_chart('Priority by Severity', priority_sev, 'bar',
                                     'Average Priority Score by Severity', 'Severity', 'Score')

        # Priority Matrix Note
        notes.append("Priority Matrix: Implemented as data table with conditional formatting")

        return notes

    def _create_efficiency_charts(self, filter_settings: Dict = None) -> List[str]:
        """Create efficiency analysis charts."""
        notes = []
        df = self._apply_filters(self.lifecycle_df, filter_settings)

        if df.empty:
            return notes

        # Resolution Velocity (Line)
        if 'status' in df.columns and 'last_seen' in df.columns:
            remediated = df[df['status'] == 'Remediated'].copy()
            if not remediated.empty:
                remediated['last_seen'] = pd.to_datetime(remediated['last_seen'])
                remediated['month'] = remediated['last_seen'].dt.to_period('M').astype(str)
                velocity = remediated.groupby('month').size().reset_index()
                velocity.columns = ['Month', 'Resolutions']
                self._add_data_and_chart('Resolution Velocity', velocity, 'line',
                                         'Resolution Velocity Over Time', 'Month', 'Resolutions')

        # Reappearance Analysis (Bar)
        if 'reappearances' in df.columns:
            reapp_data = df[df['reappearances'] > 0]['reappearances'].value_counts().head(10).sort_index().reset_index()
            reapp_data.columns = ['Reappearance Count', 'Findings']
            if not reapp_data.empty:
                self._add_data_and_chart('Reappearances', reapp_data, 'bar',
                                         'Finding Reappearance Distribution', 'Times Reappeared', 'Count')

        # Host Vulnerability Burden (Bar)
        if 'host_name' in df.columns:
            active = df[df['status'] == 'Active'] if 'status' in df.columns else df
            if not active.empty:
                burden = active.groupby('host_name').size().describe().reset_index()
                burden.columns = ['Statistic', 'Value']
                self._add_data_and_chart('Host Burden Stats', burden, 'bar',
                                         'Host Vulnerability Burden Statistics', 'Statistic', 'Value')

        return notes

    def _apply_filters(self, df: pd.DataFrame, filter_settings: Dict = None) -> pd.DataFrame:
        """Apply filter settings to DataFrame."""
        if df.empty or filter_settings is None:
            return df.copy() if not df.empty else df

        filtered = df.copy()

        if 'severity' in filter_settings and filter_settings['severity'] != 'All':
            if 'severity' in filtered.columns:
                filtered = filtered[filtered['severity'] == filter_settings['severity']]

        if 'status' in filter_settings and filter_settings['status'] != 'All':
            if 'status' in filtered.columns:
                filtered = filtered[filtered['status'] == filter_settings['status']]

        if 'environment' in filter_settings and filter_settings['environment'] != 'All':
            if 'environment' in filtered.columns:
                filtered = filtered[filtered['environment'] == filter_settings['environment']]

        return filtered

    def export_single_visualization(self, viz_name: str, filepath: str,
                                     filter_settings: Dict = None) -> bool:
        """
        Export a single visualization to Excel.

        Args:
            viz_name: Name of the visualization
            filepath: Output file path
            filter_settings: Optional filter settings

        Returns:
            True if successful
        """
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        viz_map = {
            'severity_distribution': lambda: self._create_overview_charts(filter_settings),
            'cvss_distribution': lambda: self._create_severity_charts(filter_settings),
            'remediation_rate': lambda: self._create_metrics_charts(filter_settings),
            'mttr': lambda: self._create_metrics_charts(filter_settings),
            'timeline': lambda: self._create_timeline_charts(filter_settings),
            'opdir': lambda: self._create_opdir_charts(filter_settings),
            'sla': lambda: self._create_sla_charts(filter_settings),
            'network': lambda: self._create_network_charts(filter_settings),
            'plugins': lambda: self._create_plugin_charts(filter_settings),
            'hosts': lambda: self._create_host_tracking_charts(filter_settings),
            'priority': lambda: self._create_priority_charts(filter_settings),
            'efficiency': lambda: self._create_efficiency_charts(filter_settings),
        }

        create_func = viz_map.get(viz_name.lower())
        if create_func:
            try:
                create_func()
                self.workbook.save(filepath)
                return True
            except Exception as e:
                print(f"Error exporting {viz_name}: {e}")
                return False

        return False


def export_visualizations_to_excel(lifecycle_df: pd.DataFrame,
                                   filepath: str,
                                   historical_df: pd.DataFrame = None,
                                   opdir_df: pd.DataFrame = None,
                                   host_presence_df: pd.DataFrame = None,
                                   filter_settings: Dict = None) -> Tuple[bool, List[str]]:
    """
    Convenience function to export all visualizations to Excel.

    Args:
        lifecycle_df: Finding lifecycle DataFrame
        filepath: Output Excel file path
        historical_df: Optional historical findings DataFrame
        opdir_df: Optional OPDIR mapping DataFrame
        host_presence_df: Optional host presence DataFrame
        filter_settings: Optional filter settings

    Returns:
        Tuple of (success, list of unsupported chart notes)
    """
    exporter = ExcelVisualizationExporter(
        lifecycle_df=lifecycle_df,
        historical_df=historical_df,
        opdir_df=opdir_df,
        host_presence_df=host_presence_df
    )
    return exporter.export_all_visualizations(filepath, filter_settings)
