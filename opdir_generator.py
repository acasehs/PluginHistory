import os
import re
import zipfile
import tarfile
import tempfile
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import json
import configparser  # pip install configparser (built-in)
from datetime import datetime
import ipaddress
from collections import defaultdict
import logging
import shutil
from typing import Dict, List, Set, Tuple, Optional
import subprocess
import sys
import time
import glob
import gzip
import threading
from pathlib import Path
import csv
import requests

# Configure logging with enhanced output
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('poam_generator.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =================================================
# OPDIR STANDARDIZATION
# =================================================
def standardize_opdir_format(opdir_number):
    """
    Standardize OPDIR number to XXXX-XX format (4 digits - 2 digits).
    Adds leading zeros where necessary.
    
    Examples:
        '1234-25' -> '1234-25'
        '234-25' -> '0234-25'
        '34-5' -> '0034-05'
        '1234-5' -> '1234-05'
    """
    if not opdir_number or opdir_number == 'xxxx-xx':
        return opdir_number
    
    try:
        # Split on hyphen
        parts = opdir_number.split('-')
        if len(parts) != 2:
            logger.warning(f"OPDIR format invalid (no hyphen): {opdir_number}")
            return opdir_number
        
        # Extract numeric parts
        first_part = parts[0].strip()
        second_part = parts[1].strip()
        
        # Pad with leading zeros
        first_part_padded = first_part.zfill(4)  # Ensure 4 digits
        second_part_padded = second_part.zfill(2)  # Ensure 2 digits
        
        standardized = f"{first_part_padded}-{second_part_padded}"
        
        if standardized != opdir_number:
            logger.info(f"Standardized OPDIR: {opdir_number} -> {standardized}")
        
        return standardized
        
    except Exception as e:
        logger.error(f"Error standardizing OPDIR {opdir_number}: {e}")
        return opdir_number

# =================================================
# OPDIR LOOKUP MANAGER
# =================================================
class OPDIRLookupManager:
    """Manages OPDIR number lookup from reference file."""
    
    def __init__(self):
        self.opdir_mapping = {}  # Maps IAVA/B to dict with OPDIR number and dates
        self.reference_file = None
        self.loaded = False
    
    def _normalize_iavm(self, iavm_string):
        """
        Normalize IAVM to standard format.
        Supports both:
        - Full format: YYYY-A-NNNN (preferred)
        - Suffix only: A-NNNN or B-NNNN (for backward compatibility)
        Removes prefixes (IAVA #, IAVB:) and suffixes (-S, -C).
        Returns normalized IAVM or None if can't parse.
        """
        # First try: Full format with year (YYYY-A-NNNN)
        full_pattern = r'(?:IAVA?[:\s#]*)?(\d{4}-[AB]-\d{4})(?:-[A-Z])?'
        match = re.search(full_pattern, iavm_string, re.IGNORECASE)
        if match:
            core = match.group(1)
            parts = core.split('-')
            if len(parts) == 3:
                return f"{parts[0]}-{parts[1].upper()}-{parts[2]}"
        
        # Second try: Suffix only format (A-NNNN or B-NNNN)
        # This handles reference files with just "B-0201" instead of "YYYY-B-0201"
        suffix_pattern = r'(?:IAVA?[:\s#]*)?([AB]-\d{4})(?:-[A-Z])?'
        match = re.search(suffix_pattern, iavm_string, re.IGNORECASE)
        if match:
            core = match.group(1)
            parts = core.split('-')
            if len(parts) == 2:
                return f"{parts[0].upper()}-{parts[1]}"
        
        return None
    
    def _infer_year_from_opdir(self, opdir_number):
        """
        Extract year from OPDIR number.
        Format: XXXX-YY where YY is last 2 digits of year.
        Example: 1234-24 → 2024, 5678-25 → 2025
        Returns full 4-digit year or None if can't extract.
        """
        try:
            # Extract last part after dash
            parts = opdir_number.split('-')
            if len(parts) >= 2:
                year_suffix = parts[-1].strip()
                # Convert to int to validate
                year_int = int(year_suffix)
                # Assume 2000s if 00-99
                full_year = 2000 + year_int
                return str(full_year)
        except (ValueError, IndexError):
            pass
        return None
    
    def _enhance_iavm_with_year(self, iavm_string, opdir_number):
        """
        If IAVM is suffix-only (B-0201), prepend year from OPDIR.
        If IAVM already has year (2024-B-0201), return as-is.
        
        Args:
            iavm_string: IAVA/B value from reference file (e.g., "B-0201")
            opdir_number: OPDIR number (e.g., "1234-24")
        
        Returns:
            Enhanced IAVM with year (e.g., "2024-B-0201")
        """
        # Check if already in full format (has year)
        if re.match(r'\d{4}-[AB]-\d{4}', iavm_string, re.IGNORECASE):
            # Already has year, return as-is
            return iavm_string
        
        # Check if suffix-only format (B-0201)
        if re.match(r'[AB]-\d{4}', iavm_string, re.IGNORECASE):
            # Extract year from OPDIR
            year = self._infer_year_from_opdir(opdir_number)
            if year:
                # Prepend year to create full format
                enhanced = f"{year}-{iavm_string}"
                logger.debug(f"Enhanced IAVM: '{iavm_string}' + OPDIR year {year} → '{enhanced}'")
                return enhanced
        
        # Return original if can't enhance
        return iavm_string
    
    def load_reference_file(self, file_path):
        """Load reference file (Excel or CSV) and build lookup dictionary."""
        try:
            logger.info(f"Loading OPDIR reference file: {file_path}")
            self.reference_file = file_path
            self.opdir_mapping.clear()
            
            # Determine file type and load accordingly
            if file_path.lower().endswith(('.xlsx', '.xls')):
                self._load_excel(file_path)
            elif file_path.lower().endswith('.csv'):
                self._load_csv(file_path)
            else:
                raise ValueError("Unsupported file type. Use .xlsx, .xls, or .csv")
            
            logger.info(f"Loaded {len(self.opdir_mapping)} OPDIR mappings")
            self.loaded = True
            return True
            
        except Exception as e:
            logger.error(f"Error loading reference file: {e}")
            messagebox.showerror("Error", f"Failed to load reference file:\n{str(e)}")
            return False
    
    def _load_excel(self, file_path):
        """Load Excel file and parse OPDIR mappings with dates."""
        try:
            import openpyxl  # pip install openpyxl
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl --break-system-packages")
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        # Find header row and columns
        header_row = None
        opdir_col = None
        iava_col = None
        poam_due_col = None
        final_due_col = None
        
        # Search for headers in first 5 rows
        for row_idx in range(1, min(6, sheet.max_row + 1)):
            row_values = [cell.value for cell in sheet[row_idx]]
            
            for col_idx, value in enumerate(row_values, 1):
                if value and isinstance(value, str):
                    value_upper = value.upper().strip()
                    if 'OPDIR' in value_upper and 'NUMBER' in value_upper:
                        opdir_col = col_idx
                    elif 'IAVA' in value_upper or 'IAVB' in value_upper:
                        iava_col = col_idx
                    elif 'POA&M' in value_upper and 'DUE' in value_upper:
                        poam_due_col = col_idx
                    elif 'FINAL' in value_upper and 'DUE' in value_upper:
                        final_due_col = col_idx
            
            if opdir_col and iava_col:
                header_row = row_idx
                break
        
        if not (opdir_col and iava_col):
            raise ValueError("Could not find OPDIR NUMBER and IAVA/B columns in Excel file")
        
        logger.info(f"Found headers at row {header_row}: OPDIR col={opdir_col}, IAVA/B col={iava_col}")
        if poam_due_col:
            logger.info(f"  POA&M DUE DATE col={poam_due_col}")
        if final_due_col:
            logger.info(f"  FINAL DUE DATE col={final_due_col}")
        
        # Parse data rows
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            opdir_value = sheet.cell(row_idx, opdir_col).value
            iava_value = sheet.cell(row_idx, iava_col).value
            
            if opdir_value and iava_value:
                # Clean and standardize values
                opdir_num = str(opdir_value).strip()
                iava_ref = str(iava_value).strip()
                
                # Smart enhancement: If IAVA/B is suffix-only (B-0201), 
                # infer year from OPDIR and prepend it (B-0201 → 2024-B-0201)
                enhanced_iavm = self._enhance_iavm_with_year(iava_ref, opdir_num)
                
                # Normalize IAVM to standard format (YYYY-A-NNNN)
                normalized_iavm = self._normalize_iavm(enhanced_iavm)
                if not normalized_iavm:
                    logger.warning(f"Could not normalize IAVM: {iava_ref} (enhanced: {enhanced_iavm})")
                    continue
                
                # Get dates if columns exist
                poam_due_date = None
                final_due_date = None
                
                if poam_due_col:
                    poam_val = sheet.cell(row_idx, poam_due_col).value
                    if poam_val:
                        # Handle date objects or string dates
                        if hasattr(poam_val, 'strftime'):
                            poam_due_date = poam_val.strftime('%m/%d/%Y')
                        else:
                            poam_due_date = str(poam_val).strip()
                
                if final_due_col:
                    final_val = sheet.cell(row_idx, final_due_col).value
                    if final_val:
                        # Handle date objects or string dates
                        if hasattr(final_val, 'strftime'):
                            final_due_date = final_val.strftime('%m/%d/%Y')
                        else:
                            final_due_date = str(final_val).strip()
                
                # Store in mapping with all info (standardize OPDIR format)
                # Use normalized IAVM as key to ensure consistent lookups
                self.opdir_mapping[normalized_iavm] = {
                    'opdir': standardize_opdir_format(opdir_num),
                    'poam_due_date': poam_due_date,
                    'final_due_date': final_due_date
                }
                if enhanced_iavm != iava_ref:
                    logger.info(f"Enhanced & Mapped: '{iava_ref}' → '{normalized_iavm}' (inferred year from OPDIR {opdir_num}) → OPDIR {standardize_opdir_format(opdir_num)}")
                else:
                    logger.debug(f"Mapped {normalized_iavm} (from {iava_ref}) -> {standardize_opdir_format(opdir_num)} (POA&M Due: {poam_due_date}, Final Due: {final_due_date})")
        
        wb.close()
    
    def _load_csv(self, file_path):
        """Load CSV file and parse OPDIR mappings with dates."""
        with open(file_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            
            # Find header row
            header_row = None
            opdir_col = None
            iava_col = None
            poam_due_col = None
            final_due_col = None
            
            for row_idx, row in enumerate(reader):
                if row_idx < 5:  # Check first 5 rows for headers
                    for col_idx, value in enumerate(row):
                        if value and isinstance(value, str):
                            value_upper = value.upper().strip()
                            if 'OPDIR' in value_upper and 'NUMBER' in value_upper:
                                opdir_col = col_idx
                            elif 'IAVA' in value_upper or 'IAVB' in value_upper:
                                iava_col = col_idx
                            elif 'POA&M' in value_upper and 'DUE' in value_upper:
                                poam_due_col = col_idx
                            elif 'FINAL' in value_upper and 'DUE' in value_upper:
                                final_due_col = col_idx
                    
                    if opdir_col is not None and iava_col is not None:
                        header_row = row_idx
                        break
            
            if opdir_col is None or iava_col is None:
                raise ValueError("Could not find OPDIR NUMBER and IAVA/B columns in CSV file")
        
        logger.info(f"Found headers: OPDIR col={opdir_col}, IAVA/B col={iava_col}")
        if poam_due_col is not None:
            logger.info(f"  POA&M DUE DATE col={poam_due_col}")
        if final_due_col is not None:
            logger.info(f"  FINAL DUE DATE col={final_due_col}")
        
        # Re-read to parse data
        with open(file_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader):
                if row_idx > header_row:
                    max_col = max(c for c in [opdir_col, iava_col, poam_due_col, final_due_col] if c is not None)
                    if len(row) > max_col:
                        opdir_value = row[opdir_col].strip()
                        iava_value = row[iava_col].strip()
                        
                        if opdir_value and iava_value:
                            # Smart enhancement: If IAVA/B is suffix-only (B-0201), 
                            # infer year from OPDIR and prepend it (B-0201 → 2024-B-0201)
                            enhanced_iavm = self._enhance_iavm_with_year(iava_value, opdir_value)
                            
                            # Normalize IAVM to standard format (YYYY-A-NNNN)
                            normalized_iavm = self._normalize_iavm(enhanced_iavm)
                            if not normalized_iavm:
                                logger.warning(f"Could not normalize IAVM: {iava_value} (enhanced: {enhanced_iavm})")
                                continue
                            
                            # Get dates if columns exist
                            poam_due_date = None
                            final_due_date = None
                            
                            if poam_due_col is not None and len(row) > poam_due_col:
                                poam_due_date = row[poam_due_col].strip() or None
                            
                            if final_due_col is not None and len(row) > final_due_col:
                                final_due_date = row[final_due_col].strip() or None
                            
                            # Store in mapping with all info (standardize OPDIR format)
                            # Use normalized IAVM as key to ensure consistent lookups
                            self.opdir_mapping[normalized_iavm] = {
                                'opdir': standardize_opdir_format(opdir_value),
                                'poam_due_date': poam_due_date,
                                'final_due_date': final_due_date
                            }
                            if enhanced_iavm != iava_value:
                                logger.info(f"Enhanced & Mapped: '{iava_value}' → '{normalized_iavm}' (inferred year from OPDIR {opdir_value}) → OPDIR {standardize_opdir_format(opdir_value)}")
                            else:
                                logger.debug(f"Mapped {normalized_iavm} (from {iava_value}) -> {standardize_opdir_format(opdir_value)} (POA&M Due: {poam_due_date}, Final Due: {final_due_date})")
    
    def lookup_opdir(self, iavm_id):
        """
        Look up OPDIR info for given IAVM ID.
        First tries full IAVM format (YYYY-A-NNNN) to distinguish years,
        then falls back to suffix only (A-NNNN) for backward compatibility.
        Returns dict with 'opdir', 'poam_due_date', 'final_due_date' or defaults if not found.
        """
        default_result = {
            'opdir': 'xxxx-xx',
            'poam_due_date': None,
            'final_due_date': None
        }
        
        if not self.loaded:
            logger.warning("OPDIR reference file not loaded, using defaults")
            return default_result
        
        # Extract IAVA/B reference from IAVM ID
        # Handle formats like: "IAVA:2025-A-4567", "IAVB:2021-A-0573", "2025-A-4567", etc.
        iavm_clean = iavm_id.split(":")[-1].strip()
        
        # Normalize to get clean IAVM (removes prefixes, suffixes, standardizes case)
        # e.g., "IAVB:2021-A-0573-S" -> "2021-A-0573"
        normalized_iavm = self._normalize_iavm(iavm_clean)
        
        if not normalized_iavm:
            logger.warning(f"Could not parse IAVA/B format from {iavm_id}")
            return default_result
        
        # Strategy 1: Try lookup with FULL IAVM (with year) - PREFERRED
        # This allows distinguishing 2021-A-0573 from 2024-A-0573
        if normalized_iavm in self.opdir_mapping:
            opdir_info = self.opdir_mapping[normalized_iavm]
            logger.info(f"OPDIR lookup (full): {iavm_id} -> {normalized_iavm} -> {opdir_info['opdir']} (Due: {opdir_info.get('poam_due_date', 'N/A')})")
            return opdir_info
        
        # Strategy 2: Try lookup with just suffix (A-NNNN) - FALLBACK for old format
        # This maintains backward compatibility with reference files that only have "A-4567"
        match = re.match(r'\d{4}-([A-Z]-\d+)', normalized_iavm)
        if match:
            suffix_only = match.group(1)  # Extract "A-4567" part
            
            if suffix_only in self.opdir_mapping:
                opdir_info = self.opdir_mapping[suffix_only]
                logger.info(f"OPDIR lookup (suffix): {iavm_id} -> {suffix_only} -> {opdir_info['opdir']} (Due: {opdir_info.get('poam_due_date', 'N/A')})")
                logger.warning(f"Matched using suffix only ({suffix_only}). Consider updating reference file to use full IAVM format (YYYY-A-NNNN) to avoid ambiguity.")
                return opdir_info
        
        logger.warning(f"No OPDIR mapping found for {normalized_iavm} (from {iavm_id})")
        return default_result
    
    
    def get_stats(self):
        """Get statistics about loaded mappings."""
        return {
            'loaded': self.loaded,
            'reference_file': self.reference_file,
            'mapping_count': len(self.opdir_mapping)
        }

# Global OPDIR lookup manager
opdir_lookup_manager = OPDIRLookupManager()

def get_opdir_info(iavm_id=None, config_manager=None):
    """
    Get full OPDIR info including dates.
    Returns dict with 'opdir', 'poam_due_date', 'final_due_date'.
    OPDIR numbers are automatically standardized to XXXX-XX format.
    """
    # If IAVM ID provided and reference file loaded, do lookup
    if iavm_id and opdir_lookup_manager.loaded:
        opdir_info = opdir_lookup_manager.lookup_opdir(iavm_id)
        # Standardize the OPDIR number
        opdir_info['opdir'] = standardize_opdir_format(opdir_info['opdir'])
        return opdir_info
    
    # Otherwise return defaults
    default_opdir = config_manager.get_config('TECHNICAL', 'opdir_release', 'xxxx-xx') if config_manager else 'xxxx-xx'
    return {
        'opdir': standardize_opdir_format(default_opdir),
        'poam_due_date': None,
        'final_due_date': None
    }

def get_opdir_release(config_manager=None, iavm_id=None):
    """
    Get OPDIR release number only (for backward compatibility).
    If iavm_id is provided and reference file is loaded, returns looked-up OPDIR.
    Otherwise returns configured default.
    """
    opdir_info = get_opdir_info(iavm_id, config_manager)
    return opdir_info['opdir']
    return 'xxxx-xx'

# Try to import required libraries for remote Ollama
try:
    import requests
    OLLAMA_AVAILABLE = True
    logger.info("Requests library available for remote Ollama AI contextualization")
except ImportError:
    OLLAMA_AVAILABLE = False
    logger.info("Requests library not available. Install with: pip install requests")

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    import fitz  # PyMuPDF
    PDF_AVAILABLE = True
    logger.info("PDF libraries available")
except ImportError as e:
    PDF_AVAILABLE = False
    logger.error(f"Required PDF libraries not installed: {e}")

# =================================================
# PDF FIELD MAPPING DOCUMENTATION
# =================================================
PDF_FIELD_MAPPING = {
    # ============ PAGE 1 FIELDS ============
    "IAVM": {
        "source": "Parsed from Nessus scan + Plugin DB",
        "type": "Auto-generated", 
        "predefined": False,
        "blank_behavior": "Extracted IAVA/IAVB references or None if not found"
    },
    "Date Submitted": {
        "source": "Auto-generated current date",
        "type": "Auto-generated",
        "predefined": False, 
        "blank_behavior": "Current date in MM/DD/YYYY format"
    },
    "Date Due": {
        "source": "NO VALUE - User must fill manually",
        "type": "Manual entry required",
        "predefined": False,
        "blank_behavior": "Always blank - user fills after generation"
    },
    "Number of Submission": {
        "source": "Hardcoded to '1st'",
        "type": "Hardcoded",
        "predefined": False,
        "blank_behavior": "Always '1st' as specified"
    },
    "RNOSC": {
        "source": "config_manager.get('rnosc')",
        "type": "Predefined configurable",
        "predefined": True,
        "blank_behavior": "Uses configured value or blank if not set"
    }
}

# =================================================
# CONSOLIDATED LOGGING FUNCTIONS
# =================================================
def log_with_prefix(prefix, message, level="info"):
    """Consolidated logging function with customizable prefix and level."""
    log_func = getattr(logger, level.lower(), logger.info)
    log_func(f"[{prefix.upper()}] {message}")

# Convenience functions for common log types
def log_success(message): log_with_prefix("SUCCESS", message)
def log_error(message): log_with_prefix("ERROR", message, "error")
def log_warning(message): log_with_prefix("WARNING", message, "warning")
def log_info(message): log_with_prefix("INFO", message)
def log_action(message): log_with_prefix("ACTION", message)
def log_status(message): log_with_prefix("STATUS", message)
def log_progress(message): log_with_prefix("PROGRESS", message)
def log_file(message): log_with_prefix("FILE", message)
def log_document(message): log_with_prefix("DOCUMENT", message)
def log_search(message): log_with_prefix("SEARCH", message)
def log_plugin(message): log_with_prefix("PLUGIN", message)
def log_ai(message): log_with_prefix("AI", message)
def log_config(message): log_with_prefix("CONFIG", message)
def log_ui(message): log_with_prefix("UI", message)
def log_generation(message): log_with_prefix("GENERATION", message)
def log_cleanup(message): log_with_prefix("CLEANUP", message)
def log_validation(message): log_with_prefix("VALIDATION", message)
def log_startup(message): log_with_prefix("STARTUP", message)
def log_exit(message): log_with_prefix("EXIT", message)
# =================================================
# IAVM TRACKING AND LOGGING
# =================================================
class IAVMTracker:
    """Track generated IAVMs with timestamps and metadata."""
    
    def __init__(self):
        self.script_dir = Path(__file__).parent
        self.tracking_file = self.script_dir / "iavm_generation_log.csv"
        self._ensure_tracking_file()
    
    def _ensure_tracking_file(self):
        """Ensure tracking file exists with proper headers."""
        if not self.tracking_file.exists():
            with open(self.tracking_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    'timestamp', 'iavm_id', 'hosts_affected', 'scan_file', 
                    'output_file', 'status', 'ai_enhanced', 'plugins_enriched',
                    'vulnerable_software_count'
                ])
            logger.info(f"Created IAVM tracking file: {self.tracking_file}")
    
    def log_generation(self, iavm_id, hosts_affected, scan_file, output_file, 
                      status='success', ai_enhanced=False, plugins_enriched=False,
                      vulnerable_software_count=0):
        """Log an IAVM generation event."""
        try:
            with open(self.tracking_file, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    iavm_id,
                    hosts_affected,
                    scan_file,
                    output_file,
                    status,
                    ai_enhanced,
                    plugins_enriched,
                    vulnerable_software_count
                ])
            logger.info(f"Logged generation: {iavm_id} -> {status} (vulnerable software: {vulnerable_software_count})")
        except Exception as e:
            logger.error(f"Error logging IAVM generation: {e}")
    
    def get_generation_history(self, days=30):
        """Get recent generation history."""
        try:
            if not self.tracking_file.exists():
                return []
            
            history = []
            with open(self.tracking_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    history.append(row)
            
            # Return all history entries (no limit)
            logger.info(f"Loaded {len(history)} generation history entries")
            return history
        except Exception as e:
            logger.error(f"Error reading generation history: {e}")
            return []

# Global IAVM tracker
iavm_tracker = IAVMTracker()

# =================================================
# OLLAMA SERVICE CLASS
# =================================================
class OllamaService:
    """Service for AI-enhanced contextualization using remote Ollama."""
    
    def __init__(self):
        self.available = OLLAMA_AVAILABLE
        self.models = []
        self.selected_model = None
        self.server_url = "http://172.16.27.122:11434"  # Remote server
        self.connected = False
        
        if self.available:
            self.test_connectivity()
    
    def test_connectivity(self):
        """Test connectivity to Ollama server."""
        if not self.available:
            return False
        
        try:
            import requests
            response = requests.get(f"{self.server_url}/api/tags", timeout=5)
            if response.status_code == 200:
                self.connected = True
                logger.info(f"Successfully connected to Ollama server at {self.server_url}")
                self.refresh_models()
                return True
            else:
                self.connected = False
                logger.error(f"Ollama server returned status {response.status_code}")
                return False
        except Exception as e:
            self.connected = False
            logger.error(f"Failed to connect to Ollama server: {e}")
            return False
        
    def refresh_models(self):
        """Get list of available Ollama models from remote server."""
        if not self.available or not self.connected:
            return []
        
        try:
            import requests
            response = requests.get(f"{self.server_url}/api/tags", timeout=10)
            if response.status_code == 200:
                data = response.json()
                self.models = [model['name'] for model in data.get('models', [])]
                logger.info(f"Found {len(self.models)} Ollama models: {self.models}")
                return self.models
            else:
                logger.error(f"Error fetching models: HTTP {response.status_code}")
                return []
        except Exception as e:
            logger.error(f"Error querying Ollama models: {e}")
            self.connected = False
            return []
    
    def set_model(self, model_name):
        """Set the model to use for AI enhancement."""
        if model_name in self.models:
            self.selected_model = model_name
            logger.info(f"Selected Ollama model: {model_name}")
            return True
        else:
            logger.error(f"Model {model_name} not available in Ollama")
            return False
    
    def enhance_narrative(self, field_name, plugin_data, context_data=None):
        """Use AI to enhance narrative fields based on vulnerability data."""
        if not self.available or not self.connected or not self.selected_model:
            logger.warning("AI enhancement requested but Ollama not available or model not set")
            return None
        
        try:
            # Build context prompt based on field type
            prompt_builders = {
                "Reason Cannot Complete": self._build_reason_prompt,
                "Operational Impact": self._build_impact_prompt,
                "Plan of Action": self._build_plan_prompt,
                "Timeline Milestones": self._build_timeline_prompt,
                "Temporary Mitigations": self._build_mitigations_prompt
            }
            
            if field_name not in prompt_builders:
                logger.warning(f"Unknown field for AI enhancement: {field_name}")
                return None
            
            prompt = prompt_builders[field_name](plugin_data, context_data)
            
            logger.info(f"Enhancing {field_name} with Ollama model {self.selected_model}")
            
            import requests
            response = requests.post(
                f"{self.server_url}/api/generate",
                json={
                    "model": self.selected_model,
                    "prompt": prompt,
                    "stream": False,
                    "options": {"temperature": 0.3}
                },
                timeout=30
            )
            
            if response.status_code == 200:
                data = response.json()
                enhanced_text = data.get('response', '').strip()
                
                if enhanced_text:
                    logger.info(f"Successfully enhanced {field_name} ({len(enhanced_text)} chars)")
                    return enhanced_text
                else:
                    logger.warning(f"Empty response from Ollama for {field_name}")
                    return None
            else:
                logger.error(f"Ollama API error: HTTP {response.status_code}")
                return None
                
        except Exception as e:
            logger.error(f"Error enhancing {field_name} with Ollama: {e}")
            return None
    
    def _build_reason_prompt(self, plugin_data, context_data):
        """Build prompt for 'Reason Cannot Complete' field."""
        system_descriptions = context_data.get('system_descriptions', []) if context_data else []
        
        prompt = f"""You are a cybersecurity analyst writing a POA&M (Plan of Action & Milestones) for a vulnerability management program.

VULNERABILITY DETAILS:
IAVM: {plugin_data.get('iavm_id', 'Unknown')}
Affected Systems: {len(plugin_data.get('hosts', {}))} hosts
Vulnerability Descriptions: {' | '.join(list(plugin_data.get('descriptions', []))[:3])}
Solutions Available: {' | '.join(list(plugin_data.get('solutions', []))[:2])}

SYSTEM CONTEXT:
{chr(10).join(system_descriptions) if system_descriptions else 'Standard enterprise systems'}

Write a professional, concise explanation (2-3 sentences) for why the required patching actions cannot be completed by the due date. Focus on:
- Testing and validation requirements specific to these systems
- Operational coordination needs
- System availability constraints
- Risk mitigation during patching

Keep it professional and specific to the vulnerability and system context. Do not use generic boilerplate language."""
        
        return prompt
    
    def _build_impact_prompt(self, plugin_data, context_data):
        """Build prompt for 'Operational Impact' field."""
        system_descriptions = context_data.get('system_descriptions', []) if context_data else []
        
        prompt = f"""You are a cybersecurity analyst writing a POA&M operational impact assessment.

VULNERABILITY DETAILS:
IAVM: {plugin_data.get('iavm_id', 'Unknown')}
Affected Systems: {len(plugin_data.get('hosts', {}))} hosts
System Types: {context_data.get('os_summary', 'Various systems') if context_data else 'Various systems'}

SYSTEM CONTEXT:
{chr(10).join(system_descriptions) if system_descriptions else 'Standard enterprise systems'}

Write a professional assessment (2-3 sentences) of the operational impact if these vulnerable assets were disconnected from the network. Consider the specific systems affected and their roles:
- Mission-critical services that might be affected
- Business continuity implications based on system functions
- User productivity impact
- Service availability concerns

Be specific to the systems described above but avoid overstating the impact. Keep it professional and factual."""
        
        return prompt
    
    def _build_plan_prompt(self, plugin_data, context_data):
        """Build prompt for 'Plan of Action' field."""
        prompt = f"""You are a cybersecurity analyst creating a remediation plan for a vulnerability.

VULNERABILITY DETAILS:
IAVM: {plugin_data.get('iavm_id', 'Unknown')}
Affected Systems: {len(plugin_data.get('hosts', {}))} hosts
Available Solutions: {' | '.join(list(plugin_data.get('solutions', []))[:2])}

Create a professional, step-by-step plan of action (4-6 numbered steps) to remediate this vulnerability. Include:
1. Initial coordination and planning
2. Testing procedures
3. Implementation steps
4. Verification activities
5. Documentation requirements

Keep each step concise but actionable. Focus on best practices for enterprise vulnerability management."""
        
        return prompt
    
    def _build_timeline_prompt(self, plugin_data, context_data):
        """Build prompt for 'Timeline Milestones' field."""
        prompt = f"""You are a cybersecurity analyst creating a timeline for vulnerability remediation.

VULNERABILITY DETAILS:
IAVM: {plugin_data.get('iavm_id', 'Unknown')}
Affected Systems: {len(plugin_data.get('hosts', {}))} hosts

Create a realistic timeline with weekly milestones for remediating this vulnerability. Format as:
Week X: [Activity description]

Include 4-10 milestones covering:
- Planning and coordination
- Testing phase
- Implementation windows
- Verification and validation
- Final documentation

Be realistic about timeframes for enterprise environments. Each milestone should be specific and measurable. Also consider that Oracle vendor patches are notoriously slow to release due to certification requirements and independent release channels for products"""
        
        return prompt
    
    def _build_mitigations_prompt(self, plugin_data, context_data):
        """Build prompt for 'Temporary Mitigations' field."""
        prompt = f"""You are a cybersecurity analyst documenting temporary security mitigations.

VULNERABILITY DETAILS:
IAVM: {plugin_data.get('iavm_id', 'Unknown')}
Affected Systems: {len(plugin_data.get('hosts', {}))} hosts
CVEs: {', '.join(list(plugin_data.get('cves', []))[:3])}

Describe professional temporary mitigations (3-4 sentences) that are in place to protect vulnerable assets until patching is complete. Consider:
- Network segmentation controls
- Access restrictions
- Monitoring enhancements
- Compensating controls

Be specific about security measures without revealing sensitive security details. Focus on standard enterprise security practices."""
        
        return prompt

# Global Ollama service instance
ollama_service = OllamaService()

# =================================================
# SYSTEM DESCRIPTIONS MANAGER
# =================================================
class SystemDescriptionsManager:
    """Manage system descriptions based on hostname patterns."""
    
    def __init__(self, config_manager):
        self.config_manager = config_manager
    
    def get_system_descriptions(self):
        """Get system descriptions dictionary."""
        if self.config_manager.config.has_section('SYSTEM_DESCRIPTIONS'):
            return dict(self.config_manager.config['SYSTEM_DESCRIPTIONS'])
        return {}
    
    def set_system_descriptions(self, descriptions_dict):
        """Set system descriptions dictionary."""
        if not self.config_manager.config.has_section('SYSTEM_DESCRIPTIONS'):
            self.config_manager.config.add_section('SYSTEM_DESCRIPTIONS')
        
        # Clear existing entries
        self.config_manager.config.remove_section('SYSTEM_DESCRIPTIONS')
        self.config_manager.config.add_section('SYSTEM_DESCRIPTIONS')
        
        # Add new entries
        for pattern, description in descriptions_dict.items():
            self.config_manager.config.set('SYSTEM_DESCRIPTIONS', pattern, description)
    
    def get_matching_descriptions(self, hostnames):
        """Get descriptions for hostnames that match patterns."""
        descriptions = []
        system_descriptions = self.get_system_descriptions()
        
        for pattern, description in system_descriptions.items():
            # Check if any hostname matches this pattern
            pattern_lower = pattern.lower()
            for hostname in hostnames:
                hostname_lower = hostname.lower()
                
                # Support wildcards and partial matches
                if ('*' in pattern_lower and 
                    re.match(pattern_lower.replace('*', '.*'), hostname_lower)) or \
                   (pattern_lower in hostname_lower) or \
                   (hostname_lower.startswith(pattern_lower)):
                    descriptions.append(f"{pattern}: {description}")
                    break
        
        return descriptions

# =================================================
# COMMON FILE HANDLING FUNCTIONS
# =================================================
def show_open_file_dialog(title="Select file", filetypes=[("All files", "*.*")]):
    """Show an open file dialog and return the selected file path."""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.update()
    
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes,
        parent=root
    )
    
    root.destroy()
    return file_path

def show_save_file_dialog(title="Save file as", defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")]):
    """Show a save file dialog and return the selected file path."""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.update()
    
    file_path = filedialog.asksaveasfilename(
        title=title,
        defaultextension=defaultextension,
        filetypes=filetypes,
        parent=root
    )
    
    root.destroy()
    return file_path

def safe_file_operation(operation, *args, max_retries=3, **kwargs):
    """Safely perform file operations with retry logic."""
    for attempt in range(max_retries):
        try:
            return operation(*args, **kwargs)
        except PermissionError as e:
            if attempt < max_retries - 1:
                logger.warning(f"Permission error on attempt {attempt + 1}: {e}")
                user_input = input(f"File may be open. Close it and press Enter to retry (attempt {attempt + 1}/{max_retries})...")
            else:
                logger.error(f"Failed after {max_retries} attempts due to permission error: {e}")
                raise
        except Exception as e:
            if attempt < max_retries - 1:
                logger.warning(f"Error on attempt {attempt + 1}: {e}")
                time.sleep(1)  # Brief delay before retry
            else:
                logger.error(f"Failed after {max_retries} attempts: {e}")
                raise
    
    return None

def enhanced_progress_reporting(current, total, operation_name, start_time=None):
    """Enhanced progress reporting."""
    if total > 0:
        percent = (current / total) * 100
        if start_time:
            elapsed = time.time() - start_time
            rate = current / elapsed if elapsed > 0 else 0
            eta = (total - current) / rate if rate > 0 else 0
            logger.info(f"{operation_name}: {current}/{total} ({percent:.1f}%) - {rate:.1f}/sec - ETA: {eta:.1f}s")
        else:
            logger.info(f"{operation_name}: {current}/{total} ({percent:.1f}%)")

# =================================================
# PLUGIN DATABASE FUNCTIONS
# =================================================
def extract_feed_timestamp(xml_file):
    """Extract feed timestamp from plugins.xml file."""
    try:
        logger.info(f"Extracting timestamp from {xml_file}...")
        file_size = os.path.getsize(xml_file)
        logger.info(f"File size: {file_size} bytes")
        
        # Read the beginning of the file to check structure
        with open(xml_file, 'r', encoding='utf-8', errors='ignore') as f:
            header = f.read(10000)  # Read first 10KB which should contain metadata
            logger.info(f"Reading first 10KB for timestamp...")
        
        # Look for <feed_timestamp> element
        feed_timestamp_match = re.search(r'<feed_timestamp>(\d+)</feed_timestamp>', header)
        if feed_timestamp_match:
            timestamp = int(feed_timestamp_match.group(1))
            timestamp_str = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
            logger.info(f"Found <feed_timestamp> element: {timestamp} ({timestamp_str})")
            return timestamp, timestamp_str
        
        # Look for <xml_timestamp> element as alternative
        xml_timestamp_match = re.search(r'<xml_timestamp>(\d+)</xml_timestamp>', header)
        if xml_timestamp_match:
            timestamp = int(xml_timestamp_match.group(1))
            timestamp_str = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
            logger.info(f"Found <xml_timestamp> element: {timestamp} ({timestamp_str})")
            return timestamp, timestamp_str
        
        # Look for feed_timestamp attribute on any element
        feed_timestamp_attr_match = re.search(r'feed_timestamp="(\d+)"', header)
        if feed_timestamp_attr_match:
            timestamp = int(feed_timestamp_attr_match.group(1))
            timestamp_str = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
            logger.info(f"Found feed_timestamp attribute: {timestamp} ({timestamp_str})")
            return timestamp, timestamp_str
        
        # Use file modification time as fallback
        mod_time = os.path.getmtime(xml_file)
        mod_time_str = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
        logger.info(f"Using file modification time as fallback: {int(mod_time)} ({mod_time_str})")
        return int(mod_time), mod_time_str
    
    except Exception as e:
        logger.error(f"Error extracting feed timestamp: {e}")
        # Always return a timestamp even in case of errors
        try:
            mod_time = os.path.getmtime(xml_file)
            mod_time_str = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
            logger.info(f"Using file modification time after error: {int(mod_time)} ({mod_time_str})")
            return int(mod_time), f"Error recovery - {mod_time_str}"
        except:
            current_time = int(time.time())
            current_time_str = datetime.fromtimestamp(current_time).strftime('%Y-%m-%d %H:%M:%S')
            return current_time, f"Error fallback - {current_time_str}"

def extract_plugins_xml_from_archive(archive_file, output_dir):
    """Extract plugins.xml from a compressed archive file."""
    logger.info(f"Attempting to extract plugins.xml from {archive_file}...")
    
    try:
        temp_dir = tempfile.mkdtemp(dir=output_dir)
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Check if plugins.xml already exists in the script directory
        existing_plugins_path = os.path.join(script_dir, "plugins.xml")
        existing_plugins_timestamp = None
        existing_timestamp_str = None
        
        if os.path.exists(existing_plugins_path):
            existing_plugins_timestamp, existing_timestamp_str = extract_feed_timestamp(existing_plugins_path)
            
            if existing_plugins_timestamp:
                date_str = datetime.fromtimestamp(existing_plugins_timestamp).strftime('%m%d%y')
                new_filename = f"plugins{date_str}.xml"
                new_filepath = os.path.join(script_dir, new_filename)
                
                logger.info(f"Renaming existing plugins.xml to {new_filename} (Feed date: {existing_timestamp_str})")
                try:
                    shutil.copy2(existing_plugins_path, new_filepath)
                except Exception as e:
                    logger.error(f"Error renaming existing plugins.xml: {e}")
        
        extracted_xml = None
        feed_timestamp = None
        feed_timestamp_str = None
        
        if archive_file.endswith('.tar.gz') or archive_file.endswith('.tgz'):
            try:
                with tarfile.open(archive_file, 'r:gz') as tar_ref:
                    tar_contents = tar_ref.getnames()
                    logger.info(f"TAR.GZ archive contains {len(tar_contents)} files/directories")
                    
                    plugins_xml_gz_files = [f for f in tar_contents if f.endswith('plugins.xml.gz')]
                    
                    if plugins_xml_gz_files:
                        logger.info(f"Found plugins.xml.gz files: {plugins_xml_gz_files}")
                        for plugins_xml_gz in plugins_xml_gz_files:
                            logger.info(f"Extracting {plugins_xml_gz}...")
                            tar_ref.extract(plugins_xml_gz, temp_dir)
                    else:
                        logger.info("No plugins.xml.gz found directly, extracting all files...")
                        tar_ref.extractall(temp_dir)
                
                for plugins_xml_gz in plugins_xml_gz_files:
                    extracted_gz_path = os.path.join(temp_dir, plugins_xml_gz)
                    if os.path.exists(extracted_gz_path):
                        logger.info(f"Processing {plugins_xml_gz}...")
                        plugins_xml_path = os.path.join(temp_dir, "plugins.xml")
                        
                        try:
                            import gzip
                            with gzip.open(extracted_gz_path, 'rb') as f_in:
                                with open(plugins_xml_path, 'wb') as f_out:
                                    logger.info(f"Decompressing to {plugins_xml_path}...")
                                    shutil.copyfileobj(f_in, f_out)
                            
                            if os.path.exists(plugins_xml_path) and os.path.getsize(plugins_xml_path) > 0:
                                logger.info(f"Successfully extracted to {plugins_xml_path} ({os.path.getsize(plugins_xml_path)} bytes)")
                                
                                try:
                                    feed_timestamp, feed_timestamp_str = extract_feed_timestamp(plugins_xml_path)
                                    if feed_timestamp:
                                        logger.info(f"Found feed timestamp: {feed_timestamp_str}")
                                        extracted_xml = plugins_xml_path
                                        final_path = os.path.join(script_dir, "plugins.xml")
                                        logger.info(f"Copying to {final_path}...")
                                        shutil.copy2(plugins_xml_path, final_path)
                                        extracted_xml = final_path
                                        break
                                except Exception as e:
                                    logger.error(f"Error extracting timestamp: {e}")
                        
                        except Exception as e:
                            logger.error(f"Error processing {plugins_xml_gz}: {e}")
            
            except Exception as e:
                logger.error(f"Error during tar.gz extraction: {e}")
        
        return extracted_xml, feed_timestamp, feed_timestamp_str
    
    except Exception as e:
        logger.error(f"Error extracting plugins.xml from archive: {e}")
        return None, None, None

def parse_plugins_xml(xml_file):
    """Parse a Nessus plugins.xml file and extract relevant information."""
    try:
        logger.info(f"Loading plugins XML from {xml_file}...")
        
        plugins_dict = {}
        start_time = time.time()
        
        feed_timestamp, feed_timestamp_str = extract_feed_timestamp(xml_file)
        
        import xml.etree.ElementTree as ET
        
        logger.info("Examining XML structure...")
        context = ET.iterparse(xml_file, events=('start',))
        for event, elem in context:
            root_tag = elem.tag
            logger.info(f"Root element: <{root_tag}>")
            break
        
        plugin_tag = 'nasl' if root_tag == 'nasl_plugins' else 'ReportItem'
        logger.info(f"Found <{root_tag}> format, looking for <{plugin_tag}> elements")
        
        plugin_count = 0
        progress_interval = 5000
        last_progress_time = start_time
        
        context = ET.iterparse(xml_file, events=('end',))
        
        for event, elem in context:
            if elem.tag == plugin_tag:
                try:
                    script_id_elem = elem.find('script_id')
                    if script_id_elem is not None and script_id_elem.text:
                        plugin_id = script_id_elem.text.strip()
                        
                        plugin_entry = {'plugin_id': plugin_id}
                        
                        if feed_timestamp:
                            plugin_entry['feed_timestamp'] = feed_timestamp
                            plugin_entry['feed_timestamp_str'] = feed_timestamp_str
                        
                        # Extract direct child elements
                        for child in elem:
                            if child.tag != 'attributes' and child.text:
                                if child.tag == 'script_name':
                                    plugin_entry['name'] = child.text.strip()
                                elif child.tag == 'script_family':
                                    plugin_entry['family'] = child.text.strip()
                                else:
                                    plugin_entry[child.tag] = child.text.strip()
                        
                        # Handle CVEs
                        cves_elem = elem.find('cves')
                        if cves_elem is not None:
                            cves = []
                            for cve_elem in cves_elem.findall('cve'):
                                if cve_elem.text and cve_elem.text.strip():
                                    cves.append(cve_elem.text.strip())
                            
                            if cves:
                                plugin_entry['cves'] = "\n".join(cves)
                        
                        # Handle cross-references
                        xrefs_elem = elem.find('xrefs')
                        if xrefs_elem is not None:
                            iavx_refs = []
                            other_refs = []
                            
                            for xref_elem in xrefs_elem.findall('xref'):
                                if xref_elem.text and xref_elem.text.strip():
                                    if any(x in xref_elem.text for x in ["IAVA:", "IAVB:", "IATM:"]):
                                        iavx_refs.append(xref_elem.text.strip())
                                    else:
                                        other_refs.append(xref_elem.text.strip())
                            
                            if iavx_refs:
                                plugin_entry['iavx'] = "\n".join(iavx_refs)
                            
                            if other_refs:
                                plugin_entry['cross_references'] = "\n".join(other_refs)
                        
                        # Process attributes
                        attributes_elem = elem.find('attributes')
                        if attributes_elem is not None:
                            for attr_elem in attributes_elem.findall('attribute'):
                                name_elem = attr_elem.find('name')
                                value_elem = attr_elem.find('value')
                                
                                if name_elem is not None and name_elem.text and value_elem is not None and value_elem.text:
                                    attr_name = name_elem.text.strip()
                                    attr_value = value_elem.text.strip()
                                    
                                    attr_prefix_name = f"atb_{attr_name}"
                                    plugin_entry[attr_prefix_name] = attr_value
                                    
                                    # Map key attributes
                                    if attr_name == "description":
                                        plugin_entry['description'] = attr_value
                                    elif attr_name == "solution":
                                        plugin_entry['solution'] = attr_value
                                    elif attr_name == "see_also":
                                        plugin_entry['see_also'] = attr_value
                                    elif attr_name == "iava":
                                        plugin_entry['iava'] = attr_value
                                        if 'iavx' not in plugin_entry:
                                            plugin_entry['iavx'] = f"IAVA:{attr_value}"
                                        else:
                                            plugin_entry['iavx'] += f"\nIAVA:{attr_value}"
                        
                        plugins_dict[plugin_id] = plugin_entry
                        plugin_count += 1
                        
                        current_time = time.time()
                        if plugin_count % progress_interval == 0 or (current_time - last_progress_time) >= 10:
                            elapsed = current_time - start_time
                            plugins_per_sec = plugin_count / elapsed if elapsed > 0 else 0
                            
                            logger.info(f"Progress: {plugin_count} plugins processed - {plugins_per_sec:.1f} plugins/sec")
                            last_progress_time = current_time
                
                except Exception as e:
                    logger.error(f"Error processing plugin: {str(e)}")
                    continue
                
                elem.clear()
        
        logger.info(f"Found {plugin_count} plugins in XML file")
        elapsed = time.time() - start_time
        logger.info(f"Successfully loaded {plugin_count} plugins from XML in {elapsed:.1f} seconds.")
        return plugins_dict
    
    except Exception as e:
        logger.error(f"Error loading plugins XML: {e}")
        return None

def load_plugins_database_json(plugins_file):
    """Load the Nessus plugins database from a JSON file."""
    try:
        logger.info(f"Loading plugins database from {plugins_file}...")
        import json
        
        plugins_dict = {}
        plugin_count = 0
        start_time = time.time()
        
        try:
            with open(plugins_file, 'r', encoding='utf-8', errors='ignore') as f:
                logger.info("Parsing JSON file - this may take a while for large files...")
                data = json.load(f)
                logger.info("JSON parsing complete, extracting plugin data...")
                
                feed_timestamp = None
                if 'feed_timestamp' in data:
                    try:
                        feed_timestamp = int(data['feed_timestamp'])
                        feed_timestamp_date = datetime.fromtimestamp(feed_timestamp)
                        logger.info(f"Plugin feed timestamp: {feed_timestamp} ({feed_timestamp_date.strftime('%Y-%m-%d %H:%M:%S')})")
                    except (ValueError, TypeError) as e:
                        logger.error(f"Error parsing feed timestamp: {e}")
                
                feed_timestamp_str = ""
                if feed_timestamp:
                    feed_timestamp_str = datetime.fromtimestamp(feed_timestamp).strftime('%Y-%m-%d %H:%M:%S')
                
                if 'nasl' in data and isinstance(data['nasl'], list):
                    total_plugins = len(data['nasl'])
                    logger.info(f"Processing {total_plugins} plugins from structured JSON...")
                    
                    for i, plugin in enumerate(data['nasl']):
                        try:
                            if 'script_id' in plugin:
                                plugin_id = str(plugin['script_id'])
                                
                                plugin_entry = {'plugin_id': plugin_id}
                                
                                if feed_timestamp:
                                    plugin_entry['feed_timestamp'] = feed_timestamp
                                    plugin_entry['feed_timestamp_str'] = feed_timestamp_str
                                
                                for key in plugin:
                                    if key != 'attributes':
                                        plugin_entry[key] = plugin[key]
                                
                                if 'script_name' in plugin:
                                    plugin_entry['name'] = plugin['script_name']
                                    
                                if 'script_family' in plugin:
                                    plugin_entry['family'] = plugin['script_family']
                                
                                # Process attributes
                                if 'attributes' in plugin and isinstance(plugin['attributes'], dict) and 'attribute' in plugin['attributes']:
                                    attributes = plugin['attributes']['attribute']
                                    if isinstance(attributes, list):
                                        for attr in attributes:
                                            if 'name' in attr and 'value' in attr:
                                                attr_name = f"atb_{attr['name']}"
                                                plugin_entry[attr_name] = attr['value']
                                                
                                                if attr['name'] == "description":
                                                    plugin_entry['description'] = attr['value']
                                                elif attr['name'] == "solution":
                                                    plugin_entry['solution'] = attr['value']
                                                elif attr['name'] == "see_also":
                                                    plugin_entry['see_also'] = attr['value']
                                                elif attr['name'] == "iava":
                                                    plugin_entry['iava'] = attr['value']
                                                    if 'iavx' not in plugin_entry:
                                                        plugin_entry['iavx'] = f"IAVA:{attr['value']}"
                                                    else:
                                                        plugin_entry['iavx'] += f"\nIAVA:{attr['value']}"
                                
                                # Extract CVEs
                                if 'cves' in plugin and plugin['cves']:
                                    cves = []
                                    if isinstance(plugin['cves'], dict) and 'cve' in plugin['cves']:
                                        cve_data = plugin['cves']['cve']
                                        if isinstance(cve_data, list):
                                            cves.extend([cve for cve in cve_data if cve])
                                        elif cve_data:
                                            cves.append(str(cve_data))
                                    
                                    if cves:
                                        plugin_entry['cves'] = "\n".join(cves)
                                
                                # Extract cross references
                                if 'xrefs' in plugin and plugin['xrefs']:
                                    iavx_refs = []
                                    other_refs = []
                                    
                                    if isinstance(plugin['xrefs'], dict) and 'xref' in plugin['xrefs']:
                                        xref_data = plugin['xrefs']['xref']
                                        if isinstance(xref_data, list):
                                            for xref in xref_data:
                                                if xref:
                                                    if any(x in xref for x in ["IAVA:", "IAVB:", "IATM:"]):
                                                        iavx_refs.append(xref)
                                                    else:
                                                        other_refs.append(xref)
                                        elif xref_data:
                                            if any(x in str(xref_data) for x in ["IAVA:", "IAVB:", "IATM:"]):
                                                iavx_refs.append(str(xref_data))
                                            else:
                                                other_refs.append(str(xref_data))
                                    
                                    if other_refs:
                                        plugin_entry['cross_references'] = "\n".join(other_refs)
                                    
                                    if iavx_refs:
                                        if 'iavx' not in plugin_entry:
                                            plugin_entry['iavx'] = "\n".join(iavx_refs)
                                        else:
                                            plugin_entry['iavx'] += "\n" + "\n".join(iavx_refs)
                                
                                plugins_dict[plugin_id] = plugin_entry
                                plugin_count += 1
                                
                        except Exception as e:
                            logger.error(f"Error processing plugin {i}: {str(e)}")
                            continue
                
                if plugin_count > 0:
                    elapsed = time.time() - start_time
                    logger.info(f"Successfully loaded {plugin_count} plugins from database in {elapsed:.1f} seconds.")
                    return plugins_dict
            
            logger.info("No plugins found in the standard format. The file may have a different structure.")
            return None
            
        except json.JSONDecodeError:
            logger.error("JSON parsing error. The file may be too large or have an invalid format.")
            return None
        except Exception as e:
            logger.error(f"Error loading plugins database: {e}")
            return None
    
    except Exception as e:
        logger.error(f"Error accessing plugins database file: {e}")
        return None

def load_plugins_database(plugins_file=None):
    """Load the Nessus plugins database for enriching finding information."""
    if not plugins_file:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        default_plugins_file = os.path.join(script_dir, "plugins.xml")
        existing_plugins_feed_timestamp = None
        existing_plugins_feed_timestamp_str = "Unknown"
        
        plugins_pattern = os.path.join(script_dir, "plugins*.xml")
        all_plugins_files = glob.glob(plugins_pattern)
        
        plugins_timestamps = {}
        
        for plugin_file in all_plugins_files:
            file_timestamp, file_timestamp_str = extract_feed_timestamp(plugin_file)
            if file_timestamp:
                logger.info(f"Found plugins file: {os.path.basename(plugin_file)}, timestamp: {file_timestamp_str}")
                plugins_timestamps[plugin_file] = (file_timestamp, False)
        
        sc_plugins_pattern = os.path.join(script_dir, "CM-*-sc-plugins.tar.gz")
        sc_plugins_files = glob.glob(sc_plugins_pattern)
        
        if sc_plugins_files:
            sc_plugins_files.sort(key=os.path.getmtime, reverse=True)
            archive_file = sc_plugins_files[0]
            archive_date = os.path.getmtime(archive_file)
            logger.info(f"Found plugins archive at {archive_file}")
            logger.info(f"Archive file date: {datetime.fromtimestamp(archive_date).strftime('%Y-%m-%d %H:%M:%S')}")
            
            extracted_xml, archive_feed_timestamp, archive_feed_timestamp_str = extract_plugins_xml_from_archive(archive_file, tempfile.mkdtemp())
            
            if extracted_xml and archive_feed_timestamp:
                logger.info(f"Extracted plugins.xml feed timestamp: {archive_feed_timestamp_str} (Unix: {archive_feed_timestamp})")
                
                plugins_timestamps[extracted_xml] = (archive_feed_timestamp, True)
                
                newest_file = None
                newest_timestamp = 0
                is_fresh = False
                
                if plugins_timestamps:
                    sorted_files = sorted(plugins_timestamps.items(), 
                                          key=lambda x: (x[1][0], x[1][1]), 
                                          reverse=True)
                    newest_file, (newest_timestamp, is_fresh) = sorted_files[0]
                
                if newest_timestamp > archive_feed_timestamp or (newest_timestamp == archive_feed_timestamp and is_fresh):
                    logger.info(f"Found existing plugins file with newer or same timestamp: {os.path.basename(newest_file)}")
                    logger.info(f"Using existing file: {newest_file}")
                    plugins_file = newest_file
                else:
                    logger.info(f"Archive has newest feed timestamp. Copying extracted plugins.xml to {default_plugins_file}")
                    shutil.copy2(extracted_xml, default_plugins_file)
                    plugins_file = default_plugins_file
                    
                    plugins_timestamps[default_plugins_file] = (archive_feed_timestamp, True)
            else:
                logger.info("Could not extract plugins.xml or determine feed timestamp from archive.")
                if plugins_timestamps:
                    sorted_files = sorted(plugins_timestamps.items(), 
                                         key=lambda x: (x[1][0], x[1][1]), 
                                         reverse=True)
                    newest_file, _ = sorted_files[0]
                    logger.info(f"Using existing plugins file: {newest_file}")
                    plugins_file = newest_file
                else:
                    return None
        elif plugins_timestamps:
            sorted_files = sorted(plugins_timestamps.items(), 
                                 key=lambda x: x[1][0], 
                                 reverse=True)
            newest_file, _ = sorted_files[0]
            logger.info(f"Using newest existing plugins file: {newest_file}")
            plugins_file = newest_file
        else:
            return None
    
    try:
        if plugins_file.lower().endswith('.xml'):
            return parse_plugins_xml(plugins_file)
        elif plugins_file.lower().endswith('.json'):
            return load_plugins_database_json(plugins_file)
        else:
            with open(plugins_file, 'r', encoding='utf-8', errors='ignore') as f:
                first_line = f.readline().strip()
                if first_line.startswith('{') or first_line.startswith('['):
                    return load_plugins_database_json(plugins_file)
                elif first_line.startswith('<?xml') or first_line.startswith('<'):
                    return parse_plugins_xml(plugins_file)
                else:
                    logger.error(f"Unknown file format for {plugins_file}. File should be XML or JSON.")
                    return None
    
    except Exception as e:
        logger.error(f"Error accessing plugins database file: {e}")
        return None

# =================================================
# NESSUS PARSING FUNCTIONS
# =================================================
def extract_nested_zips(zip_file_path, extraction_dir):
    """Extract a zip file that may contain other zip files."""
    logger.info(f"Extracting zip files from {zip_file_path}")
    
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        zip_ref.extractall(extraction_dir)
    
    for root, _, files in os.walk(extraction_dir):
        for file in files:
            if file.endswith('.zip'):
                zip_path = os.path.join(root, file)
                nested_dir = os.path.join(root, os.path.splitext(file)[0])
                os.makedirs(nested_dir, exist_ok=True)
                
                logger.info(f"Extracting nested zip: {file}")
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(nested_dir)

def find_nessus_files(directory):
    """Find all .nessus files in a directory (recursively)."""
    nessus_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.nessus'):
                nessus_files.append(os.path.join(root, file))
    logger.info(f"Found {len(nessus_files)} .nessus files")
    return nessus_files

def extract_hostname_from_plugins(host):
    """Extract hostname from specific plugins if available."""
    hostname = None
    
    # Try plugin 55472 (Hostname)
    for item in host.findall(".//ReportItem[@pluginID='55472']"):
        plugin_output = item.find("plugin_output")
        if plugin_output is not None and plugin_output.text:
            lines = plugin_output.text.strip().split("\n")
            for line in lines:
                if "hostname command" in line:
                    hostname = line.split()[0].split(".")[0]
                    return hostname

    # Try plugin 12053 (DNS Resolution)
    for item in host.findall(".//ReportItem[@pluginID='12053']"):
        plugin_output = item.find("plugin_output")
        if plugin_output is not None and plugin_output.text:
            match = re.search(r"resolves as ([\w.-]+)", plugin_output.text)
            if match:
                hostname = match.group(1).split(".")[0]
                return hostname
    
    return hostname

def extract_versions_from_text(text, plugin_name=""):
    """Extract version information from text using multiple patterns."""
    versions = []
    
    version_patterns = [
        r'([A-Za-z][A-Za-z\s&\-\.]+?)\s+(?:version\s+)?v?([0-9]+(?:\.[0-9]+){1,3}(?:\.[0-9]+)?)',
        r'version\s+([0-9]+(?:\.[0-9]+){1,4})',
        r'v\.?\s*([0-9]+(?:\.[0-9]+){1,4})',
        r'build\s+([0-9]+(?:\.[0-9]+)*)',
        r'service\s+pack\s+([0-9]+)',
        r'sp\s*([0-9]+)',
        r'\(version\s+([0-9]+(?:\.[0-9]+)*)\)',
        r'([0-9]+\.[0-9]+\.[0-9]+\.[0-9]+)',
        r'([0-9]+\.[0-9]+\.[0-9]+)',
        r'([0-9]+\.[0-9]+)',
    ]
    
    text_lower = text.lower()
    
    for pattern in version_patterns:
        matches = re.finditer(pattern, text, re.IGNORECASE)
        for match in matches:
            if len(match.groups()) == 2:
                software, version = match.groups()
                software = software.strip()
                version = version.strip()
                
                if (len(software) > 2 and 
                    not any(word in software.lower() for word in ['the', 'and', 'or', 'in', 'on', 'at']) and
                    version != '0.0'):
                    version_str = f"{software} {version}"
                    if version_str not in versions:
                        versions.append(version_str)
            else:
                version = match.group(1).strip()
                if version and version != '0.0':
                    if plugin_name and plugin_name != "Unknown Plugin":
                        plugin_software = extract_software_name_from_plugin(plugin_name)
                        if plugin_software:
                            version_str = f"{plugin_software} {version}"
                        else:
                            version_str = version
                    else:
                        version_str = version
                    
                    if version_str not in versions:
                        versions.append(version_str)
    
    return versions

def extract_software_name_from_plugin(plugin_name):
    """Extract likely software name from plugin name."""
    software_patterns = [
        r'^([A-Za-z][A-Za-z\s&\-\.]+?)\s+(?:Multiple\s+)?Vulnerabilit',
        r'^([A-Za-z][A-Za-z\s&\-\.]+?)\s+(?:Security\s+)?Update',
        r'^([A-Za-z][A-Za-z\s&\-\.]+?)\s+(?:Remote\s+)?Code',
        r'^([A-Za-z][A-Za-z\s&\-\.]+?)\s+(?:Buffer\s+)?Overflow',
        r'^([A-Za-z][A-Za-z\s&\-\.]+?)\s+(?:Denial\s+of\s+Service|DoS)',
        r'^([A-Za-z][A-Za-z\s&\-\.]+?)\s+(?:Information\s+)?Disclosure',
        r'^([A-Za-z][A-Za-z\s&\-\.]+?)\s+\w+\s+Detection',
        r'^(Microsoft\s+[A-Za-z\s]+?)\s+',
        r'^(Adobe\s+[A-Za-z\s]+?)\s+',
        r'^(Apache\s+[A-Za-z\s]+?)\s+',
        r'^(Oracle\s+[A-Za-z\s]+?)\s+',
    ]
    
    for pattern in software_patterns:
        match = re.search(pattern, plugin_name, re.IGNORECASE)
        if match:
            software = match.group(1).strip()
            software = re.sub(r'\s+(Server|Client|Service|Application)$', '', software, flags=re.IGNORECASE)
            if len(software) > 2:
                return software
    
    return None

def format_iavm_software_versions(iavm_data):
    """Format IAVM-specific software versions for display in POA&M."""
    if 'software_versions' not in iavm_data or not iavm_data['software_versions']:
        return "Software version information not available for this IAVM"
    
    version_summary = defaultdict(lambda: {'hosts': set(), 'plugin_names': set()})
    
    for host_ip, plugin_versions in iavm_data['software_versions'].items():
        hostname = iavm_data['hosts'].get(host_ip, {}).get('hostname', host_ip)
        host_display = f"{hostname} ({host_ip})" if hostname != host_ip else host_ip
        
        for plugin_id, plugin_data in plugin_versions.items():
            plugin_name = plugin_data.get('plugin_name', f'Plugin {plugin_id}')
            versions = plugin_data.get('versions', [])
            
            for version in versions:
                version_summary[version]['hosts'].add(host_display)
                version_summary[version]['plugin_names'].add(plugin_name)
    
    if not version_summary:
        return "No vulnerable software versions identified for this IAVM"
    
    formatted_lines = []
    
    for version, data in sorted(version_summary.items()):
        hosts = sorted(list(data['hosts']))
        
        if len(hosts) <= 3:
            host_str = ", ".join(hosts)
        else:
            host_str = f"{', '.join(hosts[:3])} ... (+{len(hosts)-3} more hosts)"
        
        formatted_lines.append(f"• {version}")
        formatted_lines.append(f"  Affected hosts: {host_str}")
    
    return "\n".join(formatted_lines)

def format_iavm_software_with_plugin_details(iavm_data):
    """
    Format IAVM software versions with plugin ID, plugin name, and plugin output.
    Removes duplicate outputs per host.
    """
    # Try vulnerable_software first
    if 'vulnerable_software' in iavm_data and iavm_data['vulnerable_software']:
        formatted_lines = []
        
        # Process each host
        for host_ip in sorted(iavm_data['vulnerable_software'].keys()):
            hostname = iavm_data['hosts'].get(host_ip, {}).get('hostname', host_ip)
            host_display = f"{hostname} ({host_ip})" if hostname != host_ip else host_ip
            
            formatted_lines.append(f"\n=== {host_display} ===")
            
            plugin_software = iavm_data['vulnerable_software'][host_ip]
            
            # Process each plugin for this host
            for plugin_id in sorted(plugin_software.keys(), key=lambda x: int(x) if x.isdigit() else 0):
                plugin_data = plugin_software[plugin_id]
                plugin_name = plugin_data.get('plugin_name', f'Plugin {plugin_id}')
                vulnerable_software = plugin_data.get('vulnerable_software', [])
                
                # Get unique outputs (deduplicated)
                software_list = []
                
                for software in vulnerable_software:
                    package_name = software.get('package_name', 'Unknown')
                    installed_ver = software.get('installed_version', 'Unknown')
                    should_be_ver = software.get('should_be_version', 'N/A')
                    
                    software_info = f"{package_name} {installed_ver}"
                    if should_be_ver and should_be_ver != 'N/A':
                        software_info += f" (should be {should_be_ver})"
                    
                    software_list.append(software_info)
                
                # Remove duplicates while preserving order
                seen = set()
                unique_software = []
                for item in software_list:
                    if item not in seen:
                        seen.add(item)
                        unique_software.append(item)
                
                if unique_software:
                    formatted_lines.append(f"\nPlugin ID {plugin_id}: {plugin_name}")
                    for sw in unique_software:
                        formatted_lines.append(f"  • {sw}")
        
        if len(formatted_lines) > 1:
            return "\n".join(formatted_lines)
    
    # Fallback: Try to get plugin names and basic info
    if 'plugins' in iavm_data and iavm_data['plugins']:
        logger.warning(f"No vulnerable_software data, falling back to plugin list for display")
        formatted_lines = []
        
        for host_ip in sorted(iavm_data['hosts'].keys()):
            hostname = iavm_data['hosts'].get(host_ip, {}).get('hostname', host_ip)
            host_display = f"{hostname} ({host_ip})" if hostname != host_ip else host_ip
            
            formatted_lines.append(f"\n=== {host_display} ===")
            
            # List plugins without detailed software versions
            plugin_list = sorted(iavm_data['plugins'])[:10]  # Limit to first 10 for readability
            for plugin_id in plugin_list:
                formatted_lines.append(f"Plugin ID {plugin_id}")
            
            if len(iavm_data['plugins']) > 10:
                formatted_lines.append(f"... and {len(iavm_data['plugins']) - 10} more plugins")
        
        if len(formatted_lines) > 1:
            return "\n".join(formatted_lines)
    
    logger.warning(f"No software version data available in iavm_data. Keys present: {list(iavm_data.keys())}")
    return "Software version information not available for this IAVM"

def extract_os_information(host):
    """Extract OS information from various plugins."""
    os_info = "Unknown"
    
    # Plugin 11936: OS Identification
    for item in host.findall(".//ReportItem[@pluginID='11936']"):
        plugin_output = item.find("plugin_output")
        if plugin_output is not None and plugin_output.text:
            first_line = plugin_output.text.strip().split('\n')[0]
            if first_line and "operating system" in first_line.lower():
                os_info = first_line
                break
            elif first_line:
                os_info = first_line
                break
    
    # Plugin 45590: Common Platform Enumeration
    if os_info == "Unknown":
        for item in host.findall(".//ReportItem[@pluginID='45590']"):
            plugin_output = item.find("plugin_output")
            if plugin_output is not None and plugin_output.text:
                lines = plugin_output.text.strip().split('\n')
                for line in lines:
                    if "cpe:/" in line.lower():
                        cpe_match = re.search(r'cpe:/o:([^:]+):([^:]+)', line)
                        if cpe_match:
                            vendor = cpe_match.group(1)
                            product = cpe_match.group(2)
                            os_info = f"{vendor.title()} {product.title()}"
                            break
                if os_info != "Unknown":
                    break
    
    return os_info

def extract_vulnerable_software_from_plugin_output(plugin_output_text, plugin_name=""):
    """Extract vulnerable software information from plugin output."""
    if not plugin_output_text:
        return []
    
    should_be_pattern = r'(?:should be:|Should be\s+:?\s+)([^\r\n]+?)(?:\s+or\s+later)?(?=\s|$|\r|\n|\.)'
    installed_pattern = r'(?:installed:|Remote package installed\s+:?\s+)([^\r\n]+?)(?=\s*(?:should be|Should be)|$|\r|\n|\.|\s*,)'
    
    direct_installed_pattern = r'Remote package installed\s+:\s+([\w\d\.-]+)'
    direct_should_be_pattern = r'Should be\s+:\s+([\w\d\.-]+)'
    
    package_pattern = r'(?:package:|Package:)([^\r\n]+?)(?=\s*(?:installed|Remote package installed)|$|\r|\n)'
    
    vulnerable_software = []
    
    direct_installed = re.findall(direct_installed_pattern, plugin_output_text)
    direct_should_be = re.findall(direct_should_be_pattern, plugin_output_text)
    
    if direct_installed and direct_should_be:
        for i in range(min(len(direct_installed), len(direct_should_be))):
            installed_version = direct_installed[i]
            should_be_version = direct_should_be[i]
            
            package_name = extract_package_name_from_version(installed_version)
            
            vulnerable_software.append({
                'package_name': package_name,
                'installed_version': clean_version_string(installed_version),
                'should_be_version': clean_version_string(should_be_version),
                'extraction_method': 'direct_pattern',
                'plugin_name': plugin_name
            })
    
    elif "should be:" in plugin_output_text.lower() or "Should be" in plugin_output_text:
        should_be_matches = re.findall(should_be_pattern, plugin_output_text, re.IGNORECASE)
        installed_matches = re.findall(installed_pattern, plugin_output_text, re.IGNORECASE)
        package_matches = re.findall(package_pattern, plugin_output_text, re.IGNORECASE)
        
        max_matches = max(len(should_be_matches), len(installed_matches), 1)
        
        for i in range(max_matches):
            should_be = should_be_matches[i].strip() if i < len(should_be_matches) else ""
            installed = installed_matches[i].strip() if i < len(installed_matches) else ""
            package = package_matches[i].strip() if i < len(package_matches) else ""
            
            should_be = clean_version_string(should_be)
            installed = clean_version_string(installed)
            
            if not package and installed:
                package = extract_package_name_from_version(installed)
            
            if not package:
                package = extract_software_name_from_plugin(plugin_name)
            
            if should_be or installed:
                vulnerable_software.append({
                    'package_name': package or "Unknown Package",
                    'installed_version': installed,
                    'should_be_version': should_be,
                    'extraction_method': 'general_pattern',
                    'plugin_name': plugin_name
                })
    
    return vulnerable_software

def extract_package_name_from_version(version_string):
    """Extract package name from version string."""
    if not version_string:
        return ""
    
    package_parts = version_string.split('-')
    if len(package_parts) >= 2:
        return '-'.join(package_parts[:-1])
    
    for sep in ['_', '.']:
        if sep in version_string:
            parts = version_string.split(sep)
            if len(parts) >= 2:
                return parts[0]
    
    return ""

def clean_version_string(version_text):
    """Clean up version strings by removing common suffixes and extra text."""
    if not version_text:
        return ""
    
    cleaned = re.sub(r'\s+or\s+later.*$', '', version_text, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s+or\s+higher.*$', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s+and\s+above.*$', '', cleaned, flags=re.IGNORECASE)
    
    cleaned = re.sub(r'[.,;:\s]+$', '', cleaned)
    
    return cleaned.strip()

def format_vulnerable_software_for_poam(vulnerable_software_list):
    """Format vulnerable software information for POA&M display."""
    if not vulnerable_software_list:
        return "No vulnerable software versions identified"
    
    formatted_lines = []
    
    software_by_package = defaultdict(list)
    for software in vulnerable_software_list:
        package_name = software['package_name']
        software_by_package[package_name].append(software)
    
    for package_name, software_entries in software_by_package.items():
        formatted_lines.append(f"• {package_name}")
        
        for entry in software_entries:
            installed = entry['installed_version']
            should_be = entry['should_be_version']
            
            if installed and should_be:
                formatted_lines.append(f"  Installed: {installed} → Should be: {should_be}")
            elif installed:
                formatted_lines.append(f"  Installed: {installed}")
            elif should_be:
                formatted_lines.append(f"  Should be: {should_be}")
    
    return "\n".join(formatted_lines)

def extract_iavm_specific_software_versions_enhanced(host, iavm_plugin_ids, plugins_dict=None):
    """Extract software version information specific to the IAVM plugins."""
    iavm_software = {}
    
    for plugin_id in iavm_plugin_ids:
        vulnerable_software = []
        plugin_name = "Unknown Plugin"
        
        for item in host.findall(f".//ReportItem[@pluginID='{plugin_id}']"):
            plugin_name = item.attrib.get('pluginName', 'Unknown Plugin')
            
            plugin_output = item.find("plugin_output")
            if plugin_output is not None and plugin_output.text:
                output_text = plugin_output.text.strip()
                extracted_software = extract_vulnerable_software_from_plugin_output(output_text, plugin_name)
                vulnerable_software.extend(extracted_software)
            
            description = item.find("description")
            if description is not None and description.text:
                desc_text = description.text.strip()
                extracted_software = extract_vulnerable_software_from_plugin_output(desc_text, plugin_name)
                vulnerable_software.extend(extracted_software)
        
        if plugins_dict and plugin_id in plugins_dict:
            plugin_info = plugins_dict[plugin_id]
            
            if 'name' in plugin_info:
                plugin_name = plugin_info['name']
            
            if 'description' in plugin_info:
                db_software = extract_vulnerable_software_from_plugin_output(plugin_info['description'], plugin_name)
                vulnerable_software.extend(db_software)
            
            if 'solution' in plugin_info:
                solution_software = extract_vulnerable_software_from_plugin_output(plugin_info['solution'], plugin_name)
                vulnerable_software.extend(solution_software)
        
        unique_software = []
        seen = set()
        for software in vulnerable_software:
            key = (software['package_name'], software['installed_version'], software['should_be_version'])
            if key not in seen:
                seen.add(key)
                unique_software.append(software)
        
        if unique_software:
            iavm_software[plugin_id] = {
                'plugin_name': plugin_name,
                'vulnerable_software': unique_software
            }
    
    return iavm_software

def format_iavm_vulnerable_software(iavm_data):
    """Format IAVM-specific vulnerable software for display in POA&M."""
    if 'vulnerable_software' not in iavm_data or not iavm_data['vulnerable_software']:
        return "Vulnerable software information not available for this IAVM"
    
    software_summary = defaultdict(lambda: {
        'hosts': set(), 
        'installed_versions': set(), 
        'should_be_versions': set(),
        'plugin_names': set()
    })
    
    for host_ip, plugin_software in iavm_data['vulnerable_software'].items():
        hostname = iavm_data['hosts'].get(host_ip, {}).get('hostname', host_ip)
        host_display = f"{hostname} ({host_ip})" if hostname != host_ip else host_ip
        
        for plugin_id, plugin_data in plugin_software.items():
            plugin_name = plugin_data.get('plugin_name', f'Plugin {plugin_id}')
            vulnerable_software = plugin_data.get('vulnerable_software', [])
            
            for software in vulnerable_software:
                package_name = software['package_name']
                software_summary[package_name]['hosts'].add(host_display)
                software_summary[package_name]['plugin_names'].add(plugin_name)
                
                if software['installed_version']:
                    software_summary[package_name]['installed_versions'].add(software['installed_version'])
                if software['should_be_version']:
                    software_summary[package_name]['should_be_versions'].add(software['should_be_version'])
    
    if not software_summary:
        return "No vulnerable software versions identified for this IAVM"
    
    formatted_lines = []
    
    for package_name, data in sorted(software_summary.items()):
        hosts = sorted(list(data['hosts']))
        installed_versions = sorted(list(data['installed_versions']))
        should_be_versions = sorted(list(data['should_be_versions']))
        
        if len(hosts) <= 3:
            host_str = ", ".join(hosts)
        else:
            host_str = f"{', '.join(hosts[:3])} ... (+{len(hosts)-3} more hosts)"
        
        formatted_lines.append(f"• {package_name}")
        formatted_lines.append(f"  Affected hosts: {host_str}")
        
        if installed_versions:
            version_str = ", ".join(installed_versions[:3])
            if len(installed_versions) > 3:
                version_str += f" ... (+{len(installed_versions)-3} more)"
            formatted_lines.append(f"  Installed versions: {version_str}")
        
        if should_be_versions:
            version_str = ", ".join(should_be_versions[:3])
            if len(should_be_versions) > 3:
                version_str += f" ... (+{len(should_be_versions)-3} more)"
            formatted_lines.append(f"  Should be: {version_str}")
    
    return "\n".join(formatted_lines)

def parse_nessus_for_iavm_enhanced(nessus_file, zip_context=None, plugins_dict=None):
    """Enhanced version of parse_nessus_for_iavm that includes vulnerable software detection."""
    logger.info(f"Parsing {nessus_file} for IAVM findings with enhanced software detection")
    
    try:
        tree = ET.parse(nessus_file)
        root = tree.getroot()
        
        original_filename = os.path.basename(nessus_file)
        report_name = original_filename
        report_element = root.find(".//Report")
        if report_element is not None and 'name' in report_element.attrib:
            report_name = report_element.attrib['name']
        
        scan_file_ref = original_filename
        if zip_context:
            scan_file_ref = f"{zip_context} > {original_filename}"
        
        iavm_findings = defaultdict(lambda: {
            'hosts': {},
            'plugins': set(),
            'plugin_ids': set(),
            'descriptions': set(),
            'solutions': set(),
            'cves': set(),
            'vulnerable_software': {},
            'report_name': report_name,
            'scan_file_ref': scan_file_ref
        })
        
        hosts = root.findall(".//ReportHost")
        logger.info(f"Processing {len(hosts)} hosts for IAVM findings with vulnerable software detection")
        
        total_plugins = 0
        enriched_plugins = 0
        missing_plugins = set()
        
        for host in hosts:
            host_name = host.attrib.get('name', 'Unknown')
            hostname = extract_hostname_from_plugins(host) or host_name
            os_info = extract_os_information(host)
            
            host_info = {
                'ip': host_name,
                'hostname': hostname,
                'os': os_info,
                'software_version': 'Determined per IAVM'
            }
            
            host_iavm_plugins = defaultdict(set)
            
            items = host.findall(".//ReportItem")
            
            for item in items:
                plugin_id = item.attrib.get('pluginID', '')
                plugin_name = item.attrib.get('pluginName', 'Unknown')
                total_plugins += 1
                
                description = ""
                description_elem = item.find("description")
                if description_elem is not None and description_elem.text:
                    description = description_elem.text.strip()
                
                solution = ""
                solution_elem = item.find("solution")
                if solution_elem is not None and solution_elem.text:
                    solution = solution_elem.text.strip()
                
                cves = set()
                cve_elements = item.findall("cve") or item.findall(".//cve")
                for cve_elem in cve_elements:
                    if cve_elem.text and cve_elem.text.strip():
                        cves.add(cve_elem.text.strip())
                
                iavx_refs = set()
                
                see_also_elem = item.find("see_also")
                if see_also_elem is not None and see_also_elem.text:
                    for line in see_also_elem.text.split('\n'):
                        line = line.strip()
                        if line and ("IAVA:" in line or "IAVB:" in line):
                            iavx_refs.add(line)
                
                xref_elements = item.findall("xref") or item.findall(".//xref")
                for xref_elem in xref_elements:
                    if xref_elem.text and xref_elem.text.strip():
                        xref_text = xref_elem.text.strip()
                        if "IAVA:" in xref_text or "IAVB:" in xref_text:
                            iavx_refs.add(xref_text)
                
                plugin_found = False
                if plugins_dict is not None:
                    if plugin_id in plugins_dict:
                        plugin_info = plugins_dict[plugin_id]
                        plugin_found = True
                    elif plugin_id.lstrip('0') in plugins_dict:
                        plugin_info = plugins_dict[plugin_id.lstrip('0')]
                        plugin_found = True
                    elif str(int(plugin_id)) in plugins_dict:
                        plugin_info = plugins_dict[str(int(plugin_id))]
                        plugin_found = True
                        
                    if plugin_found:
                        enriched_plugins += 1
                        
                        if not description and 'description' in plugin_info:
                            plugin_desc = plugin_info['description']
                            if isinstance(plugin_desc, str):
                                description = plugin_desc
                        
                        if not solution and 'solution' in plugin_info:
                            plugin_sol = plugin_info['solution']
                            if isinstance(plugin_sol, str):
                                solution = plugin_sol
                        
                        if plugin_name == 'Unknown' and 'name' in plugin_info:
                            plugin_name = plugin_info['name']
                        
                        if not cves and 'cves' in plugin_info:
                            if isinstance(plugin_info['cves'], str):
                                for cve in plugin_info['cves'].split('\n'):
                                    if cve.strip():
                                        cves.add(cve.strip())
                            elif isinstance(plugin_info['cves'], list):
                                for cve in plugin_info['cves']:
                                    if cve and isinstance(cve, str) and cve.strip():
                                        cves.add(cve.strip())
                        
                        if 'iavx' in plugin_info:
                            if isinstance(plugin_info['iavx'], str):
                                for iavx in plugin_info['iavx'].split('\n'):
                                    if iavx.strip():
                                        iavx_refs.add(iavx.strip())
                            elif isinstance(plugin_info['iavx'], list):
                                for iavx in plugin_info['iavx']:
                                    if iavx and isinstance(iavx, str) and iavx.strip():
                                        iavx_refs.add(iavx.strip())
                        
                        if 'iava' in plugin_info:
                            iava_value = plugin_info['iava']
                            if iava_value:
                                iavx_refs.add(f"IAVA:{iava_value}")
                    else:
                        missing_plugins.add(plugin_id)
                
                for iavx_ref in iavx_refs:
                    if "IAVA:" in iavx_ref or "IAVB:" in iavx_ref:
                        iavm_match = re.search(r'(IAVA?:\S+|IAVB:\S+)', iavx_ref)
                        if iavm_match:
                            iavm_id = iavm_match.group(1)
                            
                            host_iavm_plugins[iavm_id].add(plugin_id)
                            iavm_findings[iavm_id]['plugin_ids'].add(plugin_id)
                            
                            iavm_findings[iavm_id]['hosts'][host_name] = host_info
                            iavm_findings[iavm_id]['plugins'].add(f"{plugin_id} - {plugin_name}")
                            
                            if description:
                                iavm_findings[iavm_id]['descriptions'].add(description)
                            if solution:
                                iavm_findings[iavm_id]['solutions'].add(solution)
                            if cves:
                                iavm_findings[iavm_id]['cves'].update(cves)
            
            for iavm_id, plugin_ids in host_iavm_plugins.items():
                iavm_vulnerable_software = extract_iavm_specific_software_versions_enhanced(host, plugin_ids, plugins_dict)
                
                if iavm_vulnerable_software:
                    if host_name not in iavm_findings[iavm_id]['vulnerable_software']:
                        iavm_findings[iavm_id]['vulnerable_software'][host_name] = {}
                    
                    iavm_findings[iavm_id]['vulnerable_software'][host_name].update(iavm_vulnerable_software)
        
        if plugins_dict is not None:
            enrichment_rate = (enriched_plugins / total_plugins * 100) if total_plugins > 0 else 0
            logger.info(f"Plugin database enrichment: {enriched_plugins}/{total_plugins} plugins enriched ({enrichment_rate:.1f}%)")
            if missing_plugins:
                logger.info(f"{len(missing_plugins)} plugin IDs not found in database (showing first 10): {list(missing_plugins)[:10]}")
        
        logger.info(f"Found {len(iavm_findings)} IAVM references in {nessus_file}")
        
        for iavm_id, data in iavm_findings.items():
            vulnerable_software_count = sum(len(host_software) for host_software in data['vulnerable_software'].values())
            if vulnerable_software_count > 0:
                logger.info(f"IAVM {iavm_id}: {vulnerable_software_count} vulnerable software instances across {len(data['vulnerable_software'])} hosts")
        
        return dict(iavm_findings)
        
    except ET.ParseError as e:
        logger.error(f"Error parsing {nessus_file}: {e}")
        return {}
    except Exception as e:
        logger.error(f"Error processing {nessus_file}: {e}")
        return {}

# =================================================
# IP RANGE CALCULATION
# =================================================
def calculate_ip_ranges(ip_list, default_subnet_mask=24):
    """Calculate IP ranges/subnets from a list of IP addresses."""
    if not ip_list:
        return []
    
    logger.info(f"Calculating IP ranges for {len(ip_list)} addresses")
    
    networks = set()
    
    for ip_str in ip_list:
        try:
            ip = ipaddress.ip_address(ip_str)
            if ip.version == 4:
                network = ipaddress.ip_network(f"{ip}/{default_subnet_mask}", strict=False)
            else:
                network = ipaddress.ip_network(f"{ip}/64", strict=False)
            networks.add(network)
        except ValueError as e:
            logger.warning(f"Invalid IP address {ip_str}: {e}")
            continue
    
    sorted_networks = sorted(networks)
    ranges = [str(network) for network in sorted_networks]
    
    logger.info(f"Calculated {len(ranges)} IP ranges: {ranges}")
    return ranges

# =================================================
# ASSET DESCRIPTION GENERATION
# =================================================
def get_asset_description(host_info, config_manager):
    """Generate asset description with custom overrides."""
    hostname = host_info.get('hostname', 'Unknown')
    os_info = host_info.get('os', 'Unknown OS')
    ip = host_info.get('ip', 'Unknown IP')
    
    generic_desc = f"System: {hostname} ({ip})\nOperating System: {os_info}"
    
    asset_descriptions = config_manager.get_asset_descriptions()
    custom_desc = ""
    
    for ip_range_str, description in asset_descriptions.items():
        try:
            if ip == ip_range_str:
                custom_desc = description
                break
            
            try:
                if '/' in ip_range_str:
                    network = ipaddress.ip_network(ip_range_str, strict=False)
                    if ipaddress.ip_address(ip) in network:
                        custom_desc = description
                        break
                else:
                    if ip == ip_range_str:
                        custom_desc = description
                        break
            except (ValueError, ipaddress.AddressValueError):
                if ip_range_str in ip or ip in ip_range_str:
                    custom_desc = description
                    break
        except Exception as e:
            logger.warning(f"Error matching IP range {ip_range_str}: {e}")
            continue
    
    if custom_desc:
        return f"{generic_desc}\n{custom_desc}"
    else:
        return generic_desc

# =================================================
# ENHANCED CONFIGURATION MANAGER
# =================================================
class ConfigManager:
    """Enhanced configuration manager using INI files for better organization."""
    
    def __init__(self):
        self.script_dir = Path(__file__).parent
        self.predefined_file = self.script_dir / "predefined.ini"
        self.config_file = self.script_dir / "config.ini"
        
        self.predefined = configparser.ConfigParser()
        self.config = configparser.ConfigParser()
        
        self.load_configs()
    
    def load_configs(self):
        """Load both configuration files."""
        if self.predefined_file.exists():
            try:
                self.predefined.read(self.predefined_file)
                logger.info("Loaded predefined organizational settings")
            except Exception as e:
                logger.error(f"Error loading predefined config: {e}")
                self._create_default_predefined()
        else:
            self._create_default_predefined()
        
        if self.config_file.exists():
            try:
                self.config.read(self.config_file)
                logger.info("Loaded user configuration settings")
            except Exception as e:
                logger.error(f"Error loading config: {e}")
                self._create_default_config()
        else:
            self._create_default_config()
    
    def _create_default_predefined(self):
        """Create default predefined organizational settings."""
        self.predefined['POC_INFO'] = {
            'rnosc': '',
            'command_unit': '',
            'requestor': '',
            'requestor_phone': '',
            'requestor_email': '',
            'local_iam': '',
            'local_iam_phone': '',
            'local_iam_email': '',
            'regional_iam': '',
            'regional_iam_phone': '',
            'regional_iam_email': ''
        }
        
        self.predefined['PROGRAM_INFO'] = {
            'program_name': '',
            'program_manager': '',
            'program_manager_phone': '',
            'program_manager_email': '',
            'centrally_managed_program': 'N/A',
            'testing_patching_guidance': 'N/A'
        }
        
        self.predefined['AA_INFO'] = {
            'ato_approved': 'N/A',
            'ato_date': '',
            'atc_approved': 'N/A',
            'atc_date': '',
            'hbss_installed': 'N/A',
            'epo_server_name': '',
            'epo_server_ip': '',
            'hbss_explanation': ''
        }
        
        self.save_predefined()
    
    def _create_default_config(self):
        """Create default user configuration settings."""
        self.config['TECHNICAL'] = {
            'subnet_mask': '24',
            'use_ip_ranges': 'true',
            'default_output_dir': str(self.script_dir / "output")
        }
        
        self.config['NARRATIVE_TEMPLATES'] = {
            'reason_cannot_complete': 'Patching requires thorough testing and coordination with operational requirements to ensure system availability during critical mission periods.',
            'operational_impact': 'Disconnecting these assets from the MCEN would significantly impact mission operations and disrupt critical business functions.',
            'plan_of_action': '1. Coordinate with system owners to schedule maintenance window\n2. Test patches in development environment\n3. Apply patches during approved maintenance window\n4. Verify system functionality post-patching\n5. Conduct vulnerability scan to confirm remediation',
            'timeline_milestones': 'Week 1: Coordinate maintenance scheduling\nWeek 2-3: Test patches in development environment\nWeek 4: Apply patches during maintenance window\nWeek 5: Post-patch verification and vulnerability scanning',
            'vulnerability_detection_method': 'Continuous vulnerability scanning using Nessus and regular compliance assessments. System monitoring through HBSS and network monitoring tools for signs of compromise.',
            'temporary_mitigations': 'Network segmentation and access controls are in place. Intrusion detection systems are monitoring for suspicious activity. Patch management process ensures timely application of security updates.'
        }
        
        self.config['AI_SETTINGS'] = {
            'ollama_enabled': 'false',
            'ollama_model': '',
            'ai_enhance_reason': 'false',
            'ai_enhance_impact': 'false',
            'ai_enhance_plan': 'false',
            'ai_enhance_timeline': 'false',
            'ai_enhance_mitigations': 'false'
        }
        
        self.config['LAST_USED_FILES'] = {
            'plugins_file': '',
            'template_file': '',
            'scan_file': '',
            'output_directory': str(self.script_dir / "output"),
            'reference_file': ''
        }
        
        self.config['ASSET_DESCRIPTIONS'] = {}
        self.config['SYSTEM_DESCRIPTIONS'] = {}
        
        self.save_config()
    
    def save_predefined(self):
        """Save predefined organizational settings."""
        try:
            with open(self.predefined_file, 'w') as f:
                self.predefined.write(f)
            logger.info("Saved predefined organizational settings")
        except Exception as e:
            logger.error(f"Error saving predefined config: {e}")
    
    def save_config(self):
        """Save user configuration settings."""
        try:
            with open(self.config_file, 'w') as f:
                self.config.write(f)
            logger.info("Saved user configuration settings")
        except Exception as e:
            logger.error(f"Error saving config: {e}")
    
    def get_predefined(self, section, key, fallback=''):
        """Get predefined organizational setting."""
        return self.predefined.get(section, key, fallback=fallback)
    
    def set_predefined(self, section, key, value):
        """Set predefined organizational setting."""
        if not self.predefined.has_section(section):
            self.predefined.add_section(section)
        self.predefined.set(section, key, str(value))
    
    def get_config(self, section, key, fallback=''):
        """Get user configuration setting."""
        return self.config.get(section, key, fallback=fallback)
    
    def set_config(self, section, key, value):
        """Set user configuration setting."""
        if not self.config.has_section(section):
            self.config.add_section(section)
        self.config.set(section, key, str(value))
    
    def get_bool(self, section, key, fallback=False):
        """Get boolean configuration value."""
        return self.config.getboolean(section, key, fallback=fallback)
    
    def set_bool(self, section, key, value):
        """Set boolean configuration value."""
        self.set_config(section, key, str(value).lower())
    
    def get_asset_descriptions(self):
        """Get asset descriptions dictionary."""
        if self.config.has_section('ASSET_DESCRIPTIONS'):
            return dict(self.config['ASSET_DESCRIPTIONS'])
        return {}
    
    def set_asset_descriptions(self, descriptions_dict):
        """Set asset descriptions dictionary."""
        if not self.config.has_section('ASSET_DESCRIPTIONS'):
            self.config.add_section('ASSET_DESCRIPTIONS')
        
        self.config.remove_section('ASSET_DESCRIPTIONS')
        self.config.add_section('ASSET_DESCRIPTIONS')
        
        for ip_range, description in descriptions_dict.items():
            self.config.set('ASSET_DESCRIPTIONS', ip_range, description)

# =================================================
# PDF FORM FILLING
# =================================================
def fill_pdf_form(template_path, output_path, iavm_id, iavm_data, config_manager, system_descriptions_manager):
    """Fill the PDF form with IAVM data by populating form fields in the template PDF."""
    logger.info(f"Filling PDF form for {iavm_id}")
    
    try:
        doc = fitz.open(template_path)
        
        current_date = datetime.now().strftime("%m/%d/%Y")
        
        host_ips = list(iavm_data['hosts'].keys())
        host_info_list = list(iavm_data['hosts'].values())
        
        use_ip_ranges = config_manager.get_bool('TECHNICAL', 'use_ip_ranges', True)
        subnet_mask = int(config_manager.get_config('TECHNICAL', 'subnet_mask', '24'))
        
        if use_ip_ranges:
            ip_ranges = calculate_ip_ranges(host_ips, subnet_mask)
            ip_range_str = "\n".join(ip_ranges) if ip_ranges else "\n".join(host_ips)
        else:
            ip_range_str = "\n".join(host_ips)
        
        hostnames = [info['hostname'] for info in host_info_list]
        system_descriptions = system_descriptions_manager.get_matching_descriptions(hostnames)
        
        asset_descriptions = []
        for host_info in host_info_list:
            desc = get_asset_description(host_info, config_manager)
            asset_descriptions.append(desc)
        
        unique_descriptions = list(set(asset_descriptions))
        individual_systems_description = "\n\n".join(unique_descriptions)
        
        if system_descriptions:
            combined_asset_description = "System Descriptions:\n" + "\n".join(system_descriptions)
            combined_asset_description += "\n\nAffected Systems:\n" + individual_systems_description
        else:
            combined_asset_description = individual_systems_description
        
        workstation_names = []
        os_versions = []
        
        for host_info in host_info_list:
            workstation_names.append(f"{host_info['hostname']} ({host_info['ip']})")
            os_versions.append(host_info['os'])
        
        workstation_names_str = "\n".join(workstation_names)
        os_str = "\n".join(sorted(set(os_versions)))
        
        software_version_str = format_iavm_software_with_plugin_details(iavm_data)

        logger.info(f"IAVM {iavm_id} software versions: {len(iavm_data.get('software_versions', {}))} hosts with version data")
        if iavm_data.get('software_versions'):
            for host_ip, versions in list(iavm_data['software_versions'].items())[:3]:
                logger.info(f"  {host_ip}: {len(versions)} plugins with versions")
        
        context_data = {
            'os_summary': os_str,
            'host_count': len(host_info_list),
            'workstation_summary': workstation_names_str,
            'system_descriptions': system_descriptions,
            'software_summary': software_version_str,
            'ip_summary': ip_range_str
        }
        
        ai_enhanced = False
        
        if config_manager.get_bool('AI_SETTINGS', 'ollama_enabled') and config_manager.get_config('AI_SETTINGS', 'ollama_model'):
            logger.info("AI enhancement enabled, processing narrative fields...")
            
            model_name = config_manager.get_config('AI_SETTINGS', 'ollama_model')
            if ollama_service.set_model(model_name):
                ai_enhancements = {
                    'ai_enhance_reason': 'Reason Cannot Complete',
                    'ai_enhance_impact': 'Operational Impact',
                    'ai_enhance_plan': 'Plan of Action',
                    'ai_enhance_timeline': 'Timeline Milestones',
                    'ai_enhance_mitigations': 'Temporary Mitigations'
                }
                
                for setting_key, field_name in ai_enhancements.items():
                    if config_manager.get_bool('AI_SETTINGS', setting_key):
                        enhanced = ollama_service.enhance_narrative(field_name, iavm_data, context_data)
                        if enhanced:
                            # Update the narrative template with AI enhanced content
                            config_key = setting_key.replace('ai_enhance_', '').replace('reason', 'reason_cannot_complete').replace('impact', 'operational_impact').replace('plan', 'plan_of_action').replace('timeline', 'timeline_milestones').replace('mitigations', 'temporary_mitigations')
                            config_manager.set_config('NARRATIVE_TEMPLATES', config_key, enhanced)
                            ai_enhanced = True
                            logger.info(f"Enhanced '{field_name}' with AI")
            else:
                logger.warning(f"Could not set Ollama model {model_name}")
        
        all_widgets = []
        for page_num in range(doc.page_count):
            page = doc[page_num]
            widgets = list(page.widgets())
            all_widgets.extend(widgets) 

        total_affected = len(iavm_data['hosts'])
        
        # Get OPDIR info including dates
        opdir_info = get_opdir_info(iavm_id, config_manager)
        
        # Prepare Plan of Action with CVEs appended
        plan_of_action_text = config_manager.get_config('NARRATIVE_TEMPLATES', 'plan_of_action')
        if iavm_data['cves']:
            cve_list = '\n'.join(sorted(iavm_data['cves']))
            plan_of_action_text += f"\n\nRelated CVEs:\n{cve_list}"

        form_data = {
            'IAVM': iavm_id.split(":", 1)[-1] if ":" in iavm_id else iavm_id,
            'OpDir': opdir_info['opdir'],
            'DateSubmitted': current_date,
            'DateDue': opdir_info.get('poam_due_date') or opdir_info.get('final_due_date') or '',
            'NumberofSubmission': '1st',
            'RNOSC': config_manager.get_predefined('POC_INFO', 'rnosc'),
            'CommandUnit': config_manager.get_predefined('POC_INFO', 'command_unit'),
            'Requestor': config_manager.get_predefined('POC_INFO', 'requestor'),
            'RequestorPhone': config_manager.get_predefined('POC_INFO', 'requestor_phone'),
            'RequestorEmail': config_manager.get_predefined('POC_INFO', 'requestor_email'),
            'LocalIAM': config_manager.get_predefined('POC_INFO', 'local_iam'),
            'LocalIAMPhone': config_manager.get_predefined('POC_INFO', 'local_iam_phone'),
            'LocalIAMEmail': config_manager.get_predefined('POC_INFO', 'local_iam_email'),
            'RegionalIAM': config_manager.get_predefined('POC_INFO', 'regional_iam'),
            'RegionalIAMPhone': config_manager.get_predefined('POC_INFO', 'regional_iam_phone'),
            'RegionalIAMEmail': config_manager.get_predefined('POC_INFO', 'regional_iam_email'),
            'AdditionalNotes': '** ePO Server Names and IPs. This information did not fit in the box \n NENT10QUANVS551.mcdsus.mcds.usmc.mil   138.156.149.125 \n NENT10QUANVS552.mcdsus.mcds.usmc.mil   138.156.149.124',
            
            'CentrallyManagedProgram': config_manager.get_predefined('PROGRAM_INFO', 'centrally_managed_program'),
            'ProgramName': config_manager.get_predefined('PROGRAM_INFO', 'program_name'),
            'ProgramManager': config_manager.get_predefined('PROGRAM_INFO', 'program_manager'),
            'ProgramManagerPhone': config_manager.get_predefined('PROGRAM_INFO', 'program_manager_phone'),
            'ProgramManagerEmail': config_manager.get_predefined('PROGRAM_INFO', 'program_manager_email'),
            'TestingPatchingGuidance': config_manager.get_predefined('PROGRAM_INFO', 'testing_patching_guidance'),
                        
            'NIPRATODate': config_manager.get_predefined('AA_INFO', 'ato_date'),
            'NIPRATCDate': config_manager.get_predefined('AA_INFO', 'atc_date'),
            
            'HBSS Installed': config_manager.get_predefined('AA_INFO', 'hbss_installed'),
            'ePOServer': config_manager.get_predefined('AA_INFO', 'epo_server_name'),
            'HBSSExplanation': config_manager.get_predefined('AA_INFO', 'epo_server_ip'),
            'AdditionalDetail': config_manager.get_predefined('AA_INFO', 'additional_detail'),
            
            'NIPRAffected': str(total_affected),
            'NIPRPatched': '0',
            'NIPRNotPatched': str(total_affected),
            'SIPRAffected': '0',
            'SIPRPatched': '0',
            'SIPRNotPatched': '0',
            
            'IPAddressRange': ip_range_str,
            'ScanFile1': iavm_data.get('scan_file_ref', 'scan_results.nessus'),
            
            'Workstation': workstation_names_str,
            'OS': os_str,
            'CurrentVersionSW2': software_version_str,
            
            'AssetDescription': combined_asset_description,
            'ReasonNotCompletion': config_manager.get_config('NARRATIVE_TEMPLATES', 'reason_cannot_complete'),
            'OperationalImpact': config_manager.get_config('NARRATIVE_TEMPLATES', 'operational_impact'),
            'PlanOfAction': plan_of_action_text,
            'Timeline': config_manager.get_config('NARRATIVE_TEMPLATES', 'timeline_milestones'),
            'VulnerabilityMethod': config_manager.get_config('NARRATIVE_TEMPLATES', 'vulnerability_detection_method'),
            'TemporaryMitigation': config_manager.get_config('NARRATIVE_TEMPLATES', 'temporary_mitigations'),
           
            'CVEs': '\n'.join(sorted(iavm_data['cves'])) if iavm_data['cves'] else '',
            'Related Plugins': '\n'.join(sorted(iavm_data['plugins']))[:1000] + '\n...' if len('\n'.join(sorted(iavm_data['plugins']))) > 1000 else '\n'.join(sorted(iavm_data['plugins']))
        }
        
        print(f"[DEBUG] PDF has {doc.page_count} pages")
        
        fields_filled = 0
        fields_not_found = []
        
        for page_num in range(doc.page_count):
            page = doc[page_num]
            
            widgets = list(page.widgets())
            print(f"[DEBUG] Page {page_num + 1}: {len(widgets)} form fields")
            
            for widget in widgets:
                field_name = widget.field_name
                print(f"[DEBUG]   Found field: '{field_name}' ({widget.field_type_string})")
                
                if field_name in form_data:
                    widget.field_value = form_data[field_name]
                    widget.update()
                    fields_filled += 1
                    value_preview = str(form_data[field_name])[:50]
                    if len(str(form_data[field_name])) > 50:
                        value_preview += "..."
                    print(f"[FILL] Filled '{field_name}' = '{value_preview}'")
                else:
                    fields_not_found.append(field_name)
                    print(f"[WARNING] Field '{field_name}' not found in form_data")
        
        print(f"[SUMMARY] Fields filled: {fields_filled}/{fields_filled + len(fields_not_found)}")
        
        if fields_not_found:
            print(f"[WARNING] Fields not mapped (first 20):")
            for field in fields_not_found[:20]:
                print(f"  '{field}'")
            
            # Special check for common misspellings or variations
            if 'TemporaryMitigations' in form_data:
                possible_matches = [f for f in fields_not_found if 'mitigation' in f.lower() or 'temporary' in f.lower()]
                if possible_matches:
                    print(f"[INFO] Possible field name matches for TemporaryMitigations:")
                    for match in possible_matches:
                        print(f"    Possible: '{match}'")
        
        # Log if TemporaryMitigations was in form_data but not filled
        if 'TemporaryMitigations' in form_data and 'TemporaryMitigations' in fields_not_found:
            print(f"[WARNING] TemporaryMitigations field not found in PDF!")
            print(f"[INFO] Value to fill: '{form_data['TemporaryMitigations'][:100]}...'")
            print(f"[INFO] Check PDF template - field may have different name")
        
        doc.save(output_path)
        doc.close()
        
        logger.info(f"Successfully created {output_path}")
        return True, ai_enhanced
        
    except Exception as e:
        logger.error(f"Error filling PDF form: {e}")
        return False, False

def debug_pdf_fields(template_path):
    """Debug function to print all field names in a PDF form."""
    try:
        doc = fitz.open(template_path)
        print("="*60)
        print("PDF FORM FIELD ANALYSIS")
        print("="*60)
        
        all_fields = []
        for page_num in range(doc.page_count):
            page = doc[page_num]
            widgets = list(page.widgets())
            print(f"\nPage {page_num + 1}: {len(widgets)} fields")
            print("-" * 40)
            
            for i, widget in enumerate(widgets, 1):
                field_info = f"{i:2d}. '{widget.field_name}' ({widget.field_type_string})"
                if widget.field_value:
                    field_info += f" = '{widget.field_value}'"
                print(field_info)
                all_fields.append(widget.field_name)
        
        print(f"\nTotal fields: {len(all_fields)}")
        print("\nAll field names for copy-paste:")
        print("-" * 40)
        for field in all_fields:
            print(f'"{field}": "",')
        
        doc.close()
        
    except Exception as e:
        print(f"Error analyzing PDF: {e}")

# =================================================
# BASE DIALOG CLASS FOR CONFIGURATION DIALOGS
# =================================================
class BaseConfigDialog:
    """Base class for configuration dialogs to reduce code duplication."""
    
    def __init__(self, parent, config_manager, title, geometry="500x400"):
        self.parent = parent
        self.config_manager = config_manager
        self.result = False
        self.title = title
        self.geometry = geometry
        self.entries = {}
        
    def setup_dialog(self):
        """Setup common dialog properties."""
        self.dialog = tk.Toplevel(self.parent)
        self.dialog.title(self.title)
        self.dialog.geometry(self.geometry)
        self.dialog.resizable(False, True)
        self.dialog.transient(self.parent)
        self.dialog.grab_set()
        
        self.dialog.configure(bg='#2b2b2b')
        
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() - self.dialog.winfo_width()) // 2
        y = (self.dialog.winfo_screenheight() - self.dialog.winfo_height()) // 2
        self.dialog.geometry(f"+{x}+{y}")
        
        return self.dialog
    
    def add_buttons(self, parent_frame):
        """Add standard Save/Cancel buttons - fixed to ensure visibility."""
        button_frame = tk.Frame(parent_frame, bg='#2b2b2b')
        button_frame.pack(side="bottom", fill="x", pady=(10, 0))
        
        tk.Button(button_frame, text="Save", command=self._save,
                bg='#404040', fg='white', font=("Arial", 11, "bold")).pack(side="left", padx=5)
        tk.Button(button_frame, text="Cancel", command=self._cancel,
                bg='#404040', fg='white', font=("Arial", 11)).pack(side="left", padx=5)
    
    def show(self):
        """Show the dialog. Subclasses should override create_content()."""
        self.setup_dialog()
        self.create_content()
        self.dialog.wait_window()
        return self.result
    
    def create_content(self):
        """Override this method in subclasses to create dialog-specific content."""
        raise NotImplementedError("Subclasses must implement create_content()")
    
    def _save(self):
        """Override this method in subclasses for specific save logic."""
        self.result = True
        self.dialog.destroy()
    
    def _cancel(self):
        """Cancel without saving."""
        self.result = False
        self.dialog.destroy()

# =================================================
# IAVM SELECTION DIALOG
# =================================================
class IAVMSelectionDialog:
    """Enhanced GUI dialog for selecting which IAVMs to generate POA&Ms for with filter and status."""
    
    def __init__(self, iavm_list):
        self.iavm_list = iavm_list
        self.selected_iavms = []
        self.checkbox_vars = {}
        self.checkbox_frames = {}
        self.filter_var = None
        self.filtered_out = set()
        self.dialog_closed = False
        
    def show(self, parent=None):
        """Show the enhanced IAVM selection dialog with filter and status."""
        if not self.iavm_list:
            return []
        
        if parent:
            root = tk.Toplevel(parent)
            root.transient(parent)
            root.grab_set()
        else:
            root = tk.Tk()
            
        root.title("Select IAVMs to Generate")
        
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        
        num_items = len(self.iavm_list)
        estimated_height = min(700, max(400, 100 + (num_items * 25)))
        estimated_width = min(1100, max(800, 900))
        
        max_width = int(screen_width * 0.9)
        max_height = int(screen_height * 0.9)
        
        final_width = min(estimated_width, max_width)
        final_height = min(estimated_height, max_height)
        
        root.geometry(f"{final_width}x{final_height}")
        root.resizable(True, True)
        
        root.configure(bg='#2b2b2b')
        
        self.selected_iavms = []
        self.dialog_closed = False
        
        main_frame = tk.Frame(root, bg='#2b2b2b')
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        instructions = tk.Label(main_frame, 
                            text=f"Select which IAVMs you want to generate POA&Ms for ({len(self.iavm_list)} found):", 
                            font=("Arial", 12, "bold"),
                            bg='#2b2b2b', fg='white')
        instructions.pack(pady=(0, 10))
        
        filter_frame = tk.Frame(main_frame, bg='#2b2b2b')
        filter_frame.pack(fill="x", pady=(0, 10))
        
        tk.Label(filter_frame, text="Filter:", bg='#2b2b2b', fg='white', font=("Arial", 10)).pack(side="left")
        self.filter_var = tk.StringVar()
        self.filter_var.trace('w', lambda *args: self.apply_filter())
        filter_entry = tk.Entry(filter_frame, textvariable=self.filter_var, 
                            bg='#404040', fg='white', insertbackground='white', width=30)
        filter_entry.pack(side="left", padx=(5, 10))
        
        self.filter_count_label = tk.Label(filter_frame, text="", bg='#2b2b2b', fg='#cccccc', font=("Arial", 9))
        self.filter_count_label.pack(side="left")
        
        button_top_frame = tk.Frame(main_frame, bg='#2b2b2b')
        button_top_frame.pack(fill="x", pady=(0, 10))
        
        def select_all():
            print("[ACTION] Selecting all visible IAVMs")
            for iavm_id, var in self.checkbox_vars.items():
                if iavm_id not in self.filtered_out:
                    var.set(True)
                    
        def select_none():
            print("[ACTION] Deselecting all visible IAVMs")
            for iavm_id, var in self.checkbox_vars.items():
                if iavm_id not in self.filtered_out:
                    var.set(False)
        
        select_all_btn = tk.Button(button_top_frame, text="Select All Visible", command=select_all,
                                bg='#404040', fg='white', font=("Arial", 10))
        select_all_btn.pack(side="left", padx=(0, 10))
        
        select_none_btn = tk.Button(button_top_frame, text="Deselect All Visible", command=select_none,
                                bg='#404040', fg='white', font=("Arial", 10))
        select_none_btn.pack(side="left")
        
        legend_frame = tk.Frame(main_frame, bg='#2b2b2b')
        legend_frame.pack(fill="x", pady=(0, 5))
        
        tk.Label(legend_frame, text="Legend:", bg='#2b2b2b', fg='white', font=("Arial", 9, "bold")).pack(side="left")
        tk.Label(legend_frame, text="New", bg='#2b2b2b', fg='#90EE90', font=("Arial", 9)).pack(side="left", padx=(10, 5))
        tk.Label(legend_frame, text="Previously Generated", bg='#2b2b2b', fg='#FFB6C1', font=("Arial", 9)).pack(side="left", padx=(5, 0))
        
        checkbox_frame = tk.Frame(main_frame, bg='#2b2b2b')
        checkbox_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        canvas = tk.Canvas(checkbox_frame, bg='#2b2b2b', highlightthickness=0)
        scrollbar = tk.Scrollbar(checkbox_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#2b2b2b')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        generation_history = iavm_tracker.get_generation_history()
        generated_iavms = set()
        for entry in generation_history:
            if entry.get('status') == 'success':
                generated_iavms.add(entry.get('iavm_id', ''))
        
        print(f"[STATUS] Found {len(generated_iavms)} previously generated IAVMs in history")
        
        for i, (iavm_id, iavm_data) in enumerate(self.iavm_list.items()):
            var = tk.BooleanVar(value=True)
            self.checkbox_vars[iavm_id] = var
            
            cb_frame = tk.Frame(scrollable_frame, bg='#2b2b2b')
            cb_frame.pack(fill="x", padx=15, pady=2)
            self.checkbox_frames[iavm_id] = cb_frame
            
            host_count = len(iavm_data['hosts'])
            scan_ref = iavm_data.get('scan_file_ref', 'Unknown scan')
            
            was_generated = iavm_id in generated_iavms
            status_text = " [PREVIOUSLY GENERATED]" if was_generated else " [NEW]"
            checkbox_text = f"{iavm_id} ({host_count} hosts) - {scan_ref}{status_text}"
            
            text_color = '#FFB6C1' if was_generated else '#90EE90'
            
            checkbox = tk.Checkbutton(cb_frame, text=checkbox_text, variable=var, 
                                    font=("Arial", 10), anchor="w",
                                    bg='#2b2b2b', fg=text_color, selectcolor='#404040',
                                    activebackground='#404040', activeforeground=text_color)
            checkbox.pack(fill="x")
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        print(f"[DEBUG] Created {len(self.checkbox_vars)} checkboxes")
        checked_initially = sum(1 for var in self.checkbox_vars.values() if var.get())
        print(f"[DEBUG] Initially checked: {checked_initially}")
        
        sample_iavms = list(self.checkbox_vars.keys())[:3]
        for iavm_id in sample_iavms:
            var_value = self.checkbox_vars[iavm_id].get()
            print(f"[DEBUG] Sample checkbox {iavm_id}: {var_value}")
        
        button_frame = tk.Frame(main_frame, bg='#2b2b2b')
        button_frame.pack(pady=(15, 0))
        
        def on_generate():
            print("[DEBUG] Generate Selected button clicked")
            self.selected_iavms = []
            checked_count = 0
            visible_count = 0
            
            for iavm_id, var in self.checkbox_vars.items():
                is_checked = var.get()
                is_visible = iavm_id not in self.filtered_out
                
                if is_checked:
                    checked_count += 1
                if is_visible:
                    visible_count += 1
                    
                if is_checked and is_visible:
                    self.selected_iavms.append(iavm_id)
            
            print(f"[DEBUG] Total checkboxes: {len(self.checkbox_vars)}")
            print(f"[DEBUG] Checked checkboxes: {checked_count}")
            print(f"[DEBUG] Visible checkboxes: {visible_count}")
            print(f"[DEBUG] Filtered out: {len(self.filtered_out)}")
            print(f"[DEBUG] Selected for generation: {len(self.selected_iavms)}")
            
            if len(self.selected_iavms) <= 5:
                print(f"[ACTION] Selected IAVMs: {self.selected_iavms}")
            else:
                print(f"[ACTION] Selected IAVMs (first 5): {self.selected_iavms[:5]}...")
            
            self.dialog_closed = True
            print("[DEBUG] Marked dialog as closed, about to destroy window")
            
            try:
                root.destroy()
                print("[DEBUG] Window destroyed successfully")
            except Exception as e:
                print(f"[ERROR] Error destroying window: {e}")

        def on_generate_all():
            print("[DEBUG] Generate All button clicked")
            self.selected_iavms = list(self.iavm_list.keys())
            print(f"[ACTION] Generate All selected - {len(self.selected_iavms)} IAVMs")
            self.dialog_closed = True
            try:
                root.destroy()
                print("[DEBUG] Window destroyed (Generate All)")
            except Exception as e:
                print(f"[ERROR] Error destroying window (Generate All): {e}")
        
        def on_cancel():
            print("[DEBUG] Cancel button clicked")
            print("[ACTION] IAVM selection cancelled")
            self.selected_iavms = []
            self.dialog_closed = True
            try:
                root.destroy()
                print("[DEBUG] Window destroyed (Cancel)")
            except Exception as e:
                print(f"[ERROR] Error destroying window (Cancel): {e}")
        
        def on_window_close():
            print("[DEBUG] Window close event")
            self.selected_iavms = []
            self.dialog_closed = True
            try:
                root.destroy()
            except Exception as e:
                print(f"[ERROR] Error in window close handler: {e}")
        
        root.protocol("WM_DELETE_WINDOW", on_window_close)
        
        generate_btn = tk.Button(button_frame, text="Generate Selected", command=on_generate,
                                bg='#404040', fg='white', font=("Arial", 11, "bold"))
        generate_btn.pack(side="left", padx=(0, 10))
        
        generate_all_btn = tk.Button(button_frame, text="Generate All", command=on_generate_all,
                                    bg='#505050', fg='white', font=("Arial", 11, "bold"))
        generate_all_btn.pack(side="left", padx=(0, 15))
        
        cancel_btn = tk.Button(button_frame, text="Cancel", command=on_cancel,
                            bg='#404040', fg='white', font=("Arial", 11))
        cancel_btn.pack(side="left")
        
        def debug_selection():
            checked = sum(1 for var in self.checkbox_vars.values() if var.get())
            visible = len(self.checkbox_vars) - len(self.filtered_out)
            selected = sum(1 for iavm_id, var in self.checkbox_vars.items() 
                        if var.get() and iavm_id not in self.filtered_out)
            print(f"[DEBUG] Current state - Checked: {checked}, Visible: {visible}, Would select: {selected}")
        
        debug_btn = tk.Button(button_frame, text="Debug", command=debug_selection,
                            bg='#404040', fg='white', font=("Arial", 11))
        debug_btn.pack(side="left", padx=(15, 0))
        
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        self.canvas = canvas
        self.scrollable_frame = scrollable_frame
        
        self.update_filter_count()
        
        root.update_idletasks()
        x = (root.winfo_screenwidth() - root.winfo_width()) // 2
        y = (root.winfo_screenheight() - root.winfo_height()) // 2
        root.geometry(f"+{x}+{y}")
        
        print(f"[UI] Showing IAVM selection dialog with {len(self.iavm_list)} IAVMs")
        
        try:
            print("[DEBUG] Using wait_window for modal dialog...")
            root.wait_window()
            print("[DEBUG] wait_window completed")
        except Exception as e:
            print(f"[ERROR] wait_window error: {e}")
        
        result = self.selected_iavms if hasattr(self, 'selected_iavms') else []
        print(f"[DEBUG] Returning result: {len(result)} IAVMs")
        
        return result

    def apply_filter(self):
        """Apply filter to IAVM list."""
        if not hasattr(self, 'checkbox_frames'):
            return
            
        filter_text = self.filter_var.get().lower() if self.filter_var else ""
        visible_count = 0
        self.filtered_out.clear()
        
        for iavm_id, frame in self.checkbox_frames.items():
            should_show = filter_text == "" or filter_text in iavm_id.lower()
            
            if should_show:
                frame.pack(fill="x", padx=15, pady=2)
                visible_count += 1
            else:
                frame.pack_forget()
                self.filtered_out.add(iavm_id)
        
        self.update_filter_count(visible_count)
        
        if hasattr(self, 'canvas'):
            self.scrollable_frame.update_idletasks()
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
    
    def update_filter_count(self, visible_count=None):
        """Update the filter count label."""
        if not hasattr(self, 'filter_count_label'):
            return
            
        if visible_count is None:
            visible_count = len(self.iavm_list)
        
        total_count = len(self.iavm_list)
        if visible_count == total_count:
            self.filter_count_label.config(text=f"Showing all {total_count} items")
        else:
            self.filter_count_label.config(text=f"Showing {visible_count} of {total_count} items")

# =================================================
# CONFIGURATION DIALOG CLASSES
# =================================================
class POCInfoDialog(BaseConfigDialog):
    """Dialog for Point of Contact information."""
    
    def __init__(self, parent, config_manager):
        super().__init__(parent, config_manager, "POC Information", "500x600")
        
    def create_content(self):
        """Create POC-specific content."""
        canvas = tk.Canvas(self.dialog, bg='#2b2b2b', highlightthickness=0)
        scrollbar = tk.Scrollbar(self.dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#2b2b2b')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        row = 0
        
        tk.Label(scrollable_frame, text="Point of Contact Information", 
                font=("Arial", 14, "bold"), bg='#2b2b2b', fg='white').grid(row=row, column=0, columnspan=2, pady=(10, 20))
        row += 1
        
        fields = [
            ("RNOSC:", "rnosc"),
            ("Command/Unit:", "command_unit"),
            ("Requestor:", "requestor"),
            ("Requestor Phone:", "requestor_phone"),
            ("Requestor Email:", "requestor_email"),
            ("Local IAM:", "local_iam"),
            ("Local IAM Phone:", "local_iam_phone"),
            ("Local IAM Email:", "local_iam_email"),
            ("Regional IAM:", "regional_iam"),
            ("Regional IAM Phone:", "regional_iam_phone"),
            ("Regional IAM Email:", "regional_iam_email"),
        ]
        
        for label_text, config_key in fields:
            tk.Label(scrollable_frame, text=label_text, bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", padx=(10, 10), pady=5)
            entry = tk.Entry(scrollable_frame, width=40, bg='#404040', fg='white', insertbackground='white')
            entry.insert(0, self.config_manager.get_predefined('POC_INFO', config_key))
            entry.grid(row=row, column=1, sticky="ew", padx=(0, 10), pady=5)
            self.entries[config_key] = entry
            row += 1
        
        scrollable_frame.columnconfigure(1, weight=1)
        
        self.add_buttons(scrollable_frame)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def _save(self):
        """Save POC information."""
        for key, entry in self.entries.items():
            self.config_manager.set_predefined('POC_INFO', key, entry.get())
        
        self.config_manager.save_predefined()
        logger.info("Saved POC information")
        super()._save()

class ProgramInfoDialog(BaseConfigDialog):
    """Dialog for Program Information."""
    
    def __init__(self, parent, config_manager):
        super().__init__(parent, config_manager, "Program Information", "500x400")
        
    def create_content(self):
        """Create POC-specific content."""
        # Create main frame that will contain everything
        main_frame = tk.Frame(self.dialog, bg='#2b2b2b')
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Create scrollable area
        canvas = tk.Canvas(main_frame, bg='#2b2b2b', highlightthickness=0)
        scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#2b2b2b')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        row = 0
        
        tk.Label(scrollable_frame, text="Point of Contact Information", 
                font=("Arial", 14, "bold"), bg='#2b2b2b', fg='white').grid(row=row, column=0, columnspan=2, pady=(10, 20))
        row += 1
        
        fields = [
            ("RNOSC:", "rnosc"),
            ("Command/Unit:", "command_unit"),
            ("Requestor:", "requestor"),
            ("Requestor Phone:", "requestor_phone"),
            ("Requestor Email:", "requestor_email"),
            ("Local IAM:", "local_iam"),
            ("Local IAM Phone:", "local_iam_phone"),
            ("Local IAM Email:", "local_iam_email"),
            ("Regional IAM:", "regional_iam"),
            ("Regional IAM Phone:", "regional_iam_phone"),
            ("Regional IAM Email:", "regional_iam_email"),
        ]
        
        for label_text, config_key in fields:
            tk.Label(scrollable_frame, text=label_text, bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", padx=(10, 10), pady=5)
            entry = tk.Entry(scrollable_frame, width=40, bg='#404040', fg='white', insertbackground='white')
            entry.insert(0, self.config_manager.get_predefined('POC_INFO', config_key))
            entry.grid(row=row, column=1, sticky="ew", padx=(0, 10), pady=5)
            self.entries[config_key] = entry
            row += 1
        
        scrollable_frame.columnconfigure(1, weight=1)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Add buttons to the main frame (not scrollable frame)
        button_frame = tk.Frame(main_frame, bg='#2b2b2b')
        button_frame.pack(side="bottom", fill="x", pady=(10, 0))
        
        tk.Button(button_frame, text="Save", command=self._save,
                bg='#404040', fg='white', font=("Arial", 11, "bold")).pack(side="left", padx=5)
        tk.Button(button_frame, text="Cancel", command=self._cancel,
                bg='#404040', fg='white', font=("Arial", 11)).pack(side="left", padx=5)
        
    def _save(self):
        """Save program information."""
        for key, widget in self.entries.items():
            self.config_manager.set_predefined('PROGRAM_INFO', key, widget.get())
        
        self.config_manager.save_predefined()
        logger.info("Saved program information")
        super()._save()

class AAInfoDialog(BaseConfigDialog):
    """Dialog for A&A Information."""
    
    def __init__(self, parent, config_manager):
        super().__init__(parent, config_manager, "A&A Information", "500x500")
        
    def create_content(self):
        """Create A&A-specific content."""
        main_frame = tk.Frame(self.dialog, bg='#2b2b2b')
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        row = 0
        
        tk.Label(main_frame, text="Assessment & Authorization Information", 
                font=("Arial", 14, "bold"), bg='#2b2b2b', fg='white').grid(row=row, column=0, columnspan=2, pady=(0, 20))
        row += 1
        
        tk.Label(main_frame, text="ATO Approved (NIPR):", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", pady=5)
        ato_var = tk.StringVar(value=self.config_manager.get_predefined('AA_INFO', 'ato_approved', 'N/A'))
        ato_combo = ttk.Combobox(main_frame, textvariable=ato_var, values=["Yes", "No", "N/A"], width=37)
        ato_combo.grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        self.entries['ato_approved'] = ato_combo
        row += 1
        
        tk.Label(main_frame, text="ATO Date:", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", pady=5)
        ato_date_entry = tk.Entry(main_frame, width=40, bg='#404040', fg='white', insertbackground='white')
        ato_date_entry.insert(0, self.config_manager.get_predefined('AA_INFO', 'ato_date'))
        ato_date_entry.grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        self.entries['ato_date'] = ato_date_entry
        row += 1
        
        tk.Label(main_frame, text="ATC Approved (NIPR):", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", pady=5)
        atc_var = tk.StringVar(value=self.config_manager.get_predefined('AA_INFO', 'atc_approved', 'N/A'))
        atc_combo = ttk.Combobox(main_frame, textvariable=atc_var, values=["Yes", "No", "N/A"], width=37)
        atc_combo.grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        self.entries['atc_approved'] = atc_combo
        row += 1
        
        tk.Label(main_frame, text="ATC Date:", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", pady=5)
        atc_date_entry = tk.Entry(main_frame, width=40, bg='#404040', fg='white', insertbackground='white')
        atc_date_entry.insert(0, self.config_manager.get_predefined('AA_INFO', 'atc_date'))
        atc_date_entry.grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        self.entries['atc_date'] = atc_date_entry
        row += 1
        
        tk.Label(main_frame, text="HBSS Installed:", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", pady=5)
        hbss_var = tk.StringVar(value=self.config_manager.get_predefined('AA_INFO', 'hbss_installed', 'N/A'))
        hbss_combo = ttk.Combobox(main_frame, textvariable=hbss_var, values=["Yes", "No", "N/A"], width=37)
        hbss_combo.grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        self.entries['hbss_installed'] = hbss_combo
        row += 1
        
        tk.Label(main_frame, text="ePO Server Name:", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", pady=5)
        epo_name_entry = tk.Entry(main_frame, width=40, bg='#404040', fg='white', insertbackground='white')
        epo_name_entry.insert(0, self.config_manager.get_predefined('AA_INFO', 'epo_server_name'))
        epo_name_entry.grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        self.entries['epo_server_name'] = epo_name_entry
        row += 1
        
        tk.Label(main_frame, text="ePO Server IP:", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", pady=5)
        epo_ip_entry = tk.Entry(main_frame, width=40, bg='#404040', fg='white', insertbackground='white')
        epo_ip_entry.insert(0, self.config_manager.get_predefined('AA_INFO', 'epo_server_ip'))
        epo_ip_entry.grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        self.entries['epo_server_ip'] = epo_ip_entry
        row += 1
        
        tk.Label(main_frame, text="HBSS Explanation:", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="nw", pady=5)
        hbss_explanation_text = tk.Text(main_frame, height=4, width=40, wrap=tk.WORD,
                                       bg='#404040', fg='white', insertbackground='white')
        hbss_explanation_text.insert(tk.END, self.config_manager.get_predefined('AA_INFO', 'hbss_explanation'))
        hbss_explanation_text.grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        self.entries['hbss_explanation'] = hbss_explanation_text
        row += 1
        
        main_frame.columnconfigure(1, weight=1)
        
        self.add_buttons(main_frame)
    
    def _save(self):
        """Save A&A information."""
        for key, widget in self.entries.items():
            if isinstance(widget, tk.Text):
                value = widget.get(1.0, tk.END).strip()
            else:
                value = widget.get()
            self.config_manager.set_predefined('AA_INFO', key, value)
        
        self.config_manager.save_predefined()
        logger.info("Saved A&A information")
        super()._save()

class NarrativeTemplatesDialog(BaseConfigDialog):
    """Dialog for narrative template configuration."""
    
    def __init__(self, parent, config_manager):
        super().__init__(parent, config_manager, "Narrative Templates", "700x800")
        
    def create_content(self):
        """Create narrative templates content."""
        canvas = tk.Canvas(self.dialog, bg='#2b2b2b', highlightthickness=0)
        scrollbar = tk.Scrollbar(self.dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#2b2b2b')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        row = 0
        
        tk.Label(scrollable_frame, text="Narrative Templates", 
                font=("Arial", 14, "bold"), bg='#2b2b2b', fg='white').grid(row=row, column=0, columnspan=2, pady=(10, 20))
        row += 1
        
        templates = [
            ("Reason Cannot Complete:", "reason_cannot_complete", 4),
            ("Operational Impact:", "operational_impact", 4),
            ("Plan of Action:", "plan_of_action", 6),
            ("Timeline/Milestones:", "timeline_milestones", 6),
            ("Vulnerability Detection Method:", "vulnerability_detection_method", 4),
            ("Temporary Mitigations:", "temporary_mitigations", 4),
        ]
        
        for label_text, config_key, height in templates:
            tk.Label(scrollable_frame, text=label_text, bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="nw", padx=(10, 10), pady=5)
            text_widget = tk.Text(scrollable_frame, height=height, width=60, wrap=tk.WORD,
                                 bg='#404040', fg='white', insertbackground='white')
            text_widget.insert(tk.END, self.config_manager.get_config('NARRATIVE_TEMPLATES', config_key))
            text_widget.grid(row=row, column=1, sticky="ew", padx=(0, 10), pady=5)
            self.entries[config_key] = text_widget
            row += 1
        
        scrollable_frame.columnconfigure(1, weight=1)
        
        # Pack canvas and scrollbar first
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Add buttons to the main dialog (not scrollable frame)
        button_frame = tk.Frame(self.dialog, bg='#2b2b2b')
        button_frame.pack(side="bottom", fill="x", padx=15, pady=15)

        tk.Button(button_frame, text="Save", command=self._save,
                bg='#404040', fg='white', font=("Arial", 11, "bold")).pack(side="left", padx=5)
        tk.Button(button_frame, text="Reset to Defaults", command=self._reset_defaults,
                bg='#404040', fg='white', font=("Arial", 11)).pack(side="left", padx=5)
        tk.Button(button_frame, text="Cancel", command=self._cancel,
                bg='#404040', fg='white', font=("Arial", 11)).pack(side="left", padx=5)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    def _save(self):
        """Save narrative templates."""
        for key, text_widget in self.entries.items():
            content = text_widget.get(1.0, tk.END).strip()
            self.config_manager.set_config('NARRATIVE_TEMPLATES', key, content)
        
        self.config_manager.save_config()
        logger.info("Saved narrative templates")
        super()._save()
    
    def _reset_defaults(self):
        """Reset to default templates."""
        defaults = {
            'reason_cannot_complete': 'Patching requires thorough testing and coordination with operational requirements to ensure system availability during critical mission periods.',
            'operational_impact': 'Disconnecting these assets from the MCEN would significantly impact mission operations and disrupt critical business functions.',
            'plan_of_action': '1. Coordinate with system owners to schedule maintenance window\n2. Test patches in development environment\n3. Apply patches during approved maintenance window\n4. Verify system functionality post-patching\n5. Conduct vulnerability scan to confirm remediation',
            'timeline_milestones': 'Week 1: Coordinate maintenance scheduling\nWeek 2-3: Test patches in development environment\nWeek 4: Apply patches during maintenance window\nWeek 5: Post-patch verification and vulnerability scanning',
            'vulnerability_detection_method': 'Continuous vulnerability scanning using Nessus and regular compliance assessments. System monitoring through HBSS and network monitoring tools for signs of compromise.',
            'temporary_mitigations': 'Network segmentation and access controls are in place. Intrusion detection systems are monitoring for suspicious activity. Patch management process ensures timely application of security updates.'
        }
        
        for key, default_text in defaults.items():
            if key in self.entries:
                self.entries[key].delete(1.0, tk.END)
                self.entries[key].insert(tk.END, default_text)

class SystemDescriptionsDialog(BaseConfigDialog):
    """Dialog for configuring system descriptions based on hostname patterns."""
    
    def __init__(self, parent, config_manager):
        super().__init__(parent, config_manager, "System Descriptions", "800x600")
        self.system_manager = SystemDescriptionsManager(config_manager)
        
    def create_content(self):
        """Create system descriptions content."""
        main_frame = tk.Frame(self.dialog, bg='#2b2b2b')
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        tk.Label(main_frame, text="System Descriptions Configuration", 
                font=("Arial", 14, "bold"), bg='#2b2b2b', fg='white').pack(pady=(0, 20))
        
        instructions = tk.Label(main_frame, 
                               text="Define system descriptions based on hostname patterns.\n" +
                                    "Use * for wildcards (e.g., WEB* matches WEB01, WEB02, etc.)\n" +
                                    "These descriptions will be included in POA&M asset descriptions and AI context.",
                               font=("Arial", 10), bg='#2b2b2b', fg='#cccccc')
        instructions.pack(pady=(0, 20))
        
        canvas = tk.Canvas(main_frame, bg='#2b2b2b', highlightthickness=0)
        scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#2b2b2b')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        header_frame = tk.Frame(scrollable_frame, bg='#2b2b2b')
        header_frame.pack(fill="x", pady=(0, 10))
        
        tk.Label(header_frame, text="Hostname Pattern", font=("Arial", 11, "bold"),
                bg='#2b2b2b', fg='white', width=20).grid(row=0, column=0, padx=(0, 20), sticky="w")
        tk.Label(header_frame, text="System Description", font=("Arial", 11, "bold"),
                bg='#2b2b2b', fg='white').grid(row=0, column=1, sticky="w")
        
        existing_descriptions = self.system_manager.get_system_descriptions()
        
        self.entry_rows = []
        
        def add_entry_row(pattern="", description=""):
            row_frame = tk.Frame(scrollable_frame, bg='#2b2b2b')
            row_frame.pack(fill="x", pady=2)
            
            pattern_entry = tk.Entry(row_frame, width=25, bg='#404040', fg='white',
                                   insertbackground='white')
            pattern_entry.insert(0, pattern)
            pattern_entry.grid(row=0, column=0, padx=(0, 10), sticky="ew")
            
            desc_entry = tk.Entry(row_frame, width=60, bg='#404040', fg='white',
                                insertbackground='white')
            desc_entry.insert(0, description)
            desc_entry.grid(row=0, column=1, padx=(0, 10), sticky="ew")
            
            remove_btn = tk.Button(row_frame, text="Remove", 
                                 command=lambda: self.remove_row(row_frame),
                                 bg='#804040', fg='white', font=("Arial", 9))
            remove_btn.grid(row=0, column=2)
            
            row_frame.columnconfigure(1, weight=1)
            
            self.entry_rows.append((row_frame, pattern_entry, desc_entry))
            
            return pattern_entry, desc_entry
        
        for pattern, description in existing_descriptions.items():
            add_entry_row(pattern, description)
        
        for _ in range(5):
            add_entry_row()
        
        add_button_frame = tk.Frame(scrollable_frame, bg='#2b2b2b')
        add_button_frame.pack(fill="x", pady=10)
        
        tk.Button(add_button_frame, text="Add Row", 
                 command=lambda: add_entry_row(),
                 bg='#404040', fg='white', font=("Arial", 10)).pack(side="left")
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Replace the self.add_buttons(main_frame) call with:
        # Pack canvas and scrollbar first
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Add buttons to main_frame (not scrollable)
        button_frame = tk.Frame(main_frame, bg='#2b2b2b')
        button_frame.pack(side="bottom", fill="x", pady=(10, 0))

        tk.Button(button_frame, text="Save", command=self._save,
                bg='#404040', fg='white', font=("Arial", 11, "bold")).pack(side="left", padx=5)
        tk.Button(button_frame, text="Cancel", command=self._cancel,
                bg='#404040', fg='white', font=("Arial", 11)).pack(side="left", padx=5)
        
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    def remove_row(self, row_frame):
        """Remove an entry row."""
        self.entry_rows = [(rf, pe, de) for rf, pe, de in self.entry_rows if rf != row_frame]
        row_frame.destroy()
    
    def _save(self):
        """Save system descriptions."""
        descriptions = {}
        
        for row_frame, pattern_entry, desc_entry in self.entry_rows:
            pattern = pattern_entry.get().strip()
            description = desc_entry.get().strip()
            
            if pattern and description:
                descriptions[pattern] = description
        
        self.system_manager.set_system_descriptions(descriptions)
        self.config_manager.save_config()
        
        logger.info(f"Saved {len(descriptions)} system descriptions")
        super()._save()

class AISettingsDialog(BaseConfigDialog):
    """Enhanced dialog for AI enhancement settings with remote server support."""
    
    def __init__(self, parent, config_manager):
        super().__init__(parent, config_manager, "AI Enhancement Settings", "700x600")
        
    def show(self):
        """Show the AI settings dialog with connectivity testing."""
        if not OLLAMA_AVAILABLE:
            messagebox.showinfo("AI Enhancement", 
                            "Requests library not available.\n\nInstall with: pip install requests\n\nThen restart the application.")
            return False
        
        return super().show()
        
    def create_content(self):
        """Create AI settings content."""
        main_frame = tk.Frame(self.dialog, bg='#2b2b2b')
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        row = 0
        
        tk.Label(main_frame, text="AI Enhancement Settings", 
                font=("Arial", 14, "bold"), bg='#2b2b2b', fg='white').grid(row=row, column=0, columnspan=3, pady=(0, 20))
        row += 1
        
        tk.Label(main_frame, text="Server Status:", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", pady=5)
        
        connected = ollama_service.test_connectivity()
        status_text = f"Connected to {ollama_service.server_url}" if connected else f"Disconnected from {ollama_service.server_url}"
        status_color = 'green' if connected else 'red'
        
        self.status_label = tk.Label(main_frame, text=status_text, bg='#2b2b2b', fg=status_color)
        self.status_label.grid(row=row, column=1, sticky="w", pady=5)
        
        tk.Button(main_frame, text="Test Connection", command=self.test_connection,
                 bg='#404040', fg='white').grid(row=row, column=2, pady=5)
        row += 1
        
        tk.Label(main_frame, text="Enable AI Enhancement:", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", pady=5)
        ollama_enabled_var = tk.BooleanVar(value=self.config_manager.get_bool('AI_SETTINGS', 'ollama_enabled'))
        ollama_enabled_check = tk.Checkbutton(main_frame, variable=ollama_enabled_var,
                                            bg='#2b2b2b', fg='white', selectcolor='#404040')
        ollama_enabled_check.grid(row=row, column=1, sticky="w", pady=5)
        self.entries['ollama_enabled'] = ollama_enabled_var
        row += 1
        
        tk.Label(main_frame, text="Ollama Model:", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", pady=5)
        
        model_frame = tk.Frame(main_frame, bg='#2b2b2b')
        model_frame.grid(row=row, column=1, columnspan=2, sticky="ew", pady=5)
        
        available_models = ollama_service.models if connected else ["No models available - check connection"]
        
        ollama_model_var = tk.StringVar(value=self.config_manager.get_config('AI_SETTINGS', 'ollama_model'))
        
        style = ttk.Style()
        style.configure('Dark.TCombobox', fieldbackground='#404040', foreground='white')
        
        self.ollama_model_combo = ttk.Combobox(model_frame, textvariable=ollama_model_var, 
                                         values=available_models, width=25, style='Dark.TCombobox')
        self.ollama_model_combo.pack(side="left", padx=(0, 5))
        
        tk.Button(model_frame, text="Refresh Models", command=self.refresh_models,
                 bg='#404040', fg='white').pack(side="left")
        self.entries['ollama_model'] = self.ollama_model_combo
        row += 1
        
        model_count = len(available_models) if connected and available_models[0] != "No models available - check connection" else 0
        self.model_count_label = tk.Label(main_frame, text=f"Available models: {model_count}", 
                                         bg='#2b2b2b', fg='#cccccc', font=("Arial", 9))
        self.model_count_label.grid(row=row, column=1, sticky="w", pady=2)
        row += 1
        
        tk.Label(main_frame, text="AI Enhance Fields:", 
                font=("Arial", 12, "bold"), bg='#2b2b2b', fg='white').grid(row=row, column=0, columnspan=3, sticky="w", pady=(20, 10))
        row += 1
        
        ai_fields = [
            ("ai_enhance_reason", "Reason Cannot Complete"),
            ("ai_enhance_impact", "Operational Impact"),
            ("ai_enhance_plan", "Plan of Action"),
            ("ai_enhance_timeline", "Timeline/Milestones"),
            ("ai_enhance_mitigations", "Temporary Mitigations")
        ]
        
        for config_key, display_name in ai_fields:
            tk.Label(main_frame, text=f"  {display_name}:", bg='#2b2b2b', fg='white').grid(row=row, column=0, sticky="w", padx=(20, 10), pady=2)
            ai_var = tk.BooleanVar(value=self.config_manager.get_bool('AI_SETTINGS', config_key))
            ai_check = tk.Checkbutton(main_frame, variable=ai_var,
                                    bg='#2b2b2b', fg='white', selectcolor='#404040')
            ai_check.grid(row=row, column=1, sticky="w", pady=2)
            self.entries[config_key] = ai_var
            row += 1
        
        main_frame.columnconfigure(1, weight=1)
        
        self.add_buttons(main_frame)
    
    def test_connection(self):
        """Test connection to Ollama server."""
        connected = ollama_service.test_connectivity()
        
        if connected:
            status_text = f"Connected to {ollama_service.server_url}"
            status_color = 'green'
            messagebox.showinfo("Connection Test", f"Successfully connected to Ollama server!\n\nServer: {ollama_service.server_url}\nModels: {len(ollama_service.models)}")
            self.refresh_models()
        else:
            status_text = f"Disconnected from {ollama_service.server_url}"
            status_color = 'red'
            messagebox.showerror("Connection Test", f"Failed to connect to Ollama server.\n\nServer: {ollama_service.server_url}\n\nCheck that:\n- Server is running\n- Network connectivity is available\n- Firewall allows connection")
        
        self.status_label.config(text=status_text, fg=status_color)
    
    def refresh_models(self):
        """Refresh the list of available models."""
        try:
            if ollama_service.connected:
                models = ollama_service.refresh_models()
                if models:
                    self.ollama_model_combo['values'] = models
                    self.model_count_label.config(text=f"Available models: {len(models)}")
                    messagebox.showinfo("Models Refreshed", f"Found {len(models)} models:\n" + "\n".join(models[:5]) + ("..." if len(models) > 5 else ""))
                    logger.info(f"Refreshed Ollama models: {models}")
                else:
                    self.ollama_model_combo['values'] = ["No models available"]
                    self.model_count_label.config(text="Available models: 0")
                    messagebox.showwarning("No Models", "No Ollama models found on server.")
                    logger.warning("No Ollama models found during refresh")
            else:
                messagebox.showerror("Connection Error", "Not connected to Ollama server. Test connection first.")
        except Exception as e:
            messagebox.showerror("Error", f"Error refreshing models: {e}")
            logger.error(f"Error refreshing Ollama models: {e}")
    
    def _save(self):
        """Save AI settings."""
        for key, widget in self.entries.items():
            if isinstance(widget, tk.BooleanVar):
                self.config_manager.set_bool('AI_SETTINGS', key, widget.get())
            else:
                self.config_manager.set_config('AI_SETTINGS', key, widget.get())
        
        self.config_manager.save_config()
        logger.info("Saved AI enhancement settings")
        super()._save()

class GenerationHistoryDialog:
    """Dialog to view IAVM generation history."""
    
    def __init__(self, parent):
        self.parent = parent
        
    def show(self):
        """Show the generation history dialog."""
        dialog = tk.Toplevel(self.parent)
        dialog.title("IAVM Generation History")
        dialog.geometry("1000x600")
        dialog.resizable(True, True)
        dialog.transient(self.parent)
        dialog.grab_set()
        
        dialog.configure(bg='#2b2b2b')
        
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
        y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        main_frame = tk.Frame(dialog, bg='#2b2b2b')
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        tk.Label(main_frame, text="IAVM Generation History", 
                font=("Arial", 14, "bold"), bg='#2b2b2b', fg='white').pack(pady=(0, 20))
        
        columns = ('Timestamp', 'IAVM ID', 'Hosts', 'Scan File', 'Status', 'AI Enhanced')
        tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=20)
        
        tree.heading('Timestamp', text='Timestamp')
        tree.heading('IAVM ID', text='IAVM ID')
        tree.heading('Hosts', text='Hosts Affected')
        tree.heading('Scan File', text='Scan File')
        tree.heading('Status', text='Status')
        tree.heading('AI Enhanced', text='AI Enhanced')
        
        tree.column('Timestamp', width=150)
        tree.column('IAVM ID', width=120)
        tree.column('Hosts', width=80)
        tree.column('Scan File', width=200)
        tree.column('Status', width=100)
        tree.column('AI Enhanced', width=100)
        
        scrollbar_tree = ttk.Scrollbar(main_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar_tree.set)
        
        history = iavm_tracker.get_generation_history()
        
        logger.info(f"Loading {len(history)} generation history entries")
        
        for entry in reversed(history):
            tree.insert('', 'end', values=(
                entry.get('timestamp', ''),
                entry.get('iavm_id', ''),
                entry.get('hosts_affected', ''),
                entry.get('scan_file', ''),
                entry.get('status', ''),
                'Yes' if entry.get('ai_enhanced', '').lower() == 'true' else 'No'
            ))
        
        tree.pack(side="left", fill="both", expand=True)
        scrollbar_tree.pack(side="right", fill="y")
        
        tk.Button(main_frame, text="Close", command=dialog.destroy,
                 bg='#404040', fg='white', font=("Arial", 11)).pack(pady=(20, 0))
        
        dialog.wait_window()

# =================================================
# MAIN APPLICATION CLASS
# =================================================
class MainApplication:
    """Enhanced main application window with all requested features."""
    
    def __init__(self):
        """Enhanced main application constructor with console logging."""
        print("[STARTUP] Initializing Enhanced POA&M Generator v2.0...")
        
        self.root = tk.Tk()
        self.root.title("POA&M Generator - Enhanced Edition v2.0")
        self.root.geometry("900x750")
        self.root.resizable(True, True)
        
        print("[UI] Applying dark theme...")
        self.setup_dark_theme()
        
        print("[CONFIG] Initializing configuration manager...")
        self.config_manager = ConfigManager()
        self.system_descriptions_manager = SystemDescriptionsManager(self.config_manager)
        
        self.plugins_file = tk.StringVar()
        self.scan_file = tk.StringVar()
        self.template_file = tk.StringVar()
        self.reference_file = tk.StringVar()
        self.output_dir = tk.StringVar(value=self.config_manager.get_config('TECHNICAL', 'default_output_dir'))
        self.filtered_iavms = None  # Will hold set of IAVMs to filter by, or None for all
                
        print("[UI] Setting up user interface...")
        self.setup_ui()
        
        print("[STARTUP] Loading last used file paths...")
        self.load_last_used_files()
        
        print("[STARTUP] Checking for plugins files...")
        self.check_plugins_file()
        
        log_startup("Enhanced POA&M Generator v2.0 started")
        
    def setup_dark_theme(self):
        """Setup modern dark theme."""
        self.root.configure(bg='#2b2b2b')
        
        style = ttk.Style()
        style.theme_use('clam')
        
        bg_color = '#2b2b2b'
        fg_color = '#ffffff'
        
        style.configure('TFrame', background=bg_color)
        style.configure('TLabel', background=bg_color, foreground=fg_color)
        style.configure('TButton', background='#404040', foreground=fg_color)
        style.map('TButton', background=[('active', '#505050')])
        style.configure('TEntry', fieldbackground='#404040', foreground=fg_color)
        style.configure('TCombobox', fieldbackground='#404040', foreground=fg_color)
        style.configure('TLabelframe', background=bg_color, foreground=fg_color)
        style.configure('TLabelframe.Label', background=bg_color, foreground=fg_color)
        
    def setup_ui(self):
        """Setup the enhanced UI components."""
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        title_label = ttk.Label(main_frame, text="Opdir POA&M Generator - Enhanced Edition v2.0", 
                               font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 20))
        
        config_frame = ttk.LabelFrame(main_frame, text="Configuration", padding=10)
        config_frame.pack(fill="x", pady=(0, 10))
        
        config_buttons_frame = ttk.Frame(config_frame)
        config_buttons_frame.pack(fill="x")
        
        ttk.Button(config_buttons_frame, text="POC Information", 
                  command=self.open_poc_dialog).pack(side="left", padx=5)
        ttk.Button(config_buttons_frame, text="Program Information", 
                  command=self.open_program_dialog).pack(side="left", padx=5)
        ttk.Button(config_buttons_frame, text="A&A Information", 
                  command=self.open_aa_dialog).pack(side="left", padx=5)
        ttk.Button(config_buttons_frame, text="Narrative Templates", 
                  command=self.open_narrative_dialog).pack(side="left", padx=5)
        ttk.Button(config_buttons_frame, text="System Descriptions", 
                  command=self.open_system_descriptions_dialog).pack(side="left", padx=5)
        ttk.Button(config_buttons_frame, text="AI Settings", 
                  command=self.open_ai_dialog).pack(side="left", padx=5)
        ttk.Button(config_buttons_frame, text="Generation History", 
                  command=self.show_generation_history).pack(side="left", padx=5)
        
        # File Selection - Condensed to 3 rows with 50/50 split
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding=10)
        file_frame.pack(fill="x", pady=(0, 10))
        
        # Configure grid columns for 50/50 split
        file_frame.columnconfigure(0, weight=1)
        file_frame.columnconfigure(1, weight=1)
        
        # Row 1: Plugins (left 50%) and Template (right 50%)
        # Plugins section (left)
        plugins_frame = ttk.Frame(file_frame)
        plugins_frame.grid(row=0, column=0, sticky="ew", padx=(0, 5), pady=2)
        ttk.Label(plugins_frame, text="Plugins:").pack(side="left", padx=(0, 5))
        self.plugins_status = ttk.Label(plugins_frame, text="Not loaded", foreground="orange")
        self.plugins_status.pack(side="left", padx=(0, 5), fill="x", expand=True)
        ttk.Button(plugins_frame, text="Load", command=self.load_plugins_file, width=8).pack(side="right")
        
        # Template section (right)
        template_frame = ttk.Frame(file_frame)
        template_frame.grid(row=0, column=1, sticky="ew", padx=(5, 0), pady=2)
        ttk.Label(template_frame, text="Template:").pack(side="left", padx=(0, 5))
        self.template_status = ttk.Label(template_frame, text="Not selected", foreground="orange")
        self.template_status.pack(side="left", padx=(0, 5), fill="x", expand=True)
        ttk.Button(template_frame, text="Select", command=self.select_template_file, width=8).pack(side="right")
        
        # Row 2: Scan (left 50%) and Reference (right 50%)
        # Scan section (left)
        scan_frame = ttk.Frame(file_frame)
        scan_frame.grid(row=1, column=0, sticky="ew", padx=(0, 5), pady=2)
        ttk.Label(scan_frame, text="Scan File:").pack(side="left", padx=(0, 5))
        self.scan_status = ttk.Label(scan_frame, text="Not selected", foreground="orange")
        self.scan_status.pack(side="left", padx=(0, 5), fill="x", expand=True)
        ttk.Button(scan_frame, text="Select", command=self.select_scan_file, width=8).pack(side="right")
        
        # Reference section (right)
        reference_frame = ttk.Frame(file_frame)
        reference_frame.grid(row=1, column=1, sticky="ew", padx=(5, 0), pady=2)
        ttk.Label(reference_frame, text="Reference:").pack(side="left", padx=(0, 5))
        self.reference_status = ttk.Label(reference_frame, text="Not loaded (optional)", foreground="gray")
        self.reference_status.pack(side="left", padx=(0, 5), fill="x", expand=True)
        ttk.Button(reference_frame, text="Select", command=self.select_reference_file, width=8).pack(side="right")
        
        # Row 3: Output Directory (full width)
        output_frame = ttk.Frame(file_frame)
        output_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=2)
        ttk.Label(output_frame, text="Output Dir:").pack(side="left", padx=(0, 5))
        self.output_entry = ttk.Entry(output_frame, textvariable=self.output_dir)
        self.output_entry.pack(side="left", padx=(0, 5), fill="x", expand=True)
        ttk.Button(output_frame, text="Browse", command=self.select_output_dir, width=8).pack(side="right")
        
        # Run Options - Condensed to 1 row
        options_frame = ttk.LabelFrame(main_frame, text="Run Options", padding=10)
        options_frame.pack(fill="x", pady=(0, 10))
        
        options_row = ttk.Frame(options_frame)
        options_row.pack(fill="x")
        
        ttk.Label(options_row, text="IP Format:").pack(side="left")
        self.ip_format_var = tk.StringVar(value="IP Ranges" if self.config_manager.get_bool('TECHNICAL', 'use_ip_ranges', True) else "Individual IPs")
        ip_combo = ttk.Combobox(options_row, textvariable=self.ip_format_var, 
                               values=["IP Ranges", "Individual IPs"], width=15)
        ip_combo.pack(side="left", padx=(5, 20))
        
        ttk.Label(options_row, text="Subnet:").pack(side="left")
        self.subnet_var = tk.StringVar(value=self.config_manager.get_config('TECHNICAL', 'subnet_mask', '24'))
        subnet_entry = ttk.Entry(options_row, textvariable=self.subnet_var, width=8)
        subnet_entry.pack(side="left", padx=(5, 20))

        ttk.Label(options_row, text="OPDIR:").pack(side="left")
        self.opdir_var = tk.StringVar(value=self.config_manager.get_config('TECHNICAL', 'opdir_release', '0676-25'))
        opdir_entry = ttk.Entry(options_row, textvariable=self.opdir_var, width=12)
        opdir_entry.pack(side="left", padx=(5, 0))
        
        # IAVM Filter Frame
        filter_frame = ttk.LabelFrame(main_frame, text="IAVM Filter (Optional)", padding=10)
        filter_frame.pack(fill="x", pady=(0, 10))
        
        filter_info = ttk.Label(filter_frame, 
                               text="Paste IAVM numbers to generate only specific POA&Ms. Supports multiple formats, removes duplicates.",
                               foreground="gray")
        filter_info.pack(anchor="w", pady=(0, 5))
        
        filter_input_frame = ttk.Frame(filter_frame)
        filter_input_frame.pack(fill="both", expand=True)
        
        ttk.Label(filter_input_frame, text="Paste IAVMs:").pack(anchor="w")
        
        # Text area with scrollbar
        text_frame = ttk.Frame(filter_input_frame)
        text_frame.pack(fill="both", expand=True, pady=5)
        
        self.iavm_filter_text = tk.Text(text_frame, height=4, width=70, wrap=tk.WORD)
        self.iavm_filter_text.pack(side="left", fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(text_frame, command=self.iavm_filter_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.iavm_filter_text.config(yscrollcommand=scrollbar.set)
        
        filter_button_frame = ttk.Frame(filter_frame)
        filter_button_frame.pack(fill="x", pady=(5, 0))
        
        ttk.Button(filter_button_frame, text="Apply Filter", 
                  command=self.apply_iavm_filter).pack(side="left", padx=(0, 5))
        ttk.Button(filter_button_frame, text="Clear Filter", 
                  command=self.clear_iavm_filter).pack(side="left", padx=5)
        
        self.filter_status = ttk.Label(filter_button_frame, text="No filter active", foreground="gray")
        self.filter_status.pack(side="left", padx=(10, 0))
        
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding=10)
        progress_frame.pack(fill="x", pady=(0, 10))
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress_bar.pack(fill="x", pady=(0, 5))
        
        self.status_label = ttk.Label(progress_frame, text="Ready to generate POA&Ms")
        self.status_label.pack()
        
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=10)
        
        self.generate_button = ttk.Button(button_frame, text="Generate POA&Ms", 
                                         command=self.generate_poams)
        self.generate_button.pack(side="right", padx=5)
        
        ttk.Button(button_frame, text="Help", command=self.show_help).pack(side="left", padx=5)
        ttk.Button(button_frame, text="About", command=self.show_about).pack(side="left", padx=5)
    
    def load_last_used_files(self):
        """Load and populate last used file paths from config."""
        # Load plugins file
        plugins_path = self.config_manager.get_config('LAST_USED_FILES', 'plugins_file')
        if plugins_path and os.path.exists(plugins_path):
            self.plugins_file.set(plugins_path)
            self.plugins_status.config(text=f"Loaded: {Path(plugins_path).name}", foreground="green")
            print(f"[STARTUP] Restored plugins file: {Path(plugins_path).name}")
        
        # Load template file
        template_path = self.config_manager.get_config('LAST_USED_FILES', 'template_file')
        if template_path and os.path.exists(template_path):
            self.template_file.set(template_path)
            self.template_status.config(text=f"Selected: {Path(template_path).name}", foreground="green")
            print(f"[STARTUP] Restored template file: {Path(template_path).name}")
        
        # Load scan file
        scan_path = self.config_manager.get_config('LAST_USED_FILES', 'scan_file')
        if scan_path and os.path.exists(scan_path):
            self.scan_file.set(scan_path)
            self.scan_status.config(text=f"Selected: {Path(scan_path).name}", foreground="green")
            print(f"[STARTUP] Restored scan file: {Path(scan_path).name}")
        
        # Load reference file
        reference_path = self.config_manager.get_config('LAST_USED_FILES', 'reference_file')
        if reference_path and os.path.exists(reference_path):
            # Try to load the reference file
            if opdir_lookup_manager.load_reference_file(reference_path):
                self.reference_file.set(reference_path)
                stats = opdir_lookup_manager.get_stats()
                self.reference_status.config(
                    text=f"Loaded: {Path(reference_path).name} ({stats['mapping_count']} mappings)",
                    foreground="green"
                )
                print(f"[STARTUP] Restored reference file: {Path(reference_path).name}")
                print(f"[INFO] OPDIR mappings loaded: {stats['mapping_count']}")
        
        # Output directory is already loaded via tk.StringVar initialization
        output_path = self.output_dir.get()
        if output_path:
            print(f"[STARTUP] Using output directory: {output_path}")
    
    def check_plugins_file(self):
        """Check for plugins file on startup with console logging."""
        print("[ACTION] Checking for plugins files...")
        script_dir = Path(__file__).parent
        
        sc_plugins_files = list(script_dir.glob("CM-*-sc-plugins.tar.gz"))
        if sc_plugins_files:
            latest_file = max(sc_plugins_files, key=lambda f: f.stat().st_mtime)
            self.plugins_file.set(str(latest_file))
            self.plugins_status.config(text=f"Auto-detected: {latest_file.name}", foreground="green")
            print(f"[SUCCESS] Auto-detected plugins archive: {latest_file.name}")
            return
        
        xml_files = list(script_dir.glob("plugins*.xml"))
        if xml_files:
            latest_file = max(xml_files, key=lambda f: f.stat().st_mtime)
            self.plugins_file.set(str(latest_file))
            self.plugins_status.config(text=f"Auto-detected: {latest_file.name}", foreground="green")
            print(f"[SUCCESS] Auto-detected plugins XML: {latest_file.name}")
            return
        
        self.plugins_status.config(text="Not found - click Load Plugins", foreground="orange")
        print("[WARNING] No plugins files found - user must select manually")
    
    def open_poc_dialog(self):
        """Open POC information dialog with console logging."""
        print("[ACTION] Opening POC information dialog...")
        dialog = POCInfoDialog(self.root, self.config_manager)
        result = dialog.show()
        if result:
            print("[SUCCESS] POC information updated")

    def open_program_dialog(self):
        """Open program information dialog with console logging."""
        print("[ACTION] Opening program information dialog...")
        dialog = ProgramInfoDialog(self.root, self.config_manager)
        result = dialog.show()
        if result:
            print("[SUCCESS] Program information updated")

    def open_aa_dialog(self):
        """Open A&A information dialog with console logging."""
        print("[ACTION] Opening A&A information dialog...")
        dialog = AAInfoDialog(self.root, self.config_manager)
        result = dialog.show()
        if result:
            print("[SUCCESS] A&A information updated")

    def open_narrative_dialog(self):
        """Open narrative templates dialog with console logging."""
        print("[ACTION] Opening narrative templates dialog...")
        dialog = NarrativeTemplatesDialog(self.root, self.config_manager)
        result = dialog.show()
        if result:
            print("[SUCCESS] Narrative templates updated")

    def open_system_descriptions_dialog(self):
        """Open system descriptions dialog with console logging."""
        print("[ACTION] Opening system descriptions dialog...")
        dialog = SystemDescriptionsDialog(self.root, self.config_manager)
        result = dialog.show()
        if result:
            print("[SUCCESS] System descriptions updated")

    def open_ai_dialog(self):
        """Open AI settings dialog with console logging."""
        print("[ACTION] Opening AI settings dialog...")
        dialog = AISettingsDialog(self.root, self.config_manager)
        result = dialog.show()
        if result:
            print("[SUCCESS] AI settings updated")

    def show_generation_history(self):
        """Show generation history dialog with console logging."""
        print("[ACTION] Opening generation history dialog...")
        dialog = GenerationHistoryDialog(self.root)
        dialog.show()
    
    def load_plugins_file(self):
        """Load plugins database file with console logging."""
        print("[ACTION] User selecting plugins file...")
        
        # Get initial directory from last used path
        initial_dir = None
        last_path = self.config_manager.get_config('LAST_USED_FILES', 'plugins_file')
        if last_path and os.path.exists(os.path.dirname(last_path)):
            initial_dir = os.path.dirname(last_path)
        
        file_path = filedialog.askopenfilename(
            title="Select Nessus Plugins File",
            initialdir=initial_dir,
            filetypes=[
                ("XML files", "*.xml"),
                ("Archive files", "*.tar.gz *.zip"),
                ("JSON files", "*.json"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.plugins_file.set(file_path)
            self.plugins_status.config(text=f"Selected: {Path(file_path).name}", foreground="green")
            print(f"[FILE] Selected plugins file: {Path(file_path).name}")
            # Save to config
            self.config_manager.set_config('LAST_USED_FILES', 'plugins_file', file_path)
            self.config_manager.save_config()
    
    def select_template_file(self):
        """Select POA&M template file with console logging."""
        print("[ACTION] User selecting template file...")
        if not PDF_AVAILABLE:
            print("[ERROR] PDF libraries not available")
            messagebox.showerror("PDF Libraries Missing", 
                            "Required PDF libraries not installed.\n\nInstall with:\npip install reportlab PyMuPDF")
            return
        
        # Get initial directory from last used path
        initial_dir = None
        last_path = self.config_manager.get_config('LAST_USED_FILES', 'template_file')
        if last_path and os.path.exists(os.path.dirname(last_path)):
            initial_dir = os.path.dirname(last_path)
        
        file_path = filedialog.askopenfilename(
            title="Select POA&M Template PDF",
            initialdir=initial_dir,
            filetypes=[("PDF files", "*.pdf")]
        )
        
        if file_path:
            self.template_file.set(file_path)
            self.template_status.config(text=f"Selected: {Path(file_path).name}", foreground="green")
            print(f"[FILE] Selected template: {Path(file_path).name}")
            # Save to config
            self.config_manager.set_config('LAST_USED_FILES', 'template_file', file_path)
            self.config_manager.save_config()

    def select_scan_file(self):
        """Select Nessus scan file with console logging."""
        print("[ACTION] User selecting scan file...")
        
        # Get initial directory from last used path
        initial_dir = None
        last_path = self.config_manager.get_config('LAST_USED_FILES', 'scan_file')
        if last_path and os.path.exists(os.path.dirname(last_path)):
            initial_dir = os.path.dirname(last_path)
        
        file_path = filedialog.askopenfilename(
            title="Select Nessus Scan File",
            initialdir=initial_dir,
            filetypes=[
                ("Zip files", "*.zip"),
                ("Nessus files", "*.nessus"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.scan_file.set(file_path)
            self.scan_status.config(text=f"Selected: {Path(file_path).name}", foreground="green")
            print(f"[FILE] Selected scan file: {Path(file_path).name}")
            # Save to config
            self.config_manager.set_config('LAST_USED_FILES', 'scan_file', file_path)
            self.config_manager.save_config()

    def select_reference_file(self):
        """Select OPDIR reference file (Excel or CSV) with console logging."""
        print("[ACTION] User selecting OPDIR reference file...")
        try:
            import openpyxl  # pip install openpyxl
            xlsx_available = True
        except ImportError:
            xlsx_available = False
        
        # Get initial directory from last used path
        initial_dir = None
        last_path = self.config_manager.get_config('LAST_USED_FILES', 'reference_file')
        if last_path and os.path.exists(os.path.dirname(last_path)):
            initial_dir = os.path.dirname(last_path)
        
        if xlsx_available:
            file_types = [
                ("Excel files", "*.xlsx;*.xls"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ]
        else:
            file_types = [
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ]
            messagebox.showwarning("Limited Support", 
                "openpyxl not installed. Only CSV files are supported.\n\nInstall Excel support with:\npip install openpyxl --break-system-packages")
        
        file_path = filedialog.askopenfilename(
            title="Select OPDIR Reference File",
            initialdir=initial_dir,
            filetypes=file_types
        )
        
        if file_path:
            # Try to load the reference file
            if opdir_lookup_manager.load_reference_file(file_path):
                self.reference_file.set(file_path)
                stats = opdir_lookup_manager.get_stats()
                self.reference_status.config(
                    text=f"Loaded: {Path(file_path).name} ({stats['mapping_count']} mappings)", 
                    foreground="green"
                )
                print(f"[FILE] Loaded reference file: {Path(file_path).name}")
                print(f"[INFO] OPDIR mappings loaded: {stats['mapping_count']}")
                # Save to config
                self.config_manager.set_config('LAST_USED_FILES', 'reference_file', file_path)
                self.config_manager.save_config()
            else:
                self.reference_status.config(text="Failed to load", foreground="red")
                print("[ERROR] Failed to load reference file")

    def select_output_dir(self):
        """Select output directory with console logging."""
        print("[ACTION] User selecting output directory...")
        
        # Get initial directory from last used path
        initial_dir = self.output_dir.get()
        if not initial_dir or not os.path.exists(initial_dir):
            initial_dir = self.config_manager.get_config('LAST_USED_FILES', 'output_directory')
        
        dir_path = filedialog.askdirectory(
            title="Select Output Directory",
            initialdir=initial_dir if initial_dir and os.path.exists(initial_dir) else None
        )
        
        if dir_path:
            self.output_dir.set(dir_path)
            # Save to both locations for compatibility
            self.config_manager.set_config('TECHNICAL', 'default_output_dir', dir_path)
            self.config_manager.set_config('LAST_USED_FILES', 'output_directory', dir_path)
            self.config_manager.save_config()
            print(f"[FILE] Selected output directory: {dir_path}")

    def parse_iavms_from_text(self, text):
        """
        Parse IAVM numbers from pasted text.
        Handles multiple formats, removes duplicates, supports comma-separated and multi-line.
        
        Examples:
            'IAVA #2024-A-1234' -> '2024-A-1234'
            'IAVA #2024-A-1234,IAVA #2024-A-5678' -> ['2024-A-1234', '2024-A-5678']
            Multiple on one line, duplicates, etc.
        """
        import re
        
        # Pattern to match IAVM numbers: YYYY-A-NNNN or YYYY-B-NNNN
        # May be preceded by "IAVA #" or "IAVB #" or just standalone
        pattern = r'(?:IAVA?\s*#?\s*)?(\d{4}-[AB]-\d{4})'
        
        # Find all matches
        matches = re.findall(pattern, text, re.IGNORECASE)
        
        # Remove duplicates while preserving order (using dict)
        unique_iavms = list(dict.fromkeys(matches))
        
        # Standardize format to YYYY-A-NNNN
        standardized = []
        for iavm in unique_iavms:
            # Ensure uppercase A or B
            parts = iavm.split('-')
            if len(parts) == 3:
                standardized.append(f"{parts[0]}-{parts[1].upper()}-{parts[2]}")
        
        return standardized
    
    def normalize_iavm_for_matching(self, iavm_string):
        """
        Normalize IAVM string to core format for matching.
        
        Handles formats like:
        - 'IAVA #2024-A-1234' -> '2024-A-1234'
        - 'IAVB:2023-B-0029-S' -> '2023-B-0029'
        - 'IAVB:0001-B-0521' -> '0001-B-0521'
        - '2024-A-1234' -> '2024-A-1234'
        
        Returns: Normalized IAVM number (YYYY-X-NNNN format)
        """
        import re
        
        # Remove common prefixes and extract core IAVM
        # Pattern matches: optional prefix + YYYY-X-NNNN + optional suffix
        pattern = r'(?:IAVA?[:\s#]*)?(\d{4}-[AB]-\d{4})(?:-[A-Z])?'
        
        match = re.search(pattern, iavm_string, re.IGNORECASE)
        if match:
            core = match.group(1)
            # Ensure uppercase A or B
            parts = core.split('-')
            if len(parts) == 3:
                return f"{parts[0]}-{parts[1].upper()}-{parts[2]}"
        
        return None
    
    def iavms_match(self, filter_iavm, scan_iavm):
        """
        Check if two IAVM strings match, handling format differences.
        
        Examples:
        - '2024-A-1234' matches 'IAVA #2024-A-1234'
        - '2023-B-0029' matches 'IAVB:2023-B-0029-S'
        - '0001-B-0521' matches 'IAVB:0001-B-0521'
        """
        norm_filter = self.normalize_iavm_for_matching(filter_iavm)
        norm_scan = self.normalize_iavm_for_matching(scan_iavm)
        
        if norm_filter and norm_scan:
            return norm_filter == norm_scan
        
        return False
    
    def apply_iavm_filter(self):
        """Apply IAVM filter from text input."""
        print("[ACTION] Applying IAVM filter...")
        
        # Get text from input
        text = self.iavm_filter_text.get("1.0", tk.END).strip()
        
        if not text:
            messagebox.showwarning("No Input", "Please paste IAVM numbers in the text area.")
            return
        
        # Parse IAVMs
        iavms = self.parse_iavms_from_text(text)
        
        if not iavms:
            messagebox.showwarning("No IAVMs Found", 
                                 "Could not find any valid IAVM numbers in the pasted text.\n\n" +
                                 "Expected format: YYYY-A-NNNN or YYYY-B-NNNN\n" +
                                 "Example: 2024-A-1234")
            return
        
        # Store filtered list
        self.filtered_iavms = set(iavms)
        
        # Update status
        status_text = f"Filter active: {len(iavms)} unique IAVM(s)"
        self.filter_status.config(text=status_text, foreground="green")
        
        print(f"[FILTER] Applied IAVM filter: {len(iavms)} unique IAVMs")
        print(f"[FILTER] IAVMs: {', '.join(sorted(iavms))}")
        
        messagebox.showinfo("Filter Applied", 
                          f"Filter applied successfully!\n\n" +
                          f"Will generate POA&Ms for {len(iavms)} IAVM(s):\n" +
                          ', '.join(sorted(iavms)[:10]) + 
                          (f"\n... and {len(iavms)-10} more" if len(iavms) > 10 else ""))
    
    def clear_iavm_filter(self):
        """Clear IAVM filter."""
        print("[ACTION] Clearing IAVM filter...")
        
        self.filtered_iavms = None
        self.iavm_filter_text.delete("1.0", tk.END)
        self.filter_status.config(text="No filter active", foreground="gray")
        
        print("[FILTER] IAVM filter cleared - will generate all IAVMs")
        messagebox.showinfo("Filter Cleared", "IAVM filter cleared. Will generate POA&Ms for all IAVMs in scan.")

    def update_status(self, message):
        """Update status label from worker thread with console logging."""
        def update():
            self.status_label.config(text=message)
            print(f"[STATUS] {message}")
        self.root.after(0, update)

    def update_progress(self, value):
        """Update progress bar from worker thread."""
        def update():
            self.progress_bar.config(value=value)
        self.root.after(0, update)

    def generate_poams(self):
        """Generate POA&M documents with enhanced validation and console logging."""
        print("[ACTION] Starting POA&M generation process...")
        
        if not self.template_file.get():
            print("[VALIDATION] Missing template file")
            messagebox.showerror("Missing Template", "Please select a POA&M template file.")
            return
        
        if not self.scan_file.get():
            print("[VALIDATION] Missing scan file")
            messagebox.showerror("Missing Scan File", "Please select a Nessus scan file.")
            return
        
        if not self.output_dir.get():
            print("[VALIDATION] Missing output directory")
            messagebox.showerror("Missing Output Directory", "Please select an output directory.")
            return
        
        if not PDF_AVAILABLE:
            print("[VALIDATION] PDF libraries not available")
            messagebox.showerror("PDF Libraries Missing", 
                            "Required PDF libraries not installed.\n\nInstall with:\npip install reportlab PyMuPDF")
            return
        
        print("[VALIDATION] All required inputs validated successfully")
        
        self.config_manager.set_bool('TECHNICAL', 'use_ip_ranges', self.ip_format_var.get() == "IP Ranges")
        self.config_manager.set_config('TECHNICAL', 'subnet_mask', self.subnet_var.get())
        self.config_manager.save_config()
        
        print("[ACTION] Configuration saved, starting generation process")

        # Add this after the config saving in generate_poams():
        self.config_manager.set_config('TECHNICAL', 'opdir_release', self.opdir_var.get())
        
        self.generate_button.config(state="disabled")
        self.status_label.config(text="Processing scan files...")
        
        generation_thread = threading.Thread(target=self._generate_poams_worker)
        generation_thread.daemon = True
        generation_thread.start()

    def _generate_poams_worker(self):
        """Worker thread to generate POA&M documents."""
        try:
            start_time = time.time()
            
            scan_file = self.scan_file.get()
            plugins_file = self.plugins_file.get() if self.plugins_file.get() else None
            
            # FIX: Get OPDIR release from the UI
            opdir_release = self.opdir_var.get()
            
            print(f"[FILE] Scan file: {scan_file}")
            print(f"[PLUGIN] Plugins file: {plugins_file or 'Auto-detect'}")
            print(f"[CONFIG] OPDIR Release: {opdir_release}")
            
            self.update_status("Loading plugins database...")
            self.update_progress(10)
            
            plugins_dict = None
            if plugins_file:
                print(f"[ACTION] Loading plugins database from {plugins_file}")
                plugins_dict = load_plugins_database(plugins_file)
                if plugins_dict:
                    print(f"[SUCCESS] Loaded {len(plugins_dict)} plugins from database")
                    self.update_status(f"Loaded {len(plugins_dict)} plugins from database")
                else:
                    print("[WARNING] Could not load plugins database")
                    self.update_status("Warning: Could not load plugins database")
            else:
                print("[ACTION] No plugins file specified, attempting auto-detection...")
                plugins_dict = load_plugins_database()
                if plugins_dict:
                    print(f"[SUCCESS] Auto-loaded {len(plugins_dict)} plugins")
                    self.update_status(f"Auto-loaded {len(plugins_dict)} plugins")
                else:
                    print("[WARNING] No plugins database available")
                    self.update_status("Warning: No plugins database available")
            
            self.update_progress(20)
            self.update_status("Extracting scan files...")
            print("[ACTION] Starting scan file extraction...")
            
            temp_dir = tempfile.mkdtemp()
            print(f"[FILE] Created temporary directory: {temp_dir}")
            
            try:
                if scan_file.endswith('.zip'):
                    print("[ACTION] Extracting zip file...")
                    extract_nested_zips(scan_file, temp_dir)
                    nessus_files = find_nessus_files(temp_dir)
                elif scan_file.endswith('.nessus'):
                    print("[ACTION] Processing single nessus file...")
                    nessus_files = [scan_file]
                else:
                    raise ValueError(f"Unsupported file format: {scan_file}")
                
                if not nessus_files:
                    raise ValueError("No .nessus files found in the selected file")
                
                print(f"[SUCCESS] Found {len(nessus_files)} .nessus files")
                self.update_status(f"Found {len(nessus_files)} .nessus files")
                self.update_progress(30)
                
                all_iavm_findings = defaultdict(lambda: {
                    'hosts': {},
                    'plugins': set(),
                    'descriptions': set(),
                    'solutions': set(),
                    'cves': set(),
                    'vulnerable_software': {},
                    'report_name': '',
                    'scan_file_ref': ''
                })
                
                zip_contexts = {}
                
                for nessus_file in nessus_files:
                    zip_context = None
                    if scan_file.endswith('.zip'):
                        rel_path = os.path.relpath(nessus_file, temp_dir)
                        path_parts = rel_path.split(os.sep)
                        
                        for part in path_parts:
                            if part.endswith('.zip') or (len(path_parts) > 1 and not part.endswith('.nessus')):
                                zip_context = part.replace('.zip', '') if part.endswith('.zip') else part
                                break
                    
                    zip_contexts[nessus_file] = zip_context
                
                self.update_status("Parsing Nessus files for IAVM findings...")
                print("[ACTION] Starting IAVM parsing from Nessus files...")
                
                for i, nessus_file in enumerate(nessus_files, 1):
                    print(f"[PROGRESS] Processing file {i}/{len(nessus_files)}: {os.path.basename(nessus_file)}")
                    
                    progress = 30 + (20 * i / len(nessus_files))
                    self.update_progress(progress)
                    self.update_status(f"Processing {os.path.basename(nessus_file)} ({i}/{len(nessus_files)})")
                    
                    zip_context = zip_contexts.get(nessus_file)
                    iavm_findings = parse_nessus_for_iavm_enhanced(nessus_file, zip_context, plugins_dict)
                    
                    for iavm_id, data in iavm_findings.items():
                        all_iavm_findings[iavm_id]['hosts'].update(data['hosts'])
                        all_iavm_findings[iavm_id]['plugins'].update(data['plugins'])
                        all_iavm_findings[iavm_id]['descriptions'].update(data['descriptions'])
                        all_iavm_findings[iavm_id]['solutions'].update(data['solutions'])
                        all_iavm_findings[iavm_id]['cves'].update(data['cves'])
                        
                        # Merge vulnerable software
                        for host_ip, host_vuln_software in data['vulnerable_software'].items():
                            if host_ip not in all_iavm_findings[iavm_id]['vulnerable_software']:
                                all_iavm_findings[iavm_id]['vulnerable_software'][host_ip] = {}
                            all_iavm_findings[iavm_id]['vulnerable_software'][host_ip].update(host_vuln_software)
                        
                        if not all_iavm_findings[iavm_id]['report_name']:
                            all_iavm_findings[iavm_id]['report_name'] = data['report_name']
                        if not all_iavm_findings[iavm_id]['scan_file_ref']:
                            all_iavm_findings[iavm_id]['scan_file_ref'] = data['scan_file_ref']
                
                print(f"[SUCCESS] Found {len(all_iavm_findings)} unique IAVM references")
                self.update_status(f"Found {len(all_iavm_findings)} unique IAVM references")
                self.update_progress(50)
                
                if not all_iavm_findings:
                    self.update_status("No IAVM findings found")
                    raise ValueError("No IAVM findings found in the scan files")
                                
                print(f"[STATUS] IAVM Summary:")
                for iavm_id, data in all_iavm_findings.items():
                    host_count = len(data['hosts'])
                    sample_hosts = list(data['hosts'].keys())[:3]
                    hosts_preview = ", ".join(sample_hosts)
                    if host_count > 3:
                        hosts_preview += f"... (+{host_count-3} more)"
                    print(f"  {iavm_id}: {host_count} hosts affected ({hosts_preview})")
                
                def show_dialog():
                    try:
                        print("[UI] Creating IAVM selection dialog...")
                        iavm_selector = IAVMSelectionDialog(all_iavm_findings)
                        selected_iavms = iavm_selector.show(parent=self.root)
                        
                        print(f"[UI] Dialog completed successfully")
                        print(f"[UI] Final selection: {len(selected_iavms)} IAVMs")
                        
                        if len(selected_iavms) <= 10:
                            print(f"[UI] Selected IAVMs: {selected_iavms}")
                        else:
                            print(f"[UI] Selected IAVMs (first 10): {selected_iavms[:10]}...")
                        
                        return selected_iavms
                            
                    except Exception as e:
                        print(f"[ERROR] Dialog error: {e}")
                        import traceback
                        traceback.print_exc()
                        return []
                
                selected_iavms = []
                dialog_exception = [None]
                
                # Check if IAVM filter is active
                if self.filtered_iavms:
                    print(f"[FILTER] IAVM filter active with {len(self.filtered_iavms)} IAVMs")
                    
                    # Filter all_iavm_findings to only include filtered IAVMs
                    filtered_findings = {}
                    not_found = []
                    
                    for filter_iavm in self.filtered_iavms:
                        # Try to find match using flexible matching
                        found = False
                        for iavm_id in all_iavm_findings.keys():
                            if self.iavms_match(filter_iavm, iavm_id):
                                filtered_findings[iavm_id] = all_iavm_findings[iavm_id]
                                selected_iavms.append(iavm_id)
                                print(f"[FILTER] Matched filter '{filter_iavm}' to scan '{iavm_id}'")
                                found = True
                                break
                        
                        if not found:
                            not_found.append(filter_iavm)
                    
                    if not_found:
                        print(f"[FILTER] Warning: {len(not_found)} IAVMs not found in scan:")
                        for iavm in not_found[:10]:
                            print(f"  - {iavm}")
                        if len(not_found) > 10:
                            print(f"  ... and {len(not_found)-10} more")
                    
                    if not selected_iavms:
                        error_msg = f"None of the filtered IAVMs were found in the scan.\n\n"
                        error_msg += f"Filtered IAVMs ({len(self.filtered_iavms)}): {', '.join(list(self.filtered_iavms)[:5])}\n\n"
                        error_msg += f"Scan IAVMs ({len(all_iavm_findings)}): {', '.join(list(all_iavm_findings.keys())[:5])}\n\n"
                        error_msg += "Note: Scan IAVMs may have different formats (e.g., 'IAVB:2023-B-0029-S')\n"
                        error_msg += "Try checking the console output for exact scan IAVM formats."
                        
                        def show_error():
                            messagebox.showerror("No Matching IAVMs", error_msg)
                        
                        self.root.after(0, show_error)
                        raise ValueError("No matching IAVMs found between filter and scan")
                    
                    print(f"[FILTER] Matched {len(selected_iavms)}/{len(self.filtered_iavms)} filtered IAVMs")
                    print(f"[FILTER] Will generate: {', '.join(sorted(selected_iavms))}")
                    
                else:
                    # No filter active - show dialog as before
                    print("[ACTION] No IAVM filter active, showing selection dialog...")
                    
                    def dialog_wrapper():
                        try:
                            selected_iavms[:] = show_dialog()
                        except Exception as e:
                            dialog_exception[0] = e
                    
                    self.root.after(0, dialog_wrapper)
                    
                    import threading
                    while len(selected_iavms) == 0 and dialog_exception[0] is None:
                        time.sleep(0.01)
                        self.root.update()
                    
                    if dialog_exception[0]:
                        raise dialog_exception[0]
                
                if not selected_iavms:
                    self.update_status("No IAVMs selected")
                    print("[ACTION] No IAVMs selected for generation")
                    raise ValueError("No IAVMs selected for generation")

                print(f"[ACTION] Proceeding with {len(selected_iavms)} selected IAVMs")
                self.update_status(f"Generating {len(selected_iavms)} POA&M documents...")
                
                template_file = self.template_file.get()
                output_dir = self.output_dir.get()
                
                os.makedirs(output_dir, exist_ok=True)
                print(f"[FILE] Ensured output directory exists: {output_dir}")
                
                successful_count = 0
                failed_count = 0
                ai_enhanced_count = 0
                
                generation_start_time = time.time()
                print(f"[GENERATION] Starting POA&M generation for {len(selected_iavms)} IAVMs...")
                
                for i, iavm_id in enumerate(selected_iavms, 1):
                    iavm_data = all_iavm_findings[iavm_id]
                    print(f"[GENERATION] Generating POA&M {i}/{len(selected_iavms)}: {iavm_id} ({len(iavm_data['hosts'])} hosts affected)")
                    
                    progress = 50 + (40 * i / len(selected_iavms))
                    self.update_progress(progress)
                    self.update_status(f"Generating {iavm_id} ({i}/{len(selected_iavms)})")
                    
                    enhanced_progress_reporting(i, len(selected_iavms), "POA&M Generation", generation_start_time)
                    
                    clean_iavm_id = re.sub(r'[^\w\-_]', '_', iavm_id)
                    # Look up OPDIR number for this specific IAVM
                    iavm_opdir = get_opdir_release(self.config_manager, iavm_id)
                    output_filename = f"OPDIR_{iavm_opdir}_{clean_iavm_id}.pdf"
                    iavm_opdir = get_opdir_release(self.config_manager, iavm_id)
                    output_filename = f"OPDIR_{iavm_opdir}_{clean_iavm_id}.pdf"
                    output_path = os.path.join(output_dir, output_filename)
                    
                    print(f"[FILE] Creating: {output_filename}")
                    
                    try:
                        result, ai_enhanced = safe_file_operation(
                            fill_pdf_form, 
                            template_file, output_path, iavm_id, iavm_data, 
                            self.config_manager, self.system_descriptions_manager
                        )
                        
                        if result:
                            successful_count += 1
                            if ai_enhanced:
                                ai_enhanced_count += 1
                            print(f"[SUCCESS] Created: {output_filename}")
                            
                            iavm_tracker.log_generation(
                                iavm_id=iavm_id,
                                hosts_affected=len(iavm_data['hosts']),
                                scan_file=os.path.basename(scan_file),
                                output_file=output_filename,
                                status='success',
                                ai_enhanced=ai_enhanced,
                                plugins_enriched=plugins_dict is not None,
                                vulnerable_software_count=sum(len(host_software) for host_software in iavm_data.get('vulnerable_software', {}).values())
                            )
                        else:
                            failed_count += 1
                            print(f"[ERROR] Failed: {output_filename}")
                            
                            iavm_tracker.log_generation(
                                iavm_id=iavm_id,
                                hosts_affected=len(iavm_data['hosts']),
                                scan_file=os.path.basename(scan_file),
                                output_file=output_filename,
                                status='failed',
                                ai_enhanced=False,
                                plugins_enriched=plugins_dict is not None,
                                vulnerable_software_count=0
                            )
                            
                    except Exception as e:
                        failed_count += 1
                        print(f"[ERROR] Failed: {output_filename} - {str(e)}")
                        
                        iavm_tracker.log_generation(
                            iavm_id=iavm_id,
                            hosts_affected=len(iavm_data['hosts']),
                            scan_file=os.path.basename(scan_file),
                            output_file=output_filename,
                            status=f'error: {str(e)[:50]}',
                            ai_enhanced=False,
                            plugins_enriched=plugins_dict is not None,
                            vulnerable_software_count=0
                        )
                
                self.update_progress(100)
                total_time = time.time() - start_time
                
                print("=" * 60)
                print("[COMPLETE] Processing complete!")
                print(f"[STATS] Total processing time: {total_time:.1f} seconds")
                print(f"[STATS] Successfully generated: {successful_count} POA&M forms")
                if failed_count > 0:
                    print(f"[STATS] Failed to generate: {failed_count} POA&M forms")
                
                if plugins_dict:
                    print(f"[PLUGIN] Plugin database enrichment enabled - complete IAVM coverage")
                else:
                    print(f"[PLUGIN] Plugin database not loaded - may have missed some IAVM references")
                
                if ai_enhanced_count > 0:
                    print(f"[AI] AI enhanced {ai_enhanced_count} POA&M forms with contextualized narratives")
                elif self.config_manager.get_bool('AI_SETTINGS', 'ollama_enabled'):
                    print(f"[AI] AI enhancement enabled but no fields selected for enhancement")
                
                print(f"[FILE] Output directory: {output_dir}")
                print("=" * 60)
                
                def show_completion():
                    message = f"POA&M generation complete!\n\n"
                    message += f"Successfully generated: {successful_count} forms\n"
                    if failed_count > 0:
                        message += f"Failed to generate: {failed_count} forms\n"
                    if ai_enhanced_count > 0:
                        message += f"AI enhanced: {ai_enhanced_count} forms\n"
                    if plugins_dict:
                        message += f"Plugin database: Complete IAVM coverage\n"
                    else:
                        message += f"Plugin database: Not loaded\n"
                    message += f"Processing time: {total_time:.1f} seconds\n"
                    message += f"\nOutput directory:\n{output_dir}"
                    
                    messagebox.showinfo("Processing Complete", message)
                    
                    if successful_count > 0:
                        try:
                            if os.name == 'nt':
                                os.startfile(output_dir)
                            elif os.name == 'posix':
                                subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', output_dir])
                        except Exception as e:
                            print(f"[WARNING] Could not open output directory: {e}")
                    
                    self.generate_button.config(state="normal")
                    self.progress_bar.config(value=0)
                    self.status_label.config(text="Ready to generate POA&Ms")
                
                self.root.after(0, show_completion)
                
            finally:
                print("[CLEANUP] Cleaning up temporary files...")
                shutil.rmtree(temp_dir, ignore_errors=True)
        
        except Exception as error_obj:
            print(f"[ERROR] Error during POA&M generation: {error_obj}")
            
            # Capture error in local variable for closure
            error_message = f"Error during generation:\n\n{str(error_obj)}"
            
            def show_error(err_msg=error_message):
                try:
                    messagebox.showerror("Generation Error", err_msg)
                except Exception as display_error:
                    print(f"Error displaying error dialog: {display_error}")
                    print(f"Original error message: {err_msg}")
                finally:
                    self.generate_button.config(state="normal")
                    self.progress_bar.config(value=0)
                    self.status_label.config(text="Ready to generate POA&Ms")
            
            self.root.after(0, show_error)
            
    def show_help(self):
        """Show enhanced help information."""
        help_text = """POA&M Generator - Enhanced Edition v2.0

NEW FEATURES IN v2.0:
• IAVM tracking and generation history
• Fixed select all/deselect all buttons in IAVM selection
• Enhanced console logging during generation process
• Improved modal dialog sizing and autofit
• System descriptions configuration by hostname patterns
• Full Ollama AI integration for contextual analysis
• Dark theme throughout application

SYSTEM DESCRIPTIONS:
Configure system descriptions based on hostname patterns:
• Use wildcards like WEB* to match WEB01, WEB02, etc.
• Descriptions are included in POA&M asset descriptions
• AI enhancement uses these descriptions for better context
• Access via "System Descriptions" button in Configuration section

AI ENHANCEMENT:
• Configure Ollama models for narrative enhancement
• AI analyzes system context and vulnerability data
• Generates contextual content for operational impact, plans, etc.
• Enable/disable enhancement per field type
• Access via "AI Settings" button in Configuration section

GENERATION TRACKING:
• All generated POA&Ms are logged with timestamps
• Track which IAVMs were processed and when
• View generation history with success/failure status
• AI enhancement status is tracked per generation
• Access via "Generation History" button

GETTING STARTED:
1. Configure organizational information using config buttons
2. Set up system descriptions for your environment
3. Configure AI settings if Ollama is available
4. Load plugins database (CM-*-sc-plugins.tar.gz recommended)
5. Select POA&M template and scan files
6. Generate POA&Ms with enhanced tracking

TROUBLESHOOTING:
• Check console logs for detailed processing information
• Use Generation History to see past successes/failures
• Ensure all required files are selected before generation
• For AI issues, verify Ollama installation and model availability

REQUIRED LIBRARIES:
• reportlab: pip install reportlab
• PyMuPDF: pip install PyMuPDF
• ollama: pip install ollama (optional, for AI enhancement)

For support, check the generation logs and history for troubleshooting."""
        
        help_window = tk.Toplevel(self.root)
        help_window.title("Help - Enhanced Features")
        help_window.geometry("800x700")
        help_window.resizable(True, True)
        help_window.configure(bg='#2b2b2b')
        
        frame = tk.Frame(help_window, bg='#2b2b2b')
        frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        text_widget = tk.Text(frame, wrap=tk.WORD, padx=10, pady=10,
                             bg='#2b2b2b', fg='white', insertbackground='white')
        scrollbar = tk.Scrollbar(frame, orient="vertical", command=text_widget.yview)
        
        text_widget.config(yscrollcommand=scrollbar.set)
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        text_widget.insert(tk.END, help_text)
        text_widget.config(state=tk.DISABLED)
    
    def show_about(self):
        """Show enhanced about information."""
        about_text = """POA&M Generator - Enhanced Edition v2.0

NEW FEATURES IN v2.0:
• IAVM generation tracking and history (CSV-based)
• Fixed UI issues (select all/none buttons now work)
• Enhanced console logging during processing
• Improved modal dialog sizing and responsiveness
• System descriptions by hostname patterns
• Full Ollama AI integration for contextual analysis
• Dark theme throughout application
• Comprehensive error handling and retry logic

CORE FUNCTIONALITY PRESERVED:
• Parse Nessus scan files for IAVM references
• Generate individual POA&M documents per finding
• AI-enhanced narrative generation with system context
• Automatic plugin database enrichment
• Configurable organizational templates
• Multi-threaded processing with progress tracking
• All original 1700+ lines of functionality retained

TRACKING & LOGGING:
• CSV-based generation history tracking
• Detailed console logging for troubleshooting
• Success/failure status for each IAVM
• AI enhancement tracking per generation
• Plugin database enrichment status

AI ENHANCEMENT:
• Ollama integration for contextual narrative generation
• System-aware impact assessments
• Contextualized remediation plans
• Field-specific AI enhancement controls
• Uses system descriptions for better context

TECHNICAL REQUIREMENTS:
• Python 3.7+
• reportlab, PyMuPDF (required for PDF generation)
• ollama (optional for AI features)

ARCHITECTURE:
• Complete restoration of all original functionality
• Enhanced with new tracking and AI capabilities
• Dark theme UI with proper modal sizing
• Thread-safe processing with progress tracking
• INI-based configuration management

Built with enhanced UI, comprehensive logging, and full functionality preservation.
All original features + new enhancements = Complete solution."""
        
        messagebox.showinfo("About - Enhanced Edition v2.0", about_text)
    
    def run(self):
        """Run the enhanced application with console logging."""
        print("[STARTUP] Starting Enhanced POA&M Generator application")
        self.root.mainloop()
        print("[EXIT] Application closed")

def verify_missing_functions_check():
    """Verification function to ensure all critical functions are included."""
    required_functions = [
        'extract_feed_timestamp',
        'extract_plugins_xml_from_archive', 
        'parse_plugins_xml',
        'load_plugins_database_json',
        'load_plugins_database',
        'extract_nested_zips',
        'find_nessus_files', 
        'extract_hostname_from_plugins',
        'extract_os_information',
        'extract_iavm_specific_software_versions_enhanced',
        'extract_versions_from_text',
        'extract_software_name_from_plugin',
        'format_iavm_software_versions',
        'parse_nessus_for_iavm_enhanced',
        'calculate_ip_ranges',
        'get_asset_description',
        'fill_pdf_form',
        'safe_file_operation',
        'enhanced_progress_reporting'
    ]
    
    current_module = sys.modules[__name__]
    missing_functions = []
    
    for func_name in required_functions:
        if not hasattr(current_module, func_name):
            missing_functions.append(func_name)
    
    if missing_functions:
        logger.error(f"Missing critical functions: {missing_functions}")
        return False
    else:
        logger.info("All critical functions are included")
        return True

def print_field_mapping_analysis():
    """Print comprehensive field mapping analysis."""
    print("\n" + "=" * 80)
    print("PDF FIELD MAPPING ANALYSIS")
    print("=" * 80)
    
    predefined_fields = []
    auto_generated_fields = []
    manual_fields = []
    hardcoded_fields = []
    
    for field_name, info in PDF_FIELD_MAPPING.items():
        if info["predefined"]:
            predefined_fields.append(field_name)
        elif "Auto-generated" in info["type"]:
            auto_generated_fields.append(field_name)
        elif "Manual" in info["type"]:
            manual_fields.append(field_name)
        elif "Hardcoded" in info["type"]:
            hardcoded_fields.append(field_name)
    
    print(f"\nCONFIG PREDEFINED CONFIGURABLE FIELDS ({len(predefined_fields)}):")
    for field in predefined_fields:
        print(f"  * {field}")
    
    print(f"\nAUTO AUTO-GENERATED FROM SCAN DATA ({len(auto_generated_fields)}):")
    for field in auto_generated_fields:
        print(f"  * {field}")
    
    print(f"\nHARDCODE HARDCODED VALUES ({len(hardcoded_fields)}):")
    for field in hardcoded_fields:
        print(f"  * {field}")
    
    print(f"\nMANUAL MANUAL ENTRY REQUIRED ({len(manual_fields)}):")
    for field in manual_fields:
        print(f"  * {field}")
    
    print(f"\nTOTAL FIELDS MAPPED: {len(PDF_FIELD_MAPPING)}")
    print("=" * 80)

def main():
    """Enhanced main function with comprehensive feature set."""
    print("=" * 60)
    print("POA&M Generator - Enhanced Edition v2.0")
    print("=" * 60)
    print("NEW FEATURES:")
    print("  SUCCESS IAVM tracking and generation history")
    print("  SUCCESS Fixed select all/deselect all buttons")
    print("  SUCCESS Enhanced logging during generation")
    print("  SUCCESS Improved modal dialog sizing with autofit")
    print("  SUCCESS System descriptions by hostname patterns")
    print("  SUCCESS Full Ollama AI integration")
    print("  SUCCESS Dark theme throughout")
    print("  SUCCESS Filter functionality in IAVM selection")
    print("  SUCCESS Generation status display (new vs previously generated)")
    print("  SUCCESS Console logging for all actions")
    print("  SUCCESS Generate All button functionality")
    print("  SUCCESS ALL ORIGINAL FUNCTIONALITY PRESERVED")
    print("=" * 60)
    
    if OLLAMA_AVAILABLE:
        print("AI Ollama available for AI-enhanced narratives")
    else:
        print("WARNING Install 'pip install ollama' for AI-enhanced narratives")
    
    if PDF_AVAILABLE:
        print("DOCUMENT PDF libraries available")
    else:
        print("ERROR Install: pip install reportlab PyMuPDF")
        return
    
    print("=" * 60)
    log_startup("Starting Enhanced POA&M Generator v2.0")
    
    if not verify_missing_functions_check():
        print("ERROR Critical functions missing! Check logs for details.")
        return
    
    print_field_mapping_analysis()
    
    try:
        app = MainApplication()
        app.run()
    except Exception as e:
        log_error(f"Application error: {e}")
        print(f"ERROR Application error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()